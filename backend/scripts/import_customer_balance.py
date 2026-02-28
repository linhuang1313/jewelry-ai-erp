# -*- coding: utf-8 -*-
"""
导入客户期初余额数据（从旧系统 Excel "客户实时账目查询.xlsx" 迁移）

用法:
  python backend/scripts/import_customer_balance.py                  # dry-run 模式（默认）
  python backend/scripts/import_customer_balance.py --execute        # 正式执行
  python backend/scripts/import_customer_balance.py --verify         # 仅验证已导入数据
  python backend/scripts/import_customer_balance.py --db-url "postgresql://..."  # 指定数据库

Excel 结构（6 列）:
  序号 | 客户名称 | 结算点 | 欠料重量 | 欠款金额 | 业务

同一客户可能占 1-2 行（欠料和欠款分行记录），脚本会自动合并。
"""
import sys
import os
import re
import math
import uuid
from datetime import datetime, timezone, timedelta

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import sessionmaker

from app.models import (
    Customer, Salesperson,
    CustomerGoldDeposit, CustomerGoldDepositTransaction,
    CustomerTransaction,
)

CHINA_TZ = timezone(timedelta(hours=8))
REMOTE_URL = "postgresql://postgres:weiwenhao520@47.76.120.172:5432/railway"


def china_now():
    return datetime.now(CHINA_TZ)


def safe_float(val):
    try:
        f = float(val) if val is not None else 0.0
        return 0.0 if math.isnan(f) or math.isinf(f) else f
    except (ValueError, TypeError):
        return 0.0


def clean_name(raw):
    """清洗客户/业务员名称：strip + 压缩连续空格 + 统一括号为全角"""
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace('(', '\uff08').replace(')', '\uff09')
    return s


def find_excel_file():
    """在桌面自动查找 '客户实时账目查询.xlsx'"""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    for entry in os.scandir(desktop):
        if entry.name.endswith('.xlsx') and not entry.name.startswith('~'):
            try:
                wb = openpyxl.load_workbook(entry.path, read_only=True, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                wb.close()
                if headers and '客户名称' in str(headers):
                    return entry.path
            except Exception:
                continue
    return None


def read_excel(path):
    """读取 Excel 并合并同一客户的多行数据，返回 {cleaned_name: {weight, amount, sales}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    customers = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        seq, name, point, weight, amount, sales = row
        if name is None:
            continue

        cleaned = clean_name(name)
        if not cleaned:
            continue

        if cleaned not in customers:
            customers[cleaned] = {
                'seq': seq,
                'weight': None,
                'amount': None,
                'sales': '',
                'raw_name': str(name).strip(),
            }

        if weight is not None:
            customers[cleaned]['weight'] = safe_float(weight)
        if amount is not None:
            customers[cleaned]['amount'] = safe_float(amount)
        if sales and str(sales).strip():
            customers[cleaned]['sales'] = clean_name(sales)

    wb.close()
    return customers


def generate_customer_no():
    """生成客户编号，与线上系统一致: KH + 年月日时分秒"""
    return f"KH{china_now().strftime('%Y%m%d%H%M%S')}"


def get_db_session(db_url=None):
    """创建数据库会话"""
    url = db_url or REMOTE_URL
    engine = create_engine(url, connect_args={"connect_timeout": 10})
    Session = sessionmaker(bind=engine)
    return Session()


def dry_run(customers):
    """仅输出将要导入的数据，不写入数据库"""
    print("=" * 80)
    print("DRY RUN - 以下数据将被导入（不会写入数据库）")
    print("=" * 80)

    salespersons = set()
    total_weight_pos = 0.0
    total_weight_neg = 0.0
    total_amount_pos = 0.0
    total_amount_neg = 0.0

    for i, (name, data) in enumerate(sorted(customers.items()), 1):
        w = data['weight']
        a = data['amount']
        s = data['sales']

        w_str = f"{w:>12.3f}g" if w is not None else f"{'':>13}"
        a_str = f"{a:>14.2f}元" if a is not None else f"{'':>15}"

        if w is not None:
            if w > 0:
                total_weight_pos += w
            else:
                total_weight_neg += w
        if a is not None:
            if a > 0:
                total_amount_pos += a
            else:
                total_amount_neg += a

        if s:
            salespersons.add(s)

        print(f"  {i:>3}. {name:<40} 欠料={w_str}  欠款={a_str}  业务={s}")

    print()
    print("-" * 80)
    print(f"客户总数: {len(customers)}")
    print(f"业务员: {len(salespersons)} 人 -> {sorted(salespersons)}")
    print(f"欠料汇总: 正值(存料)={total_weight_pos:,.3f}g  负值(欠料)={total_weight_neg:,.3f}g  净值={total_weight_pos + total_weight_neg:,.3f}g")
    print(f"欠款汇总: 正值(应收)={total_amount_pos:,.2f}元  负值(应付)={total_amount_neg:,.2f}元  净值={total_amount_pos + total_amount_neg:,.2f}元")
    print("-" * 80)

    # 数据清洗报告
    cleaned_count = 0
    for name, data in sorted(customers.items()):
        if name != data['raw_name']:
            if cleaned_count == 0:
                print("\n数据清洗记录:")
            cleaned_count += 1
            print(f"  [{data['raw_name']}] -> [{name}]")
    if cleaned_count:
        print(f"  共 {cleaned_count} 条名称被清洗")


def execute(customers, db_url=None):
    """正式执行导入"""
    db = get_db_session(db_url)
    try:
        stats = {
            'customers_created': 0,
            'customers_matched': 0,
            'customers_duplicate_warn': 0,
            'salespersons_created': 0,
            'deposits_created': 0,
            'deposits_updated': 0,
            'transactions_created': 0,
        }

        # Step 1: 创建/匹配业务员
        print("\n[Step 1] 创建/匹配业务员...")
        all_sales = set(d['sales'] for d in customers.values() if d['sales'])
        for sp_name in sorted(all_sales):
            existing = db.query(Salesperson).filter(Salesperson.name == sp_name).first()
            if not existing:
                new_sp = Salesperson(name=sp_name, status="active")
                db.add(new_sp)
                stats['salespersons_created'] += 1
                print(f"  新建业务员: {sp_name}")
            else:
                print(f"  已存在业务员: {sp_name}")
        db.flush()

        # Step 2: 创建/匹配客户 + 初始化余额
        print("\n[Step 2] 创建/匹配客户并初始化余额...")
        import time
        seq_counter = 0

        for name, data in sorted(customers.items()):
            w = data['weight']
            a = data['amount']
            s = data['sales']

            # 按名称查找客户（只匹配 active 状态）
            matches = db.query(Customer).filter(
                Customer.name == name,
                Customer.status == 'active'
            ).all()

            if len(matches) == 0:
                # 新客户 - 生成编号并创建
                seq_counter += 1
                time.sleep(0.01)  # 确保时间戳不重复
                customer_no = generate_customer_no()
                # 追加序号防碰撞
                customer_no = f"{customer_no}{seq_counter:04d}"

                remark_parts = ["期初数据迁移"]
                if s:
                    remark_parts.append(f"业务员: {s}")
                customer = Customer(
                    customer_no=customer_no,
                    name=name,
                    customer_type="个人",
                    status="active",
                    remark="；".join(remark_parts),
                )
                db.add(customer)
                db.flush()
                stats['customers_created'] += 1
                print(f"  新建客户: {name} ({customer_no})")
            elif len(matches) == 1:
                customer = matches[0]
                stats['customers_matched'] += 1
                if s and not customer.remark:
                    customer.remark = f"业务员: {s}"
                print(f"  匹配已有客户: {name} (id={customer.id}, no={customer.customer_no})")
            else:
                # 多个同名客户 - 取第一个（id最小的），记录警告
                customer = sorted(matches, key=lambda c: c.id)[0]
                stats['customers_matched'] += 1
                stats['customers_duplicate_warn'] += 1
                print(f"  [警告] 同名客户 {len(matches)} 个: {name}, 使用 id={customer.id} (no={customer.customer_no})")

            # Step 3: 初始化金料余额
            if w is not None and w != 0:
                deposit = db.query(CustomerGoldDeposit).filter(
                    CustomerGoldDeposit.customer_id == customer.id
                ).first()

                if not deposit:
                    deposit = CustomerGoldDeposit(
                        customer_id=customer.id,
                        customer_name=name,
                        current_balance=round(w, 2),
                        total_deposited=0,
                        total_used=0,
                    )
                    db.add(deposit)
                    db.flush()
                    stats['deposits_created'] += 1
                else:
                    deposit.current_balance = round(w, 2)
                    deposit.customer_name = name
                    stats['deposits_updated'] += 1

                deposit_tx = CustomerGoldDepositTransaction(
                    customer_id=customer.id,
                    customer_name=name,
                    transaction_type='deposit' if w > 0 else 'use',
                    amount=round(abs(w), 2),
                    balance_before=0,
                    balance_after=round(w, 2),
                    status='active',
                    created_by='系统迁移',
                    remark=f"期初余额迁移（来自旧系统），原始值: {w:.3f}克",
                )
                db.add(deposit_tx)

                print(f"    金料余额: {w:+.3f}g -> 存入 {round(w, 2)}g (精度: 2位小数)")

            # Step 4: 初始化现金余额
            if a is not None and a != 0:
                existing_tx = db.query(CustomerTransaction).filter(
                    CustomerTransaction.customer_id == customer.id,
                    CustomerTransaction.transaction_type == 'initial_balance',
                ).first()

                if existing_tx:
                    print(f"    [跳过] 已有期初现金交易 (id={existing_tx.id}, amount={existing_tx.amount})")
                else:
                    tx = CustomerTransaction(
                        customer_id=customer.id,
                        customer_name=name,
                        transaction_type='initial_balance',
                        amount=round(a, 2),
                        gold_weight=0,
                        gold_due_before=0,
                        gold_due_after=0,
                        status='active',
                        remark=f"期初余额迁移（来自旧系统），原始值: {a:.2f}元",
                    )
                    db.add(tx)
                    stats['transactions_created'] += 1
                    print(f"    现金余额: {a:+,.2f}元")

        db.commit()
        print("\n" + "=" * 80)
        print("导入完成!")
        print(f"  客户: 新建 {stats['customers_created']}, 匹配已有 {stats['customers_matched']}")
        if stats['customers_duplicate_warn']:
            print(f"  [注意] {stats['customers_duplicate_warn']} 个客户存在同名，已使用最早创建的记录")
        print(f"  业务员: 新建 {stats['salespersons_created']}")
        print(f"  金料余额: 新建 {stats['deposits_created']}, 更新 {stats['deposits_updated']}")
        print(f"  现金交易: 新建 {stats['transactions_created']}")
        print("=" * 80)

    except Exception as e:
        db.rollback()
        print(f"\n导入失败，已回滚: {e}")
        raise
    finally:
        db.close()


def verify(customers, db_url=None):
    """验证数据库中的数据与 Excel 是否一致"""
    db = get_db_session(db_url)
    try:
        print("\n" + "=" * 80)
        print("验证导入结果")
        print("=" * 80)

        mismatches = []
        matched = 0

        for name, data in sorted(customers.items()):
            expected_w = data['weight']
            expected_a = data['amount']

            customer = db.query(Customer).filter(
                Customer.name == name,
                Customer.status == 'active'
            ).first()
            if not customer:
                mismatches.append((name, "客户不存在", None, None))
                continue

            # 验证金料余额（数据库精度为2位小数）
            if expected_w is not None and expected_w != 0:
                deposit = db.query(CustomerGoldDeposit).filter(
                    CustomerGoldDeposit.customer_id == customer.id
                ).first()
                expected_rounded = round(expected_w, 2)
                if not deposit:
                    mismatches.append((name, "金料余额记录不存在", f"期望={expected_rounded:.2f}", "实际=无"))
                elif abs(float(deposit.current_balance) - expected_rounded) > 0.01:
                    mismatches.append((name, "金料余额不匹配",
                                      f"期望={expected_rounded:.2f}", f"实际={float(deposit.current_balance):.2f}"))

            # 验证现金余额
            if expected_a is not None and expected_a != 0:
                tx = db.query(CustomerTransaction).filter(
                    CustomerTransaction.customer_id == customer.id,
                    CustomerTransaction.transaction_type == 'initial_balance',
                ).first()
                if not tx:
                    mismatches.append((name, "期初现金交易不存在", f"期望={expected_a:.2f}", "实际=无"))
                elif abs(float(tx.amount) - expected_a) > 0.01:
                    mismatches.append((name, "现金余额不匹配",
                                      f"期望={expected_a:.2f}", f"实际={float(tx.amount):.2f}"))

            if not any(m[0] == name for m in mismatches):
                matched += 1

        # 汇总验证
        db_deposit_total = 0.0
        for d in db.query(CustomerGoldDeposit).all():
            db_deposit_total += float(d.current_balance or 0)

        excel_weight_total = sum(round(d['weight'], 2) for d in customers.values() if d['weight'] is not None)

        db_tx_total = 0.0
        for tx in db.query(CustomerTransaction).filter(
            CustomerTransaction.transaction_type == 'initial_balance'
        ).all():
            db_tx_total += float(tx.amount or 0)

        excel_amount_total = sum(d['amount'] for d in customers.values() if d['amount'] is not None)

        print(f"\n逐客户验证: {matched}/{len(customers)} 通过")
        if mismatches:
            print(f"\n不匹配项 ({len(mismatches)}):")
            for name, issue, expected, actual in mismatches:
                print(f"  {name}: {issue} | {expected or ''} | {actual or ''}")

        print(f"\n金料余额汇总: Excel(2位)={excel_weight_total:,.2f}g  数据库={db_deposit_total:,.2f}g  差异={db_deposit_total - excel_weight_total:,.2f}g")
        print(f"现金余额汇总: Excel={excel_amount_total:,.2f}元  数据库={db_tx_total:,.2f}元  差异={db_tx_total - excel_amount_total:,.2f}元")

        if not mismatches and abs(db_deposit_total - excel_weight_total) < 0.01 and abs(db_tx_total - excel_amount_total) < 0.01:
            print("\n全部验证通过!")
        else:
            print("\n存在差异，请检查!")

    finally:
        db.close()


def main():
    mode = "dry-run"
    db_url = None
    auto_confirm = False

    args = sys.argv[1:]
    i = 0
    excel_path = None
    while i < len(args):
        if args[i] == "--execute":
            mode = "execute"
        elif args[i] == "--verify":
            mode = "verify"
        elif args[i] == "--yes":
            auto_confirm = True
        elif args[i] == "--db-url" and i + 1 < len(args):
            db_url = args[i + 1]
            i += 1
        elif not args[i].startswith("--") and os.path.exists(args[i]):
            excel_path = args[i]
        i += 1

    if not excel_path:
        excel_path = find_excel_file()

    if not excel_path:
        print("未找到 Excel 文件，请指定路径或将文件放在桌面")
        sys.exit(1)

    target_url = db_url or REMOTE_URL
    print(f"目标数据库: {target_url.split('@')[1] if '@' in target_url else target_url}")
    print(f"读取文件: {excel_path}")
    customers = read_excel(excel_path)
    print(f"解析到 {len(customers)} 个客户")

    if mode == "dry-run":
        dry_run(customers)
        # 额外检查远程数据库匹配情况
        print("\n" + "=" * 80)
        print("远程数据库匹配预检...")
        db = get_db_session(db_url)
        try:
            new_count = 0
            match_count = 0
            dup_count = 0
            for name in sorted(customers.keys()):
                matches = db.query(Customer).filter(
                    Customer.name == name,
                    Customer.status == 'active'
                ).count()
                if matches == 0:
                    new_count += 1
                elif matches == 1:
                    match_count += 1
                else:
                    dup_count += 1
            print(f"  将匹配已有客户: {match_count}")
            print(f"  将新建客户: {new_count}")
            if dup_count:
                print(f"  [注意] 同名客户: {dup_count} (将使用最早创建的记录)")
            existing_deposits = db.query(CustomerGoldDeposit).count()
            existing_init_tx = db.query(CustomerTransaction).filter(
                CustomerTransaction.transaction_type == 'initial_balance'
            ).count()
            print(f"  现有金料余额记录: {existing_deposits}")
            print(f"  现有期初交易记录: {existing_init_tx}")
        finally:
            db.close()
        print("\n要正式执行导入，请运行: python backend/scripts/import_customer_balance.py --execute")
    elif mode == "execute":
        dry_run(customers)
        if not auto_confirm:
            print("\n" + "=" * 80)
            confirm = input("确认导入以上数据到远程数据库？(yes/no): ").strip().lower()
            if confirm != "yes":
                print("已取消")
                return
        else:
            print("\n[--yes] 自动确认，开始导入...")
        execute(customers, db_url)
        verify(customers, db_url)
    elif mode == "verify":
        verify(customers, db_url)


if __name__ == "__main__":
    main()

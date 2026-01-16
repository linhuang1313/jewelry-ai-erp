from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from datetime import datetime, timezone, timedelta
import logging

# 中国时区 UTC+8
CHINA_TZ = timezone(timedelta(hours=8))

def china_now() -> datetime:
    """获取中国时间（UTC+8）"""
    return datetime.now(CHINA_TZ)

def to_china_time(dt: datetime) -> datetime:
    """将任意datetime转换为中国时间"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        # 假设无时区的时间是UTC
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(CHINA_TZ)
import tempfile
import os
import json
import asyncio
from typing import List, Dict, Any, Optional

from .database import get_db, init_db
from .schemas import (
    AIRequest, InboundOrderCreate, BatchInboundCreate, BatchInboundItem,
    InboundOrderResponse, InboundDetailResponse, InventoryResponse,
    CustomerCreate, CustomerResponse,
    SupplierCreate, SupplierResponse,
    SalesOrderCreate, SalesOrderResponse, SalesDetailResponse, SalesDetailItem,
    SalespersonCreate, SalespersonResponse
)
from .models import InboundOrder, InboundDetail, Inventory, Customer, SalesOrder, SalesDetail, Supplier, ChatLog, Location, LocationInventory, Salesperson, ReturnOrder
from .ai_parser import parse_user_message
from .utils import to_pinyin_initials
from . import context_manager as ctx
from .routers import finance_router
from .routers.warehouse import router as warehouse_router
from .routers.settlement import router as settlement_router
from .routers.suppliers import router as suppliers_router
from .routers.customers import router as customers_router
from .routers.returns import router as returns_router
from .routers.analytics import router as analytics_router
from .routers.gold_material import router as gold_material_router
from .routers.product_codes import router as product_codes_router
from .routers.salespersons import router as salespersons_router
from .routers.sales import router as sales_router
from .routers.export import router as export_router
from .ocr_parser import OCR_AVAILABLE

# 百度云 OCR（云端可用）
from .baidu_ocr import is_ocr_configured as is_baidu_ocr_configured, extract_text_from_image as baidu_extract_text

# OCR 功能状态：优先使用百度云 OCR，其次使用本地 PaddleOCR
BAIDU_OCR_ENABLED = is_baidu_ocr_configured()
LOCAL_OCR_ENABLED = OCR_AVAILABLE

if LOCAL_OCR_ENABLED:
    from .ocr_parser import extract_text_from_image as local_extract_text
else:
    local_extract_text = None

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="AI-ERP珠宝入库BETA测试")

# ========== 配置CORS（必须在路由注册之前）==========
# 配置CORS - 支持所有来源（包括 Vercel 和 Railway）
# 注意：allow_credentials=True 时不能使用 allow_origins=["*"]
# 所以这里使用 allow_credentials=False
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 允许所有域名
    allow_credentials=False,  # 设为 False 才能使用 allow_origins=["*"]
    allow_methods=["*"],  # 允许所有方法
    allow_headers=["*"],  # 允许所有头
    expose_headers=["*"],
    max_age=3600,  # 预检请求缓存时间
)

# 注册财务对账路由
app.include_router(finance_router)

# 注册仓库管理路由
app.include_router(warehouse_router)

# 注册供应商管理路由
app.include_router(suppliers_router)

# 注册客户管理路由
app.include_router(customers_router)

# 注册退货管理路由
app.include_router(returns_router)

# 注册结算管理路由
app.include_router(settlement_router)

# 注册数据分析路由
app.include_router(analytics_router)

# 注册金料管理路由
app.include_router(gold_material_router)

# 注册商品编码管理路由
app.include_router(product_codes_router)

# 注册业务员管理路由
app.include_router(salespersons_router)

# 注册销售单管理路由
app.include_router(sales_router)

# 注册数据导出路由
app.include_router(export_router)


# 初始化数据库
@app.on_event("startup")
async def startup_event():
    init_db()
    logger.info("数据库初始化完成")
    
    # 执行数据库迁移 - 添加缺失的列
    from .database import SessionLocal, engine
    from sqlalchemy import text
    db = SessionLocal()
    try:
        # inbound_details 表的列
        inbound_columns = [
            ("product_code", "VARCHAR(20) NULL"),
            ("piece_count", "INTEGER DEFAULT 0"),
            ("piece_labor_cost", "FLOAT DEFAULT 0.0"),
        ]
        
        for col_name, col_type in inbound_columns:
            result = db.execute(text(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'inbound_details' AND column_name = '{col_name}'
            """))
            if not result.fetchone():
                db.execute(text(f"""
                    ALTER TABLE inbound_details 
                    ADD COLUMN {col_name} {col_type}
                """))
                db.commit()
                logger.info(f"已添加 {col_name} 列到 inbound_details 表")
        
        # settlement_orders 表的客户余额字段和混合支付字段
        settlement_columns = [
            ("previous_cash_debt", "FLOAT DEFAULT 0.0"),
            ("previous_gold_debt", "FLOAT DEFAULT 0.0"),
            ("gold_deposit_balance", "FLOAT DEFAULT 0.0"),
            ("cash_deposit_balance", "FLOAT DEFAULT 0.0"),
            # 混合支付专用字段
            ("gold_payment_weight", "FLOAT NULL"),
            ("cash_payment_weight", "FLOAT NULL"),
            # 灵活支付字段
            ("payment_difference", "FLOAT DEFAULT 0.0"),
            ("payment_status", "VARCHAR(20) DEFAULT 'full'"),
        ]
        
        for col_name, col_type in settlement_columns:
            result = db.execute(text(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'settlement_orders' AND column_name = '{col_name}'
            """))
            if not result.fetchone():
                db.execute(text(f"""
                    ALTER TABLE settlement_orders 
                    ADD COLUMN {col_name} {col_type}
                """))
                db.commit()
                logger.info(f"已添加 {col_name} 列到 settlement_orders 表")
        
        # sales_details 表的件数和件工费字段
        sales_detail_columns = [
            ("piece_count", "INTEGER NULL"),
            ("piece_labor_cost", "FLOAT NULL"),
        ]
        
        for col_name, col_type in sales_detail_columns:
            result = db.execute(text(f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'sales_details' AND column_name = '{col_name}'
            """))
            if not result.fetchone():
                db.execute(text(f"""
                    ALTER TABLE sales_details 
                    ADD COLUMN {col_name} {col_type}
                """))
                db.commit()
                logger.info(f"已添加 {col_name} 列到 sales_details 表")
    except Exception as e:
        logger.warning(f"迁移检查: {e}")
        db.rollback()
    finally:
        db.close()
    
    # 确保 product_codes 表存在
    from sqlalchemy import inspect
    from .models import ProductCode
    inspector = inspect(engine)
    if 'product_codes' not in inspector.get_table_names():
        ProductCode.__table__.create(bind=engine)
        logger.info("已创建 product_codes 表")
    
    # 初始化预定义商品编码
    from .init_product_codes import init_product_codes
    from .database import SessionLocal
    db = SessionLocal()
    try:
        count = init_product_codes(db)
        if count > 0:
            logger.info(f"已初始化 {count} 个预定义商品编码")
    except Exception as e:
        logger.error(f"初始化商品编码失败: {e}")
    finally:
        db.close()
    
    # 初始化默认位置
    from .database import SessionLocal
    db = SessionLocal()
    try:
        existing = db.query(Location).first()
        if not existing:
            default_locations = [
                Location(code="warehouse", name="商品部仓库", location_type="warehouse", description="总仓库，入库货品存放处"),
                Location(code="showroom", name="展厅", location_type="showroom", description="销售展厅"),
            ]
            for loc in default_locations:
                db.add(loc)
            db.commit()
            logger.info("默认位置初始化完成")
    except Exception as e:
        logger.error(f"初始化默认位置失败: {e}")
    finally:
        db.close()

@app.get("/")
async def root():
    import sys
    import os
    try:
        import paddle
        paddle_available = True
        paddle_path = paddle.__file__
    except ImportError:
        paddle_available = False
        paddle_path = None
    
    return {
        "message": "珠宝ERP系统API", 
        "status": "running",
        "python_version": sys.version,
        "python_executable": sys.executable,
        "paddle_available": paddle_available,
        "paddle_path": paddle_path
    }

# ============= 对话日志辅助函数 =============

def log_chat_message(
    db: Session,
    session_id: str,
    user_role: str,
    message_type: str,
    content: str,
    intent: str = None,
    entities: dict = None,
    response_time_ms: int = None,
    is_successful: bool = True,
    error_message: str = None
):
    """记录对话日志到数据库"""
    try:
        chat_log = ChatLog(
            session_id=session_id or f"session_{datetime.now().strftime('%Y%m%d%H%M%S')}",
            user_role=user_role or "sales",
            message_type=message_type,
            content=content[:10000] if content else "",  # 限制内容长度
            intent=intent,
            entities=json.dumps(entities, ensure_ascii=False) if entities else None,
            response_time_ms=response_time_ms,
            is_successful=1 if is_successful else 0,
            error_message=error_message
        )
        db.add(chat_log)
        db.commit()
        logger.info(f"[对话日志] 已记录: role={user_role}, type={message_type}, intent={intent}")
    except Exception as e:
        logger.error(f"[对话日志] 记录失败: {e}")
        db.rollback()


@app.post("/api/recognize-inbound-sheet")
async def recognize_inbound_sheet(file: UploadFile = File(...)):
    """
    识别入库单图片，提取文字内容
    优先使用百度云 OCR，其次使用本地 PaddleOCR
    """
    # 检查 OCR 功能是否可用
    if not BAIDU_OCR_ENABLED and not LOCAL_OCR_ENABLED:
        return {
            "success": False,
            "message": "OCR 功能未配置。请设置百度云 OCR 密钥（BAIDU_OCR_API_KEY 和 BAIDU_OCR_SECRET_KEY）或在本地安装 PaddleOCR。",
            "recognized_text": "",
            "thinking_steps": ["OCR 功能未配置"]
        }
    
    try:
        logger.info(f"收到图片上传请求：{file.filename}, 类型：{file.content_type}")
        
        # 验证文件类型
        if not file.content_type or not file.content_type.startswith('image/'):
            return {
                "success": False,
                "message": "请上传图片文件（jpg、png等格式）",
                "recognized_text": ""
            }
        
            # 读取上传的文件内容
            content = await file.read()
        
        # 确定使用哪种 OCR
        ocr_method = "百度云 OCR" if BAIDU_OCR_ENABLED else "本地 PaddleOCR"
        logger.info(f"使用 {ocr_method} 进行识别")
        
        try:
            if BAIDU_OCR_ENABLED:
                # 使用百度云 OCR（直接传入图片字节）
                recognized_text = baidu_extract_text(image_bytes=content)
            else:
                # 本地 OCR 已禁用
                return {
                    "success": False,
                    "message": "OCR 功能未配置，请联系管理员",
                    "recognized_text": ""
                }
            
            logger.info(f"OCR识别完成，识别到 {len(recognized_text)} 个字符")
            
            if not recognized_text or len(recognized_text.strip()) == 0:
                return {
                    "success": False,
                    "message": "未能识别出文字内容，请检查图片是否清晰",
                    "recognized_text": "",
                    "thinking_steps": ["OCR识别完成，但未识别到任何文字"]
                }
            
            return {
                "success": True,
                "message": f"图片识别成功（{ocr_method}）",
                "recognized_text": recognized_text,
                "thinking_steps": [
                    f"已识别图片：{file.filename}",
                    f"使用：{ocr_method}",
                    f"识别出 {len(recognized_text.split(chr(10)))} 行文字",
                    "请检查识别内容是否正确"
                ]
            }
        
        except Exception as ocr_error:
            logger.error(f"OCR识别失败：{ocr_error}", exc_info=True)
            return {
                "success": False,
                "message": f"OCR识别失败：{str(ocr_error)}",
                "recognized_text": "",
                "thinking_steps": [f"OCR识别过程出错：{str(ocr_error)}"]
            }
    
    except Exception as e:
        logger.error(f"处理图片上传时出错：{e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理图片时出错：{str(e)}",
            "recognized_text": "",
            "thinking_steps": [f"处理过程出错：{str(e)}"]
        }

@app.post("/api/chat")
async def chat(request: AIRequest, db: Session = Depends(get_db)):
    """处理用户聊天消息 - AI驱动架构"""
    try:
        logger.info(f"收到用户消息: {request.message}")
        
        # 使用AI解析用户消息
        ai_response = parse_user_message(request.message)
        logger.info(f"AI解析结果: action={ai_response.action}, products={ai_response.products}")
        
        # ========== 业务员角色限制：禁止写操作 ==========
        user_role = request.user_role or 'sales'
        if user_role == 'sales':
            # 业务员禁止执行的写操作
            forbidden_actions = ["入库", "创建客户", "创建供应商", "创建销售单", "创建转移单", "退货"]
            if ai_response.action in forbidden_actions:
                return {
                    "success": False,
                    "message": "⚠️ 您是业务员角色，只能查询客户相关信息（销售、退货、欠款、往来账目）。\n\n如需执行入库、开单、退货等操作，请联系相应岗位人员。"
                }
        
        # ========== 写操作：先检查权限，再执行 ==========
        from .middleware.permissions import check_action_permission
        
        if ai_response.action == "入库":
            has_perm, perm_error = check_action_permission(request.user_role or 'sales', ai_response.action)
            if not has_perm:
                return {"success": False, "message": perm_error}
            return await handle_inbound(ai_response, db)
        elif ai_response.action == "创建客户":
            has_perm, perm_error = check_action_permission(request.user_role or 'sales', ai_response.action)
            if not has_perm:
                return {"success": False, "message": perm_error}
            return await handle_create_customer(ai_response, db)
        elif ai_response.action == "创建供应商":
            has_perm, perm_error = check_action_permission(request.user_role or 'sales', ai_response.action)
            if not has_perm:
                return {"success": False, "message": perm_error}
            return await handle_create_supplier(ai_response, db)
        elif ai_response.action == "创建销售单":
            has_perm, perm_error = check_action_permission(request.user_role or 'sales', ai_response.action)
            if not has_perm:
                return {"success": False, "message": perm_error}
            return await handle_create_sales_order(ai_response, db)
        elif ai_response.action == "创建转移单":
            has_perm, perm_error = check_action_permission(request.user_role or 'sales', ai_response.action)
            if not has_perm:
                return {"success": False, "message": perm_error}
            return await handle_create_transfer(ai_response, db)
        elif ai_response.action == "退货":
            # 退货操作：商品专员退给供应商，或柜台退给商品部
            return await handle_return(ai_response, db, request.user_role or 'product')
        
        # ========== 查询和分析操作：使用AI分析引擎 ==========
        else:
            from .ai_analyzer import ai_analyzer
            
            # 收集所有相关数据（传入入库单号、销售单号和用户角色）
            # 用户角色决定显示哪个仓位的库存
            data = ai_analyzer.collect_all_data(
                ai_response.action,
                request.message,
                db,
                order_no=ai_response.order_no,  # RK开头的入库单号
                sales_order_no=ai_response.sales_order_no,  # XS开头的销售单号
                user_role=user_role  # 传入用户角色以确定显示哪个仓位的库存
            )
            
            # 使用AI进行分析
            analysis_result = ai_analyzer.analyze(
                request.message,
                ai_response.action,
                data
            )
            
            # 检测图表需求并生成图表数据
            chart_keywords = ["图表", "chart", "可视化", "用图表", "画图", "画个图", "生成图表", "给我看图表", "用图标", "饼图", "柱状图", "折线图", "趋势图", "分布图", "占比"]
            needs_chart = (
                ai_response.action == "生成图表" or 
                any(keyword in request.message for keyword in chart_keywords)
            )
            
            chart_data_result = {}
            if needs_chart:
                logger.info("[普通] 检测到图表需求，生成图表数据")
                chart_data_result = ai_analyzer.generate_chart_data(ai_response.action, data)
                if chart_data_result:
                    logger.info(f"[普通] 已生成图表数据：柱状图={bool(chart_data_result.get('chart_data'))}, 饼图={bool(chart_data_result.get('pie_data'))}")
            
            # 返回结果（同时返回原始数据，供前端需要时使用）
            return {
                "success": True,
                "message": analysis_result,
                "raw_data": data,  # 可选：前端可以用于图表等
                "action": ai_response.action,
                "chart_data": chart_data_result.get('chart_data'),
                "pie_data": chart_data_result.get('pie_data')
            }
    
    except Exception as e:
        logger.error(f"处理消息时出错: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理消息时出错: {str(e)}"
        }

@app.options("/api/chat-stream")
async def chat_stream_options():
    """处理CORS预检请求"""
    from fastapi.responses import Response
    logger.info("[流式] 收到OPTIONS预检请求")
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "POST, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type",
            "Access-Control-Max-Age": "3600",
        }
    )

@app.post("/api/chat-stream")
async def chat_stream(request: AIRequest, db: Session = Depends(get_db)):
    """流式响应聊天接口 - 显示思考过程和内容逐字生成"""
    
    logger.info(f"[流式] 收到请求: {request.message}, 角色: {request.user_role}")
    
    # 记录用户消息到日志
    start_time = datetime.now()
    session_id = request.session_id or f"session_{start_time.strftime('%Y%m%d%H%M%S%f')}"
    
    log_chat_message(
        db=db,
        session_id=session_id,
        user_role=request.user_role,
        message_type="user",
        content=request.message
    )
    
    # ========== 上下文工程：Read-Before-Decide ==========
    # 加载会话上下文（文件系统作为外部记忆）
    session_context = ctx.load_session_context(session_id)
    context_summary = ctx.generate_context_summary(session_id)
    knowledge_base = ctx.load_knowledge_base()
    
    if context_summary:
        logger.info(f"[Context] 已加载会话上下文，当前目标: {session_context.get('goal', '无')}")
    
    # ========== 查询最近的对话历史（用于上下文理解）==========
    conversation_history = []
    try:
        # 查询最近的5条对话记录（按时间倒序，然后反转）
        recent_logs = db.query(ChatLog).filter(
            ChatLog.user_role == request.user_role
        ).order_by(desc(ChatLog.created_at)).limit(10).all()
        
        # 反转顺序（从旧到新）并格式化
        for log in reversed(recent_logs):
            if log.content:
                conversation_history.append({
                    "role": "user" if log.message_type == "user" else "assistant",
                    "content": log.content
                })
        logger.info(f"[流式] 加载了 {len(conversation_history)} 条历史对话作为上下文")
    except Exception as e:
        logger.warning(f"[流式] 加载对话历史失败: {e}")
    
    async def generate():
        try:
            logger.info("[流式] 开始生成响应")
            
            # ========== 初始填充数据（强制代理开始转发）==========
            # 发送约2KB的初始填充，某些代理需要接收到一定量数据才会开始转发
            initial_padding = ": " + "." * 2048 + "\n\n"
            yield initial_padding
            logger.info("[流式] 已发送初始填充数据")
            
            # ========== 阶段1: 意图解析（立即开始）==========
            first_chunk = f"data: {json.dumps({'type': 'thinking', 'step': '意图解析', 'message': '正在理解您的问题...', 'progress': 10}, ensure_ascii=False)}\n\n"
            logger.info(f"[流式] 发送第一个数据块: {len(first_chunk)} 字节")
            yield first_chunk
            await asyncio.sleep(0.05)
            
            # 构建增强的消息（包含上下文和知识库）
            enhanced_message = request.message
            if context_summary or knowledge_base:
                context_parts = []
                if knowledge_base:
                    context_parts.append(f"【业务规则参考】\n{knowledge_base[:1000]}...")  # 截取前1000字符
                if context_summary:
                    context_parts.append(f"【会话上下文】\n{context_summary}")
                context_parts.append(f"【用户请求】\n{request.message}")
                enhanced_message = "\n\n".join(context_parts)
            
            # 传递对话历史给 AI 解析器
            ai_response = parse_user_message(enhanced_message, conversation_history)
            logger.info(f"[流式] 识别到意图: {ai_response.action}")
            
            # 更新上下文中的实体信息
            entities_to_save = {}
            if ai_response.product_name:
                entities_to_save["last_product"] = ai_response.product_name
            if ai_response.customer_name:
                entities_to_save["last_customer"] = ai_response.customer_name
            if ai_response.supplier_name or ai_response.supplier:
                entities_to_save["last_supplier"] = ai_response.supplier_name or ai_response.supplier
            if entities_to_save:
                ctx.update_entities(session_id, entities_to_save)
            
            yield f"data: {json.dumps({'type': 'thinking', 'step': '意图解析', 'message': f'已识别意图：{ai_response.action}', 'progress': 20, 'status': 'complete'}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # ========== 业务员角色限制：禁止写操作 ==========
            user_role = request.user_role or 'sales'
            if user_role == 'sales':
                # 业务员禁止执行的写操作
                forbidden_actions = ["入库", "创建客户", "创建供应商", "创建销售单", "创建转移单", "退货"]
                if ai_response.action in forbidden_actions:
                    logger.warning(f"[流式] 业务员尝试执行写操作: {ai_response.action}")
                    yield f"data: {json.dumps({'type': 'thinking', 'step': '权限检查', 'message': '权限验证失败', 'progress': 25, 'status': 'error'}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(0.1)
                    error_msg = "⚠️ 您是业务员角色，只能查询客户相关信息（销售、退货、欠款、往来账目）。\n\n如需执行入库、开单、退货等操作，请联系相应岗位人员。"
                    yield f"data: {json.dumps({'type': 'complete', 'data': {'success': False, 'message': error_msg}}, ensure_ascii=False)}\n\n"
                    
                    # 记录业务员权限限制的回复到日志
                    end_time = datetime.now()
                    response_time_ms = int((end_time - start_time).total_seconds() * 1000)
                    log_chat_message(
                        db=db,
                        session_id=session_id,
                        user_role=request.user_role,
                        message_type="assistant",
                        content=error_msg[:5000],
                        intent=ai_response.action,
                        response_time_ms=response_time_ms,
                        is_successful=False,
                        error_message="业务员权限限制"
                    )
                    
                    return
            
            # ========== 写操作：先检查权限，再执行 ==========
            if ai_response.action in ["入库", "创建客户", "创建供应商", "创建销售单", "创建转移单", "退货"]:
                # 导入权限检查模块
                from .middleware.permissions import check_action_permission, get_permission_denied_message
                
                # 检查操作权限
                user_role = request.user_role or 'sales'
                has_perm, perm_error = check_action_permission(user_role, ai_response.action)
                
                if not has_perm:
                    logger.warning(f"[流式] 权限不足: 角色={user_role}, 操作={ai_response.action}")
                    yield f"data: {json.dumps({'type': 'thinking', 'step': '权限检查', 'message': '权限验证失败', 'progress': 25, 'status': 'error'}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(0.1)
                    yield f"data: {json.dumps({'type': 'complete', 'data': {'success': False, 'message': perm_error}}, ensure_ascii=False)}\n\n"
                    
                    # 记录权限不足的回复到日志
                    end_time = datetime.now()
                    response_time_ms = int((end_time - start_time).total_seconds() * 1000)
                    log_chat_message(
                        db=db,
                        session_id=session_id,
                        user_role=request.user_role,
                        message_type="assistant",
                        content=perm_error[:5000] if perm_error else "权限不足",
                        intent=ai_response.action,
                        response_time_ms=response_time_ms,
                        is_successful=False,
                        error_message="权限不足"
                    )
                    
                    return
                
                yield f"data: {json.dumps({'type': 'thinking', 'step': '执行操作', 'message': f'正在执行{ai_response.action}操作...', 'progress': 30}, ensure_ascii=False)}\n\n"
                await asyncio.sleep(0.05)
                
                try:
                    if ai_response.action == "入库":
                        result = await handle_inbound(ai_response, db)
                    elif ai_response.action == "创建客户":
                        result = await handle_create_customer(ai_response, db)
                    elif ai_response.action == "创建供应商":
                        result = await handle_create_supplier(ai_response, db)
                    elif ai_response.action == "创建销售单":
                        result = await handle_create_sales_order(ai_response, db)
                    elif ai_response.action == "创建转移单":
                        result = await handle_create_transfer(ai_response, db)
                    elif ai_response.action == "退货":
                        result = await handle_return(ai_response, db, request.user_role or 'product')
                    elif ai_response.action == "查询客户账务":
                        # 特殊处理：查询客户账务，需要走AI分析流程
                        pass  # 继续执行后续的AI分析流程
                    elif ai_response.action == "登记收款":
                        # 登记收款：返回确认卡片数据
                        result = await handle_payment_registration(ai_response, db)
                        if result.get("success"):
                            # 返回确认卡片数据
                            yield f"data: {json.dumps({'type': 'payment_confirm', 'data': result}, ensure_ascii=False)}\n\n"
                            return
                        else:
                            # 返回错误信息
                            yield f"data: {json.dumps({'type': 'complete', 'data': result}, ensure_ascii=False)}\n\n"
                            return
                    
                    # 【上下文工程】记录成功操作（Append-Only）
                    action_desc = f"{ai_response.action}"
                    if ai_response.product_name:
                        action_desc += f" - {ai_response.product_name}"
                    if ai_response.weight:
                        action_desc += f" {ai_response.weight}g"
                    ctx.append_action(session_id, action_desc, result.get("message", "成功"), success=True)
                    
                    logger.info(f"[流式] {ai_response.action}操作完成，准备返回结果")
                    logger.info(f"[流式] result 包含的字段: {list(result.keys()) if isinstance(result, dict) else type(result)}")
                    
                    # 详细日志：检查 all_products
                    if isinstance(result, dict):
                        has_all_products = 'all_products' in result
                        logger.info(f"[流式][重要] result 是否包含 all_products: {has_all_products}")
                        if has_all_products:
                            logger.info(f"[流式][重要] all_products 数量: {len(result['all_products'])}")
                            for i, p in enumerate(result['all_products']):
                                logger.info(f"[流式][重要] 商品{i+1}: {p}")
                        else:
                            logger.warning(f"[流式][警告] result 中没有 all_products 字段！")
                            logger.warning(f"[流式][警告] result 完整内容: {result}")
                    
                    # 确保 result 可以被 JSON 序列化
                    result_json = json.dumps({'type': 'complete', 'data': result}, ensure_ascii=False, default=str)
                    logger.info(f"[流式] 序列化后的JSON长度: {len(result_json)} 字符")
                    yield f"data: {result_json}\n\n"
                    logger.info("[流式] 已发送完成事件")
                    
                    # 记录 AI 回复到日志（操作类消息）
                    end_time = datetime.now()
                    response_time_ms = int((end_time - start_time).total_seconds() * 1000)
                    log_chat_message(
                        db=db,
                        session_id=session_id,
                        user_role=request.user_role,
                        message_type="assistant",
                        content=result.get("message", "操作完成")[:5000] if isinstance(result, dict) else "操作完成",
                        intent=ai_response.action,
                        response_time_ms=response_time_ms,
                        is_successful=result.get("success", True) if isinstance(result, dict) else True
                    )
                    
                    return
                except Exception as op_error:
                    logger.error(f"[流式] 执行{ai_response.action}操作时出错: {op_error}", exc_info=True)
                    error_msg = f"执行{ai_response.action}操作失败: {str(op_error)}"
                    
                    # 【上下文工程】记录错误（Failure Traces）
                    ctx.record_error(
                        session_id, 
                        error_type=f"{ai_response.action}失败",
                        error_detail=str(op_error),
                        context_info=request.message
                    )
                    
                    # 记录错误回复到日志
                    end_time = datetime.now()
                    response_time_ms = int((end_time - start_time).total_seconds() * 1000)
                    log_chat_message(
                        db=db,
                        session_id=session_id,
                        user_role=request.user_role,
                        message_type="assistant",
                        content=error_msg[:5000],
                        intent=ai_response.action,
                        response_time_ms=response_time_ms,
                        is_successful=False,
                        error_message=str(op_error)
                    )
                    
                    yield f"data: {json.dumps({'type': 'error', 'message': error_msg}, ensure_ascii=False)}\n\n"
                    return
            
            # ========== 阶段2: 数据收集（分步骤发送）==========
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': '正在从数据库收集相关数据...', 'progress': 30}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            from .ai_analyzer import ai_analyzer
            
            # 分步骤收集数据，每完成一部分就发送更新
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': '正在收集库存数据...', 'progress': 35}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # 开始收集数据
            data = {}
            data['context'] = {
                'intent': ai_response.action,
                'user_message': request.message,
                'timestamp': datetime.now().isoformat(),
                'order_no': ai_response.order_no if hasattr(ai_response, 'order_no') else None,  # 添加入库单号（RK开头）
                'sales_order_no': ai_response.sales_order_no if hasattr(ai_response, 'sales_order_no') else None  # 添加销售单号（XS开头）
            }
            
            # 收集库存数据（根据用户角色显示对应仓位的库存）
            # 管理层：显示总库存
            # 商品专员：只显示商品部仓库库存
            # 柜台/结算：只显示展厅库存
            if user_role == "manager":
                # 管理层看总库存
                inventories = db.query(Inventory).all()
                data['inventory'] = [
                    {
                        'product_name': inv.product_name,
                        'total_weight': inv.total_weight,
                        'location': '全部仓位',
                        'last_update': str(inv.last_update) if inv.last_update else None
                    }
                    for inv in inventories
                ]
                data['inventory_location'] = '全部仓位（总库存）'
            else:
                # 根据角色确定仓位
                if user_role == "product":
                    location_code = "warehouse"
                    location_display = "商品部仓库"
                elif user_role in ["counter", "settlement"]:
                    location_code = "showroom"
                    location_display = "展厅"
                else:
                    # 其他角色（如业务员）看总库存
                    location_code = None
                    location_display = "全部仓位"
                
                if location_code:
                    # 查询对应仓位的库存
                    location = db.query(Location).filter(Location.code == location_code).first()
                    if location:
                        location_inventories = db.query(LocationInventory).filter(
                            LocationInventory.location_id == location.id,
                            LocationInventory.weight > 0
                        ).all()
                        data['inventory'] = [
                            {
                                'product_name': li.product_name,
                                'total_weight': li.weight,
                                'location': location_display,
                                'last_update': str(li.last_update) if li.last_update else None
                            }
                            for li in location_inventories
                        ]
                    else:
                        data['inventory'] = []
                    data['inventory_location'] = location_display
                else:
                    # 没有指定仓位，显示总库存
                    inventories = db.query(Inventory).all()
                    data['inventory'] = [
                        {
                            'product_name': inv.product_name,
                            'total_weight': inv.total_weight,
                            'location': '全部仓位',
                            'last_update': str(inv.last_update) if inv.last_update else None
                        }
                        for inv in inventories
                    ]
                    data['inventory_location'] = '全部仓位（总库存）'
            
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': '正在收集供应商数据...', 'progress': 45}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # 收集供应商数据
            suppliers = db.query(Supplier).filter(Supplier.status == "active").all()
            data['suppliers'] = []
            for s in suppliers:
                supplier_name = s.name
                product_count = db.query(func.count(func.distinct(InboundDetail.product_name))).filter(
                    InboundDetail.supplier == supplier_name
                ).scalar() or 0
                data['suppliers'].append({
                    'name': s.name,
                    'supplier_no': s.supplier_no,
                    'phone': s.phone,
                    'address': s.address,
                    'contact_person': s.contact_person,
                    'total_cost': float(s.total_supply_amount) if s.total_supply_amount else 0,
                    'total_weight': float(s.total_supply_weight) if s.total_supply_weight else 0,
                    'supply_count': s.total_supply_count,
                    'last_supply_time': str(s.last_supply_time) if s.last_supply_time else None,
                    'product_count': product_count
                })
            
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': '正在收集客户数据...', 'progress': 55}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # 收集客户数据
            customers = db.query(Customer).filter(Customer.status == "active").all()
            data['customers'] = [
                {
                    'name': c.name,
                    'phone': c.phone,
                    'total_purchase_amount': c.total_purchase_amount,
                    'total_purchase_count': c.total_purchase_count,
                    'last_purchase_time': str(c.last_purchase_time) if c.last_purchase_time else None,
                    'customer_type': c.customer_type
                }
                for c in customers
            ]
            
            # 如果是查询客户账务，收集账务数据
            if ai_response.action == "查询客户账务":
                debt_customer_name = getattr(ai_response, 'debt_customer_name', None)
                if debt_customer_name:
                    yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': f'正在查询 {debt_customer_name} 的账务信息...', 'progress': 60}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(0.05)
                    
                    # 调用聊天查询API
                    from .routers.customers import chat_debt_query
                    debt_result = await chat_debt_query(
                        customer_name=debt_customer_name,
                        query_type=getattr(ai_response, 'debt_query_type', 'all') or 'all',
                        date_start=getattr(ai_response, 'date_start', None),
                        date_end=getattr(ai_response, 'date_end', None),
                        db=db
                    )
                    data['customer_debt'] = debt_result
                    logger.info(f"[流式] 已收集客户账务数据: {debt_customer_name}")
            
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': '正在收集订单数据...', 'progress': 65}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # 收集入库单数据（如果指定了入库单号，只查询该入库单）
            order_no = ai_response.order_no if hasattr(ai_response, 'order_no') else None
            if order_no:
                # 精确查询指定入库单
                order = db.query(InboundOrder).filter(InboundOrder.order_no == order_no).first()
                if order:
                    details = db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all()
                    data['inbound_orders'] = [{
                        'order_id': order.id,  # 添加 order_id 用于下载和打印
                        'order_no': order.order_no,
                        'create_time': str(order.create_time) if order.create_time else None,
                        'status': order.status,
                        'details': [
                            {
                                'product_name': d.product_name,
                                'weight': d.weight,
                                'labor_cost': d.labor_cost,
                                'piece_count': d.piece_count if hasattr(d, 'piece_count') and d.piece_count else None,  # 添加件数
                                'piece_labor_cost': d.piece_labor_cost if hasattr(d, 'piece_labor_cost') and d.piece_labor_cost else None,  # 添加件工费
                                'supplier': d.supplier,
                                'total_cost': d.total_cost
                            }
                            for d in details
                        ]
                    }]
                else:
                    # 入库单不存在
                    data['inbound_orders'] = []
            else:
                # 查询最近的入库单（最多50个）
                inbound_orders = db.query(InboundOrder).order_by(desc(InboundOrder.create_time)).limit(50).all()
                data['inbound_orders'] = []
                for order in inbound_orders:
                    details = db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all()
                    data['inbound_orders'].append({
                        'order_no': order.order_no,
                        'create_time': str(order.create_time) if order.create_time else None,
                        'status': order.status,
                        'details': [
                            {
                                'product_name': d.product_name,
                                'weight': d.weight,
                                'labor_cost': d.labor_cost,
                                'supplier': d.supplier,
                                'total_cost': d.total_cost
                            }
                            for d in details
                        ]
                    })
            
            # 收集销售单数据（如果指定了销售单号，只查询该销售单）
            sales_order_no = ai_response.sales_order_no if hasattr(ai_response, 'sales_order_no') else None
            if sales_order_no:
                # 精确查询指定销售单
                so = db.query(SalesOrder).filter(SalesOrder.order_no == sales_order_no).first()
                if so:
                    details = db.query(SalesDetail).filter(SalesDetail.order_id == so.id).all()
                    data['sales_orders'] = [{
                        'order_no': so.order_no,
                        'customer_name': so.customer_name,
                        'salesperson': so.salesperson,
                        'store_code': so.store_code,
                        'total_labor_cost': so.total_labor_cost,
                        'total_weight': so.total_weight,
                        'status': so.status,
                        'order_date': str(so.order_date) if so.order_date else None,
                        'details': [
                            {
                                'product_name': d.product_name,
                                'weight': d.weight,
                                'labor_cost': d.labor_cost,
                                'total_labor_cost': d.total_labor_cost
                            }
                            for d in details
                        ]
                    }]
                else:
                    # 销售单不存在
                    data['sales_orders'] = []
            else:
                # 查询最近的销售单（最多50个）
                sales_orders = db.query(SalesOrder).order_by(desc(SalesOrder.order_date)).limit(50).all()
                data['sales_orders'] = []
                for so in sales_orders:
                    details = db.query(SalesDetail).filter(SalesDetail.order_id == so.id).all()
                    data['sales_orders'].append({
                        'order_no': so.order_no,
                        'customer_name': so.customer_name,
                        'salesperson': so.salesperson,
                        'store_code': so.store_code,
                        'total_labor_cost': so.total_labor_cost,
                        'total_weight': so.total_weight,
                        'status': so.status,
                        'order_date': str(so.order_date) if so.order_date else None,
                        'details': [
                            {
                                'product_name': d.product_name,
                                'weight': d.weight,
                                'labor_cost': d.labor_cost,
                                'total_labor_cost': d.total_labor_cost
                            }
                            for d in details
                        ]
                    })
            
            # 收集总体统计（根据角色计算库存总重量）
            # 库存总重量使用已收集的库存数据计算，这样就是角色对应仓位的库存
            inventory_total = sum(inv.get('total_weight', 0) for inv in data.get('inventory', []))
            
            data['statistics'] = {
                'total_inventory_weight': float(inventory_total),
                'inventory_location': data.get('inventory_location', '全部仓位'),
                'total_inbound_cost': float(db.query(func.sum(InboundDetail.total_cost)).scalar() or 0),
                'total_suppliers': len(data['suppliers']),
                'total_customers': len(data['customers']),
                'total_inbound_orders': db.query(func.count(InboundOrder.id)).scalar() or 0,
                'total_sales_orders': db.query(func.count(SalesOrder.id)).scalar() or 0,
                'total_products': len(data['inventory'])
            }
            
            inventory_count = len(data.get("inventory", []))
            suppliers_count = len(data.get("suppliers", []))
            customers_count = len(data.get("customers", []))
            
            # 数据收集完成消息（不显示具体数字，保护业务数据隐私）
            message_text = '已加载库存、客户、订单等数据'
            
            yield f"data: {json.dumps({'type': 'thinking', 'step': '数据收集', 'message': message_text, 'progress': 70, 'status': 'complete'}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # ========== 阶段3: AI流式生成（关键部分）==========
            yield f"data: {json.dumps({'type': 'thinking', 'step': 'AI分析', 'message': '正在使用AI进行智能分析...', 'progress': 75}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(0.05)
            
            # 标记开始生成内容
            yield f"data: {json.dumps({'type': 'content_start'})}\n\n"
            
            # 使用流式AI分析
            full_response = ""
            chunk_count = 0
            try:
                for text_chunk in ai_analyzer.analyze_stream(
                    request.message,
                    ai_response.action,
                    data
                ):
                    full_response += text_chunk
                    chunk_count += 1
                    # 立即发送每个文本块
                    yield f"data: {json.dumps({'type': 'content', 'chunk': text_chunk}, ensure_ascii=False)}\n\n"
                    
                    # 每5个块发送一个心跳注释，强制刷新代理缓冲
                    if chunk_count % 5 == 0:
                        yield ": heartbeat\n\n"
            except Exception as e:
                logger.error(f"流式AI分析失败: {e}", exc_info=True)
                error_msg = f"AI分析过程中出现错误：{str(e)}。请稍后重试。"
                yield f"data: {json.dumps({'type': 'content', 'chunk': error_msg}, ensure_ascii=False)}\n\n"
                full_response = error_msg
            
            # ========== 检测图表需求并生成图表数据 ==========
            chart_keywords = ["图表", "chart", "可视化", "用图表", "画图", "画个图", "生成图表", "给我看图表", "用图标", "饼图", "柱状图", "折线图", "趋势图", "分布图", "占比"]
            needs_chart = (
                ai_response.action == "生成图表" or 
                any(keyword in request.message for keyword in chart_keywords)
            )
            
            chart_data_result = {}
            if needs_chart:
                logger.info("[流式] 检测到图表需求，生成图表数据")
                chart_data_result = ai_analyzer.generate_chart_data(ai_response.action, data)
                if chart_data_result:
                    logger.info(f"[流式] 已生成图表数据：柱状图={bool(chart_data_result.get('chart_data'))}, 饼图={bool(chart_data_result.get('pie_data'))}")
            
            # ========== 入库单查询后处理：确保包含完整信息和隐藏标记 ==========
            supplement_text = ""
            order_no_from_response = ai_response.order_no if hasattr(ai_response, 'order_no') else None
            logger.info(f"[流式后处理] action={ai_response.action}, order_no={order_no_from_response}")
            
            # 检查是否是入库单查询（通过 action 或消息内容判断）
            is_inbound_query = (
                ai_response.action == "查询入库单" or 
                (order_no_from_response and order_no_from_response.startswith('RK')) or
                ('RK' in request.message and '入库单' in request.message)
            )
            
            if is_inbound_query:
                inbound_orders = data.get('inbound_orders', [])
                target_order = None
                
                # 尝试从消息中提取入库单号（如果 order_no 为空）
                query_order_no = order_no_from_response
                if not query_order_no:
                    import re
                    match = re.search(r'(RK[A-Z0-9]+)', request.message)
                    if match:
                        query_order_no = match.group(1)
                        logger.info(f"[流式后处理] 从消息中提取入库单号: {query_order_no}")
                
                # 查找目标入库单
                for order in inbound_orders:
                    if order.get('order_no') == query_order_no:
                        target_order = order
                        break
                
                if target_order:
                    # 检查响应中是否包含隐藏标记
                    has_marker = '<!-- INBOUND_ORDER:' in full_response
                    order_id = target_order.get('order_id')
                    
                    # 检查响应中是否包含供应商和克工费信息
                    details = target_order.get('details', [])
                    if details:
                        # 检查第一个商品的供应商和克工费是否在响应中
                        first_detail = details[0]
                        supplier = first_detail.get('supplier', '')
                        labor_cost = first_detail.get('labor_cost', 0)
                        
                        # 总是补充详细信息，确保用户能看到完整的入库单明细
                        # 原因：AI 可能只返回简短回复，缺少供应商和克工费
                        needs_supplement = True  # 强制补充详细信息
                        
                        if needs_supplement or not has_marker:
                            # 构建补充信息
                            supplement_text = "\n\n**📋 入库单明细：**\n"
                            total_weight = 0
                            total_cost = 0
                            total_piece_count = 0
                            
                            for idx, detail in enumerate(details, 1):
                                product_name = detail.get('product_name', '')
                                weight = detail.get('weight', 0)
                                labor_cost_item = detail.get('labor_cost', 0)
                                piece_count = detail.get('piece_count')
                                piece_labor_cost = detail.get('piece_labor_cost')
                                supplier_item = detail.get('supplier', '')
                                total_cost_item = detail.get('total_cost', 0)
                                
                                supplement_text += f"{idx}. **{product_name}**\n"
                                supplement_text += f"   ⚖️ 重量：{weight}克\n"
                                supplement_text += f"   💰 克工费：¥{labor_cost_item}/g\n"
                                if piece_count and piece_count > 0:
                                    supplement_text += f"   📦 件数：{piece_count}件\n"
                                    if piece_labor_cost:
                                        supplement_text += f"   💵 件工费：¥{piece_labor_cost}/件\n"
                                    total_piece_count += piece_count
                                if supplier_item:
                                    supplement_text += f"   🏭 供应商：{supplier_item}\n"
                                supplement_text += f"   💸 总成本：¥{total_cost_item:.2f}\n\n"
                                
                                total_weight += weight
                                total_cost += total_cost_item
                            
                            supplement_text += f"**📊 总计：**\n"
                            supplement_text += f"- 总重量：{total_weight:.2f}克\n"
                            if total_piece_count > 0:
                                supplement_text += f"- 总件数：{total_piece_count}件\n"
                            supplement_text += f"- 总成本：¥{total_cost:.2f}\n"
                            
                            # 发送补充信息作为额外的content chunk
                            if supplement_text:
                                yield f"data: {json.dumps({'type': 'content', 'chunk': supplement_text}, ensure_ascii=False)}\n\n"
                                full_response += supplement_text
                            
                            # 添加隐藏标记（如果还没有）
                            if not has_marker and order_id:
                                marker = f"\n\n<!-- INBOUND_ORDER:{order_id}:{query_order_no} -->"
                                yield f"data: {json.dumps({'type': 'content', 'chunk': marker}, ensure_ascii=False)}\n\n"
                                full_response += marker
            
            # ========== 最终完成 ==========
            yield f"data: {json.dumps({'type': 'complete', 'data': {'success': True, 'message': full_response, 'raw_data': data, 'action': ai_response.action, 'chart_data': chart_data_result.get('chart_data'), 'pie_data': chart_data_result.get('pie_data')}, 'progress': 100}, ensure_ascii=False)}\n\n"
            
            # 记录 AI 回复到日志
            end_time = datetime.now()
            response_time_ms = int((end_time - start_time).total_seconds() * 1000)
            log_chat_message(
                db=db,
                session_id=session_id,
                user_role=request.user_role,
                message_type="assistant",
                content=full_response[:5000] if full_response else "",  # 限制长度
                intent=ai_response.action,
                response_time_ms=response_time_ms,
                is_successful=True
            )
            
        except Exception as e:
            logger.error(f"流式处理出错: {e}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': f'处理出错：{str(e)}'}, ensure_ascii=False)}\n\n"
            
            # 记录错误日志
            log_chat_message(
                db=db,
                session_id=session_id,
                user_role=request.user_role,
                message_type="assistant",
                content="",
                is_successful=False,
                error_message=str(e)
            )
    
    logger.info("[流式] 创建StreamingResponse")
    return StreamingResponse(
        generate(), 
        media_type="text/event-stream",
        headers={
            # 禁用所有缓存
            "Cache-Control": "no-cache, no-store, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
            # 保持连接
            "Connection": "keep-alive",
            # 禁用各层代理缓冲
            "X-Accel-Buffering": "no",  # Nginx
            "X-Content-Type-Options": "nosniff",
            # CORS 头
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "POST, OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type",
        }
    )

async def handle_inbound(ai_response, db: Session) -> Dict[str, Any]:
    """处理入库操作"""
    logger.info(f"[入库] 开始处理入库，AI解析出 {len(ai_response.products) if ai_response.products else 0} 个商品")
    if ai_response.products:
        for i, p in enumerate(ai_response.products):
            logger.info(f"[入库] AI解析商品{i+1}: name={p.product_name}, weight={p.weight}, labor_cost={p.labor_cost}, supplier={p.supplier}")
    
    if not ai_response.products:
        return {
            "success": False,
            "action": "入库",
            "need_form": True,  # 告诉前端需要弹出表单
            "message": "📝 请在弹出的表格中填写入库信息",
            "hint": "请提供商品名称、重量、工费和供应商"
        }
    
    # 先验证所有商品，确保数据完整性
    validated_products = []
    validation_errors = []
    
    # 提取共享的供应商（从任何一个商品中提取）
    shared_supplier = None
    for product in ai_response.products:
        if product.supplier:
            shared_supplier = product.supplier
            break
    
    for idx, product in enumerate(ai_response.products):
        # 如果当前商品没有供应商，使用共享供应商
        product_supplier = product.supplier or shared_supplier
        
        # 验证必填字段
        if not product.product_name or not product.weight or product.labor_cost is None:
            validation_errors.append({
                "index": idx + 1,
                "message": f"商品信息不完整: {product.model_dump()}",
                "product": product.model_dump()
            })
            continue
        
        if not product_supplier:
            validation_errors.append({
                "index": idx + 1,
                "message": f"缺少供应商信息: {product.model_dump()}",
                "product": product.model_dump()
            })
            continue
        
        # 更新商品的供应商字段（使用共享供应商）
        product.supplier = product_supplier
        
        # 验证并转换数值类型
        try:
            weight = float(product.weight)
            labor_cost = float(product.labor_cost)
        except (ValueError, TypeError) as e:
            validation_errors.append({
                "index": idx + 1,
                "message": f"数值格式错误: {str(e)}",
                "product": product.model_dump()
            })
            continue
        
        # 验证数值范围
        if weight <= 0:
            validation_errors.append({
                "index": idx + 1,
                "message": f"重量必须大于0: {weight}",
                "product": product.model_dump()
            })
            continue
        
        if labor_cost < 0:
            validation_errors.append({
                "index": idx + 1,
                "message": f"工费不能为负数: {labor_cost}",
                "product": product.model_dump()
            })
            continue
        
        # 处理件数和件工费（可选字段）
        piece_count = None
        piece_labor_cost = None
        if product.piece_count is not None:
            try:
                piece_count = int(product.piece_count)
                if piece_count < 0:
                    piece_count = None
            except (ValueError, TypeError):
                piece_count = None
        
        if product.piece_labor_cost is not None:
            try:
                piece_labor_cost = float(product.piece_labor_cost)
                if piece_labor_cost < 0:
                    piece_labor_cost = None
            except (ValueError, TypeError):
                piece_labor_cost = None
        
        # 计算总工费：克工费 + 件工费
        gram_cost = weight * labor_cost
        piece_cost = (piece_count or 0) * (piece_labor_cost or 0)
        total_cost = gram_cost + piece_cost
        
        validated_products.append({
            "product": product,
            "weight": weight,
            "labor_cost": labor_cost,
            "piece_count": piece_count,
            "piece_labor_cost": piece_labor_cost,
            "total_cost": total_cost
        })
        
        log_msg = f"[入库] 添加商品 {idx+1}: {product.product_name}, 重量={weight}g, 克工费={labor_cost}元/克"
        if piece_count and piece_labor_cost:
            log_msg += f", 件数={piece_count}, 件工费={piece_labor_cost}元/件"
        log_msg += f", 总工费={total_cost}元"
        logger.info(log_msg)
    
    # 如果有验证错误，返回错误信息
    if validation_errors:
        return {
            "success": False,
            "message": f"商品验证失败，共{len(validation_errors)}个商品有问题",
            "validation_errors": validation_errors,
            "parsed": ai_response.model_dump()
        }
    
    # 如果没有有效商品，返回错误
    if not validated_products:
        return {
            "success": False,
            "message": "没有有效的商品信息",
            "parsed": ai_response.model_dump()
        }
    
    # 检查所有商品是否来自同一供应商（业务规则：一个入库单只能有一个供应商）
    suppliers = set(p["product"].supplier for p in validated_products if p["product"].supplier)
    if len(suppliers) > 1:
        return {
            "success": False,
            "message": f"检测到多个供应商: {', '.join(suppliers)}。每张入库单只能对应一个供应商，请按供应商拆分后分别提交。",
            "suppliers": list(suppliers),
            "parsed": ai_response.model_dump()
        }
    
    # 方案B：返回待确认的卡片数据，不执行入库
    # 准备卡片数据供前端显示
    supplier_name = validated_products[0]["product"].supplier if validated_products else None
    
    # 如果只有一个商品，返回单个商品格式
    if len(validated_products) == 1:
        product = validated_products[0]
        card_data = {
            "product_name": product["product"].product_name,
            "weight": product["weight"],
            "labor_cost": product["labor_cost"],
            "piece_count": product.get("piece_count"),
            "piece_labor_cost": product.get("piece_labor_cost"),
            "supplier": supplier_name,
            "total_cost": product["total_cost"]
        }
        return {
            "success": True,
            "message": f"请核对入库信息: {product['product'].product_name} {product['weight']}克",
            "pending": True,  # 标记为待确认
            "card_data": card_data
        }
    else:
        # 多个商品的情况，返回所有商品
        first_product = validated_products[0]
        all_products_list = [
            {
                "product_name": p["product"].product_name,
                "weight": p["weight"],
                "labor_cost": p["labor_cost"],
                "piece_count": p.get("piece_count"),
                "piece_labor_cost": p.get("piece_labor_cost"),
                "supplier": p["product"].supplier,
                "total_cost": p["total_cost"]
            }
            for p in validated_products
        ]
        logger.info(f"[入库] 多商品入库，共{len(all_products_list)}个商品: {all_products_list}")
        return {
            "success": True,
            "message": f"请核对入库信息，共{len(validated_products)}个商品",
            "pending": True,  # 标记为待确认
            "card_data": {
                "product_name": first_product["product"].product_name,
                "weight": first_product["weight"],
                "labor_cost": first_product["labor_cost"],
                "piece_count": first_product.get("piece_count"),
                "piece_labor_cost": first_product.get("piece_labor_cost"),
                "supplier": supplier_name,
                "total_cost": first_product["total_cost"]
            },
            "all_products": all_products_list
        }

# ========== 以下查询和分析函数已被AI分析引擎替代，已删除 ==========
# handle_query_inventory - 已由AI分析引擎替代
# handle_query_suppliers - 已由AI分析引擎替代
# handle_supplier_analysis - 已由AI分析引擎替代
# handle_generate_chart - 已由AI分析引擎替代
# handle_query_orders - 已由AI分析引擎替代
# handle_statistics - 已由AI分析引擎替代

# ==================== 入库管理API ====================

async def execute_inbound(card_data: Dict[str, Any], db: Session) -> Dict[str, Any]:
    """执行实际的入库操作（从卡片数据）"""
    try:
        product_name = card_data.get("product_name")
        product_code = card_data.get("product_code")  # 商品编码
        weight = float(card_data.get("weight", 0))
        labor_cost = float(card_data.get("labor_cost", 0))
        piece_count = card_data.get("piece_count")  # 件数（可选）
        piece_labor_cost = card_data.get("piece_labor_cost")  # 件工费（可选）
        supplier_name = card_data.get("supplier")
        
        # 转换件数和件工费
        if piece_count is not None:
            piece_count = int(piece_count)
        if piece_labor_cost is not None:
            piece_labor_cost = float(piece_labor_cost)
        
        # ========== 商品编码与名称关联处理 ==========
        # 如果提供了商品编码，查询对应的商品名称
        if product_code:
            from .models import ProductCode as ProductCodeModel
            code_record = db.query(ProductCodeModel).filter(ProductCodeModel.code == product_code).first()
            if code_record and code_record.name:
                # 使用商品编码对应的名称作为库存标识
                product_name = code_record.name
        
        # 如果product_name看起来像是商品编码（全大写或纯数字），尝试查找对应名称
        if product_name and (product_name.isupper() or product_name.replace(' ', '').isalnum()):
            from .models import ProductCode as ProductCodeModel
            code_record = db.query(ProductCodeModel).filter(ProductCodeModel.code == product_name).first()
            if code_record and code_record.name:
                # 找到了对应的商品名称，使用它
                product_code = product_name  # 保存原始编码
                product_name = code_record.name
        
        # 验证数据
        if not product_name or weight <= 0 or labor_cost < 0:
            return {
                "success": False,
                "message": "商品信息不完整或无效",
                "error": "validation_failed"
            }
        
        # 生成入库单号（使用中国时间）
        pinyin_initials = to_pinyin_initials(product_name)
        timestamp = china_now().strftime("%Y%m%d%H%M%S")
        order_no = f"RK{pinyin_initials}{timestamp}"
        
        # 创建入库单
        order = InboundOrder(order_no=order_no, create_time=china_now())
        db.add(order)
        db.flush()
        
        # 查找或创建供应商
        supplier_id = None
        supplier_obj = None
        if supplier_name:
            supplier_obj = db.query(Supplier).filter(
                Supplier.name == supplier_name,
                Supplier.status == "active"
            ).first()
            
            if not supplier_obj:
                supplier_no = f"GYS{china_now().strftime('%Y%m%d%H%M%S')}"
                supplier_obj = Supplier(
                    supplier_no=supplier_no,
                    name=supplier_name,
                    supplier_type="个人"
                )
                db.add(supplier_obj)
                db.flush()
            
            supplier_id = supplier_obj.id
        
        # 计算总成本：克工费 + 件工费
        gram_cost = labor_cost * weight
        piece_cost = (piece_count or 0) * (piece_labor_cost or 0)
        total_cost = gram_cost + piece_cost
        
        # 创建入库明细
        detail = InboundDetail(
            order_id=order.id,
            product_name=product_name,
            product_category=card_data.get("product_category"),
            weight=weight,
            labor_cost=labor_cost,
            piece_count=piece_count,
            piece_labor_cost=piece_labor_cost,
            supplier=supplier_name,
            supplier_id=supplier_id,
            total_cost=total_cost
        )
        db.add(detail)
        
        # 更新或创建库存（总库存表）
        inventory = db.query(Inventory).filter(Inventory.product_name == product_name).first()
        if inventory:
            inventory.total_weight += weight
        else:
            inventory = Inventory(product_name=product_name, total_weight=weight)
            db.add(inventory)
        
        # 更新分仓库存（默认入库到"商品部仓库"）
        default_location = db.query(Location).filter(Location.code == "warehouse").first()
        if not default_location:
            # 如果默认位置不存在，创建它
            default_location = Location(
                code="warehouse",
                name="商品部仓库",
                location_type="warehouse",
                description="默认入库位置"
            )
            db.add(default_location)
            db.flush()
        
        # 更新或创建分仓库存记录
        location_inventory = db.query(LocationInventory).filter(
            LocationInventory.product_name == product_name,
            LocationInventory.location_id == default_location.id
        ).first()
        
        if location_inventory:
            location_inventory.weight += weight
        else:
            location_inventory = LocationInventory(
                product_name=product_name,
                location_id=default_location.id,
                weight=weight
            )
            db.add(location_inventory)
        
        # 更新供应商统计信息
        if supplier_obj:
            supplier_obj.total_supply_amount += total_cost
            supplier_obj.total_supply_weight += weight
            supplier_obj.total_supply_count += 1
            supplier_obj.last_supply_time = datetime.now()
        
        # 提交事务
        db.commit()
        
        # 刷新对象
        db.refresh(order)
        db.refresh(detail)
        db.refresh(inventory)
        
        # 构建响应
        order_response = InboundOrderResponse.model_validate(order).model_dump(mode='json')
        detail_response = InboundDetailResponse.model_validate(detail).model_dump(mode='json')
        inventory_response = InventoryResponse.model_validate(inventory).model_dump(mode='json')
        
        logger.info(f"入库成功: order_id={order.id}, order_no={order.order_no}")
        logger.info(f"返回的order响应包含字段: {list(order_response.keys())}")
        logger.info(f"返回的order.id值: {order_response.get('id')}")
        
        return {
            "success": True,
            "message": f"入库成功: {product_name} {weight}克",
            "order_id": order.id,
            "order_no": order.order_no,
            "order": order_response,
            "detail": detail_response,
            "inventory": inventory_response
        }
    
    except Exception as e:
        db.rollback()
        logger.error(f"入库操作失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"入库操作失败: {str(e)}",
            "error": str(e)
        }

@app.post("/api/inbound-orders")
async def create_inbound_order(card_data: InboundOrderCreate, db: Session = Depends(get_db)):
    """创建入库单（从卡片数据确认入库）"""
    try:
        # 转换为字典格式
        card_dict = {
            "product_name": card_data.product_name,
            "product_category": card_data.product_category,
            "weight": card_data.weight,
            "labor_cost": card_data.labor_cost,
            "supplier": card_data.supplier
        }
        
        result = await execute_inbound(card_dict, db)
        return result
    except Exception as e:
        logger.error(f"创建入库单失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"创建入库单失败: {str(e)}",
            "error": str(e)
        }


@app.post("/api/inbound-orders/batch")
async def create_batch_inbound_orders(batch_data: BatchInboundCreate, db: Session = Depends(get_db)):
    """批量创建入库单（快捷入库表格）- 创建1个入库单包含所有商品明细"""
    try:
        if not batch_data.items:
            return {
                "success": False,
                "message": "没有商品数据"
            }
        
        from .models import ProductCode as ProductCodeModel
        import uuid
        
        # 生成唯一入库单号（使用UUID后缀确保唯一性）
        timestamp = china_now().strftime("%Y%m%d%H%M%S")
        unique_suffix = uuid.uuid4().hex[:4].upper()
        order_no = f"RKPL{timestamp}{unique_suffix}"  # RKPL = 入库批量
        
        # 创建入库单
        order = InboundOrder(order_no=order_no, create_time=china_now())
        db.add(order)
        db.flush()
        
        # 查找或创建供应商
        supplier_name = batch_data.supplier
        supplier_id = None
        supplier_obj = None
        if supplier_name:
            supplier_obj = db.query(Supplier).filter(
                Supplier.name == supplier_name,
                Supplier.status == "active"
            ).first()
            
            if not supplier_obj:
                supplier_no = f"GYS{china_now().strftime('%Y%m%d%H%M%S')}"
                supplier_obj = Supplier(
                    supplier_no=supplier_no,
                    name=supplier_name,
                    supplier_type="个人"
                )
                db.add(supplier_obj)
                db.flush()
            
            supplier_id = supplier_obj.id
        
        # 获取默认入库位置
        default_location = db.query(Location).filter(Location.code == "warehouse").first()
        if not default_location:
            default_location = Location(
                code="warehouse",
                name="商品部仓库",
                location_type="warehouse",
                description="默认入库位置"
            )
            db.add(default_location)
            db.flush()
        
        results = []
        total_weight = 0
        total_cost = 0
        success_count = 0
        error_count = 0
        
        # 本地缓存：跟踪已处理的库存，避免同一事务中重复查询/插入
        inventory_cache = {}  # product_name -> Inventory object
        location_inventory_cache = {}  # product_name -> LocationInventory object
        
        for idx, item in enumerate(batch_data.items):
            try:
                product_name = item.product_name
                product_code = item.product_code
                weight = float(item.weight)
                labor_cost = float(item.labor_cost)
                piece_count = item.piece_count or 0
                piece_labor_cost = item.piece_labor_cost or 0.0
                
                # 商品编码与名称关联处理
                if product_code:
                    code_record = db.query(ProductCodeModel).filter(ProductCodeModel.code == product_code).first()
                    if code_record and code_record.name:
                        product_name = code_record.name
                
                # 验证数据
                if not product_name or weight <= 0 or labor_cost < 0:
                    error_count += 1
                    results.append({
                        "index": idx + 1,
                        "product_name": item.product_name,
                        "success": False,
                        "error": "商品信息不完整或无效"
                    })
                    continue
                
                # 计算总成本
                gram_cost = labor_cost * weight
                piece_cost = piece_count * piece_labor_cost
                item_total_cost = gram_cost + piece_cost
                
                # 创建入库明细
                detail = InboundDetail(
                    order_id=order.id,
                    product_code=product_code,  # 保存商品编码/条码
                    product_name=product_name,
                    weight=weight,
                    labor_cost=labor_cost,
                    piece_count=piece_count if piece_count > 0 else None,
                    piece_labor_cost=piece_labor_cost if piece_labor_cost > 0 else None,
                    supplier=supplier_name,
                    supplier_id=supplier_id,
                    total_cost=item_total_cost
                )
                db.add(detail)
                
                # 更新或创建总库存（使用本地缓存避免重复插入）
                if product_name in inventory_cache:
                    # 已在本批次中处理过，直接累加
                    inventory_cache[product_name].total_weight += weight
                else:
                    # 首次处理，查询数据库
                    inventory = db.query(Inventory).filter(Inventory.product_name == product_name).first()
                    if inventory:
                        inventory.total_weight += weight
                        inventory_cache[product_name] = inventory
                    else:
                        inventory = Inventory(product_name=product_name, total_weight=weight)
                        db.add(inventory)
                        inventory_cache[product_name] = inventory
                
                # 更新或创建分仓库存（使用本地缓存避免重复插入）
                if product_name in location_inventory_cache:
                    # 已在本批次中处理过，直接累加
                    location_inventory_cache[product_name].weight += weight
                else:
                    # 首次处理，查询数据库
                    location_inventory = db.query(LocationInventory).filter(
                        LocationInventory.product_name == product_name,
                        LocationInventory.location_id == default_location.id
                    ).first()
                    
                    if location_inventory:
                        location_inventory.weight += weight
                        location_inventory_cache[product_name] = location_inventory
                    else:
                        location_inventory = LocationInventory(
                            product_name=product_name,
                            location_id=default_location.id,
                            weight=weight
                        )
                        db.add(location_inventory)
                        location_inventory_cache[product_name] = location_inventory
                
                total_weight += weight
                total_cost += item_total_cost
                success_count += 1
                results.append({
                    "index": idx + 1,
                    "product_name": product_name,
                    "weight": weight,
                    "success": True
                })
                
            except Exception as e:
                error_count += 1
                results.append({
                    "index": idx + 1,
                    "product_name": item.product_name,
                    "success": False,
                    "error": str(e)
                })
        
        # 更新供应商统计信息
        if supplier_obj and success_count > 0:
            supplier_obj.total_supply_amount += total_cost
            supplier_obj.total_supply_weight += total_weight
            supplier_obj.total_supply_count += success_count
            supplier_obj.last_supply_time = datetime.now()
        
        # 提交事务
        db.commit()
        db.refresh(order)
        
        logger.info(f"批量入库成功: order_id={order.id}, order_no={order.order_no}, items={success_count}")
        
        return {
            "success": success_count > 0,
            "message": f"批量入库成功：{success_count} 件商品已入库" if error_count == 0 else f"批量入库完成：成功 {success_count} 件，失败 {error_count} 件",
            "order_id": order.id,
            "order_no": order.order_no,
            "success_count": success_count,
            "error_count": error_count,
            "total_weight": total_weight,
            "total_cost": total_cost,
            "results": results
        }
    except Exception as e:
        db.rollback()
        logger.error(f"批量入库失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"批量入库失败: {str(e)}",
            "error": str(e)
        }


@app.options("/api/inbound-orders/{order_id}/download")
async def download_inbound_order_options(order_id: int):
    """处理CORS预检请求"""
    from fastapi.responses import Response
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*",
        }
    )

@app.get("/api/inbound-orders/{order_id}/download")
async def download_inbound_order(
    order_id: int, 
    format: str = Query("pdf", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """下载或打印入库单（支持PDF和HTML格式）"""
    try:
        logger.info(f"下载入库单请求: order_id={order_id}, format={format}")
        # 查询入库单和明细
        order = db.query(InboundOrder).filter(InboundOrder.id == order_id).first()
        if not order:
            logger.warning(f"入库单不存在: order_id={order_id}")
            # 尝试查询所有入库单，看看有哪些ID
            all_orders = db.query(InboundOrder).limit(10).all()
            logger.info(f"数据库中存在的入库单ID: {[o.id for o in all_orders]}")
            raise HTTPException(status_code=404, detail="入库单不存在")
        
        logger.info(f"找到入库单: order_no={order.order_no}, status={order.status}")
        
        details = db.query(InboundDetail).filter(InboundDetail.order_id == order_id).all()
        if not details:
            logger.warning(f"入库单明细不存在: order_id={order_id}")
            raise HTTPException(status_code=404, detail="入库单明细不存在")
        
        logger.info(f"找到 {len(details)} 条入库单明细")
        
        if format == "pdf":
            try:
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import mm
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                import io
                import os
                from .timezone_utils import to_china_time, format_china_time
                
                # 自定义纸张尺寸：241mm × 140mm 横向（针式打印机常用尺寸）
                PAGE_WIDTH = 241 * mm
                PAGE_HEIGHT = 140 * mm
                
                # 生成PDF
                buffer = io.BytesIO()
                p = canvas.Canvas(buffer, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))
                width, height = PAGE_WIDTH, PAGE_HEIGHT
                
                # 使用 CID 字体（内置支持中文，无需外部字体文件）
                try:
                    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                    chinese_font = 'STSong-Light'
                    logger.info("成功注册中文CID字体: STSong-Light")
                except Exception as cid_error:
                    logger.warning(f"注册CID字体失败: {cid_error}")
                    chinese_font = None
                
                # 页边距设置（适应241mm×140mm横向纸张）
                left_margin = 10 * mm
                right_margin = width - 10 * mm
                top_margin = height - 8 * mm
                
                # 标题（居中）
                if chinese_font:
                    p.setFont(chinese_font, 14)
                else:
                    p.setFont("Helvetica-Bold", 14)
                p.drawCentredString(width / 2, top_margin, "珠宝入库单")
                
                # 入库单信息（两列布局）
                y = top_margin - 18
                if chinese_font:
                    p.setFont(chinese_font, 9)
                else:
                    p.setFont("Helvetica", 9)
                
                # 入库时间（使用时区转换）
                if order.create_time:
                    china_time = to_china_time(order.create_time)
                    create_time_str = format_china_time(china_time, '%Y-%m-%d %H:%M')
                else:
                    create_time_str = "未知"
                
                # 左侧：入库单号
                p.drawString(left_margin, y, f"单号：{order.order_no}")
                # 右侧：时间和操作员
                p.drawString(width / 2, y, f"时间：{create_time_str}  操作员：{order.operator}")
                y -= 15
                
                # 分隔线
                p.line(left_margin, y, right_margin, y)
                y -= 12
                
                # 表格标题（调整列位置适应横向布局）
                col_x = [left_margin, 70*mm, 100*mm, 125*mm, 150*mm, 175*mm, 200*mm]
                if chinese_font:
                    p.setFont(chinese_font, 8)
                else:
                    p.setFont("Helvetica-Bold", 8)
                p.drawString(col_x[0], y, "商品名称")
                p.drawString(col_x[1], y, "重量(克)")
                p.drawString(col_x[2], y, "克工费")
                p.drawString(col_x[3], y, "件数")
                p.drawString(col_x[4], y, "件工费")
                p.drawString(col_x[5], y, "总成本")
                p.drawString(col_x[6], y, "供应商")
                y -= 10
                
                # 分隔线
                p.line(left_margin, y, right_margin, y)
                y -= 10
                
                # 商品明细
                total_cost = 0
                total_weight = 0
                total_piece_count = 0
                bottom_margin = 25 * mm  # 底部留空给合计
                
                for idx, detail in enumerate(details):
                    if y < bottom_margin:  # 换页
                        p.showPage()
                        y = top_margin - 10
                        # 重新绘制表头
                        if chinese_font:
                            p.setFont(chinese_font, 8)
                        else:
                            p.setFont("Helvetica-Bold", 8)
                        p.drawString(col_x[0], y, "商品名称")
                        p.drawString(col_x[1], y, "重量(克)")
                        p.drawString(col_x[2], y, "克工费")
                        p.drawString(col_x[3], y, "件数")
                        p.drawString(col_x[4], y, "件工费")
                        p.drawString(col_x[5], y, "总成本")
                        p.drawString(col_x[6], y, "供应商")
                        y -= 10
                        p.line(left_margin, y, right_margin, y)
                        y -= 10
                    
                    # 商品信息（处理长文本，适应更紧凑的布局）
                    product_name = detail.product_name[:12] if len(detail.product_name) > 12 else detail.product_name
                    supplier_name = (detail.supplier or "-")[:6] if detail.supplier else "-"
                    piece_count = getattr(detail, 'piece_count', None) or 0
                    piece_labor_cost = getattr(detail, 'piece_labor_cost', None) or 0
                    piece_count_str = str(piece_count) if piece_count > 0 else "-"
                    piece_labor_cost_str = f"{piece_labor_cost:.1f}" if piece_count > 0 else "-"
                    
                    # 使用中文字体绘制商品名称和供应商
                    if chinese_font:
                        p.setFont(chinese_font, 8)
                        p.drawString(col_x[0], y, product_name)
                        p.setFont("Helvetica", 8)
                        p.drawString(col_x[1], y, f"{detail.weight:.2f}")
                        p.drawString(col_x[2], y, f"{detail.labor_cost:.1f}")
                        p.drawString(col_x[3], y, piece_count_str)
                        p.drawString(col_x[4], y, piece_labor_cost_str)
                        p.drawString(col_x[5], y, f"{detail.total_cost:.2f}")
                        p.setFont(chinese_font, 8)
                        p.drawString(col_x[6], y, supplier_name)
                    else:
                        p.setFont("Helvetica", 8)
                        p.drawString(col_x[0], y, product_name)
                        p.drawString(col_x[1], y, f"{detail.weight:.2f}")
                        p.drawString(col_x[2], y, f"{detail.labor_cost:.1f}")
                        p.drawString(col_x[3], y, piece_count_str)
                        p.drawString(col_x[4], y, piece_labor_cost_str)
                        p.drawString(col_x[5], y, f"{detail.total_cost:.2f}")
                        p.drawString(col_x[6], y, supplier_name)
                    
                    total_cost += detail.total_cost
                    total_weight += detail.weight
                    total_piece_count += piece_count
                    y -= 12
                
                # 总计（紧凑布局）
                y -= 5
                p.line(left_margin, y, right_margin, y)
                y -= 12
                if chinese_font:
                    p.setFont(chinese_font, 9)
                else:
                    p.setFont("Helvetica-Bold", 9)
                # 一行显示所有合计信息
                summary_text = f"合计：重量 {total_weight:.2f}克"
                if total_piece_count > 0:
                    summary_text += f"  |  件数 {total_piece_count}件"
                summary_text += f"  |  总工费 ¥{total_cost:.2f}"
                p.drawString(left_margin, y, summary_text)
                
                p.save()
                buffer.seek(0)
                
                # 使用英文文件名避免编码问题
                filename = f"inbound_order_{order.order_no}.pdf"
                pdf_content = buffer.getvalue()
                logger.info(f"PDF生成成功，文件名: {filename}, 大小: {len(pdf_content)} 字节")
                
                # 使用 Response 直接返回字节数据（BytesIO 不兼容 FileResponse）
                from fastapi.responses import Response
                response = Response(
                    content=pdf_content,
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f"attachment; filename={filename}",
                        "Content-Type": "application/pdf",
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "GET, OPTIONS",
                        "Access-Control-Allow-Headers": "*",
                    }
                )
                return response
            except ImportError as e:
                logger.error(f"reportlab未安装，无法生成PDF: {e}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"PDF生成功能需要安装reportlab库: {str(e)}")
            except Exception as pdf_error:
                import traceback
                error_trace = traceback.format_exc()
                logger.error(f"PDF生成失败: {pdf_error}", exc_info=True)
                logger.error(f"PDF生成错误堆栈: {error_trace}")
                raise HTTPException(status_code=500, detail=f"PDF生成失败: {str(pdf_error)}")
        
        elif format == "html":
            # 生成HTML（用于打印）
            from .timezone_utils import to_china_time, format_china_time
            
            # 转换入库时间为中国时间
            if order.create_time:
                china_time = to_china_time(order.create_time)
                create_time_str = format_china_time(china_time, '%Y-%m-%d %H:%M:%S')
            else:
                create_time_str = '未知'
            
            html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>入库单 - {order.order_no}</title>
    <style>
        @media print {{
            @page {{
                size: 241mm 140mm landscape;
                margin: 5mm;
            }}
            body {{
                margin: 0;
                padding: 0;
            }}
        }}
        body {{
            font-family: "Microsoft YaHei", Arial, sans-serif;
            padding: 8px;
            max-width: 235mm;
            margin: 0 auto;
            font-size: 10px;
        }}
        .header {{
            text-align: center;
            margin-bottom: 8px;
            border-bottom: 1px solid #333;
            padding-bottom: 5px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 14px;
            color: #333;
        }}
        .info {{
            margin-bottom: 8px;
            display: flex;
            justify-content: space-between;
            flex-wrap: wrap;
            font-size: 9px;
        }}
        .info-item {{
            display: inline-block;
            margin-right: 15px;
        }}
        .info-label {{
            font-weight: bold;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 5px;
            margin-bottom: 8px;
            font-size: 9px;
        }}
        th, td {{
            border: 1px solid #999;
            padding: 3px 5px;
            text-align: center;
        }}
        th {{
            background-color: #f0f0f0;
            font-weight: bold;
        }}
        .total {{
            margin-top: 5px;
            padding-top: 5px;
            border-top: 1px solid #333;
            font-size: 10px;
        }}
        .total-item {{
            display: inline-block;
            margin-right: 20px;
        }}
        .total-amount {{
            font-weight: bold;
            color: #d32f2f;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>珠宝入库单</h1>
    </div>
    <div class="info">
        <span class="info-item"><span class="info-label">单号：</span>{order.order_no}</span>
        <span class="info-item"><span class="info-label">时间：</span>{create_time_str}</span>
        <span class="info-item"><span class="info-label">操作员：</span>{order.operator}</span>
        <span class="info-item"><span class="info-label">状态：</span>{order.status}</span>
    </div>
    <table>
        <thead>
            <tr>
                <th>商品名称</th>
                <th>重量(克)</th>
                <th>克工费</th>
                <th>件数</th>
                <th>件工费</th>
                <th>总成本</th>
                <th>供应商</th>
            </tr>
        </thead>
        <tbody>
"""
            
            total_cost = 0
            total_weight = 0
            total_piece_count = 0
            for detail in details:
                piece_count = getattr(detail, 'piece_count', None) or 0
                piece_labor_cost = getattr(detail, 'piece_labor_cost', None) or 0
                piece_count_str = str(piece_count) if piece_count > 0 else '-'
                piece_labor_cost_str = f"{piece_labor_cost:.2f}" if piece_count > 0 else '-'
                
                html_content += f"""
            <tr>
                <td>{detail.product_name}</td>
                <td>{detail.weight:.2f}</td>
                <td>{detail.labor_cost:.2f}</td>
                <td>{piece_count_str}</td>
                <td>{piece_labor_cost_str}</td>
                <td>{detail.total_cost:.2f}</td>
                <td>{detail.supplier or '-'}</td>
            </tr>
"""
                total_cost += detail.total_cost
                total_weight += detail.weight
                total_piece_count += piece_count
            
            # 生成件数文本（如果有件数）
            piece_text = f"  |  件数：{total_piece_count}件" if total_piece_count > 0 else ""
            
            html_content += f"""
        </tbody>
    </table>
    <div class="total">
        <span class="total-item">合计：重量 <strong>{total_weight:.2f}</strong>克{piece_text}</span>
        <span class="total-item total-amount">总成本：¥<strong>{total_cost:.2f}</strong></span>
    </div>
</body>
</html>
"""
            
            # 创建HTML响应并添加CORS头
            response = HTMLResponse(content=html_content)
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
            response.headers["Access-Control-Allow-Headers"] = "*"
            return response
        
        else:
            raise HTTPException(status_code=400, detail="不支持的格式，请使用 pdf 或 html")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成入库单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成入库单失败: {str(e)}")

# 客户管理API 已移至 routers/customers.py


# ==================== 业务员管理API ====================

@app.get("/api/salespersons")
async def get_salespersons(db: Session = Depends(get_db)):
    """获取所有业务员列表"""
    try:
        salespersons = db.query(Salesperson).filter(
            Salesperson.status == "active"
        ).order_by(Salesperson.id).all()
        
        return {
            "success": True,
            "salespersons": [SalespersonResponse.model_validate(s).model_dump(mode='json') for s in salespersons],
            "total": len(salespersons)
        }
    except Exception as e:
        logger.error(f"获取业务员列表失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.post("/api/salespersons")
async def create_salesperson(data: SalespersonCreate, db: Session = Depends(get_db)):
    """创建业务员"""
    try:
        # 检查是否已存在
        existing = db.query(Salesperson).filter(Salesperson.name == data.name).first()
        if existing:
            if existing.status == "inactive":
                # 重新激活
                existing.status = "active"
                db.commit()
                return {
                    "success": True,
                    "message": f"业务员【{data.name}】已重新激活",
                    "salesperson": SalespersonResponse.model_validate(existing).model_dump(mode='json')
                }
            return {
                "success": False,
                "message": f"业务员【{data.name}】已存在"
            }
        
        salesperson = Salesperson(
            name=data.name,
            phone=data.phone,
            remark=data.remark
        )
        db.add(salesperson)
        db.commit()
        db.refresh(salesperson)
        
        return {
            "success": True,
            "message": f"业务员【{data.name}】创建成功",
            "salesperson": SalespersonResponse.model_validate(salesperson).model_dump(mode='json')
        }
    except Exception as e:
        db.rollback()
        logger.error(f"创建业务员失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.put("/api/salespersons/{salesperson_id}")
async def update_salesperson(salesperson_id: int, data: SalespersonCreate, db: Session = Depends(get_db)):
    """更新业务员信息"""
    try:
        salesperson = db.query(Salesperson).filter(Salesperson.id == salesperson_id).first()
        if not salesperson:
            return {"success": False, "message": "业务员不存在"}
        
        # 检查新名字是否与其他业务员重复
        if data.name != salesperson.name:
            existing = db.query(Salesperson).filter(
                Salesperson.name == data.name,
                Salesperson.id != salesperson_id
            ).first()
            if existing:
                return {"success": False, "message": f"业务员【{data.name}】已存在"}
        
        salesperson.name = data.name
        if data.phone is not None:
            salesperson.phone = data.phone
        if data.remark is not None:
            salesperson.remark = data.remark
        
        db.commit()
        db.refresh(salesperson)
        
        return {
            "success": True,
            "message": f"业务员信息已更新",
            "salesperson": SalespersonResponse.model_validate(salesperson).model_dump(mode='json')
        }
    except Exception as e:
        db.rollback()
        logger.error(f"更新业务员失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.delete("/api/salespersons/{salesperson_id}")
async def delete_salesperson(salesperson_id: int, db: Session = Depends(get_db)):
    """删除业务员（软删除）"""
    try:
        salesperson = db.query(Salesperson).filter(Salesperson.id == salesperson_id).first()
        if not salesperson:
            return {"success": False, "message": "业务员不存在"}
        
        salesperson.status = "inactive"
        db.commit()
        
        return {
            "success": True,
            "message": f"业务员【{salesperson.name}】已删除"
        }
    except Exception as e:
        db.rollback()
        logger.error(f"删除业务员失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.post("/api/salespersons/init")
async def init_salespersons(db: Session = Depends(get_db)):
    """初始化业务员数据（清除现有数据并添加新数据）"""
    try:
        # 预定义的业务员列表
        salesperson_names = [
            "郑梅", "何云波", "姚财寿", "纪鸿杰", "郑光辉",
            "魏荔岚", "林纯洁", "赵燕珠", "步昭芬", "魏瑶峰"
        ]
        
        # 将所有现有业务员设为inactive
        db.query(Salesperson).update({"status": "inactive"})
        
        # 添加或激活新的业务员
        added = []
        for name in salesperson_names:
            existing = db.query(Salesperson).filter(Salesperson.name == name).first()
            if existing:
                existing.status = "active"
                added.append(name)
            else:
                salesperson = Salesperson(name=name, status="active")
                db.add(salesperson)
                added.append(name)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"业务员数据已初始化，共{len(added)}人",
            "salespersons": added
        }
    except Exception as e:
        db.rollback()
        logger.error(f"初始化业务员失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# 智能匹配API 已移至 routers/customers.py


# ==================== 销售单管理API ====================

@app.post("/api/sales/orders")
async def create_sales_order(order_data: SalesOrderCreate, db: Session = Depends(get_db)):
    """创建销售单"""
    try:
        # ==================== 数据验证 ====================
        # 验证商品明细数据
        for item in order_data.items:
            if item.weight <= 0:
                return {
                    "success": False,
                    "message": f"商品 {item.product_name} 的重量必须大于0",
                    "validation_error": {
                        "product_name": item.product_name,
                        "field": "weight",
                        "value": item.weight
                    }
                }
            if item.labor_cost < 0:
                return {
                    "success": False,
                    "message": f"商品 {item.product_name} 的工费不能为负数",
                    "validation_error": {
                        "product_name": item.product_name,
                        "field": "labor_cost",
                        "value": item.labor_cost
                    }
                }
        # ==================== 数据验证结束 ====================
        
        # ==================== 商品编码转换 ====================
        # 如果输入的是商品编码，自动转换为商品名称
        from .models import ProductCode as ProductCodeModel
        for item in order_data.items:
            product_name = item.product_name
            # 检查是否是商品编码（全大写或包含数字）
            if product_name and (product_name.isupper() or any(c.isdigit() for c in product_name)):
                code_record = db.query(ProductCodeModel).filter(ProductCodeModel.code == product_name).first()
                if code_record and code_record.name:
                    # 找到了对应的商品名称，更新 item
                    logger.info(f"商品编码转换: {product_name} -> {code_record.name}")
                    item.product_name = code_record.name
        # ==================== 商品编码转换结束 ====================
        
        # ==================== 库存检查 ====================
        # 在创建客户之前先检查库存，避免创建了客户但销售单创建失败
        inventory_errors = []
        for item in order_data.items:
            # 查询库存（精确匹配商品名称）
            inventory = db.query(Inventory).filter(
                Inventory.product_name == item.product_name
            ).first()
            
            if not inventory:
                # 商品不存在于库存中
                inventory_errors.append({
                    "product_name": item.product_name,
                    "error": "商品不存在于库存中",
                    "required_weight": item.weight,
                    "available_weight": 0.0
                })
            else:
                # 计算可用库存：总库存 - 待结算销售单占用的库存
                # 查询该商品在待结算销售单中的总重量
                reserved_weight = db.query(func.sum(SalesDetail.weight)).join(
                    SalesOrder
                ).filter(
                    SalesDetail.product_name == item.product_name,
                    SalesOrder.status == "待结算"
                ).scalar() or 0.0
                
                available_weight = inventory.total_weight - reserved_weight
                
                if available_weight < item.weight:
                    # 库存不足（考虑待结算的销售单）
                    inventory_errors.append({
                        "product_name": item.product_name,
                        "error": "库存不足",
                        "required_weight": item.weight,
                        "available_weight": available_weight,
                        "total_weight": inventory.total_weight,
                        "reserved_weight": reserved_weight
                    })
        
        # 如果有任何商品库存不足，拒绝创建销售单
        if inventory_errors:
            return {
                "success": False,
                "message": "库存检查失败，无法创建销售单",
                "inventory_errors": inventory_errors
            }
        # ==================== 库存检查结束 ====================
        
        # 处理客户（在库存检查通过后）
        customer_id = order_data.customer_id
        customer_name = order_data.customer_name
        
        # 如果没有提供customer_id，尝试根据姓名查找
        if not customer_id:
            customer = db.query(Customer).filter(
                Customer.name == customer_name,
                Customer.status == "active"
            ).first()
            if customer:
                customer_id = customer.id
            else:
                # 客户不存在，自动创建
                customer_no = f"KH{china_now().strftime('%Y%m%d%H%M%S')}"
                customer = Customer(
                    customer_no=customer_no,
                    name=customer_name,
                    customer_type="个人"
                )
                db.add(customer)
                db.flush()
                customer_id = customer.id
        
        # 计算总工费和总克重
        total_labor_cost = sum(item.labor_cost * item.weight for item in order_data.items)
        total_weight = sum(item.weight for item in order_data.items)
        
        # 生成销售单号（使用中国时间）
        order_no = f"XS{china_now().strftime('%Y%m%d%H%M%S')}"
        
        # 创建销售单
        sales_order = SalesOrder(
            order_no=order_no,
            order_date=order_data.order_date or datetime.now(),
            customer_id=customer_id,
            customer_name=customer_name,
            salesperson=order_data.salesperson,
            store_code=order_data.store_code,
            remark=order_data.remark,
            total_labor_cost=total_labor_cost,
            total_weight=total_weight,
            status="待结算"
        )
        db.add(sales_order)
        db.flush()
        
        # 创建销售明细
        details = []
        for item in order_data.items:
            detail = SalesDetail(
                order_id=sales_order.id,
                product_name=item.product_name,
                weight=item.weight,
                labor_cost=item.labor_cost,
                total_labor_cost=item.labor_cost * item.weight
            )
            db.add(detail)
            details.append(detail)
        
        # 更新客户统计信息
        if customer_id:
            customer = db.query(Customer).filter(Customer.id == customer_id).first()
            if customer:
                customer.total_purchase_amount += total_labor_cost
                customer.total_purchase_count += 1
                customer.last_purchase_time = sales_order.order_date
        
        db.commit()
        db.refresh(sales_order)
        for detail in details:
            db.refresh(detail)
        
        # 构建响应
        order_response = SalesOrderResponse.model_validate(sales_order)
        order_response.details = [SalesDetailResponse.model_validate(d).model_dump(mode='json') for d in details]
        
        return {
            "success": True,
            "message": f"销售单创建成功：{order_no}",
            "order": order_response.model_dump(mode='json')
        }
    
    except Exception as e:
        db.rollback()
        logger.error(f"创建销售单失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"创建销售单失败: {str(e)}"
        }

@app.get("/api/sales/orders")
async def get_sales_orders(
    customer_name: Optional[str] = None,
    salesperson: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取销售单列表"""
    try:
        query = db.query(SalesOrder)
        
        if customer_name:
            query = query.filter(SalesOrder.customer_name.contains(customer_name))
        if salesperson:
            query = query.filter(SalesOrder.salesperson == salesperson)
        if start_date:
            try:
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(SalesOrder.order_date >= start_dt)
            except:
                pass
        if end_date:
            try:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(SalesOrder.order_date <= end_dt)
            except:
                pass
        
        orders = query.order_by(desc(SalesOrder.order_date)).limit(100).all()
        
        # 加载明细
        result = []
        for order in orders:
            details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
            order_response = SalesOrderResponse.model_validate(order)
            order_response.details = [SalesDetailResponse.model_validate(d).model_dump(mode='json') for d in details]
            result.append(order_response.model_dump(mode='json'))
        
        return {
            "success": True,
            "orders": result
        }
    except Exception as e:
        logger.error(f"查询销售单失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"查询销售单失败: {str(e)}"
        }

@app.get("/api/sales/orders/{order_id}")
async def get_sales_order(order_id: int, db: Session = Depends(get_db)):
    """获取销售单详情"""
    try:
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
        
        if not order:
            return {
                "success": False,
                "message": "销售单不存在"
            }
        
        details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
        order_response = SalesOrderResponse.model_validate(order)
        order_response.details = [SalesDetailResponse.model_validate(d).model_dump(mode='json') for d in details]
        
        return {
            "success": True,
            "order": order_response.model_dump(mode='json')
        }
    except Exception as e:
        logger.error(f"查询销售单详情失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"查询销售单详情失败: {str(e)}"
        }

# ==================== AI对话处理函数 ====================

async def handle_create_supplier(ai_response, db: Session) -> Dict[str, Any]:
    """处理创建供应商操作"""
    try:
        # 从AI响应中提取供应商信息
        message = ai_response.model_dump() if hasattr(ai_response, 'model_dump') else {}
        
        # 尝试从AI响应中提取供应商名称
        supplier_name = message.get('supplier_name') or message.get('name')
        if not supplier_name:
            return {
                "success": False,
                "message": "未找到供应商名称，请提供供应商名称",
                "parsed": message
            }
        
        # 创建供应商
        supplier_data = SupplierCreate(
            name=supplier_name,
            phone=message.get('phone'),
            wechat=message.get('wechat'),
            address=message.get('address'),
            contact_person=message.get('contact_person'),
            supplier_type=message.get('supplier_type', '个人'),
            remark=message.get('remark')
        )
        
        result = await create_supplier(supplier_data, db)
        return result
    except Exception as e:
        logger.error(f"处理创建供应商失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理创建供应商失败: {str(e)}"
        }

async def handle_create_customer(ai_response, db: Session) -> Dict[str, Any]:
    """处理创建客户操作"""
    try:
        # 从AI响应中提取客户信息
        # 这里需要根据AI返回的格式来解析
        # 暂时使用简单的解析逻辑
        message = ai_response.model_dump() if hasattr(ai_response, 'model_dump') else {}
        
        # 尝试从AI响应中提取客户名称
        customer_name = message.get('customer_name') or message.get('name')
        if not customer_name:
            return {
                "success": False,
                "message": "未找到客户姓名，请提供客户姓名",
                "parsed": message
            }
        
        # 创建客户
        customer_data = CustomerCreate(
            name=customer_name,
            phone=message.get('phone'),
            wechat=message.get('wechat'),
            address=message.get('address'),
            customer_type=message.get('customer_type', '个人')
        )
        
        result = await create_customer(customer_data, db)
        return result
    except Exception as e:
        logger.error(f"处理创建客户失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理创建客户失败: {str(e)}"
        }

async def handle_query_customers(ai_response, db: Session) -> Dict[str, Any]:
    """处理查询客户操作"""
    try:
        # 从AI响应中提取查询条件
        message = ai_response.model_dump() if hasattr(ai_response, 'model_dump') else {}
        customer_name = message.get('customer_name') or message.get('name')
        
        result = await get_customers(name=customer_name, db=db)
        return result
    except Exception as e:
        logger.error(f"处理查询客户失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理查询客户失败: {str(e)}"
        }

async def handle_create_sales_order(ai_response, db: Session) -> Dict[str, Any]:
    """处理创建销售单操作"""
    try:
        # 从AI响应中提取销售单信息
        message = ai_response.model_dump() if hasattr(ai_response, 'model_dump') else {}
        
        # 提取必填字段
        customer_name = message.get('customer_name')
        salesperson = message.get('salesperson')
        items_data = message.get('items', [])
        
        if not customer_name:
            return {
                "success": False,
                "message": "未找到客户姓名，请提供客户姓名",
                "parsed": message
            }
        
        # ========== 智能匹配业务员 ==========
        auto_matched_salesperson = False
        if not salesperson:
            # 根据客户名查找历史销售记录，自动匹配业务员
            latest_order = db.query(SalesOrder).filter(
                SalesOrder.customer_name == customer_name.strip(),
                SalesOrder.status != "已取消"
            ).order_by(SalesOrder.create_time.desc()).first()
            
            if latest_order and latest_order.salesperson:
                salesperson = latest_order.salesperson
                auto_matched_salesperson = True
                logger.info(f"自动匹配业务员: 客户={customer_name} -> 业务员={salesperson}")
            else:
                return {
                    "success": False,
                    "message": f"未找到业务员姓名。这是新客户【{customer_name}】的第一笔订单，请提供业务员姓名",
                    "parsed": message,
                    "hint": "新客户需要指定业务员"
                }
        # ========== 智能匹配结束 ==========
        
        if not items_data or len(items_data) == 0:
            return {
                "success": False,
                "message": "未找到商品信息，请提供商品明细",
                "parsed": message
            }
        
        # 转换商品明细
        items = []
        for item in items_data:
            if isinstance(item, dict):
                try:
                    weight = float(item.get('weight', 0))
                    labor_cost = float(item.get('labor_cost', 0))
                    
                    # 验证数据
                    if weight <= 0:
                        return {
                            "success": False,
                            "message": f"商品 {item.get('product_name')} 的重量必须大于0",
                            "parsed": message
                        }
                    if labor_cost < 0:
                        return {
                            "success": False,
                            "message": f"商品 {item.get('product_name')} 的工费不能为负数",
                            "parsed": message
                        }
                    
                    items.append(SalesDetailItem(
                        product_name=item.get('product_name'),
                        weight=weight,
                        labor_cost=labor_cost
                    ))
                except (ValueError, TypeError) as e:
                    return {
                        "success": False,
                        "message": f"商品信息格式错误: {str(e)}",
                        "parsed": message
                    }
        
        if not items:
            return {
                "success": False,
                "message": "商品信息格式错误，未找到有效商品",
                "parsed": message
            }
        
        # 解析日期（安全处理）
        order_date = None
        if message.get('order_date'):
            try:
                order_date = datetime.fromisoformat(message['order_date'].replace('Z', '+00:00'))
            except (ValueError, AttributeError):
                # 如果日期格式错误，使用当前时间
                logger.warning(f"日期格式错误，使用当前时间: {message.get('order_date')}")
                order_date = None
        
        # 创建销售单
        order_data = SalesOrderCreate(
            order_date=order_date,
            customer_name=customer_name,
            customer_id=message.get('customer_id'),
            salesperson=salesperson,
            store_code=message.get('store_code'),
            remark=message.get('remark'),
            items=items
        )
        
        result = await create_sales_order(order_data, db)
        
        # 如果业务员是自动匹配的，在消息中添加提示
        if result.get('success') and auto_matched_salesperson:
            result['message'] = f"✅ {result.get('message', '销售单创建成功')}（业务员【{salesperson}】已根据历史记录自动匹配）"
            result['auto_matched_salesperson'] = True
        
        return result
    except Exception as e:
        logger.error(f"处理创建销售单失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"处理创建销售单失败: {str(e)}"
        }


async def handle_create_transfer(ai_response, db: Session) -> Dict[str, Any]:
    """处理创建库存转移单操作"""
    try:
        # 从AI响应中提取转移信息
        transfer_product_name = ai_response.transfer_product_name
        transfer_weight = ai_response.transfer_weight
        from_location_name = ai_response.from_location or "商品部仓库"  # 默认从商品部仓库发出
        to_location_name = ai_response.to_location
        
        # 验证目标位置
        if not to_location_name:
            return {
                "success": False,
                "message": "请指定目标位置（如：展厅）",
                "hint": "例如：帮我把古法戒指转移到展厅"
            }
        
        # 验证重量
        if not transfer_weight or transfer_weight <= 0:
            return {
                "success": False,
                "message": "请指定要转移的重量",
                "hint": "例如：帮我转移100克到展厅"
            }
        
        # 查找发出位置
        from_location = db.query(Location).filter(
            Location.name.contains(from_location_name)
        ).first()
        if not from_location:
            # 尝试模糊匹配
            from_location = db.query(Location).filter(
                Location.location_type == "warehouse"
            ).first()
        
        if not from_location:
            return {
                "success": False,
                "message": f"未找到发出位置：{from_location_name}",
                "hint": "请确保仓库位置已创建"
            }
        
        # 查找目标位置
        to_location = db.query(Location).filter(
            Location.name.contains(to_location_name)
        ).first()
        if not to_location:
            # 尝试模糊匹配展厅
            if "展厅" in to_location_name:
                to_location = db.query(Location).filter(
                    Location.location_type == "showroom"
                ).first()
        
        if not to_location:
            return {
                "success": False,
                "message": f"未找到目标位置：{to_location_name}",
                "hint": "可用的位置：商品部仓库、展厅"
            }
        
        # 如果没有指定商品名称，尝试从库存中查找
        if not transfer_product_name:
            # 查找该位置有库存的商品
            location_inv = db.query(LocationInventory).filter(
                LocationInventory.location_id == from_location.id,
                LocationInventory.weight >= transfer_weight
            ).first()
            
            if location_inv:
                transfer_product_name = location_inv.product_name
                logger.info(f"自动选择商品：{transfer_product_name}")
            else:
                return {
                    "success": False,
                    "message": "请指定要转移的商品名称，或确保发出位置有足够库存",
                    "hint": "例如：帮我把古法戒指 100克转移到展厅"
                }
        
        # 验证发出位置是否有足够库存
        source_inv = db.query(LocationInventory).filter(
            LocationInventory.location_id == from_location.id,
            LocationInventory.product_name == transfer_product_name
        ).first()
        
        if not source_inv or source_inv.weight < transfer_weight:
            available = source_inv.weight if source_inv else 0
            return {
                "success": False,
                "message": f"{from_location.name} 的 {transfer_product_name} 库存不足（当前：{available}克，需要：{transfer_weight}克）",
                "hint": "请先入库或减少转移数量"
            }
        
        # 生成转移单号
        now = datetime.now()
        count = db.query(InventoryTransfer).filter(
            InventoryTransfer.transfer_no.like(f"TR{now.strftime('%Y%m%d')}%")
        ).count()
        transfer_no = f"TR{now.strftime('%Y%m%d')}{count + 1:03d}"
        
        # 创建转移单
        db_transfer = InventoryTransfer(
            transfer_no=transfer_no,
            product_name=transfer_product_name,
            weight=transfer_weight,
            from_location_id=from_location.id,
            to_location_id=to_location.id,
            status="pending",
            created_by="AI助手",
            remark=f"通过AI对话创建"
        )
        db.add(db_transfer)
        
        # 扣减发出位置库存
        source_inv.weight -= transfer_weight
        
        db.commit()
        
        logger.info(f"创建转移单成功: {transfer_no}, {transfer_product_name} {transfer_weight}克, "
                   f"{from_location.name} -> {to_location.name}")
        
        return {
            "success": True,
            "message": f"✅ 转移单创建成功！\n\n"
                      f"📋 单号：{transfer_no}\n"
                      f"📦 商品：{transfer_product_name}\n"
                      f"⚖️ 重量：{transfer_weight}克\n"
                      f"📍 从：{from_location.name}\n"
                      f"📍 到：{to_location.name}\n"
                      f"⏳ 状态：待接收\n\n"
                      f"请到【分仓库存】页面查看和接收转移单。",
            "transfer_no": transfer_no,
            "action": "创建转移单"
        }
        
    except Exception as e:
        logger.error(f"处理创建转移单失败: {e}", exc_info=True)
        db.rollback()
        return {
            "success": False,
            "message": f"创建转移单失败: {str(e)}"
        }


async def handle_return(ai_response, db: Session, user_role: str) -> Dict[str, Any]:
    """处理退货操作（商品专员退给供应商，或柜台退给商品部）"""
    try:
        from .middleware.permissions import has_permission
        
        # 从AI响应中提取退货信息
        product_name = ai_response.product_name
        weight = ai_response.weight
        supplier_name = ai_response.supplier  # 供应商名称（退给供应商时需要）
        
        # 确定退货类型
        # 商品专员默认退给供应商，柜台默认退给商品部
        if user_role == 'product':
            return_type = "to_supplier"
        elif user_role == 'counter':
            return_type = "to_warehouse"
        else:
            # 根据AI解析结果判断
            if supplier_name:
                return_type = "to_supplier"
            else:
                return_type = "to_warehouse"
        
        # 权限检查
        if return_type == "to_supplier" and not has_permission(user_role, 'can_return_to_supplier'):
            return {
                "success": False,
                "message": "权限不足：您没有【退货给供应商】的权限"
            }
        if return_type == "to_warehouse" and not has_permission(user_role, 'can_return_to_warehouse'):
            return {
                "success": False,
                "message": "权限不足：您没有【退货给商品部】的权限"
            }
        
        # 验证必填信息 - 如果信息不完整，触发前端弹出表单
        if not product_name or not weight or weight <= 0:
            return {
                "success": False,
                "action": "退货",
                "need_form": True,  # 告诉前端需要弹出表单
                "message": "📝 请在弹出的表格中填写退货信息",
                "hint": "请提供商品名称、重量和供应商"
            }
        
        # 根据用户角色确定发起位置
        if user_role == 'product':
            # 商品专员从仓库发起
            from_location = db.query(Location).filter(Location.code == "warehouse").first()
        elif user_role == 'counter':
            # 柜台从展厅发起
            from_location = db.query(Location).filter(Location.code == "showroom").first()
        else:
            # 默认从仓库发起
            from_location = db.query(Location).filter(Location.code == "warehouse").first()
        
        if not from_location:
            return {
                "success": False,
                "message": "未找到发起位置，请联系管理员初始化位置信息"
            }
        
        # 检查库存
        inventory = db.query(LocationInventory).filter(
            LocationInventory.location_id == from_location.id,
            LocationInventory.product_name == product_name
        ).first()
        
        if not inventory or inventory.weight < weight:
            available = inventory.weight if inventory else 0
            return {
                "success": False,
                "message": f"库存不足：{from_location.name} 的 {product_name} 仅有 {available}g，无法退货 {weight}g"
            }
        
        # 退给供应商时需要找到供应商
        supplier_id = None
        if return_type == "to_supplier":
            if not supplier_name:
                return {
                    "success": False,
                    "message": "退给供应商时请提供供应商名称",
                    "hint": "例如：退货古法戒指10克给金源珠宝"
                }
            
            # 查找供应商
            supplier = db.query(Supplier).filter(
                Supplier.name.contains(supplier_name),
                Supplier.status == "active"
            ).first()
            
            if not supplier:
                return {
                    "success": False,
                    "message": f"未找到供应商：{supplier_name}，请先创建供应商或检查名称"
                }
            supplier_id = supplier.id
        
        # 生成退货单号
        now = china_now()
        count = db.query(ReturnOrder).filter(
            ReturnOrder.return_no.like(f"TH{now.strftime('%Y%m%d')}%")
        ).count()
        return_no = f"TH{now.strftime('%Y%m%d')}{count + 1:03d}"
        
        # 创建退货单（直接完成，无需审批）
        return_order = ReturnOrder(
            return_no=return_no,
            return_type=return_type,
            product_name=product_name,
            return_weight=weight,
            from_location_id=from_location.id,
            supplier_id=supplier_id,
            return_reason="其他",  # 默认原因
            status="completed",
            created_by=user_role,
            completed_by=user_role,
            completed_at=now
        )
        db.add(return_order)
        
        # 扣减库存
        inventory.weight -= weight
        
        # 更新总库存
        from .models import Inventory
        total_inventory = db.query(Inventory).filter(Inventory.product_name == product_name).first()
        if total_inventory:
            total_inventory.total_weight -= weight
        
        # 如果是退给供应商，更新供应商统计
        if return_type == "to_supplier" and supplier_id:
            supplier.total_supply_weight -= weight
            if supplier.total_supply_count > 0:
                supplier.total_supply_count -= 1
        
        db.commit()
        
        # 构建返回消息
        if return_type == "to_supplier":
            message = f"✅ 退货成功！\n\n" \
                     f"📋 退货单号：{return_no}\n" \
                     f"📦 商品：{product_name}\n" \
                     f"⚖️ 退货克重：{weight}g\n" \
                     f"🏭 退给供应商：{supplier_name}\n" \
                     f"📍 从：{from_location.name}\n" \
                     f"✅ 库存已扣减"
        else:
            message = f"✅ 退货成功！\n\n" \
                     f"📋 退货单号：{return_no}\n" \
                     f"📦 商品：{product_name}\n" \
                     f"⚖️ 退货克重：{weight}g\n" \
                     f"📍 从：{from_location.name} 退回商品部\n" \
                     f"✅ 库存已扣减"
        
        logger.info(f"[退货] 退货成功: {return_no}, {product_name} {weight}g")
        
        return {
            "success": True,
            "message": message,
            "return_no": return_no,
            "action": "退货"
        }
        
    except Exception as e:
        logger.error(f"处理退货失败: {e}", exc_info=True)
        db.rollback()
        return {
            "success": False,
            "message": f"退货失败: {str(e)}"
        }


async def handle_payment_registration(ai_response, db: Session) -> Dict[str, Any]:
    """处理登记收款：返回确认数据供前端显示确认卡片"""
    try:
        customer_name = ai_response.payment_customer_name
        amount = ai_response.payment_amount
        payment_method = ai_response.payment_method or "转账"
        remark = ai_response.payment_remark or ""
        
        # 验证必填信息
        if not customer_name:
            return {
                "success": False,
                "message": "请提供收款的客户名称，例如：张老板收到5000元"
            }
        
        if not amount or amount <= 0:
            return {
                "success": False,
                "message": "请提供收款金额，例如：张老板收到5000元"
            }
        
        # 模糊查询客户
        from .models import Customer
        from .models.finance import AccountReceivable
        from sqlalchemy import func
        
        customer = db.query(Customer).filter(
            Customer.name.ilike(f"%{customer_name}%")
        ).first()
        
        if not customer:
            return {
                "success": False,
                "message": f"未找到名称包含【{customer_name}】的客户，请确认客户名称"
            }
        
        # 查询客户当前欠款
        total_debt = db.query(func.sum(AccountReceivable.unpaid_amount)).filter(
            AccountReceivable.customer_id == customer.id,
            AccountReceivable.status.in_(["unpaid", "overdue"])
        ).scalar() or 0
        
        # 查询未付清的应收账款明细
        unpaid_receivables = db.query(AccountReceivable).filter(
            AccountReceivable.customer_id == customer.id,
            AccountReceivable.status.in_(["unpaid", "overdue"]),
            AccountReceivable.unpaid_amount > 0
        ).order_by(AccountReceivable.credit_start_date.asc()).all()
        
        receivables_list = []
        for r in unpaid_receivables:
            receivables_list.append({
                "id": r.id,
                "sales_order_id": r.sales_order_id,
                "total_amount": r.total_amount,
                "received_amount": r.received_amount,
                "unpaid_amount": r.unpaid_amount,
                "credit_start_date": r.credit_start_date.strftime("%Y-%m-%d") if r.credit_start_date else None
            })
        
        # 计算收款后余额
        balance_after = max(0, total_debt - amount)
        
        return {
            "success": True,
            "action": "登记收款",
            "confirm_required": True,
            "customer": {
                "id": customer.id,
                "name": customer.name,
                "customer_no": customer.customer_no,
                "phone": customer.phone
            },
            "current_debt": round(total_debt, 2),
            "payment_amount": round(amount, 2),
            "balance_after": round(balance_after, 2),
            "payment_method": payment_method,
            "remark": remark,
            "receivables": receivables_list,
            "message": f"确认为【{customer.name}】登记收款 ¥{amount:.2f}？"
        }
        
    except Exception as e:
        logger.error(f"处理登记收款失败: {e}", exc_info=True)
        return {
            "success": False,
            "message": f"登记收款失败: {str(e)}"
        }


# handle_query_sales_orders - 已由AI分析引擎替代


# ============= 统计分析 API =============

@app.get("/api/analytics/overview")
async def get_analytics_overview(db: Session = Depends(get_db)):
    """获取对话分析概览数据"""
    try:
        from sqlalchemy import func, case, distinct
        from datetime import timedelta
        
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # 总对话数
        total_chats = db.query(func.count(ChatLog.id)).filter(
            ChatLog.message_type == "user"
        ).scalar() or 0
        
        # 本周对话数
        week_chats = db.query(func.count(ChatLog.id)).filter(
            ChatLog.message_type == "user",
            func.date(ChatLog.created_at) >= week_ago
        ).scalar() or 0
        
        # 今日对话数
        today_chats = db.query(func.count(ChatLog.id)).filter(
            ChatLog.message_type == "user",
            func.date(ChatLog.created_at) == today
        ).scalar() or 0
        
        # 按角色统计
        role_stats = db.query(
            ChatLog.user_role,
            func.count(ChatLog.id).label("count")
    ).filter(
            ChatLog.message_type == "user"
        ).group_by(ChatLog.user_role).all()
        
        role_distribution = {stat.user_role: stat.count for stat in role_stats}
        
        # 按意图统计（Top 10）
        intent_stats = db.query(
            ChatLog.intent,
            func.count(ChatLog.id).label("count")
        ).filter(
            ChatLog.message_type == "assistant",
            ChatLog.intent.isnot(None)
        ).group_by(ChatLog.intent).order_by(
            func.count(ChatLog.id).desc()
        ).limit(10).all()
        
        intent_distribution = [{"intent": stat.intent, "count": stat.count} for stat in intent_stats]
        
        # 平均响应时间
        avg_response_time = db.query(
            func.avg(ChatLog.response_time_ms)
        ).filter(
            ChatLog.message_type == "assistant",
            ChatLog.response_time_ms.isnot(None)
        ).scalar() or 0
        
        # 成功率
        total_responses = db.query(func.count(ChatLog.id)).filter(
            ChatLog.message_type == "assistant"
        ).scalar() or 1
        successful_responses = db.query(func.count(ChatLog.id)).filter(
            ChatLog.message_type == "assistant",
            ChatLog.is_successful == 1
        ).scalar() or 0
        success_rate = (successful_responses / total_responses) * 100 if total_responses > 0 else 100
        
        return {
            "success": True,
            "data": {
                "total_chats": total_chats,
                "week_chats": week_chats,
                "today_chats": today_chats,
                "role_distribution": role_distribution,
                "intent_distribution": intent_distribution,
                "avg_response_time_ms": round(avg_response_time, 2),
                "success_rate": round(success_rate, 2)
            }
        }
    except Exception as e:
        logger.error(f"获取分析概览失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/analytics/role/{role}")
async def get_role_analytics(role: str, db: Session = Depends(get_db)):
    """获取特定角色的详细分析数据"""
    try:
        from sqlalchemy import func
        from datetime import timedelta
        
        today = datetime.now().date()
        week_ago = today - timedelta(days=7)
        
        # 该角色的总对话数
        total_chats = db.query(func.count(ChatLog.id)).filter(
            ChatLog.user_role == role,
            ChatLog.message_type == "user"
        ).scalar() or 0
        
        # 本周对话数
        week_chats = db.query(func.count(ChatLog.id)).filter(
            ChatLog.user_role == role,
            ChatLog.message_type == "user",
            func.date(ChatLog.created_at) >= week_ago
        ).scalar() or 0
        
        # 该角色最常用的意图（Top 5）
        top_intents = db.query(
            ChatLog.intent,
            func.count(ChatLog.id).label("count")
        ).filter(
            ChatLog.user_role == role,
            ChatLog.message_type == "assistant",
            ChatLog.intent.isnot(None)
        ).group_by(ChatLog.intent).order_by(
            func.count(ChatLog.id).desc()
        ).limit(5).all()
        
        # 最近的对话记录（最新10条）
        recent_chats = db.query(ChatLog).filter(
            ChatLog.user_role == role,
            ChatLog.message_type == "user"
        ).order_by(ChatLog.created_at.desc()).limit(10).all()
        
        # 热门关键词提取（简单实现：从内容中提取）
        all_contents = db.query(ChatLog.content).filter(
            ChatLog.user_role == role,
            ChatLog.message_type == "user"
        ).limit(100).all()
        
        # 简单的关键词统计
        keyword_counts = {}
        keywords_to_track = ["库存", "入库", "供应商", "客户", "销售", "订单", "对账", "收款", "图表"]
        for content_row in all_contents:
            content = content_row.content or ""
            for keyword in keywords_to_track:
                if keyword in content:
                    keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1
        
        hot_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        
        return {
            "success": True,
            "data": {
                "role": role,
                "total_chats": total_chats,
                "week_chats": week_chats,
                "top_intents": [{"intent": t.intent, "count": t.count} for t in top_intents],
                "hot_keywords": [{"keyword": k, "count": c} for k, c in hot_keywords],
                "recent_chats": [
                    {
                        "content": chat.content[:100] if chat.content else "",
                        "created_at": chat.created_at.isoformat() if chat.created_at else None
                    }
                    for chat in recent_chats
                ]
            }
        }
    except Exception as e:
        logger.error(f"获取角色分析失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/analytics/daily")
async def get_daily_analytics(days: int = 7, db: Session = Depends(get_db)):
    """获取每日对话趋势数据"""
    try:
        from sqlalchemy import func
        from datetime import timedelta
        
        today = datetime.now().date()
        start_date = today - timedelta(days=days)
        
        # 按日期和角色统计
        daily_stats = db.query(
            func.date(ChatLog.created_at).label("date"),
            ChatLog.user_role,
            func.count(ChatLog.id).label("count")
    ).filter(
            ChatLog.message_type == "user",
            func.date(ChatLog.created_at) >= start_date
        ).group_by(
            func.date(ChatLog.created_at),
            ChatLog.user_role
        ).order_by(func.date(ChatLog.created_at)).all()
        
        # 整理数据
        daily_data = {}
        for stat in daily_stats:
            date_str = str(stat.date)
            if date_str not in daily_data:
                daily_data[date_str] = {"date": date_str, "total": 0, "by_role": {}}
            daily_data[date_str]["total"] += stat.count
            daily_data[date_str]["by_role"][stat.user_role] = stat.count
        
        return {
            "success": True,
            "data": list(daily_data.values())
        }
    except Exception as e:
        logger.error(f"获取每日分析失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# ============= 数据导出 API =============

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import zipfile

def create_excel_response(wb: Workbook, filename: str):
    """创建 Excel 文件响应"""
    import urllib.parse
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 使用 RFC 5987 规范对中文文件名进行 URL 编码
    encoded_filename = urllib.parse.quote(filename)
    
    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Allow-Origin": "*",
        }
    )

def style_header(ws, row=1):
    """为表头添加样式"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def auto_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@app.get("/api/export/chat-logs")
async def export_chat_logs(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出对话日志为 Excel"""
    try:
        query = db.query(ChatLog).order_by(ChatLog.created_at.desc())
        
        if start_date:
            query = query.filter(func.date(ChatLog.created_at) >= start_date)
        if end_date:
            query = query.filter(func.date(ChatLog.created_at) <= end_date)
        
        logs = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "对话日志"
        
        # 表头
        headers = ["ID", "会话ID", "用户角色", "消息类型", "内容", "意图", "响应时间(ms)", "是否成功", "创建时间"]
        ws.append(headers)
        style_header(ws)
        
        # 数据
        role_names = {"sales": "业务员", "finance": "财务", "product": "商品专员", "manager": "管理层"}
        for log in logs:
            ws.append([
                log.id,
                log.session_id,
                role_names.get(log.user_role, log.user_role),
                "用户" if log.message_type == "user" else "AI助手",
                log.content[:500] if log.content else "",
                log.intent or "",
                log.response_time_ms,
                "是" if log.is_successful else "否",
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"对话日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出对话日志失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ============= 聊天历史回溯 API =============

@app.get("/api/chat-sessions")
async def get_chat_sessions(
    user_role: str = None,
    limit: int = 30,
    db: Session = Depends(get_db)
):
    """获取对话会话列表（按会话分组）- 用于历史回溯"""
    try:
        # 查询所有不同的会话，按最新消息时间排序
        subquery = db.query(
            ChatLog.session_id,
            func.min(ChatLog.created_at).label('start_time'),
            func.max(ChatLog.created_at).label('end_time'),
            func.count(ChatLog.id).label('message_count')
        ).group_by(ChatLog.session_id)
        
        if user_role:
            subquery = subquery.filter(ChatLog.user_role == user_role)
        
        subquery = subquery.order_by(desc(func.max(ChatLog.created_at))).limit(limit)
        sessions = subquery.all()
        
        result = []
        for session in sessions:
            # 获取该会话的第一条用户消息作为摘要
            first_msg = db.query(ChatLog).filter(
                ChatLog.session_id == session.session_id,
                ChatLog.message_type == "user"
            ).order_by(ChatLog.created_at).first()
            
            # 获取最后一条消息的意图
            last_msg = db.query(ChatLog).filter(
                ChatLog.session_id == session.session_id
            ).order_by(desc(ChatLog.created_at)).first()
            
            summary = ""
            if first_msg and first_msg.content:
                summary = first_msg.content[:60] + "..." if len(first_msg.content) > 60 else first_msg.content
            
            # 查询会话的自定义名称
            from .models import ChatSessionMeta
            session_meta = db.query(ChatSessionMeta).filter(
                ChatSessionMeta.session_id == session.session_id
            ).first()
            
            custom_name = session_meta.custom_name if session_meta else None
            is_pinned = session_meta.is_pinned if session_meta else 0
            
            result.append({
                "session_id": session.session_id,
                "start_time": session.start_time.isoformat() if session.start_time else None,
                "end_time": session.end_time.isoformat() if session.end_time else None,
                "message_count": session.message_count,
                "summary": summary,
                "custom_name": custom_name,  # 用户自定义名称
                "is_pinned": is_pinned,  # 是否置顶
                "last_intent": last_msg.intent if last_msg else None,
                "user_role": last_msg.user_role if last_msg else None
            })
        
        # 按置顶状态排序（置顶的在前）
        result.sort(key=lambda x: (-x.get('is_pinned', 0), x.get('start_time', '') or ''), reverse=False)
        result.sort(key=lambda x: -x.get('is_pinned', 0))
        
        return {"success": True, "sessions": result, "total": len(result)}
    except Exception as e:
        logger.error(f"获取会话列表失败: {e}", exc_info=True)
        return {"success": False, "message": str(e), "sessions": []}


@app.put("/api/chat-sessions/{session_id}/rename")
async def rename_chat_session(
    session_id: str,
    name: str = Query(..., description="新的会话名称"),
    db: Session = Depends(get_db)
):
    """重命名对话会话"""
    try:
        from .models import ChatSessionMeta
        
        # 查找或创建会话元数据
        session_meta = db.query(ChatSessionMeta).filter(
            ChatSessionMeta.session_id == session_id
        ).first()
        
        if session_meta:
            # 更新现有记录
            session_meta.custom_name = name.strip() if name.strip() else None
        else:
            # 创建新记录
            session_meta = ChatSessionMeta(
                session_id=session_id,
                custom_name=name.strip() if name.strip() else None
            )
            db.add(session_meta)
        
        db.commit()
        
        return {
            "success": True,
            "message": "会话重命名成功",
            "session_id": session_id,
            "custom_name": session_meta.custom_name
        }
    except Exception as e:
        db.rollback()
        logger.error(f"重命名会话失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.put("/api/chat-sessions/{session_id}/pin")
async def toggle_pin_chat_session(
    session_id: str,
    pinned: bool = Query(..., description="是否置顶"),
    db: Session = Depends(get_db)
):
    """置顶/取消置顶对话会话"""
    try:
        from .models import ChatSessionMeta
        
        # 查找或创建会话元数据
        session_meta = db.query(ChatSessionMeta).filter(
            ChatSessionMeta.session_id == session_id
        ).first()
        
        if session_meta:
            session_meta.is_pinned = 1 if pinned else 0
        else:
            session_meta = ChatSessionMeta(
                session_id=session_id,
                is_pinned=1 if pinned else 0
            )
            db.add(session_meta)
        
        db.commit()
        
        return {
            "success": True,
            "message": "置顶已" + ("设置" if pinned else "取消"),
            "session_id": session_id,
            "is_pinned": session_meta.is_pinned
        }
    except Exception as e:
        db.rollback()
        logger.error(f"置顶会话失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.delete("/api/chat-sessions/{session_id}")
async def delete_chat_session(
    session_id: str,
    db: Session = Depends(get_db)
):
    """删除对话会话"""
    try:
        from .models import ChatSessionMeta
        
        # 删除会话的所有消息
        deleted_logs = db.query(ChatLog).filter(
            ChatLog.session_id == session_id
        ).delete()
        
        # 删除会话元数据
        db.query(ChatSessionMeta).filter(
            ChatSessionMeta.session_id == session_id
        ).delete()
        
        db.commit()
        
        return {
            "success": True,
            "message": f"会话已删除，共删除 {deleted_logs} 条消息",
            "session_id": session_id
        }
    except Exception as e:
        db.rollback()
        logger.error(f"删除会话失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/chat-history/{session_id}")
async def get_chat_history(
    session_id: str,
    db: Session = Depends(get_db)
):
    """获取指定会话的完整对话历史"""
    try:
        logs = db.query(ChatLog).filter(
            ChatLog.session_id == session_id
        ).order_by(ChatLog.created_at).all()
        
        if not logs:
            return {"success": False, "message": "未找到该会话", "messages": []}
        
        messages = []
        for log in logs:
            messages.append({
                "id": log.id,
                "message_type": log.message_type,
                "content": log.content,
                "intent": log.intent,
                "user_role": log.user_role,
                "is_successful": log.is_successful,
                "response_time_ms": log.response_time_ms,
                "created_at": log.created_at.isoformat() if log.created_at else None
            })
        
        return {
            "success": True,
            "session_id": session_id,
            "messages": messages,
            "total": len(messages)
        }
    except Exception as e:
        logger.error(f"获取对话历史失败: {e}", exc_info=True)
        return {"success": False, "message": str(e), "messages": []}


@app.post("/api/chat-logs/message")
async def save_chat_message(
    session_id: str,
    message_type: str,  # 'user' or 'assistant'
    content: str,
    user_role: str = 'sales',
    intent: str = None,
    db: Session = Depends(get_db)
):
    """保存单条聊天消息到历史记录 - 用于快捷操作后记录操作日志"""
    try:
        log_chat_message(
            db=db,
            session_id=session_id,
            user_role=user_role,
            message_type=message_type,
            content=content[:10000] if content else "",
            intent=intent,
            response_time_ms=None,
            is_successful=True
        )
        return {"success": True, "message": "消息保存成功"}
    except Exception as e:
        logger.error(f"保存消息失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/chat-logs/search")
async def search_chat_logs(
    keyword: str = None,
    user_role: str = None,
    intent: str = None,
    start_date: str = None,
    end_date: str = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """搜索聊天记录 - 支持关键词、角色、意图、日期范围筛选"""
    try:
        query = db.query(ChatLog).order_by(desc(ChatLog.created_at))
        
        if keyword:
            query = query.filter(ChatLog.content.contains(keyword))
        if user_role:
            query = query.filter(ChatLog.user_role == user_role)
        if intent:
            query = query.filter(ChatLog.intent == intent)
        if start_date:
            query = query.filter(func.date(ChatLog.created_at) >= start_date)
        if end_date:
            query = query.filter(func.date(ChatLog.created_at) <= end_date)
        
        total = query.count()
        logs = query.offset(offset).limit(limit).all()
        
        return {
            "success": True,
            "total": total,
            "logs": [
                {
                    "id": log.id,
                    "session_id": log.session_id,
                    "user_role": log.user_role,
                    "message_type": log.message_type,
                    "content": log.content[:200] + "..." if log.content and len(log.content) > 200 else log.content,
                    "intent": log.intent,
                    "created_at": log.created_at.isoformat() if log.created_at else None,
                    "is_successful": log.is_successful
                }
                for log in logs
            ]
        }
    except Exception as e:
        logger.error(f"搜索聊天记录失败: {e}", exc_info=True)
        return {"success": False, "message": str(e), "logs": [], "total": 0}


# 供应商管理 API 已移至 routers/suppliers.py


# ============= 上下文工程 API =============

@app.get("/api/context/{session_id}")
async def get_session_context(session_id: str):
    """获取会话上下文"""
    try:
        context = ctx.load_session_context(session_id)
        summary = ctx.generate_context_summary(session_id)
        return {
            "success": True,
            "context": context,
            "summary": summary
        }
    except Exception as e:
        logger.error(f"获取上下文失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.post("/api/context/{session_id}/goal")
async def set_session_goal(session_id: str, goal: str, phases: List[str] = None):
    """设置会话目标和阶段"""
    try:
        context = ctx.update_session_goal(session_id, goal, phases)
        return {
            "success": True,
            "message": f"目标已设置: {goal}",
            "context": context
        }
    except Exception as e:
        logger.error(f"设置目标失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.post("/api/context/{session_id}/note")
async def add_session_note(session_id: str, note: str, category: str = "general"):
    """添加会话笔记"""
    try:
        context = ctx.add_note(session_id, note, category)
        return {
            "success": True,
            "message": "笔记已添加",
            "notes_count": len(context.get("notes", []))
        }
    except Exception as e:
        logger.error(f"添加笔记失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.delete("/api/context/{session_id}")
async def clear_session_context(session_id: str):
    """清除会话上下文"""
    try:
        ctx.clear_session(session_id)
        return {"success": True, "message": f"会话 {session_id} 的上下文已清除"}
    except Exception as e:
        logger.error(f"清除上下文失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/context/list/all")
async def list_all_contexts():
    """列出所有会话上下文"""
    try:
        sessions = ctx.list_sessions()
        contexts = []
        for sid in sessions:
            c = ctx.load_session_context(sid)
            contexts.append({
                "session_id": sid,
                "goal": c.get("goal"),
                "last_updated": c.get("last_updated"),
                "actions_count": len(c.get("completed_actions", [])),
                "errors_count": len(c.get("errors", []))
            })
        return {
            "success": True,
            "sessions": contexts,
            "total": len(contexts)
        }
    except Exception as e:
        logger.error(f"列出上下文失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/knowledge-base")
async def get_knowledge_base():
    """获取业务知识库内容"""
    try:
        knowledge = ctx.load_knowledge_base()
        return {
            "success": True,
            "content": knowledge,
            "length": len(knowledge)
        }
    except Exception as e:
        logger.error(f"获取知识库失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/export/inventory")
async def export_inventory(db: Session = Depends(get_db)):
    """导出库存数据为 Excel"""
    try:
        inventories = db.query(Inventory).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "库存数据"
        
        headers = ["ID", "商品名称", "库存重量(克)", "最后更新时间"]
        ws.append(headers)
        style_header(ws)
        
        for inv in inventories:
            ws.append([
                inv.id,
                inv.product_name,
                inv.total_weight,
                inv.last_update.strftime("%Y-%m-%d %H:%M:%S") if inv.last_update else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"库存数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出库存数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/inbound")
async def export_inbound(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出入库记录为 Excel"""
    try:
        query = db.query(InboundOrder).order_by(InboundOrder.create_time.desc())
        
        if start_date:
            query = query.filter(func.date(InboundOrder.create_time) >= start_date)
        if end_date:
            query = query.filter(func.date(InboundOrder.create_time) <= end_date)
        
        orders = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "入库记录"
        
        headers = ["入库单号", "商品名称", "重量(克)", "工费(元/克)", "总成本(元)", "供应商", "入库时间", "操作员"]
        ws.append(headers)
        style_header(ws)
        
        for order in orders:
            details = db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all()
            for detail in details:
                ws.append([
                    order.order_no,
                    detail.product_name,
                    detail.weight,
                    detail.labor_cost,
                    detail.total_cost,
                    detail.supplier or "",
                    order.create_time.strftime("%Y-%m-%d %H:%M:%S") if order.create_time else "",
                    order.operator or ""
                ])
        
        auto_column_width(ws)
        
        filename = f"入库记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出入库记录失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/sales")
async def export_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出销售订单为 Excel"""
    try:
        query = db.query(SalesOrder).order_by(SalesOrder.order_date.desc())
        
        if start_date:
            query = query.filter(func.date(SalesOrder.order_date) >= start_date)
        if end_date:
            query = query.filter(func.date(SalesOrder.order_date) <= end_date)
        
        orders = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "销售订单"
        
        headers = ["订单号", "客户名称", "商品名称", "重量(克)", "工费(元/克)", "总工费(元)", "业务员", "门店代码", "订单日期"]
        ws.append(headers)
        style_header(ws)
        
        for order in orders:
            details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
            customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
            for detail in details:
                ws.append([
                    order.order_no,
                    customer.name if customer else "",
                    detail.product_name,
                    detail.weight,
                    detail.labor_cost,
                    detail.total_labor_cost,
                    order.salesperson or "",
                    order.store_code or "",
                    order.order_date.strftime("%Y-%m-%d") if order.order_date else ""
                ])
        
        auto_column_width(ws)
        
        filename = f"销售订单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出销售订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/customers")
async def export_customers(db: Session = Depends(get_db)):
    """导出客户列表为 Excel"""
    try:
        customers = db.query(Customer).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "客户列表"
        
        headers = ["ID", "客户编号", "客户名称", "电话", "地址", "创建时间"]
        ws.append(headers)
        style_header(ws)
        
        for cust in customers:
            ws.append([
                cust.id,
                cust.customer_no,
                cust.name,
                cust.phone or "",
                cust.address or "",
                cust.create_time.strftime("%Y-%m-%d %H:%M:%S") if cust.create_time else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"客户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出客户列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/suppliers")
async def export_suppliers(db: Session = Depends(get_db)):
    """导出供应商列表为 Excel"""
    try:
        suppliers = db.query(Supplier).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商列表"
        
        headers = ["ID", "供应商编号", "供应商名称", "联系人", "电话", "地址", "总供货金额(元)", "总供货重量(克)", "供货次数", "最后供货时间", "状态"]
        ws.append(headers)
        style_header(ws)
        
        for sup in suppliers:
            ws.append([
                sup.id,
                sup.supplier_no,
                sup.name,
                sup.contact_person or "",
                sup.phone or "",
                sup.address or "",
                sup.total_supply_amount or 0,
                sup.total_supply_weight or 0,
                sup.total_supply_count or 0,
                sup.last_supply_time.strftime("%Y-%m-%d %H:%M:%S") if sup.last_supply_time else "",
                "活跃" if sup.status == "active" else "停用"
            ])
        
        auto_column_width(ws)
        
        filename = f"供应商列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出供应商列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/customer-transactions/{customer_id}")
async def export_customer_transactions(
    customer_id: int,
    db: Session = Depends(get_db)
):
    """导出客户账务明细为 Excel（包含销售记录和往来明细）"""
    try:
        # 获取客户信息
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        wb = Workbook()
        
        # ========== Sheet 1: 销售记录 ==========
        ws1 = wb.active
        ws1.title = "销售记录"
        
        headers1 = ["销售单号", "销售日期", "商品名称", "重量(克)", "工费(元/克)", "总工费(元)", "业务员", "状态"]
        ws1.append(headers1)
        style_header(ws1)
        
        # 查询该客户的销售记录
        sales_orders = db.query(SalesOrder).filter(
            SalesOrder.customer_name == customer.name,
            SalesOrder.status != "已取消"
        ).order_by(SalesOrder.order_date.desc()).all()
        
        total_sales_weight = 0
        total_sales_amount = 0
        
        for order in sales_orders:
            details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
            for detail in details:
                ws1.append([
                    order.order_no,
                    order.order_date.strftime("%Y-%m-%d") if order.order_date else "",
                    detail.product_name,
                    detail.weight,
                    detail.labor_cost,
                    detail.total_labor_cost,
                    order.salesperson or "",
                    order.status
                ])
                total_sales_weight += detail.weight or 0
                total_sales_amount += detail.total_labor_cost or 0
        
        # 添加汇总行
        ws1.append([])
        ws1.append(["汇总", "", "", f"{total_sales_weight:.2f}", "", f"{total_sales_amount:.2f}", "", ""])
        
        auto_column_width(ws1)
        
        # ========== Sheet 2: 往来账目明细 ==========
        ws2 = wb.create_sheet(title="往来账目")
        
        headers2 = ["交易日期", "交易类型", "金额(元)", "金料重量(克)", "欠料余额(克)", "备注"]
        ws2.append(headers2)
        style_header(ws2)
        
        # 查询往来账记录
        from .models import CustomerTransaction, CustomerGoldDepositTransaction, AccountReceivable
        
        transactions = db.query(CustomerTransaction).filter(
            CustomerTransaction.customer_id == customer_id
        ).order_by(CustomerTransaction.created_at.desc()).all()
        
        type_labels = {
            "sales": "销售",
            "settlement": "结算",
            "gold_receipt": "收料",
            "payment": "付款"
        }
        
        for tx in transactions:
            ws2.append([
                tx.created_at.strftime("%Y-%m-%d %H:%M") if tx.created_at else "",
                type_labels.get(tx.transaction_type, tx.transaction_type),
                tx.amount or 0,
                tx.gold_weight or 0,
                tx.gold_due_after or 0,
                tx.remark or ""
            ])
        
        auto_column_width(ws2)
        
        # ========== Sheet 3: 存料记录 ==========
        ws3 = wb.create_sheet(title="存料记录")
        
        headers3 = ["交易日期", "交易类型", "金额(克)", "交易前余额(克)", "交易后余额(克)", "备注"]
        ws3.append(headers3)
        style_header(ws3)
        
        deposit_txs = db.query(CustomerGoldDepositTransaction).filter(
            CustomerGoldDepositTransaction.customer_id == customer_id
        ).order_by(CustomerGoldDepositTransaction.created_at.desc()).all()
        
        deposit_type_labels = {
            "deposit": "存入",
            "use": "使用",
            "refund": "退还"
        }
        
        for tx in deposit_txs:
            ws3.append([
                tx.created_at.strftime("%Y-%m-%d %H:%M") if tx.created_at else "",
                deposit_type_labels.get(tx.transaction_type, tx.transaction_type),
                tx.amount or 0,
                tx.balance_before or 0,
                tx.balance_after or 0,
                tx.remark or ""
            ])
        
        auto_column_width(ws3)
        
        # ========== Sheet 4: 应收账款 ==========
        ws4 = wb.create_sheet(title="应收账款")
        
        headers4 = ["销售单ID", "应收总额(元)", "已收金额(元)", "未收金额(元)", "账期开始", "到期日", "状态"]
        ws4.append(headers4)
        style_header(ws4)
        
        receivables = db.query(AccountReceivable).filter(
            AccountReceivable.customer_id == customer_id
        ).order_by(AccountReceivable.credit_start_date.desc()).all()
        
        status_labels = {
            "unpaid": "未付",
            "paid": "已付",
            "overdue": "逾期",
            "cancelled": "已取消"
        }
        
        for r in receivables:
            ws4.append([
                r.sales_order_id,
                r.total_amount or 0,
                r.received_amount or 0,
                r.unpaid_amount or 0,
                r.credit_start_date.strftime("%Y-%m-%d") if r.credit_start_date else "",
                r.due_date.strftime("%Y-%m-%d") if r.due_date else "",
                status_labels.get(r.status, r.status)
            ])
        
        auto_column_width(ws4)
        
        # 生成文件名
        safe_name = customer.name.replace("/", "_").replace("\\", "_")[:20]
        filename = f"{safe_name}_账务明细_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return create_excel_response(wb, filename)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出客户账务失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/all")
async def export_all_data(db: Session = Depends(get_db)):
    """一键导出全部数据为 ZIP 包"""
    try:
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 导出对话日志
            wb = Workbook()
            ws = wb.active
            ws.title = "对话日志"
            ws.append(["ID", "会话ID", "用户角色", "消息类型", "内容", "意图", "响应时间(ms)", "是否成功", "创建时间"])
            style_header(ws)
            role_names = {"sales": "业务员", "finance": "财务", "product": "商品专员", "manager": "管理层"}
            for log in db.query(ChatLog).order_by(ChatLog.created_at.desc()).all():
                ws.append([
                    log.id, log.session_id, role_names.get(log.user_role, log.user_role),
                    "用户" if log.message_type == "user" else "AI助手",
                    log.content[:500] if log.content else "", log.intent or "", log.response_time_ms,
                    "是" if log.is_successful else "否",
                    log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
                ])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("对话日志.xlsx", excel_buffer.getvalue())
            
            # 2. 导出库存数据
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"
            ws.append(["ID", "商品名称", "库存重量(克)", "最后更新时间"])
            style_header(ws)
            for inv in db.query(Inventory).all():
                ws.append([inv.id, inv.product_name, inv.total_weight,
                          inv.last_update.strftime("%Y-%m-%d %H:%M:%S") if inv.last_update else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("库存数据.xlsx", excel_buffer.getvalue())
            
            # 3. 导出入库记录
            wb = Workbook()
            ws = wb.active
            ws.title = "入库记录"
            ws.append(["入库单号", "商品名称", "重量(克)", "工费(元/克)", "总成本(元)", "供应商", "入库时间", "操作员"])
            style_header(ws)
            for order in db.query(InboundOrder).all():
                for detail in db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all():
                    ws.append([order.order_no, detail.product_name, detail.weight, detail.labor_cost,
                              detail.total_cost, detail.supplier or "",
                              order.create_time.strftime("%Y-%m-%d %H:%M:%S") if order.create_time else "",
                              order.operator or ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("入库记录.xlsx", excel_buffer.getvalue())
            
            # 4. 导出销售订单
            wb = Workbook()
            ws = wb.active
            ws.title = "销售订单"
            ws.append(["订单号", "客户名称", "商品名称", "重量(克)", "工费(元/克)", "总工费(元)", "业务员", "门店代码", "订单日期"])
            style_header(ws)
            for order in db.query(SalesOrder).all():
                customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
                for detail in db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all():
                    ws.append([order.order_no, customer.name if customer else "", detail.product_name,
                              detail.weight, detail.labor_cost, detail.total_labor_cost,
                              order.salesperson or "", order.store_code or "",
                              order.order_date.strftime("%Y-%m-%d") if order.order_date else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("销售订单.xlsx", excel_buffer.getvalue())
            
            # 5. 导出客户列表
            wb = Workbook()
            ws = wb.active
            ws.title = "客户列表"
            ws.append(["ID", "客户编号", "客户名称", "电话", "地址", "创建时间"])
            style_header(ws)
            for cust in db.query(Customer).all():
                ws.append([cust.id, cust.customer_no, cust.name, cust.phone or "", cust.address or "",
                          cust.create_time.strftime("%Y-%m-%d %H:%M:%S") if cust.create_time else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("客户列表.xlsx", excel_buffer.getvalue())
            
            # 6. 导出供应商列表
            wb = Workbook()
            ws = wb.active
            ws.title = "供应商列表"
            ws.append(["ID", "供应商编号", "供应商名称", "联系人", "电话", "地址", "总供货金额(元)", "总供货重量(克)", "供货次数", "最后供货时间", "状态"])
            style_header(ws)
            for sup in db.query(Supplier).all():
                ws.append([sup.id, sup.supplier_no, sup.name, sup.contact_person or "", sup.phone or "",
                          sup.address or "", sup.total_supply_amount or 0, sup.total_supply_weight or 0,
                          sup.total_supply_count or 0,
                          sup.last_supply_time.strftime("%Y-%m-%d %H:%M:%S") if sup.last_supply_time else "",
                          "活跃" if sup.status == "active" else "停用"])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("供应商列表.xlsx", excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        
        from fastapi.responses import Response
        filename = f"珠宝ERP数据备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Allow-Origin": "*",
            }
        )
        
    except Exception as e:
        logger.error(f"导出全部数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/export/stats")
async def get_export_stats(db: Session = Depends(get_db)):
    """获取可导出数据的统计信息"""
    try:
        return {
            "success": True,
            "data": {
                "chat_logs": db.query(func.count(ChatLog.id)).scalar() or 0,
                "inventory": db.query(func.count(Inventory.id)).scalar() or 0,
                "inbound_orders": db.query(func.count(InboundOrder.id)).scalar() or 0,
                "sales_orders": db.query(func.count(SalesOrder.id)).scalar() or 0,
                "customers": db.query(func.count(Customer.id)).scalar() or 0,
                "suppliers": db.query(func.count(Supplier.id)).scalar() or 0,
            }
        }
    except Exception as e:
        logger.error(f"获取导出统计失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


# ==================== 库存维护API ====================

@app.post("/api/inventory/merge-duplicates")
async def merge_duplicate_inventory(db: Session = Depends(get_db)):
    """
    合并重复的库存商品（商品编码与商品名称分开存储的情况）
    例如：将 "3DDZ" 的库存合并到 "足金3D硬金吊坠"
    """
    try:
        from .models import ProductCode as ProductCodeModel
        
        merged_count = 0
        merge_details = []
        
        # 获取所有商品编码及其对应的名称
        product_codes = db.query(ProductCodeModel).filter(
            ProductCodeModel.name.isnot(None),
            ProductCodeModel.name != ""
        ).all()
        
        for pc in product_codes:
            code = pc.code  # 如 "3DDZ"
            name = pc.name  # 如 "足金3D硬金吊坠"
            
            # 检查是否有以编码为名称的库存记录
            code_inventory = db.query(Inventory).filter(Inventory.product_name == code).first()
            if not code_inventory:
                continue
            
            # 检查是否有以名称为名称的库存记录
            name_inventory = db.query(Inventory).filter(Inventory.product_name == name).first()
            
            # 合并总库存
            if name_inventory:
                # 将编码库存合并到名称库存
                name_inventory.total_weight += code_inventory.total_weight
                db.delete(code_inventory)
            else:
                # 将编码库存重命名为名称
                code_inventory.product_name = name
            
            # 合并分仓库存 (LocationInventory)
            code_location_inventories = db.query(LocationInventory).filter(
                LocationInventory.product_name == code
            ).all()
            
            for cli in code_location_inventories:
                # 检查该位置是否已有名称库存
                name_location_inventory = db.query(LocationInventory).filter(
                    LocationInventory.product_name == name,
                    LocationInventory.location_id == cli.location_id
                ).first()
                
                if name_location_inventory:
                    # 合并库存
                    name_location_inventory.weight += cli.weight
                    db.delete(cli)
                else:
                    # 重命名
                    cli.product_name = name
            
            merged_count += 1
            merge_details.append({
                "code": code,
                "name": name,
                "merged_weight": code_inventory.total_weight if code_inventory else 0
            })
        
        db.commit()
        
        return {
            "success": True,
            "message": f"成功合并 {merged_count} 个重复商品",
            "merged_count": merged_count,
            "details": merge_details
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"合并重复库存失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.post("/api/inventory/merge-manual")
async def merge_inventory_manual(
    source_name: str = Query(..., description="要合并的源商品名称（将被删除）"),
    target_name: str = Query(..., description="目标商品名称（将保留）"),
    db: Session = Depends(get_db)
):
    """
    手动合并两个商品库存
    将 source_name 的库存合并到 target_name，然后删除 source_name
    """
    try:
        # 查找源商品
        source_inventory = db.query(Inventory).filter(Inventory.product_name == source_name).first()
        if not source_inventory:
            return {"success": False, "message": f"未找到源商品：{source_name}"}
        
        # 查找目标商品
        target_inventory = db.query(Inventory).filter(Inventory.product_name == target_name).first()
        
        source_weight = source_inventory.total_weight
        
        # 合并总库存
        if target_inventory:
            target_inventory.total_weight += source_weight
            db.delete(source_inventory)
        else:
            # 目标不存在，直接重命名
            source_inventory.product_name = target_name
        
        # 合并分仓库存
        source_location_inventories = db.query(LocationInventory).filter(
            LocationInventory.product_name == source_name
        ).all()
        
        for sli in source_location_inventories:
            target_li = db.query(LocationInventory).filter(
                LocationInventory.product_name == target_name,
                LocationInventory.location_id == sli.location_id
            ).first()
            
            if target_li:
                target_li.weight += sli.weight
                db.delete(sli)
            else:
                sli.product_name = target_name
        
        db.commit()
        
        return {
            "success": True,
            "message": f"成功将 {source_name} ({source_weight}g) 合并到 {target_name}",
            "merged_weight": source_weight
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"手动合并库存失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@app.get("/api/inventory/by-barcode")
async def get_inventory_by_barcode(
    search: Optional[str] = Query(None, description="搜索商品编码或名称"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """
    按条码查看库存 - 返回入库明细列表
    
    每条记录包含独立的商品编码、克重、工费等信息
    """
    try:
        from .timezone_utils import to_china_time, format_china_time
        
        # 查询入库明细，关联入库单获取时间和单号
        query = db.query(InboundDetail, InboundOrder).join(
            InboundOrder, InboundDetail.order_id == InboundOrder.id
        )
        
        # 搜索过滤
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                (InboundDetail.product_code.ilike(search_pattern)) |
                (InboundDetail.product_name.ilike(search_pattern))
            )
        
        # 排序：按入库时间倒序
        query = query.order_by(InboundOrder.create_time.desc())
        
        # 获取总数
        total = query.count()
        
        # 分页
        results = query.offset(skip).limit(limit).all()
        
        # 构建响应数据
        items = []
        for detail, order in results:
            china_time = to_china_time(order.create_time) if order.create_time else None
            items.append({
                "id": detail.id,
                "product_code": detail.product_code or "-",
                "product_name": detail.product_name,
                "weight": detail.weight,
                "labor_cost": detail.labor_cost,
                "piece_count": detail.piece_count,
                "piece_labor_cost": detail.piece_labor_cost,
                "total_cost": detail.total_cost,
                "supplier": detail.supplier,
                "order_no": order.order_no,
                "inbound_time": format_china_time(china_time, "%Y-%m-%d %H:%M") if china_time else None,
                "status": order.status
            })
        
        return {
            "success": True,
            "data": items,
            "total": total,
            "skip": skip,
            "limit": limit
        }
        
    except Exception as e:
        logger.error(f"按条码查询库存失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@app.post("/api/inventory/merge-product-names")
async def merge_product_names(
    old_name: str = Query(..., description="原商品名称（要被合并的）"),
    new_name: str = Query(..., description="新商品名称（合并到）"),
    db: Session = Depends(get_db)
):
    """
    合并商品名称：将 old_name 的所有数据合并到 new_name
    
    - 库存表：克重累加，删除旧记录
    - 分仓库存表：克重累加，删除旧记录
    - 入库明细表：更新商品名称
    - 销售明细表：更新商品名称
    """
    try:
        from sqlalchemy import update
        
        changes = {
            "inventory": 0,
            "location_inventory": 0,
            "inbound_details": 0,
            "sales_details": 0
        }
        
        # 1. 处理总库存表
        old_inv = db.query(Inventory).filter(Inventory.product_name == old_name).first()
        new_inv = db.query(Inventory).filter(Inventory.product_name == new_name).first()
        
        if old_inv:
            if new_inv:
                # 合并克重
                new_inv.total_weight += old_inv.total_weight
                db.delete(old_inv)
                changes["inventory"] = 1
                logger.info(f"库存合并: {old_name} ({old_inv.total_weight}g) -> {new_name}")
            else:
                # 直接改名
                old_inv.product_name = new_name
                changes["inventory"] = 1
                logger.info(f"库存改名: {old_name} -> {new_name}")
        
        # 2. 处理分仓库存表
        old_loc_invs = db.query(LocationInventory).filter(LocationInventory.product_name == old_name).all()
        for old_loc in old_loc_invs:
            new_loc = db.query(LocationInventory).filter(
                LocationInventory.product_name == new_name,
                LocationInventory.location_id == old_loc.location_id
            ).first()
            
            if new_loc:
                # 合并克重
                new_loc.weight += old_loc.weight
                db.delete(old_loc)
            else:
                # 直接改名
                old_loc.product_name = new_name
            changes["location_inventory"] += 1
        
        # 3. 更新入库明细表
        result = db.execute(
            update(InboundDetail).where(InboundDetail.product_name == old_name).values(product_name=new_name)
        )
        changes["inbound_details"] = result.rowcount
        
        # 4. 更新销售明细表
        result = db.execute(
            update(SalesDetail).where(SalesDetail.product_name == old_name).values(product_name=new_name)
        )
        changes["sales_details"] = result.rowcount
        
        db.commit()
        
        logger.info(f"商品名称合并完成: {old_name} -> {new_name}, 变更: {changes}")
        
        return {
            "success": True,
            "message": f"已将 '{old_name}' 合并到 '{new_name}'",
            "changes": changes
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"合并商品名称失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@app.get("/api/inventory/check-consistency")
async def check_inventory_consistency(db: Session = Depends(get_db)):
    """
    检查 Inventory（总库存）和 LocationInventory（分仓库存）的一致性
    """
    from sqlalchemy import func
    
    # 获取总库存表数据
    total_inventory = db.query(Inventory).all()
    
    # 获取分仓库存汇总（按商品名称分组求和）
    location_inventory_summary = db.query(
        LocationInventory.product_name,
        func.sum(LocationInventory.weight).label("total_weight")
    ).group_by(LocationInventory.product_name).all()
    
    # 转换为字典方便比对
    total_dict = {inv.product_name: inv.total_weight for inv in total_inventory}
    location_dict = {item.product_name: float(item.total_weight or 0) for item in location_inventory_summary}
    
    # 比对差异
    discrepancies = []
    all_products = set(total_dict.keys()) | set(location_dict.keys())
    
    for product in all_products:
        total_weight = total_dict.get(product, 0)
        location_weight = location_dict.get(product, 0)
        
        if abs(total_weight - location_weight) > 0.01:  # 允许 0.01 的误差
            discrepancies.append({
                "product_name": product,
                "inventory_total": total_weight,
                "location_total": location_weight,
                "difference": total_weight - location_weight
            })
    
    return {
        "success": True,
        "is_consistent": len(discrepancies) == 0,
        "total_inventory_weight": sum(total_dict.values()),
        "total_location_weight": sum(location_dict.values()),
        "difference": sum(total_dict.values()) - sum(location_dict.values()),
        "discrepancies": discrepancies,
        "product_count_in_inventory": len(total_dict),
        "product_count_in_location": len(location_dict)
    }


@app.post("/api/inventory/sync-to-location")
async def sync_inventory_to_location(db: Session = Depends(get_db)):
    """
    将 Inventory 表的数据同步到 LocationInventory 表
    以 Inventory 表为准，确保分仓库存（默认商品部仓库）与总库存一致
    """
    from sqlalchemy import func
    
    # 获取或创建默认仓库
    default_location = db.query(Location).filter(Location.code == "warehouse").first()
    if not default_location:
        default_location = Location(
            code="warehouse",
            name="商品部仓库",
            location_type="warehouse",
            description="商品部主仓库"
        )
        db.add(default_location)
        db.flush()
    
    # 获取总库存表数据
    total_inventory = db.query(Inventory).all()
    
    sync_results = []
    
    for inv in total_inventory:
        if inv.total_weight <= 0:
            continue
            
        # 获取该商品在分仓库存中的总量
        location_total = db.query(
            func.sum(LocationInventory.weight)
        ).filter(
            LocationInventory.product_name == inv.product_name
        ).scalar() or 0
        
        difference = inv.total_weight - float(location_total)
        
        if abs(difference) > 0.01:  # 有差异需要调整
            # 获取或创建该商品在默认仓库的记录
            location_inv = db.query(LocationInventory).filter(
                LocationInventory.product_name == inv.product_name,
                LocationInventory.location_id == default_location.id
            ).first()
            
            if location_inv:
                # 更新库存，补足差额
                old_weight = location_inv.weight
                location_inv.weight += difference
                sync_results.append({
                    "product_name": inv.product_name,
                    "action": "updated",
                    "old_weight": old_weight,
                    "new_weight": location_inv.weight,
                    "adjustment": difference
                })
            else:
                # 创建新记录
                new_location_inv = LocationInventory(
                    product_name=inv.product_name,
                    location_id=default_location.id,
                    weight=inv.total_weight
                )
                db.add(new_location_inv)
                sync_results.append({
                    "product_name": inv.product_name,
                    "action": "created",
                    "weight": inv.total_weight
                })
    
    db.commit()
    
    return {
        "success": True,
        "message": f"同步完成，共调整 {len(sync_results)} 项商品库存",
        "sync_results": sync_results
    }

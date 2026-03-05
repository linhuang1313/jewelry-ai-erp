"""
暂借单管理 API 路由（支持多商品 + 还货单）
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime
from pydantic import BaseModel
import logging

from ..database import get_db
from ..timezone_utils import china_now, to_china_time, format_china_time
from ..models import (
    LoanOrder, LoanDetail, LoanOrderLog, LoanReturn, LoanReturnDetail,
    Inventory, LocationInventory, Location, Customer, ProductCode
)
from ..utils.response import success_response, error_response
from ..dependencies.auth import get_current_role, require_permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/loan", tags=["暂借管理"])


# ============= Pydantic 模型 =============

class LoanItemCreate(BaseModel):
    """暂借单明细项"""
    product_name: str
    weight: float
    piece_count: Optional[int] = None


class LoanOrderCreate(BaseModel):
    """创建暂借单请求（支持多商品）"""
    customer_id: int
    items: List[LoanItemCreate]  # 多商品列表
    salesperson: str
    loan_date: datetime
    remark: Optional[str] = None
    created_by: Optional[str] = None


class LoanOrderConfirm(BaseModel):
    """确认借出请求"""
    operator: str


class LoanOrderReturn(BaseModel):
    """确认归还请求（旧接口兼容）"""
    operator: str
    remark: Optional[str] = None


class LoanOrderCancel(BaseModel):
    """撤销暂借单请求"""
    operator: str
    reason: str  # 撤销原因（必填，用于留痕）


class LoanReturnCreate(BaseModel):
    """创建还货单请求"""
    loan_id: int  # 关联暂借单ID
    detail_ids: List[int]  # 要归还的暂借明细ID列表
    operator: str
    remark: Optional[str] = None


class LoanDetailResponse(BaseModel):
    """暂借单明细响应"""
    id: int
    loan_id: int
    product_name: str
    product_code: Optional[str] = None
    weight: float
    piece_count: Optional[int] = None
    status: str
    returned_at: Optional[datetime] = None
    returned_by: Optional[str] = None

    class Config:
        from_attributes = True


class LoanOrderResponse(BaseModel):
    """暂借单响应（含明细）"""
    id: int
    loan_no: str
    customer_id: int
    customer_name: str
    total_weight: float = 0.0
    salesperson: str
    loan_date: datetime
    status: str
    created_by: Optional[str] = None
    created_at: datetime
    confirmed_at: Optional[datetime] = None
    returned_at: Optional[datetime] = None
    returned_by: Optional[str] = None
    cancelled_at: Optional[datetime] = None
    cancelled_by: Optional[str] = None
    cancel_reason: Optional[str] = None
    printed_at: Optional[datetime] = None
    remark: Optional[str] = None
    details: List[LoanDetailResponse] = []

    class Config:
        from_attributes = True


class LoanOrderLogResponse(BaseModel):
    """暂借单操作日志响应"""
    id: int
    loan_order_id: int
    action: str
    operator: str
    action_time: datetime
    old_status: Optional[str] = None
    new_status: str
    remark: Optional[str] = None

    class Config:
        from_attributes = True


class LoanReturnDetailResponse(BaseModel):
    """还货单明细响应"""
    id: int
    return_id: int
    loan_detail_id: int
    product_name: str
    weight: float

    class Config:
        from_attributes = True


class LoanReturnResponse(BaseModel):
    """还货单响应"""
    id: int
    return_no: str
    loan_id: int
    loan_no: Optional[str] = None
    customer_id: int
    customer_name: str
    total_weight: float = 0.0
    operator: Optional[str] = None
    created_at: datetime
    remark: Optional[str] = None
    printed_at: Optional[datetime] = None
    details: List[LoanReturnDetailResponse] = []

    class Config:
        from_attributes = True


# ============= 辅助函数 =============

def generate_loan_no(db: Session, max_retries: int = 3) -> str:
    """生成暂借单号：ZJ + 日期 + 4位序号"""
    today = china_now()
    date_str = today.strftime("%Y%m%d")
    full_prefix = f"ZJ{date_str}"

    for attempt in range(max_retries):
        latest_row = (
            db.query(LoanOrder.loan_no)
            .filter(LoanOrder.loan_no.like(f"{full_prefix}%"))
            .order_by(LoanOrder.loan_no.desc())
            .limit(1)
            .with_for_update()
            .scalar()
        )

        if latest_row and latest_row.startswith(full_prefix):
            try:
                seq = int(latest_row[len(full_prefix):]) + 1
            except ValueError:
                seq = 1
        else:
            seq = 1

        loan_no = f"{full_prefix}{seq:04d}"

        existing = db.query(LoanOrder).filter(LoanOrder.loan_no == loan_no).first()
        if existing is None:
            return loan_no

        logger.warning(
            "Loan number collision: %s (attempt %d/%d)",
            loan_no, attempt + 1, max_retries,
        )

    fallback_ts = today.strftime("%H%M%S")
    return f"{full_prefix}{fallback_ts}"


def generate_return_no(db: Session, max_retries: int = 3) -> str:
    """生成还货单号：HH + 日期 + 4位序号"""
    today = china_now()
    date_str = today.strftime("%Y%m%d")
    full_prefix = f"HH{date_str}"

    for attempt in range(max_retries):
        latest_row = (
            db.query(LoanReturn.return_no)
            .filter(LoanReturn.return_no.like(f"{full_prefix}%"))
            .order_by(LoanReturn.return_no.desc())
            .limit(1)
            .with_for_update()
            .scalar()
        )

        if latest_row and latest_row.startswith(full_prefix):
            try:
                seq = int(latest_row[len(full_prefix):]) + 1
            except ValueError:
                seq = 1
        else:
            seq = 1

        return_no = f"{full_prefix}{seq:04d}"

        existing = db.query(LoanReturn).filter(LoanReturn.return_no == return_no).first()
        if existing is None:
            return return_no

        logger.warning(
            "Return number collision: %s (attempt %d/%d)",
            return_no, attempt + 1, max_retries,
        )

    fallback_ts = today.strftime("%H%M%S")
    return f"{full_prefix}{fallback_ts}"


def create_log(db: Session, loan_order_id: int, action: str, operator: str, 
               old_status: Optional[str], new_status: str, remark: Optional[str] = None):
    """创建操作日志"""
    log = LoanOrderLog(
        loan_order_id=loan_order_id,
        action=action,
        operator=operator,
        action_time=china_now(),
        old_status=old_status,
        new_status=new_status,
        remark=remark
    )
    db.add(log)


def get_status_label(status: str) -> str:
    """获取状态的中文标签"""
    labels = {
        "pending": "待确认",
        "borrowed": "已借出",
        "partial_returned": "部分归还",
        "returned": "已归还",
        "cancelled": "已撤销"
    }
    return labels.get(status, status)


def build_order_response(loan_order: LoanOrder) -> dict:
    """构建暂借单响应，包含明细"""
    details = []
    for d in (loan_order.details or []):
        details.append(LoanDetailResponse.model_validate(d).model_dump(mode='json'))
    
    resp = LoanOrderResponse.model_validate(loan_order)
    resp.details = details
    return resp


def deduct_inventory(db: Session, product_name: str, weight: float):
    """扣减指定商品的库存（展厅优先，不够从仓库扣）"""
    # 扣减总库存
    inventory = db.query(Inventory).filter(
        Inventory.product_name == product_name
    ).first()
    
    if not inventory or inventory.total_weight < weight:
        available = inventory.total_weight if inventory else 0
        raise HTTPException(
            status_code=400,
            detail=f"库存不足：{product_name} 当前库存 {available:.2f}克，需要 {weight:.2f}克"
        )
    
    inventory.total_weight = round(inventory.total_weight - weight, 3)
    inventory.last_update = china_now()
    
    # 展厅优先扣减
    remaining = weight
    showroom_locations = db.query(Location).filter(
        Location.location_type == "showroom",
        Location.is_active == 1
    ).all()
    
    warehouse_locations = db.query(Location).filter(
        Location.location_type == "warehouse",
        Location.is_active == 1
    ).all()
    
    # 批量查询该商品在所有位置的库存（避免 N+1）
    all_location_ids = [loc.id for loc in showroom_locations + warehouse_locations]
    if all_location_ids:
        loc_invs = db.query(LocationInventory).filter(
            LocationInventory.location_id.in_(all_location_ids),
            LocationInventory.product_name == product_name
        ).all()
        loc_inv_map = {li.location_id: li for li in loc_invs}
    else:
        loc_inv_map = {}
    
    for location in showroom_locations:
        if remaining <= 0:
            break
        loc_inv = loc_inv_map.get(location.id)
        if loc_inv and loc_inv.weight > 0:
            deduct = min(loc_inv.weight, remaining)
            loc_inv.weight -= deduct
            loc_inv.last_update = china_now()
            remaining -= deduct
    
    # 不够从仓库扣
    if remaining > 0:
        for location in warehouse_locations:
            if remaining <= 0:
                break
            loc_inv = loc_inv_map.get(location.id)
            if loc_inv and loc_inv.weight > 0:
                deduct = min(loc_inv.weight, remaining)
                loc_inv.weight -= deduct
                loc_inv.last_update = china_now()
                remaining -= deduct


def restore_inventory(db: Session, product_name: str, weight: float):
    """恢复指定商品的库存（总库存 + 展厅）"""
    inventory = db.query(Inventory).filter(
        Inventory.product_name == product_name
    ).first()
    
    if inventory:
        inventory.total_weight = round(inventory.total_weight + weight, 3)
        inventory.last_update = china_now()
    else:
        inventory = Inventory(
            product_name=product_name,
            total_weight=weight,
            last_update=china_now()
        )
        db.add(inventory)
    
    # 恢复到默认展厅
    default_showroom = db.query(Location).filter(
        Location.location_type == "showroom",
        Location.is_active == 1
    ).first()
    
    if default_showroom:
        loc_inv = db.query(LocationInventory).filter(
            LocationInventory.location_id == default_showroom.id,
            LocationInventory.product_name == product_name
        ).first()
        if loc_inv:
            loc_inv.weight += weight
            loc_inv.last_update = china_now()
        else:
            loc_inv = LocationInventory(
                location_id=default_showroom.id,
                product_name=product_name,
                weight=weight,
                last_update=china_now()
            )
            db.add(loc_inv)


# ============= 还货单 API（静态路由放在动态路由前） =============

@router.get("/returns")
async def get_loan_returns(
    customer_name: Optional[str] = Query(None, description="客户姓名搜索"),
    loan_no: Optional[str] = Query(None, description="原暂借单号搜索"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=200, description="每页数量"),
    db: Session = Depends(get_db)
):
    """获取还货单列表（分页）"""
    import math
    query = db.query(LoanReturn)
    
    if customer_name:
        query = query.filter(LoanReturn.customer_name.contains(customer_name))
    if loan_no:
        query = query.join(LoanOrder).filter(LoanOrder.loan_no.contains(loan_no))
    
    total = query.count()
    total_pages = math.ceil(total / page_size) if total > 0 else 0
    offset = (page - 1) * page_size
    
    returns = query.options(joinedload(LoanReturn.details)).order_by(LoanReturn.created_at.desc()).offset(offset).limit(page_size).all()
    seen_ids = set()
    unique_returns = []
    for r in returns:
        if r.id not in seen_ids:
            seen_ids.add(r.id)
            unique_returns.append(r)
    returns = unique_returns
    
    # 批量查询所有关联的暂借单（避免 N+1）
    loan_ids = list(set(ret.loan_id for ret in returns))
    if loan_ids:
        loan_orders = db.query(LoanOrder).filter(LoanOrder.id.in_(loan_ids)).all()
        loan_order_map = {lo.id: lo for lo in loan_orders}
    else:
        loan_order_map = {}
    
    result = []
    for ret in returns:
        loan_order = loan_order_map.get(ret.loan_id)
        resp = {
            "id": ret.id,
            "return_no": ret.return_no,
            "loan_id": ret.loan_id,
            "loan_no": loan_order.loan_no if loan_order else "",
            "customer_id": ret.customer_id,
            "customer_name": ret.customer_name,
            "total_weight": ret.total_weight,
            "operator": ret.operator,
            "created_at": ret.created_at.isoformat() if ret.created_at else None,
            "remark": ret.remark,
            "printed_at": ret.printed_at.isoformat() if ret.printed_at else None,
            "details": [{
                "id": d.id,
                "return_id": d.return_id,
                "loan_detail_id": d.loan_detail_id,
                "product_name": d.product_name,
                "weight": d.weight,
            } for d in (ret.details or [])],
            "item_count": len(ret.details or []),
        }
        result.append(resp)
    
    return {"success": True, "data": {"returns": result, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}}


@router.post("/returns")
async def create_loan_return(
    data: LoanReturnCreate,
    db: Session = Depends(get_db)
):
    """创建还货单（创建即生效，立即恢复库存，行级锁防并发）"""
    try:
        # 行级锁：锁定暂借单防止并发还货
        loan_order = db.query(LoanOrder).filter(
            LoanOrder.id == data.loan_id
        ).with_for_update().first()
        
        if not loan_order:
            raise HTTPException(status_code=404, detail="暂借单不存在")
        
        if loan_order.status not in ["borrowed", "partial_returned"]:
            raise HTTPException(
                status_code=400,
                detail=f"暂借单状态为 {get_status_label(loan_order.status)}，无法创建还货单"
            )
        
        if not data.detail_ids or len(data.detail_ids) == 0:
            raise HTTPException(status_code=400, detail="请选择要归还的商品")
        
        # 验证明细并收集
        return_details = []
        total_weight = 0.0
        for detail_id in data.detail_ids:
            detail = db.query(LoanDetail).filter(
                LoanDetail.id == detail_id,
                LoanDetail.loan_id == data.loan_id
            ).first()
            
            if not detail:
                raise HTTPException(status_code=400, detail=f"暂借明细 ID {detail_id} 不存在或不属于该暂借单")
            
            if detail.status != "borrowed":
                raise HTTPException(
                    status_code=400,
                    detail=f"商品 {detail.product_name} 状态为 {get_status_label(detail.status)}，无法归还"
                )
            
            return_details.append(detail)
            total_weight += detail.weight
        
        # 生成还货单号
        return_no = generate_return_no(db)
        
        # 创建还货单
        loan_return = LoanReturn(
            return_no=return_no,
            loan_id=data.loan_id,
            customer_id=loan_order.customer_id,
            customer_name=loan_order.customer_name,
            total_weight=round(total_weight, 3),
            operator=data.operator,
            created_at=china_now(),
            remark=data.remark,
        )
        db.add(loan_return)
        db.flush()
        
        # 创建还货单明细 & 更新暂借明细状态 & 恢复库存
        for detail in return_details:
            # 还货单明细
            return_detail = LoanReturnDetail(
                return_id=loan_return.id,
                loan_detail_id=detail.id,
                product_name=detail.product_name,
                weight=detail.weight,
            )
            db.add(return_detail)
            
            # 更新暂借明细状态
            detail.status = "returned"
            detail.returned_at = china_now()
            detail.returned_by = data.operator
            
            # 恢复库存
            restore_inventory(db, detail.product_name, detail.weight)
        
        # 检查暂借单所有明细是否全部归还
        all_details = db.query(LoanDetail).filter(LoanDetail.loan_id == data.loan_id).all()
        all_returned = all(d.status == "returned" for d in all_details)
        any_returned = any(d.status == "returned" for d in all_details)
        
        old_status = loan_order.status
        if all_returned:
            loan_order.status = "returned"
            loan_order.returned_at = china_now()
            loan_order.returned_by = data.operator
        elif any_returned:
            loan_order.status = "partial_returned"
        
        # 操作日志
        returned_names = ", ".join([d.product_name for d in return_details])
        create_log(
            db, loan_order.id, "partial_return" if not all_returned else "return",
            data.operator, old_status, loan_order.status,
            f"还货单 {return_no}，归还: {returned_names}，共 {total_weight:.2f}克"
        )
        
        db.commit()
        db.refresh(loan_return)
        
        logger.info(f"创建还货单: {return_no}, 暂借单: {loan_order.loan_no}, 归还 {len(return_details)} 件，{total_weight:.2f}克")
        
        return {
            "success": True,
            "message": f"还货单 {return_no} 创建成功，已归还 {len(return_details)} 件商品",
            "data": {
                "id": loan_return.id,
                "return_no": return_no,
                "loan_no": loan_order.loan_no,
                "total_weight": round(total_weight, 3),
                "item_count": len(return_details),
                "loan_status": loan_order.status,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建还货单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建还货单失败: {str(e)}")


@router.get("/returns/{return_id}")
async def get_loan_return_detail(
    return_id: int,
    db: Session = Depends(get_db)
):
    """获取还货单详情"""
    loan_return = db.query(LoanReturn).options(
        joinedload(LoanReturn.details)
    ).filter(LoanReturn.id == return_id).first()
    
    if not loan_return:
        raise HTTPException(status_code=404, detail="还货单不存在")
    
    loan_order = db.query(LoanOrder).filter(LoanOrder.id == loan_return.loan_id).first()
    
    return {
        "id": loan_return.id,
        "return_no": loan_return.return_no,
        "loan_id": loan_return.loan_id,
        "loan_no": loan_order.loan_no if loan_order else "",
        "customer_id": loan_return.customer_id,
        "customer_name": loan_return.customer_name,
        "total_weight": loan_return.total_weight,
        "operator": loan_return.operator,
        "created_at": loan_return.created_at.isoformat() if loan_return.created_at else None,
        "remark": loan_return.remark,
        "printed_at": loan_return.printed_at.isoformat() if loan_return.printed_at else None,
        "details": [{
            "id": d.id,
            "return_id": d.return_id,
            "loan_detail_id": d.loan_detail_id,
            "product_name": d.product_name,
            "weight": d.weight,
        } for d in (loan_return.details or [])],
    }


# ============= 暂借单 CRUD =============

@router.get("/orders")
async def get_loan_orders(
    status: Optional[str] = Query(None, description="状态筛选"),
    customer_name: Optional[str] = Query(None, description="客户姓名搜索"),
    product_name: Optional[str] = Query(None, description="产品名称搜索"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=200, description="每页数量"),
    db: Session = Depends(get_db)
):
    """获取暂借单列表（分页，含明细）"""
    import math
    query = db.query(LoanOrder)
    
    if status:
        query = query.filter(LoanOrder.status == status)
    if customer_name:
        query = query.filter(LoanOrder.customer_name.contains(customer_name))
    if product_name:
        query = query.join(LoanDetail).filter(LoanDetail.product_name.contains(product_name))
    
    total = query.count()
    total_pages = math.ceil(total / page_size) if total > 0 else 0
    offset = (page - 1) * page_size
    
    orders = query.options(joinedload(LoanOrder.details)).order_by(LoanOrder.created_at.desc()).offset(offset).limit(page_size).all()
    # joinedload + offset/limit 可能返回重复行，去重
    seen_ids = set()
    unique_orders = []
    for o in orders:
        if o.id not in seen_ids:
            seen_ids.add(o.id)
            unique_orders.append(o)
    orders = unique_orders
    
    all_product_names = set()
    for order in orders:
        for d in (order.details or []):
            all_product_names.add(d.product_name)
    product_code_map = {}
    if all_product_names:
        codes = db.query(ProductCode).filter(ProductCode.name.in_(list(all_product_names))).all()
        product_code_map = {pc.name: pc.code for pc in codes}

    result = []
    for order in orders:
        details = [{
            "id": d.id,
            "loan_id": d.loan_id,
            "product_name": d.product_name,
            "product_code": product_code_map.get(d.product_name),
            "weight": d.weight,
            "piece_count": d.piece_count,
            "status": d.status,
            "returned_at": d.returned_at.isoformat() if d.returned_at else None,
            "returned_by": d.returned_by,
        } for d in (order.details or [])]
        
        resp = {
            "id": order.id,
            "loan_no": order.loan_no,
            "customer_id": order.customer_id,
            "customer_name": order.customer_name,
            "total_weight": order.total_weight or 0,
            "salesperson": order.salesperson,
            "loan_date": order.loan_date.isoformat() if order.loan_date else None,
            "status": order.status,
            "created_by": order.created_by,
            "created_at": order.created_at.isoformat() if order.created_at else None,
            "confirmed_at": order.confirmed_at.isoformat() if order.confirmed_at else None,
            "returned_at": order.returned_at.isoformat() if order.returned_at else None,
            "returned_by": order.returned_by,
            "cancelled_at": order.cancelled_at.isoformat() if order.cancelled_at else None,
            "cancelled_by": order.cancelled_by,
            "cancel_reason": order.cancel_reason,
            "printed_at": order.printed_at.isoformat() if order.printed_at else None,
            "remark": order.remark,
            "details": details,
            "item_count": len(details),
        }
        result.append(resp)
    
    return {"success": True, "data": {"orders": result, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}}


@router.post("/orders")
async def create_loan_order(
    data: LoanOrderCreate,
    db: Session = Depends(get_db)
):
    """创建暂借单（支持多商品）"""
    try:
        # 验证客户
        customer = db.query(Customer).filter(Customer.id == data.customer_id).first()
        if not customer:
            raise HTTPException(status_code=400, detail="客户不存在")
    
        if not data.items or len(data.items) == 0:
            raise HTTPException(status_code=400, detail="请添加至少一个商品")
        
        # 验证所有商品
        inventory_errors = []
        for item in data.items:
            if item.weight <= 0:
                raise HTTPException(status_code=400, detail=f"商品 {item.product_name} 的克重必须大于0")
            # 检查库存
            inventory = db.query(Inventory).filter(
                Inventory.product_name == item.product_name
            ).first()
            if not inventory or inventory.total_weight < item.weight:
                available = inventory.total_weight if inventory else 0
                inventory_errors.append(
                    f"{item.product_name} 库存不足（当前 {available:.2f}克，需要 {item.weight:.2f}克）"
                )
        
        if inventory_errors:
            raise HTTPException(status_code=400, detail="库存检查失败：" + "；".join(inventory_errors))
        
        total_weight = sum(item.weight for item in data.items)
        
        # 生成单号
        loan_no = generate_loan_no(db)
        
        # 创建暂借单主表
        loan_order = LoanOrder(
            loan_no=loan_no,
            customer_id=data.customer_id,
            customer_name=customer.name,
            total_weight=round(total_weight, 3),
            salesperson=data.salesperson,
            loan_date=data.loan_date,
            status="pending",
            created_by=data.created_by or "结算专员",
            created_at=china_now(),
            remark=data.remark,
        )
        db.add(loan_order)
        db.flush()
        
        for item in data.items:
            detail = LoanDetail(
                loan_id=loan_order.id,
                product_name=item.product_name,
                weight=item.weight,
                piece_count=item.piece_count,
                status="pending",
            )
            db.add(detail)
        
        # 操作日志
        product_names = ", ".join([item.product_name for item in data.items])
        create_log(db, loan_order.id, "create", loan_order.created_by, None, "pending",
                   f"创建暂借单，{len(data.items)} 件商品: {product_names}")
        
        db.commit()
        db.refresh(loan_order)
    
        logger.info(f"创建暂借单: {loan_no}, 客户: {customer.name}, {len(data.items)} 件商品, 总克重: {total_weight}")
        
        return {
            "success": True,
            "message": f"暂借单 {loan_no} 创建成功",
            "data": {
                "id": loan_order.id,
                "loan_no": loan_no,
                "item_count": len(data.items),
                "total_weight": round(total_weight, 3),
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建暂借单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建暂借单失败: {str(e)}")


# ============= 获取暂借单详情 =============

@router.get("/orders/{loan_id}")
async def get_loan_order(
    loan_id: int,
    db: Session = Depends(get_db)
):
    """获取暂借单详情（含明细和还货记录）"""
    loan_order = db.query(LoanOrder).options(
        joinedload(LoanOrder.details)
    ).filter(LoanOrder.id == loan_id).first()
    
    if not loan_order:
        raise HTTPException(status_code=404, detail="暂借单不存在")
    
    # 获取关联的还货单
    returns = db.query(LoanReturn).options(
        joinedload(LoanReturn.details)
    ).filter(LoanReturn.loan_id == loan_id).order_by(LoanReturn.created_at.desc()).all()
    
    product_names = [d.product_name for d in (loan_order.details or [])]
    product_code_map = {}
    if product_names:
        codes = db.query(ProductCode).filter(ProductCode.name.in_(product_names)).all()
        product_code_map = {pc.name: pc.code for pc in codes}

    details = [{
        "id": d.id,
        "loan_id": d.loan_id,
        "product_name": d.product_name,
        "product_code": product_code_map.get(d.product_name),
        "weight": d.weight,
        "piece_count": d.piece_count,
        "status": d.status,
        "returned_at": d.returned_at.isoformat() if d.returned_at else None,
        "returned_by": d.returned_by,
    } for d in (loan_order.details or [])]
    
    return_list = []
    for ret in returns:
        return_list.append({
            "id": ret.id,
            "return_no": ret.return_no,
            "total_weight": ret.total_weight,
            "operator": ret.operator,
            "created_at": ret.created_at.isoformat() if ret.created_at else None,
            "remark": ret.remark,
            "details": [{
                "id": rd.id,
                "product_name": rd.product_name,
                "weight": rd.weight,
            } for rd in (ret.details or [])],
        })
    
    return {
        "id": loan_order.id,
        "loan_no": loan_order.loan_no,
        "customer_id": loan_order.customer_id,
        "customer_name": loan_order.customer_name,
        "total_weight": loan_order.total_weight or 0,
        "salesperson": loan_order.salesperson,
        "loan_date": loan_order.loan_date.isoformat() if loan_order.loan_date else None,
        "status": loan_order.status,
        "created_by": loan_order.created_by,
        "created_at": loan_order.created_at.isoformat() if loan_order.created_at else None,
        "confirmed_at": loan_order.confirmed_at.isoformat() if loan_order.confirmed_at else None,
        "returned_at": loan_order.returned_at.isoformat() if loan_order.returned_at else None,
        "returned_by": loan_order.returned_by,
        "cancelled_at": loan_order.cancelled_at.isoformat() if loan_order.cancelled_at else None,
        "cancelled_by": loan_order.cancelled_by,
        "cancel_reason": loan_order.cancel_reason,
        "printed_at": loan_order.printed_at.isoformat() if loan_order.printed_at else None,
        "remark": loan_order.remark,
        "details": details,
        "item_count": len(details),
        "returns": return_list,
    }


# ============= 确认借出 =============

@router.post("/orders/{loan_id}/confirm")
async def confirm_loan_order(
    loan_id: int,
    data: LoanOrderConfirm,
    db: Session = Depends(get_db)
):
    """确认借出 - 逐项扣减库存（行级锁防并发）"""
    try:
        # 行级锁：锁定暂借单防止并发确认
        loan_order = db.query(LoanOrder).filter(
            LoanOrder.id == loan_id
        ).with_for_update().first()
        
        if not loan_order:
            raise HTTPException(status_code=404, detail="暂借单不存在")
        
        if loan_order.status != "pending":
            raise HTTPException(
                status_code=400,
                detail=f"当前状态为 {get_status_label(loan_order.status)}，无法确认借出"
            )
        
        details = db.query(LoanDetail).filter(LoanDetail.loan_id == loan_id).all()
        if not details:
            raise HTTPException(status_code=400, detail="暂借单没有明细，无法确认")
        
        # 行级锁：锁定库存行防止并发扣减
        product_names = [detail.product_name for detail in details]
        inventories = db.query(Inventory).filter(
            Inventory.product_name.in_(product_names)
        ).with_for_update().all()
        inventory_map = {inv.product_name: inv for inv in inventories}
        
        for detail in details:
            inventory = inventory_map.get(detail.product_name)
            if not inventory or inventory.total_weight < detail.weight:
                available = inventory.total_weight if inventory else 0
                raise HTTPException(
                    status_code=400,
                    detail=f"库存不足：{detail.product_name} 当前库存 {available:.2f}克，需要 {detail.weight:.2f}克"
                )
        
        # 全部检查通过，逐项扣减
        for detail in details:
            deduct_inventory(db, detail.product_name, detail.weight)
            detail.status = "borrowed"
        
        # 更新主单状态
        old_status = loan_order.status
        loan_order.status = "borrowed"
        loan_order.confirmed_at = china_now()
        
        create_log(db, loan_order.id, "confirm", data.operator, old_status, "borrowed",
                   f"确认借出，共 {len(details)} 件商品，总克重 {loan_order.total_weight:.2f}克")
        
        db.commit()
        
        logger.info(f"确认借出暂借单: {loan_order.loan_no}, {len(details)} 件, {loan_order.total_weight}克")
        
        return success_response(message=f"暂借单 {loan_order.loan_no} 已确认借出")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"确认借出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"确认借出失败: {str(e)}")


# ============= 撤销暂借单 =============

@router.post("/orders/{loan_id}/cancel")
async def cancel_loan_order(
    loan_id: int,
    data: LoanOrderCancel,
    db: Session = Depends(get_db)
):
    """撤销暂借单 - 仅允许待确认状态取消，已借出的须通过还货单归还"""
    try:
        loan_order = db.query(LoanOrder).options(
            joinedload(LoanOrder.details)
        ).filter(LoanOrder.id == loan_id).first()
        
        if not loan_order:
            raise HTTPException(status_code=404, detail="暂借单不存在")
        
        if loan_order.status != "pending":
            status_label = get_status_label(loan_order.status)
            if loan_order.status in ["borrowed", "partial_returned"]:
                raise HTTPException(
                    status_code=400,
                    detail=f"当前状态为{status_label}，已借出的商品请通过创建还货单归还，不允许直接取消"
                )
            raise HTTPException(
                status_code=400,
                detail=f"当前状态为{status_label}，无法取消"
            )
        
        old_status = loan_order.status
        details = loan_order.details or []
        
        # pending 阶段没有扣过库存，直接将所有明细标记为 cancelled
        for detail in details:
            detail.status = "cancelled"
        
        # 更新主单状态
        loan_order.status = "cancelled"
        loan_order.cancelled_at = china_now()
        loan_order.cancelled_by = data.operator
        loan_order.cancel_reason = data.reason
        
        log_remark = f"撤销原因：{data.reason}"
        
        create_log(db, loan_order.id, "cancel", data.operator, old_status, "cancelled", log_remark)
        
        db.commit()
        
        logger.info(f"撤销暂借单: {loan_order.loan_no}, 原因: {data.reason}")
        
        return success_response(message=f"暂借单 {loan_order.loan_no} 已撤销")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"撤销暂借单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"撤销失败: {str(e)}")


# ============= 整单归还（兼容旧接口）=============

@router.post("/orders/{loan_id}/return")
async def return_loan_order(
    loan_id: int,
    data: LoanOrderReturn,
    db: Session = Depends(get_db)
):
    """整单归还 - 创建还货单归还所有未还商品"""
    loan_order = db.query(LoanOrder).options(
        joinedload(LoanOrder.details)
    ).filter(LoanOrder.id == loan_id).first()
    
    if not loan_order:
        raise HTTPException(status_code=404, detail="暂借单不存在")
    
    if loan_order.status not in ["borrowed", "partial_returned"]:
        raise HTTPException(
            status_code=400,
            detail=f"当前状态为 {get_status_label(loan_order.status)}，无法归还"
        )
    
    # 收集所有未还的明细ID
    borrowed_detail_ids = [
        d.id for d in (loan_order.details or []) if d.status == "borrowed"
    ]
    
    if not borrowed_detail_ids:
        raise HTTPException(status_code=400, detail="没有未归还的商品")
    
    # 创建还货单
    return_data = LoanReturnCreate(
        loan_id=loan_id,
        detail_ids=borrowed_detail_ids,
        operator=data.operator,
        remark=data.remark,
    )
    
    return await create_loan_return(return_data, db)


# ============= 获取操作日志 =============

@router.get("/orders/{loan_id}/logs", response_model=List[LoanOrderLogResponse])
async def get_loan_order_logs(
    loan_id: int,
    db: Session = Depends(get_db)
):
    """获取暂借单操作日志"""
    loan_order = db.query(LoanOrder).filter(LoanOrder.id == loan_id).first()
    if not loan_order:
        raise HTTPException(status_code=404, detail="暂借单不存在")
    
    logs = db.query(LoanOrderLog).filter(
        LoanOrderLog.loan_order_id == loan_id
    ).order_by(LoanOrderLog.action_time.desc()).all()
    
    return [LoanOrderLogResponse.model_validate(log) for log in logs]


# ============= 标记已打印 =============

@router.post("/orders/{loan_id}/print")
async def mark_loan_printed(
    loan_id: int,
    db: Session = Depends(get_db)
):
    """标记暂借单为已打印"""
    loan_order = db.query(LoanOrder).filter(LoanOrder.id == loan_id).first()
    if not loan_order:
        raise HTTPException(status_code=404, detail="暂借单不存在")
    
    loan_order.printed_at = china_now()
    db.commit()
    
    logger.info(f"打印暂借单: {loan_order.loan_no}")
    return success_response(message="已标记打印")


@router.post("/returns/{return_id}/print")
async def mark_return_printed(
    return_id: int,
    db: Session = Depends(get_db)
):
    """标记还货单为已打印"""
    loan_return = db.query(LoanReturn).filter(LoanReturn.id == return_id).first()
    if not loan_return:
        raise HTTPException(status_code=404, detail="还货单不存在")
    
    loan_return.printed_at = china_now()
    db.commit()
    
    logger.info(f"打印还货单: {loan_return.return_no}")
    return success_response(message="已标记打印")


# ============= 下载/打印暂借单 =============

@router.get("/orders/{loan_id}/download")
async def download_loan_order(
    loan_id: int,
    format: str = Query("html", pattern="^(html|pdf)$"),
    db: Session = Depends(get_db)
):
    """下载或打印暂借单（商品借货单格式）"""
    loan_order = db.query(LoanOrder).options(
        joinedload(LoanOrder.details)
    ).filter(LoanOrder.id == loan_id).first()
    
    if not loan_order:
        raise HTTPException(status_code=404, detail="暂借单不存在")
    
    details = loan_order.details or []
    
    loan_date_str = format_china_time(to_china_time(loan_order.loan_date), '%Y-%m-%d') if loan_order.loan_date else ""
    print_time = format_china_time(china_now(), '%Y/%m/%d %H:%M')
    
    # 获取展厅名称（分销商）
    showroom = db.query(Location).filter(
        Location.location_type == "showroom",
        Location.is_active == 1
    ).first()
    distributor_name = showroom.name if showroom else ""
    
    # 反查商品编码：product_name -> ProductCode.code
    product_names = [d.product_name for d in details]
    product_code_map = {}
    if product_names:
        codes = db.query(ProductCode).filter(ProductCode.name.in_(product_names)).all()
        for pc in codes:
            product_code_map[pc.name] = pc.code
    
    # 构建商品行 HTML
    rows_html = ""
    total_pieces = 0
    for i, d in enumerate(details, 1):
        code = product_code_map.get(d.product_name, "")
        pc = getattr(d, 'piece_count', None) or ""
        if isinstance(pc, int) and pc > 0:
            total_pieces += pc
        rows_html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{code}</td>
                    <td style="text-align:left;">{d.product_name}</td>
                    <td>{pc}</td>
                    <td>{d.weight:.2f}</td>
                </tr>"""
    
    total_pieces_str = str(total_pieces) if total_pieces > 0 else ""
    rows_html += f"""
                <tr style="font-weight:bold; background:#f9f9f9;">
                    <td colspan="3" style="text-align:right;">合计（{len(details)} 件）</td>
                    <td>{total_pieces_str}</td>
                    <td>{loan_order.total_weight:.2f}</td>
                </tr>"""
    
    remark_html = loan_order.remark or ""
    
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>商品借货单 - {loan_order.loan_no}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: "Microsoft YaHei", "SimHei", Arial, sans-serif; 
            font-size: 11px; color: #333;
        }}
        .page {{ width: 241mm; min-height: 140mm; padding: 6mm 8mm; margin: 0 auto; background: white; }}
        .header {{ text-align: center; margin-bottom: 10px; border-bottom: 2px solid #333; padding-bottom: 8px; }}
        .company-name {{ font-size: 14px; font-weight: bold; letter-spacing: 2px; }}
        .doc-title {{ font-size: 18px; font-weight: bold; margin-top: 4px; letter-spacing: 4px; }}
        .info-grid {{ display: flex; flex-wrap: wrap; margin: 8px 0; font-size: 10px; }}
        .info-col {{ flex: 1; min-width: 33%; }}
        .info-item {{ margin-bottom: 4px; }}
        .info-label {{ font-weight: bold; }}
        .product-table {{ width: 100%; border-collapse: collapse; margin: 8px 0; }}
        .product-table th, .product-table td {{ border: 1px solid #333; padding: 4px 6px; text-align: center; font-size: 10px; }}
        .product-table th {{ background: #f0f0f0; font-weight: bold; }}
        .product-table .amount {{ font-weight: bold; }}
        .signature-section {{ margin-top: 20px; display: flex; justify-content: space-between; font-size: 10px; }}
        .signature-item {{ text-align: center; }}
        .signature-line {{ border-bottom: 1px solid #333; width: 120px; height: 30px; display: inline-block; margin-top: 5px; }}
        .print-btn {{ display: block; margin: 15px auto; padding: 8px 25px; background: #1a4d8c; color: white; border: none; border-radius: 5px; font-size: 12px; cursor: pointer; }}
        .footer {{ margin-top: 10px; text-align: center; font-size: 9px; color: #666; }}
        @media print {{ @page {{ size: 241mm auto; margin: 0; }} .print-btn {{ display: none; }} }}
        @media screen {{ body {{ background: #f0f0f0; padding: 15px; }} .page {{ box-shadow: 0 0 10px rgba(0,0,0,0.1); }} }}
    </style>
</head>
<body>
    <div class="page">
        <div class="header">
            <div class="company-name">深圳市梵贝琳珠宝有限公司</div>
            <div class="doc-title">商 品 借 货 单</div>
        </div>
        
        <div class="info-grid">
            <div class="info-col">
                <div class="info-item"><span class="info-label">借货单号：</span>{loan_order.loan_no}</div>
                <div class="info-item"><span class="info-label">借货日期：</span>{loan_date_str}</div>
                <div class="info-item"><span class="info-label">打印人：</span></div>
            </div>
            <div class="info-col">
                <div class="info-item"><span class="info-label">分销商：</span>{distributor_name}</div>
                <div class="info-item"><span class="info-label">售货员：</span>{loan_order.salesperson}</div>
                <div class="info-item"><span class="info-label">打印日期：</span>{print_time}</div>
            </div>
            <div class="info-col">
                <div class="info-item"><span class="info-label">客户：</span>{loan_order.customer_name}</div>
                <div class="info-item"><span class="info-label">备注：</span>{remark_html}</div>
                <div class="info-item"><span class="info-label">制单审核：</span></div>
            </div>
        </div>
        
        <table class="product-table">
            <thead>
                <tr>
                    <th>序号</th>
                    <th>商品编码</th>
                    <th>商品名称</th>
                    <th>数量</th>
                    <th>重量(g)</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        
        <div class="signature-section">
            <div class="signature-item">
                <div>借方签字：</div>
                <div class="signature-line"></div>
                </div>
            <div class="signature-item">
                <div>贷方签字：</div>
                <div class="signature-line"></div>
                </div>
            <div class="signature-item">
                <div>日期：____年____月____日</div>
                </div>
            </div>
        
        <div class="footer">打印时间：{print_time}</div>
                </div>
    
    <button class="print-btn" onclick="window.print()">打印借货单</button>
</body>
</html>
"""
    
    return HTMLResponse(content=html_content)


# ============= 还货单打印 =============

@router.get("/returns/{return_id}/download")
async def download_loan_return(
    return_id: int,
    format: str = Query("html", pattern="^(html|pdf)$"),
    db: Session = Depends(get_db)
):
    """下载或打印还货单（商品还货单格式）"""
    return_order = db.query(LoanReturn).options(
        joinedload(LoanReturn.details),
        joinedload(LoanReturn.loan_order)
    ).filter(LoanReturn.id == return_id).first()
    
    if not return_order:
        raise HTTPException(status_code=404, detail="还货单不存在")
    
    loan_order = return_order.loan_order
    details = return_order.details or []
    
    return_date_str = format_china_time(to_china_time(return_order.created_at), '%Y-%m-%d') if return_order.created_at else ""
    print_time = format_china_time(china_now(), '%Y/%m/%d %H:%M')
    
    # 获取展厅名称（分销商）
    showroom = db.query(Location).filter(
        Location.location_type == "showroom",
        Location.is_active == 1
    ).first()
    distributor_name = showroom.name if showroom else ""
    
    # 售货员从原暂借单取
    salesperson = loan_order.salesperson if loan_order else ""
    customer_name = return_order.customer_name
    loan_no = loan_order.loan_no if loan_order else ""
    
    # 反查商品编码
    product_names = [d.product_name for d in details]
    product_code_map = {}
    if product_names:
        codes = db.query(ProductCode).filter(ProductCode.name.in_(product_names)).all()
        for pc in codes:
            product_code_map[pc.name] = pc.code
    
    # 构建商品行 HTML
    rows_html = ""
    total_weight_sum = 0.0
    for i, d in enumerate(details, 1):
        code = product_code_map.get(d.product_name, "")
        total_weight_sum += float(d.weight or 0)
        rows_html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{code}</td>
                    <td style="text-align:left;">{d.product_name}</td>
                    <td></td>
                    <td>{d.weight:.2f}</td>
                    <td></td>
                </tr>"""
    
    # 合计行
    rows_html += f"""
                <tr style="font-weight:bold; background:#f9f9f9;">
                    <td colspan="3" style="text-align:right;">合计（{len(details)} 件）</td>
                    <td></td>
                    <td>{total_weight_sum:.2f}</td>
                    <td></td>
                </tr>"""
    
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>商品还货单 - {return_order.return_no}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: "Microsoft YaHei", "SimHei", Arial, sans-serif; 
            font-size: 11px; color: #333;
        }}
        .page {{ width: 241mm; min-height: 140mm; padding: 6mm 8mm; margin: 0 auto; background: white; }}
        .header {{ text-align: center; margin-bottom: 10px; border-bottom: 2px solid #333; padding-bottom: 8px; }}
        .company-name {{ font-size: 14px; font-weight: bold; letter-spacing: 2px; }}
        .doc-title {{ font-size: 18px; font-weight: bold; margin-top: 4px; letter-spacing: 4px; }}
        .info-grid {{ display: flex; flex-wrap: wrap; margin: 8px 0; font-size: 10px; }}
        .info-col {{ flex: 1; min-width: 33%; }}
        .info-item {{ margin-bottom: 4px; }}
        .info-label {{ font-weight: bold; }}
        .product-table {{ width: 100%; border-collapse: collapse; margin: 8px 0; }}
        .product-table th, .product-table td {{ border: 1px solid #333; padding: 4px 6px; text-align: center; font-size: 10px; }}
        .product-table th {{ background: #f0f0f0; font-weight: bold; }}
        .signature-section {{ margin-top: 20px; display: flex; justify-content: space-between; font-size: 10px; }}
        .signature-item {{ text-align: center; }}
        .signature-line {{ border-bottom: 1px solid #333; width: 120px; height: 30px; display: inline-block; margin-top: 5px; }}
        .print-btn {{ display: block; margin: 15px auto; padding: 8px 25px; background: #1a4d8c; color: white; border: none; border-radius: 5px; font-size: 12px; cursor: pointer; }}
        .footer {{ margin-top: 10px; text-align: center; font-size: 9px; color: #666; }}
        @media print {{ @page {{ size: 241mm auto; margin: 0; }} .print-btn {{ display: none; }} }}
        @media screen {{ body {{ background: #f0f0f0; padding: 15px; }} .page {{ box-shadow: 0 0 10px rgba(0,0,0,0.1); }} }}
    </style>
</head>
<body>
    <div class="page">
        <div class="header">
            <div class="company-name">深圳市梵贝琳珠宝有限公司</div>
            <div class="doc-title">商 品 还 货 单</div>
        </div>
        
        <div class="info-grid">
            <div class="info-col">
                <div class="info-item"><span class="info-label">还货单号：</span>{return_order.return_no}</div>
                <div class="info-item"><span class="info-label">还货日期：</span>{return_date_str}</div>
                <div class="info-item"><span class="info-label">打印人：</span></div>
            </div>
            <div class="info-col">
                <div class="info-item"><span class="info-label">分销商：</span>{distributor_name}</div>
                <div class="info-item"><span class="info-label">售货员：</span>{salesperson}</div>
                <div class="info-item"><span class="info-label">打印日期：</span>{print_time}</div>
            </div>
            <div class="info-col">
                <div class="info-item"><span class="info-label">客户：</span>{customer_name}</div>
                <div class="info-item"><span class="info-label">备注：</span>{loan_no}</div>
                <div class="info-item"><span class="info-label">制单审核：</span></div>
            </div>
        </div>
        
            <table class="product-table">
                <thead>
                    <tr>
                    <th>序号</th>
                    <th>商品编码</th>
                    <th>商品名称</th>
                    <th>数量</th>
                    <th>重量(g)</th>
                    <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                {rows_html}
                </tbody>
            </table>
        
        <div class="signature-section">
            <div class="signature-item">
                <div>还方签字：</div>
                <div class="signature-line"></div>
            </div>
            <div class="signature-item">
                <div>收方签字：</div>
                <div class="signature-line"></div>
            </div>
            <div class="signature-item">
                <div>日期：____年____月____日</div>
        </div>
    </div>
    
        <div class="footer">打印时间：{print_time}</div>
    </div>
    
    <button class="print-btn" onclick="window.print()">打印还货单</button>
</body>
</html>
"""
    
    return HTMLResponse(content=html_content)


# ============= 借货汇总 =============

@router.get("/summary")
async def get_loan_summary(db: Session = Depends(get_db)):
    """获取未归还暂借单汇总（全局，不分页）"""
    outstanding_weight = db.query(func.sum(LoanOrder.total_weight)).filter(
        LoanOrder.status.in_(["borrowed", "partial_returned"])
    ).scalar() or 0
    return {"success": True, "data": {"outstanding_weight": round(float(outstanding_weight), 3)}}


# ============= 借货统计报表 =============

@router.get("/statistics")
async def get_loan_statistics(
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """获取借货统计报表（基于主表汇总字段）"""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import and_
        
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        
        all_orders = db.query(LoanOrder).filter(
            LoanOrder.created_at >= start_dt,
            LoanOrder.created_at < end_dt
        ).all()
        
        total_borrowed_count = len([o for o in all_orders if o.status in ['borrowed', 'returned', 'partial_returned']])
        total_returned_count = len([o for o in all_orders if o.status == 'returned'])
        outstanding_count = len([o for o in all_orders if o.status in ['borrowed', 'partial_returned']])
        cancelled_count = len([o for o in all_orders if o.status == 'cancelled'])
        pending_count = len([o for o in all_orders if o.status == 'pending'])
        
        total_borrowed_weight = sum(float(o.total_weight or 0) for o in all_orders if o.status in ['borrowed', 'returned', 'partial_returned'])
        total_returned_weight = sum(float(o.total_weight or 0) for o in all_orders if o.status == 'returned')
        outstanding_weight = sum(float(o.total_weight or 0) for o in all_orders if o.status in ['borrowed', 'partial_returned'])
        
        daily_breakdown = {}
        for order in all_orders:
            if order.status in ['borrowed', 'returned', 'partial_returned', 'pending']:
                date_key = order.created_at.strftime('%Y-%m-%d')
                if date_key not in daily_breakdown:
                    daily_breakdown[date_key] = {
                        'date': date_key,
                        'borrowed_count': 0, 'returned_count': 0,
                        'borrowed_weight': 0.0, 'returned_weight': 0.0,
                    }
                
                if order.status in ['borrowed', 'returned', 'partial_returned']:
                    daily_breakdown[date_key]['borrowed_count'] += 1
                    daily_breakdown[date_key]['borrowed_weight'] += float(order.total_weight or 0)
                
                if order.status == 'returned':
                    daily_breakdown[date_key]['returned_count'] += 1
                    daily_breakdown[date_key]['returned_weight'] += float(order.total_weight or 0)
        
        daily_list = sorted(daily_breakdown.values(), key=lambda x: x['date'])
        
        return {
            "success": True,
            "data": {
                "period": {"start_date": start_date, "end_date": end_date},
                "summary": {
                    "total_borrowed_count": total_borrowed_count,
                    "total_returned_count": total_returned_count,
                    "outstanding_count": outstanding_count,
                    "pending_count": pending_count,
                    "cancelled_count": cancelled_count,
                    "total_borrowed_weight": round(total_borrowed_weight, 2),
                    "total_returned_weight": round(total_returned_weight, 2),
                    "outstanding_weight": round(outstanding_weight, 2),
                },
                "daily_breakdown": daily_list,
            }
        }
    except Exception as e:
        logger.error(f"获取借货统计失败: {e}", exc_info=True)
        return error_response(message=str(e))


@router.get("/statistics/export")
async def export_loan_statistics(
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """导出借货统计报表（Excel格式）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import io
        
        stats_response = await get_loan_statistics(start_date, end_date, db)
        if not stats_response.get("success"):
            raise HTTPException(status_code=500, detail="获取统计数据失败")
        
        stats = stats_response["data"]
        summary = stats["summary"]
        daily_breakdown = stats["daily_breakdown"]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "借货统计报表"
        
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f"借货统计报表 ({start_date} 至 {end_date})"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        
        row = 3
        ws[f'A{row}'] = "汇总统计"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        summary_data = [
            ("借出笔数", summary["total_borrowed_count"], "笔"),
            ("归还笔数", summary["total_returned_count"], "笔"),
            ("未归还笔数", summary["outstanding_count"], "笔"),
            ("借出总克重", summary["total_borrowed_weight"], "克"),
            ("归还总克重", summary["total_returned_weight"], "克"),
            ("未归还克重", summary["outstanding_weight"], "克"),
        ]
        
        for item_name, item_value, item_unit in summary_data:
            ws[f'A{row}'] = item_name
            ws[f'B{row}'] = item_value
            ws[f'C{row}'] = item_unit
            row += 1
        
        row += 1
        ws[f'A{row}'] = "每日明细"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        headers = ["日期", "借出笔数", "归还笔数", "借出克重", "归还克重"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
        
        for day in daily_breakdown:
            ws.cell(row=row, column=1, value=day['date']).border = thin_border
            ws.cell(row=row, column=2, value=day['borrowed_count']).border = thin_border
            ws.cell(row=row, column=3, value=day['returned_count']).border = thin_border
            ws.cell(row=row, column=4, value=round(day['borrowed_weight'], 2)).border = thin_border
            ws.cell(row=row, column=5, value=round(day['returned_weight'], 2)).border = thin_border
            row += 1
        
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"loan_statistics_{start_date}_{end_date}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出借货统计失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

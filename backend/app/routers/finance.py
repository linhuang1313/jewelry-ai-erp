"""
财务对账模块 - API路由
"""

import logging
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from typing import Optional, List

from ..database import get_db
from ..services.finance_service import FinanceService
from ..models import SalesOrder
from ..dependencies.auth import get_current_role, require_permission
from ..schemas.finance import (
    AccountReceivableResponse,
    PaymentRecordCreate,
    PaymentRecordResponse,
    ReminderRecordCreate,
    ReminderRecordResponse,
    ReconciliationStatementCreate,
    ReconciliationStatementResponse,
    ReconciliationSalesDetail,
    ReconciliationPaymentDetail,
    FinanceStatistics,
    ApiResponse,
    ReceivableListResponse,
    StatisticsResponse,
    CustomerReference,
    SalesOrderReference,
)
import json

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/finance", tags=["财务对账"])


def _generate_safe_no(db: Session, model, column, prefix: str, seq_digits: int = 3, max_retries: int = 3) -> str:
    from datetime import datetime
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    full_prefix = f"{prefix}{date_str}"

    for attempt in range(max_retries):
        latest_row = (
            db.query(column)
            .filter(column.like(f"{full_prefix}%"))
            .order_by(column.desc())
            .limit(1)
            .with_for_update()
            .scalar()
        )

        if latest_row and latest_row.startswith(full_prefix):
            try:
                seq = int(latest_row[len(full_prefix):]) + 1
            except ValueError:
                seq = 1
        else:
            seq = 1

        new_no = f"{full_prefix}{seq:0{seq_digits}d}"

        existing = db.query(model).filter(column == new_no).first()
        if existing is None:
            return new_no

        logger.warning(
            "Number collision: %s (attempt %d/%d)",
            new_no, attempt + 1, max_retries,
        )

    fallback_ts = now.strftime("%H%M%S")
    return f"{full_prefix}{fallback_ts}"


@router.get("/receivables", response_model=ReceivableListResponse)
async def get_receivables(
    filter_type: str = Query("all", description="筛选类型: all/unpaid/overdue/due_this_month"),
    search: Optional[str] = Query(None, description="搜索客户名称"),
    sort_by: str = Query("overdue_days", description="排序字段: amount/overdue_days/due_date"),
    sort_order: str = Query("desc", description="排序方向: asc/desc"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    sales_order_no: Optional[str] = Query(None, description="销售单号"),
    settlement_no: Optional[str] = Query(None, description="结算单号"),
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回记录数"),
    db: Session = Depends(get_db)
):
    """
    获取应收账款列表
    
    - **filter_type**: 筛选类型
        - all: 全部
        - unpaid: 未付清
        - overdue: 已逾期
        - due_this_month: 本月到期
    - **search**: 按客户名称搜索
    - **sort_by**: 排序字段 (amount/overdue_days/due_date)
    - **sort_order**: 排序方向 (asc/desc)
    - **start_date**: 开始日期 (按销售日期筛选)
    - **end_date**: 结束日期 (按销售日期筛选)
    - **sales_order_no**: 销售单号
    - **settlement_no**: 结算单号
    """
    try:
        service = FinanceService(db)
        receivables = await service.get_receivables(
            filter_type=filter_type,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
            start_date=start_date,
            end_date=end_date,
            sales_order_no=sales_order_no,
            settlement_no=settlement_no,
            skip=skip,
            limit=limit
        )
        total = await service.get_receivable_count(filter_type)
        
        # 转换为响应格式
        response_data = []
        for r in receivables:
            customer_ref = None
            if r.customer:
                customer_ref = CustomerReference(
                    id=r.customer.id,
                    customer_no=r.customer.customer_no,
                    name=r.customer.name,
                    phone=r.customer.phone,
                    wechat=r.customer.wechat
                )
            
            sales_order_ref = None
            if r.sales_order:
                sales_order_ref = SalesOrderReference(
                    id=r.sales_order.id,
                    order_no=r.sales_order.order_no,
                    order_date=r.sales_order.order_date,
                    salesperson=r.sales_order.salesperson,
                    store_code=r.sales_order.store_code,
                    total_amount=r.sales_order.total_labor_cost
                )
            
            response_data.append(AccountReceivableResponse(
                id=r.id,
                sales_order_id=r.sales_order_id,
                customer_id=r.customer_id,
                total_amount=r.total_amount,
                received_amount=r.received_amount,
                unpaid_amount=r.unpaid_amount,
                credit_days=r.credit_days,
                credit_start_date=r.credit_start_date,
                due_date=r.due_date,
                overdue_days=r.overdue_days,
                status=r.status,
                is_overdue=r.is_overdue,
                salesperson=r.salesperson,
                store_code=r.store_code,
                contract_no=r.contract_no,
                invoice_no=r.invoice_no,
                expected_payment_date=r.expected_payment_date,
                remark=r.remark,
                create_time=r.create_time,
                update_time=r.update_time,
                operator=r.operator,
                last_updater=r.last_updater,
                customer=customer_ref,
                sales_order=sales_order_ref
            ))
        
        return ReceivableListResponse(
            success=True,
            data=response_data,
            total=total
        )
        
    except Exception as e:
        logger.error(f"获取应收账款列表失败: {e}", exc_info=True)
        return ReceivableListResponse(
            success=False,
            data=[],
            total=0,
            error=str(e)
        )


@router.post("/payment", response_model=ApiResponse)
async def record_payment(
    payment_data: PaymentRecordCreate,
    db: Session = Depends(get_db)
):
    """
    记录收款
    
    在事务中处理，自动更新应收账款状态
    """
    try:
        service = FinanceService(db)
        payment = await service.record_payment(payment_data)
        
        return ApiResponse(
            success=True,
            data={
                "payment_id": payment.id,
                "amount": payment.amount,
                "message": "收款记录成功"
            },
            message="收款记录成功"
        )
        
    except ValueError as e:
        logger.warning(f"收款验证失败: {e}")
        return ApiResponse(
            success=False,
            error=str(e)
        )
    except Exception as e:
        logger.error(f"记录收款失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=f"记录收款失败: {str(e)}"
        )


@router.post("/payment/chat", response_model=ApiResponse)
async def record_payment_from_chat(
    customer_id: int,
    amount: float,
    payment_method: str = "bank_transfer",
    payment_date: str = None,
    remark: str = "",
    role: str = Depends(require_permission("can_record_payment")),
    db: Session = Depends(get_db)
):
    """
    从聊天界面登记收款
    
    创建一笔完整的收款记录，然后FIFO方式更新应收账款余额
    """
    try:
        from ..models.finance import AccountReceivable, PaymentRecord
        from ..models import Customer
        from datetime import date, datetime
        from ..timezone_utils import china_now
        
        # 验证客户存在
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return ApiResponse(
                success=False,
                error=f"客户不存在: ID={customer_id}"
            )
        
        # 解析日期
        if payment_date:
            try:
                pay_date = datetime.strptime(payment_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pay_date = date.today()
        else:
            pay_date = date.today()
        
        # 生成收款单号 SK + 时间戳
        now = china_now()
        payment_no = f"SK{now.strftime('%Y%m%d%H%M%S')}"
        
        # ========== 创建一笔完整的收款记录（财务端点直接确认） ==========
        payment = PaymentRecord(
            payment_no=payment_no,
            account_receivable_id=None,  # 不关联具体应收账款
            customer_id=customer_id,
            payment_date=pay_date,
            amount=amount,  # 实际收款金额
            payment_method=payment_method,
            remark=remark or "收款登记",
            operator="财务",
            status="confirmed",
            confirmed_by="财务",
            confirmed_at=now
        )
        db.add(payment)
        db.flush()
        
        # ========== 锁定并FIFO方式更新应收账款余额 ==========
        unpaid_receivables = db.query(AccountReceivable).filter(
            AccountReceivable.customer_id == customer_id,
            AccountReceivable.status.in_(["unpaid", "overdue"]),
            AccountReceivable.unpaid_amount > 0
        ).order_by(AccountReceivable.credit_start_date.asc()).with_for_update().all()
        
        remaining_amount = amount
        offset_details = []  # 冲抵明细
        
        for receivable in unpaid_receivables:
            if remaining_amount <= 0:
                break
            
            # 计算本次可冲抵金额
            offset_amount = min(remaining_amount, receivable.unpaid_amount)
            
            # 更新应收账款
            receivable.received_amount = float(receivable.received_amount or 0) + offset_amount
            receivable.unpaid_amount = float(receivable.total_amount or 0) - float(receivable.received_amount or 0)
            
            # 更新状态
            if receivable.unpaid_amount <= 0:
                receivable.status = "paid"
            
            offset_details.append({
                "receivable_id": receivable.id,
                "amount": offset_amount
            })
            
            remaining_amount -= offset_amount
        
        # 如果还有剩余金额（预收款），创建独立的预收款记录，不修改任何AR
        prepayment = 0
        if remaining_amount > 0:
            from ..models import CustomerTransaction
            prepayment_tx = CustomerTransaction(
                customer_id=customer_id,
                customer_name=customer.name,
                transaction_type='prepayment',
                amount=remaining_amount,
                status='active',
                remark=f"预收款：收款单 {payment_no}，超额部分 {remaining_amount:.2f}元"
            )
            db.add(prepayment_tx)
            prepayment = remaining_amount
            logger.info(f"预收款: 客户={customer.name}, 金额={remaining_amount}，已创建独立预收款记录")
        
        # ========== 记录资金流水 ==========
        try:
            from ..models.finance import BankAccount, CashFlow
            
            # 锁定默认账户防止并发余额修改
            default_account = db.query(BankAccount).filter(
                BankAccount.is_default == True,
                BankAccount.status == "active"
            ).with_for_update().first()
            
            if default_account:
                balance_before = default_account.current_balance
                default_account.current_balance += amount
                
                flow_no = _generate_safe_no(db, CashFlow, CashFlow.flow_no, "LS", seq_digits=4)
                
                cash_flow = CashFlow(
                    flow_no=flow_no,
                    account_id=default_account.id,
                    flow_type="income",
                    category="销售收款",
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=default_account.current_balance,
                    related_type="payment_record",
                    related_id=payment.id,
                    flow_date=now,
                    counterparty=customer.name,
                    remark=f"收款单：{payment_no}",
                    created_by="财务"
                )
                db.add(cash_flow)
                logger.info(f"记录资金流水: {flow_no}, 收入={amount}")
        except Exception as e:
            logger.warning(f"记录资金流水失败（不影响收款）: {e}")
        # ========== 资金流水记录结束 ==========
        
        db.commit()
        db.refresh(payment)
        
        message = f"收款登记成功，收款单号：{payment_no}，金额：¥{amount:.2f}"
        if prepayment > 0:
            message += f"（含预收款¥{prepayment:.2f}）"
        
        logger.info(f"收款成功: 客户={customer.name}, 单号={payment_no}, 金额={amount}, 冲抵{len(offset_details)}笔应收账款")
        
        return ApiResponse(
            success=True,
            data={
                "payment_id": payment.id,
                "payment_no": payment_no,
                "total_paid": amount,
                "prepayment": prepayment,
                "offset_count": len(offset_details),
                "offset_details": offset_details,
                "message": message
            },
            message="收款登记成功"
        )
        
    except Exception as e:
        db.rollback()
        logger.error(f"聊天收款失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=f"收款登记失败: {str(e)}"
        )


@router.post("/reminder", response_model=ApiResponse)
async def record_reminder(
    reminder_data: ReminderRecordCreate,
    db: Session = Depends(get_db)
):
    """
    记录催款
    """
    try:
        service = FinanceService(db)
        reminder = await service.record_reminder(reminder_data)
        
        return ApiResponse(
            success=True,
            data={
                "reminder_id": reminder.id,
                "message": "催款记录成功"
            },
            message="催款记录成功"
        )
        
    except ValueError as e:
        logger.warning(f"催款验证失败: {e}")
        return ApiResponse(
            success=False,
            error=str(e)
        )
    except Exception as e:
        logger.error(f"记录催款失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=f"记录催款失败: {str(e)}"
        )


@router.post("/statement", response_model=ApiResponse)
async def generate_statement(
    statement_data: ReconciliationStatementCreate,
    db: Session = Depends(get_db)
):
    """
    生成对账单
    
    自动计算期初欠款、本期销售、本期收款、期末欠款
    返回合并的往来账明细表
    """
    try:
        service = FinanceService(db)
        statement = await service.generate_statement(statement_data)
        
        # 解析明细数据
        sales_details = json.loads(statement.sales_details) if statement.sales_details else []
        payment_details = json.loads(statement.payment_details) if statement.payment_details else []
        
        # 获取合并的往来明细（如果存在）
        transactions = getattr(statement, '_transactions', [])
        opening_gold = getattr(statement, '_opening_gold', 0.0)
        closing_gold = getattr(statement, '_closing_gold', 0.0)
        total_gold = getattr(statement, '_total_gold', 0.0)
        total_cash = getattr(statement, '_total_cash', 0.0)
        
        # 获取客户信息
        customer_ref = None
        if statement.customer:
            customer_ref = {
                "id": statement.customer.id,
                "customer_no": statement.customer.customer_no,
                "name": statement.customer.name,
                "phone": statement.customer.phone
            }
        
        return ApiResponse(
            success=True,
            data={
                "statement_no": statement.statement_no,
                "customer": customer_ref,
                "period": {
                    "start": statement.period_start_date.isoformat(),
                    "end": statement.period_end_date.isoformat()
                },
                "summary": {
                    "openingBalance": statement.opening_balance,
                    "openingGold": round(opening_gold, 3),
                    "totalSales": statement.period_sales_amount,
                    "totalPayments": statement.period_payment_amount,
                    "totalGold": round(total_gold, 3),
                    "totalCash": round(total_cash, 2),
                    "closingBalance": statement.closing_balance,
                    "closingGold": round(closing_gold, 3)
                },
                "transactions": transactions,  # 合并的往来明细
                "salesDetails": sales_details,  # 保持向后兼容
                "paymentDetails": payment_details,  # 保持向后兼容
                "generatedAt": statement.create_time.isoformat()
            },
            message="对账单生成成功"
        )
        
    except ValueError as e:
        logger.warning(f"生成对账单验证失败: {e}")
        return ApiResponse(
            success=False,
            error=str(e)
        )
    except Exception as e:
        logger.error(f"生成对账单失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=f"生成对账单失败: {str(e)}"
        )



@router.post("/statement/excel")
async def export_statement_excel(
    customer_id: int = Query(..., description="客户ID"),
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """
    导出对账单为 Excel 文件
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return ApiResponse(success=False, error="服务器未安装 openpyxl 库")
    
    try:
        from ..models import Customer, SalesOrder, SettlementOrder
        from ..models.finance import GoldReceipt, PaymentRecord
        
        # 验证客户
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return ApiResponse(success=False, error="客户不存在")
        
        # 解析日期
        period_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        period_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        # ========== 获取往来明细 ==========
        transactions = []
        
        # 1. 销售结算记录
        settlements = db.query(SettlementOrder).join(SalesOrder).filter(
            SalesOrder.customer_id == customer_id,
            SettlementOrder.status.in_(['confirmed', 'printed']),
            SettlementOrder.created_at >= datetime.combine(period_start, datetime.min.time()),
            SettlementOrder.created_at <= datetime.combine(period_end, datetime.max.time())
        ).all()
        
        for s in settlements:
            gold_amount = 0.0
            cash_amount = 0.0
            if s.payment_method == 'physical_gold':
                gold_amount = s.physical_gold_weight or 0.0
            elif s.payment_method == 'cash_price':
                cash_amount = s.total_amount or 0.0
            elif s.payment_method == 'mixed':
                gold_amount = s.gold_payment_weight or 0.0
                cash_amount = s.cash_payment_weight or 0.0
            
            remark_parts = []
            if s.sales_order and s.sales_order.salesperson:
                remark_parts.append(s.sales_order.salesperson)
            if s.sales_order and s.sales_order.store_code:
                remark_parts.append(s.sales_order.store_code)
            
            transactions.append({
                "date": s.created_at.strftime("%Y-%m-%d") if s.created_at else "",
                "type": "销售结算",
                "order_no": s.settlement_no,
                "gold_amount": round(gold_amount, 3),
                "cash_amount": round(cash_amount, 2),
                "remark": ' '.join(remark_parts)
            })
        
        # 2. 客户来料记录
        receipts = db.query(GoldReceipt).filter(
            GoldReceipt.customer_id == customer_id,
            GoldReceipt.status == 'received',
            GoldReceipt.created_at >= datetime.combine(period_start, datetime.min.time()),
            GoldReceipt.created_at <= datetime.combine(period_end, datetime.max.time())
        ).all()
        
        for r in receipts:
            transactions.append({
                "date": r.received_at.strftime("%Y-%m-%d") if r.received_at else (r.created_at.strftime("%Y-%m-%d") if r.created_at else ""),
                "type": "客户来料",
                "order_no": r.receipt_no or "",
                "gold_amount": round(-(r.gold_weight or 0), 3),
                "cash_amount": 0.0,
                "remark": r.gold_fineness or ""
            })
        
        # 3. 客户来款记录
        payments = db.query(PaymentRecord).filter(
            PaymentRecord.customer_id == customer_id,
            PaymentRecord.payment_date >= period_start,
            PaymentRecord.payment_date <= period_end
        ).all()
        
        for p in payments:
            payment_method_label = {
                'bank_transfer': '银行转账',
                'cash': '现金',
                'wechat': '微信',
                'alipay': '支付宝',
                'card': '刷卡'
            }.get(p.payment_method, p.payment_method or "")
            
            transactions.append({
                "date": p.payment_date.strftime("%Y-%m-%d") if p.payment_date else "",
                "type": "客户来款",
                "order_no": p.payment_no or f"SK{p.id}",
                "gold_amount": 0.0,
                "cash_amount": round(-(p.amount or 0), 2),
                "remark": payment_method_label
            })
        
        # 按日期排序
        transactions.sort(key=lambda x: x.get("date") or "0000-00-00")
        
        # 计算合计
        total_gold = sum(t["gold_amount"] for t in transactions)
        total_cash = sum(t["cash_amount"] for t in transactions)
        
        # ========== 创建 Excel ==========
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "客户往来账明细表"
        
        # 样式
        header_font = Font(bold=True, size=16, color="006400")
        title_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = "客户往来账明细表"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        
        # 信息行
        ws['A3'] = f"统计日期：{start_date} 到 {end_date}"
        ws['E3'] = f"打印日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A4'] = f"往来客户：{customer.name}"
        ws['E4'] = f"客户编号：{customer.customer_no or ''}"
        
        # 表头
        headers = ['序号', '发生日期', '往来类型', '往来单号', '足金', '欠款金额', '单据备注']
        col_widths = [8, 12, 12, 20, 12, 15, 25]
        
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = title_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 数据行
        row = 7
        for i, tx in enumerate(transactions, 1):
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=tx["date"]).border = thin_border
            ws.cell(row=row, column=3, value=tx["type"]).border = thin_border
            ws.cell(row=row, column=4, value=tx["order_no"]).border = thin_border
            
            gold_cell = ws.cell(row=row, column=5, value=tx["gold_amount"] if tx["gold_amount"] != 0 else "")
            gold_cell.border = thin_border
            gold_cell.alignment = right_align
            if tx["gold_amount"] < 0:
                gold_cell.font = Font(color="0000FF")
            elif tx["gold_amount"] > 0:
                gold_cell.font = Font(color="FF8C00")
            
            cash_cell = ws.cell(row=row, column=6, value=tx["cash_amount"] if tx["cash_amount"] != 0 else "")
            cash_cell.border = thin_border
            cash_cell.alignment = right_align
            if tx["cash_amount"] < 0:
                cash_cell.font = Font(color="008000")
            elif tx["cash_amount"] > 0:
                cash_cell.font = Font(color="FF0000")
            
            ws.cell(row=row, column=7, value=tx["remark"]).border = thin_border
            row += 1
        
        # 合计行
        ws.cell(row=row, column=1, value="合计").border = thin_border
        ws.cell(row=row, column=2, value="").border = thin_border
        ws.cell(row=row, column=3, value="").border = thin_border
        ws.cell(row=row, column=4, value="").border = thin_border
        ws.cell(row=row, column=5, value=round(total_gold, 3)).border = thin_border
        ws.cell(row=row, column=6, value=round(total_cash, 2)).border = thin_border
        ws.cell(row=row, column=7, value="").border = thin_border
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).font = title_font
            ws.cell(row=row, column=col).fill = header_fill
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 使用 URL 编码处理中文文件名
        from urllib.parse import quote
        filename = f"客户往来账_{customer.name}_{start_date}_{end_date}.xlsx"
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )
        
    except Exception as e:
        logger.error(f"导出Excel失败: {e}", exc_info=True)
        return ApiResponse(success=False, error=f"导出失败: {str(e)}")


@router.get("/statistics", response_model=StatisticsResponse)
async def get_statistics(db: Session = Depends(get_db)):
    """
    获取财务统计
    
    返回:
    - 总应收账款
    - 本月回款
    - 逾期金额
    - 逾期客户数
    - 环比变化百分比
    """
    try:
        service = FinanceService(db)
        statistics = await service.get_statistics()
        
        return StatisticsResponse(
            success=True,
            data=statistics
        )
        
    except Exception as e:
        logger.error(f"获取财务统计失败: {e}", exc_info=True)
        return StatisticsResponse(
            success=False,
            error=str(e)
        )


@router.get("/receivables/{receivable_id}", response_model=ApiResponse)
async def get_receivable_detail(
    receivable_id: int,
    db: Session = Depends(get_db)
):
    """
    获取应收账款详情
    """
    try:
        from ..models.finance import AccountReceivable
        
        receivable = db.query(AccountReceivable).filter(
            AccountReceivable.id == receivable_id
        ).first()
        
        if not receivable:
            return ApiResponse(
                success=False,
                error=f"应收账款不存在: {receivable_id}"
            )
        
        customer_ref = None
        if receivable.customer:
            customer_ref = {
                "id": receivable.customer.id,
                "customer_no": receivable.customer.customer_no,
                "name": receivable.customer.name,
                "phone": receivable.customer.phone
            }
        
        sales_order_ref = None
        if receivable.sales_order:
            sales_order_ref = {
                "id": receivable.sales_order.id,
                "order_no": receivable.sales_order.order_no,
                "order_date": receivable.sales_order.order_date.isoformat(),
                "salesperson": receivable.sales_order.salesperson,
                "total_amount": receivable.sales_order.total_labor_cost
            }
        
        return ApiResponse(
            success=True,
            data={
                "id": receivable.id,
                "sales_order_id": receivable.sales_order_id,
                "customer_id": receivable.customer_id,
                "total_amount": receivable.total_amount,
                "received_amount": receivable.received_amount,
                "unpaid_amount": receivable.unpaid_amount,
                "credit_days": receivable.credit_days,
                "credit_start_date": receivable.credit_start_date.isoformat(),
                "due_date": receivable.due_date.isoformat(),
                "overdue_days": receivable.overdue_days,
                "status": receivable.status,
                "is_overdue": receivable.is_overdue,
                "salesperson": receivable.salesperson,
                "store_code": receivable.store_code,
                "customer": customer_ref,
                "sales_order": sales_order_ref
            }
        )
        
    except Exception as e:
        logger.error(f"获取应收账款详情失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=str(e)
        )


@router.get("/payments", response_model=ApiResponse)
async def get_payment_records(
    customer_id: Optional[int] = Query(None, description="客户ID"),
    receivable_id: Optional[int] = Query(None, description="应收账款ID"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    sales_order_no: Optional[str] = Query(None, description="销售单号"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """
    获取收款记录列表
    
    - **start_date**: 开始日期 (按收款日期筛选)
    - **end_date**: 结束日期 (按收款日期筛选)
    - **sales_order_no**: 销售单号
    """
    try:
        from ..models.finance import PaymentRecord, AccountReceivable
        from datetime import datetime
        
        query = db.query(PaymentRecord)
        
        if customer_id:
            query = query.filter(PaymentRecord.customer_id == customer_id)
        if receivable_id:
            query = query.filter(PaymentRecord.account_receivable_id == receivable_id)
        
        # 时间范围筛选（按收款日期）
        if start_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(PaymentRecord.payment_date >= start)
            except ValueError:
                pass
        
        if end_date:
            try:
                end = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(PaymentRecord.payment_date <= end)
            except ValueError:
                pass
        
        # 销售单号筛选（需要联表）
        if sales_order_no:
            query = query.join(
                AccountReceivable, 
                PaymentRecord.account_receivable_id == AccountReceivable.id
            ).join(
                SalesOrder,
                AccountReceivable.sales_order_id == SalesOrder.id
            ).filter(SalesOrder.order_no.contains(sales_order_no))
        
        payments = query.order_by(PaymentRecord.payment_date.desc()).offset(skip).limit(limit).all()
        
        data = []
        for p in payments:
            customer_ref = None
            if p.customer:
                customer_ref = {
                    "id": p.customer.id,
                    "customer_no": p.customer.customer_no,
                    "name": p.customer.name
                }
            
            data.append({
                "id": p.id,
                "account_receivable_id": p.account_receivable_id,
                "customer_id": p.customer_id,
                "payment_date": p.payment_date.isoformat(),
                "amount": p.amount,
                "payment_method": p.payment_method,
                "voucher_images": p.voucher_images,
                "bank_name": p.bank_name,
                "remark": p.remark,
                "operator": p.operator,
                "create_time": p.create_time.isoformat(),
                "customer": customer_ref
            })
        
        return ApiResponse(
            success=True,
            data={"records": data, "total": len(data)}
        )
        
    except Exception as e:
        logger.error(f"获取收款记录失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=str(e)
        )


@router.get("/receipts/{payment_no}")
async def get_receipt_detail(payment_no: str, db: Session = Depends(get_db)):
    """获取收据详情（标准财务收据格式，供打印/预览）"""
    try:
        from ..models.finance import PaymentRecord

        payment = db.query(PaymentRecord).filter(
            PaymentRecord.payment_no == payment_no
        ).first()
        if not payment:
            return {"success": False, "message": f"收据 {payment_no} 不存在"}

        customer_name = ""
        if payment.customer:
            customer_name = payment.customer.name

        def _amount_to_chinese(amount: float) -> str:
            """人民币金额转大写"""
            digits = "零壹贰叁肆伍陆柒捌玖"
            units_int = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]
            units_dec = ["角", "分"]

            if amount < 0.005:
                return "零元整"

            rounded = round(amount, 2)
            int_part = int(rounded)
            dec_part = round((rounded - int_part) * 100)

            result = ""
            if int_part == 0:
                result = "零"
            else:
                str_int = str(int_part)
                n = len(str_int)
                for i, ch in enumerate(str_int):
                    d = int(ch)
                    pos = n - 1 - i
                    if d != 0:
                        result += digits[d] + units_int[pos]
                    else:
                        if not result.endswith("零"):
                            result += "零"
                result = result.rstrip("零")

            result += "元"
            if dec_part == 0:
                result += "整"
            else:
                jiao = dec_part // 10
                fen = dec_part % 10
                if jiao > 0:
                    result += digits[jiao] + "角"
                elif fen > 0:
                    result += "零"
                if fen > 0:
                    result += digits[fen] + "分"

            return result

        amount = float(payment.amount or 0)
        gold = float(payment.gold_amount or 0)
        labor = float(payment.labor_amount or 0)

        method_labels = {
            "cash": "现金",
            "bank_transfer": "银行转账",
            "wechat": "微信",
            "alipay": "支付宝",
            "card": "刷卡",
            "check": "支票",
        }

        return {
            "success": True,
            "data": {
                "payment_no": payment.payment_no,
                "payment_date": payment.payment_date.isoformat() if payment.payment_date else None,
                "customer_name": customer_name,
                "amount": amount,
                "amount_chinese": _amount_to_chinese(amount),
                "gold_amount": gold,
                "labor_amount": labor,
                "receipt_reason": payment.receipt_reason or "货款",
                "payment_method": payment.payment_method,
                "payment_method_label": method_labels.get(payment.payment_method, payment.payment_method),
                "bank_name": payment.bank_name,
                "voucher_images": payment.voucher_images,
                "operator": payment.operator,
                "confirmed_by": payment.confirmed_by,
                "reviewed_by": payment.reviewed_by,
                "action_card_id": payment.action_card_id,
                "create_time": payment.create_time.isoformat() if payment.create_time else None,
            },
        }

    except Exception as e:
        logger.error(f"获取收据详情失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@router.get("/reminders", response_model=ApiResponse)
async def get_reminder_records(
    customer_id: Optional[int] = Query(None, description="客户ID"),
    receivable_id: Optional[int] = Query(None, description="应收账款ID"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """
    获取催款记录列表
    """
    try:
        from ..models.finance import ReminderRecord
        
        query = db.query(ReminderRecord)
        
        if customer_id:
            query = query.filter(ReminderRecord.customer_id == customer_id)
        if receivable_id:
            query = query.filter(ReminderRecord.account_receivable_id == receivable_id)
        
        reminders = query.order_by(ReminderRecord.reminder_date.desc()).offset(skip).limit(limit).all()
        
        data = []
        for r in reminders:
            customer_ref = None
            if r.customer:
                customer_ref = {
                    "id": r.customer.id,
                    "customer_no": r.customer.customer_no,
                    "name": r.customer.name
                }
            
            data.append({
                "id": r.id,
                "account_receivable_id": r.account_receivable_id,
                "customer_id": r.customer_id,
                "reminder_date": r.reminder_date.isoformat(),
                "reminder_person": r.reminder_person,
                "reminder_method": r.reminder_method,
                "reminder_content": r.reminder_content,
                "customer_feedback": r.customer_feedback,
                "promised_payment_date": r.promised_payment_date.isoformat() if r.promised_payment_date else None,
                "status": r.status,
                "create_time": r.create_time.isoformat(),
                "customer": customer_ref
            })
        
        return ApiResponse(
            success=True,
            data={"records": data, "total": len(data)}
        )
        
    except Exception as e:
        logger.error(f"获取催款记录失败: {e}", exc_info=True)
        return ApiResponse(
            success=False,
            error=str(e)
        )


# ==================== 应付账款管理 ====================

@router.get("/payables")
async def get_payables(
    filter_type: str = Query("all", description="筛选类型: all/unpaid/partial/paid/overdue"),
    search: Optional[str] = Query(None, description="搜索供应商名称"),
    supplier_id: Optional[int] = Query(None, description="供应商ID"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """获取应付账款列表"""
    try:
        from ..models.finance import AccountPayable
        from ..models import Supplier, InboundOrder
        from sqlalchemy.orm import joinedload
        from datetime import datetime, date
        
        query = db.query(AccountPayable).options(
            joinedload(AccountPayable.supplier),
            joinedload(AccountPayable.inbound_order)
        )
        
        # 筛选条件
        if filter_type == "unpaid":
            query = query.filter(AccountPayable.status == "unpaid")
        elif filter_type == "partial":
            query = query.filter(AccountPayable.status == "partial")
        elif filter_type == "paid":
            query = query.filter(AccountPayable.status == "paid")
        elif filter_type == "overdue":
            query = query.filter(AccountPayable.is_overdue == True)
        
        if supplier_id:
            query = query.filter(AccountPayable.supplier_id == supplier_id)
        
        if search:
            query = query.join(Supplier).filter(Supplier.name.contains(search))
        
        if start_date:
            query = query.filter(AccountPayable.credit_start_date >= datetime.strptime(start_date, "%Y-%m-%d").date())
        if end_date:
            query = query.filter(AccountPayable.credit_start_date <= datetime.strptime(end_date, "%Y-%m-%d").date())
        
        total = query.count()
        payables = query.order_by(AccountPayable.due_date.asc()).offset(skip).limit(limit).all()
        
        # 更新逾期状态
        today = date.today()
        for p in payables:
            if p.status != "paid" and p.due_date < today:
                p.is_overdue = True
                p.overdue_days = (today - p.due_date).days
        db.commit()
        
        data = []
        for p in payables:
            data.append({
                "id": p.id,
                "payable_no": p.payable_no,
                "supplier_id": p.supplier_id,
                "supplier_name": p.supplier.name if p.supplier else None,
                "inbound_order_id": p.inbound_order_id,
                "inbound_order_no": p.inbound_order.order_no if p.inbound_order else None,
                "total_amount": p.total_amount,
                "paid_amount": p.paid_amount,
                "unpaid_amount": p.unpaid_amount,
                "credit_days": p.credit_days,
                "credit_start_date": p.credit_start_date.isoformat() if p.credit_start_date else None,
                "due_date": p.due_date.isoformat() if p.due_date else None,
                "overdue_days": p.overdue_days,
                "status": p.status,
                "is_overdue": p.is_overdue,
                "remark": p.remark,
                "create_time": p.create_time.isoformat() if p.create_time else None
            })
        
        return {
            "success": True,
            "data": data,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"获取应付账款列表失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/payables/statistics")
async def get_payables_statistics(db: Session = Depends(get_db)):
    """获取应付账款统计"""
    try:
        from ..models.finance import AccountPayable
        from sqlalchemy import func
        from datetime import date
        
        # 总应付账款
        total_payable = db.query(func.sum(AccountPayable.unpaid_amount)).filter(
            AccountPayable.status.in_(["unpaid", "partial"])
        ).scalar() or 0
        
        # 本月应付
        today = date.today()
        month_start = today.replace(day=1)
        month_payable = db.query(func.sum(AccountPayable.unpaid_amount)).filter(
            AccountPayable.status.in_(["unpaid", "partial"]),
            AccountPayable.due_date >= month_start,
            AccountPayable.due_date <= today.replace(day=28)  # 简化处理
        ).scalar() or 0
        
        # 逾期金额
        overdue_amount = db.query(func.sum(AccountPayable.unpaid_amount)).filter(
            AccountPayable.status.in_(["unpaid", "partial"]),
            AccountPayable.is_overdue == True
        ).scalar() or 0
        
        # 逾期供应商数
        overdue_suppliers = db.query(func.count(func.distinct(AccountPayable.supplier_id))).filter(
            AccountPayable.status.in_(["unpaid", "partial"]),
            AccountPayable.is_overdue == True
        ).scalar() or 0
        
        return {
            "success": True,
            "data": {
                "total_payable": round(total_payable, 2),
                "month_payable": round(month_payable, 2),
                "overdue_amount": round(overdue_amount, 2),
                "overdue_suppliers": overdue_suppliers
            }
        }
        
    except Exception as e:
        logger.error(f"获取应付账款统计失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/payables/{payable_id}")
async def get_payable_detail(
    payable_id: int,
    db: Session = Depends(get_db)
):
    """获取应付账款详情"""
    try:
        from ..models.finance import AccountPayable, SupplierPayment
        from sqlalchemy.orm import joinedload
        
        payable = db.query(AccountPayable).options(
            joinedload(AccountPayable.supplier),
            joinedload(AccountPayable.inbound_order),
            joinedload(AccountPayable.supplier_payments)
        ).filter(AccountPayable.id == payable_id).first()
        
        if not payable:
            return {"success": False, "error": "应付账款不存在"}
        
        # 付款记录
        payments = []
        for p in payable.supplier_payments:
            payments.append({
                "id": p.id,
                "payment_no": p.payment_no,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "amount": p.amount,
                "payment_method": p.payment_method,
                "remark": p.remark,
                "created_by": p.created_by,
                "create_time": p.create_time.isoformat() if p.create_time else None
            })
        
        return {
            "success": True,
            "data": {
                "id": payable.id,
                "payable_no": payable.payable_no,
                "supplier_id": payable.supplier_id,
                "supplier_name": payable.supplier.name if payable.supplier else None,
                "inbound_order_id": payable.inbound_order_id,
                "inbound_order_no": payable.inbound_order.order_no if payable.inbound_order else None,
                "total_amount": payable.total_amount,
                "paid_amount": payable.paid_amount,
                "unpaid_amount": payable.unpaid_amount,
                "credit_days": payable.credit_days,
                "credit_start_date": payable.credit_start_date.isoformat() if payable.credit_start_date else None,
                "due_date": payable.due_date.isoformat() if payable.due_date else None,
                "overdue_days": payable.overdue_days,
                "status": payable.status,
                "is_overdue": payable.is_overdue,
                "remark": payable.remark,
                "create_time": payable.create_time.isoformat() if payable.create_time else None,
                "payments": payments
            }
        }
        
    except Exception as e:
        logger.error(f"获取应付账款详情失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/supplier-payment")
async def record_supplier_payment(
    supplier_id: int,
    amount: float,
    payment_method: str = Query("bank_transfer", description="付款方式: bank_transfer/cash/check/acceptance"),
    payment_date: Optional[str] = Query(None, description="付款日期 YYYY-MM-DD"),
    bank_account_id: Optional[int] = Query(None, description="付款账户ID"),
    remark: str = "",
    created_by: str = "财务专员",
    role: str = Depends(require_permission("can_record_supplier_payment")),
    db: Session = Depends(get_db)
):
    """
    登记供应商付款
    FIFO方式冲抵最早的应付账款
    """
    try:
        from ..models.finance import AccountPayable, SupplierPayment, BankAccount, CashFlow
        from ..models import Supplier
        from datetime import datetime, date, timedelta
        
        # 验证供应商
        supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
        if not supplier:
            return {"success": False, "error": "供应商不存在"}
        
        # 解析日期
        pay_date = datetime.strptime(payment_date, "%Y-%m-%d").date() if payment_date else date.today()
        
        # 生成付款单号
        now = datetime.now()
        payment_no = _generate_safe_no(db, SupplierPayment, SupplierPayment.payment_no, "FK")
        
        # 创建付款记录（财务端点直接确认）
        payment = SupplierPayment(
            payment_no=payment_no,
            supplier_id=supplier_id,
            payment_date=pay_date,
            amount=amount,
            payment_method=payment_method,
            bank_account_id=bank_account_id,
            remark=remark,
            created_by=created_by,
            status="confirmed",
            confirmed_by=created_by,
            confirmed_at=now
        )
        db.add(payment)
        db.flush()
        
        # 锁定并FIFO方式冲抵应付账款
        unpaid_payables = db.query(AccountPayable).filter(
            AccountPayable.supplier_id == supplier_id,
            AccountPayable.status.in_(["unpaid", "partial"]),
            AccountPayable.unpaid_amount > 0
        ).order_by(AccountPayable.due_date.asc()).with_for_update().all()
        
        remaining_amount = amount
        offset_details = []
        
        for payable in unpaid_payables:
            if remaining_amount <= 0:
                break
            
            offset_amount = min(remaining_amount, payable.unpaid_amount)
            payable.paid_amount += offset_amount
            payable.unpaid_amount = payable.total_amount - payable.paid_amount
            
            if payable.unpaid_amount <= 0:
                payable.status = "paid"
            else:
                payable.status = "partial"
            
            offset_details.append({
                "payable_no": payable.payable_no,
                "offset_amount": offset_amount
            })
            remaining_amount -= offset_amount
        
        # 锁定银行账户后记录资金流水
        if bank_account_id:
            account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).with_for_update().first()
            if account:
                balance_before = account.current_balance
                account.current_balance -= amount
                
                flow_no = _generate_safe_no(db, CashFlow, CashFlow.flow_no, "LS", seq_digits=4)
                
                cash_flow = CashFlow(
                    flow_no=flow_no,
                    account_id=bank_account_id,
                    flow_type="expense",
                    category="供应商付款",
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=account.current_balance,
                    related_type="supplier_payment",
                    related_id=payment.id,
                    flow_date=now,
                    counterparty=supplier.name,
                    remark=f"付款单：{payment_no}",
                    created_by=created_by
                )
                db.add(cash_flow)
        
        db.commit()
        
        logger.info(f"供应商付款: {payment_no}, 供应商={supplier.name}, 金额={amount}, 冲抵{len(offset_details)}笔应付账款")
        
        return {
            "success": True,
            "message": f"付款成功，冲抵{len(offset_details)}笔应付账款",
            "data": {
                "payment_no": payment_no,
                "amount": amount,
                "offset_details": offset_details,
                "remaining": remaining_amount if remaining_amount > 0 else 0
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"供应商付款失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/supplier-payments")
async def get_supplier_payments(
    supplier_id: Optional[int] = Query(None, description="供应商ID"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """获取供应商付款记录列表"""
    try:
        from ..models.finance import SupplierPayment
        from ..models import Supplier
        from sqlalchemy.orm import joinedload
        from datetime import datetime
        
        query = db.query(SupplierPayment).options(joinedload(SupplierPayment.supplier))
        
        if supplier_id:
            query = query.filter(SupplierPayment.supplier_id == supplier_id)
        if start_date:
            query = query.filter(SupplierPayment.payment_date >= datetime.strptime(start_date, "%Y-%m-%d").date())
        if end_date:
            query = query.filter(SupplierPayment.payment_date <= datetime.strptime(end_date, "%Y-%m-%d").date())
        
        total = query.count()
        payments = query.order_by(SupplierPayment.payment_date.desc()).offset(skip).limit(limit).all()
        
        data = []
        for p in payments:
            data.append({
                "id": p.id,
                "payment_no": p.payment_no,
                "supplier_id": p.supplier_id,
                "supplier_name": p.supplier.name if p.supplier else None,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "amount": p.amount,
                "payment_method": p.payment_method,
                "remark": p.remark,
                "created_by": p.created_by,
                "create_time": p.create_time.isoformat() if p.create_time else None
            })
        
        return {
            "success": True,
            "data": data,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"获取供应商付款记录失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


# ==================== 资金流水管理 ====================

@router.get("/accounts")
async def get_bank_accounts(
    account_type: Optional[str] = Query(None, description="账户类型: bank/cash/alipay/wechat"),
    status: str = Query("active", description="状态: active/inactive/all"),
    db: Session = Depends(get_db)
):
    """获取银行账户列表"""
    try:
        from ..models.finance import BankAccount
        
        query = db.query(BankAccount)
        
        if account_type:
            query = query.filter(BankAccount.account_type == account_type)
        if status != "all":
            query = query.filter(BankAccount.status == status)
        
        accounts = query.order_by(BankAccount.is_default.desc(), BankAccount.id.asc()).all()
        
        data = []
        for a in accounts:
            data.append({
                "id": a.id,
                "account_name": a.account_name,
                "account_no": a.account_no,
                "bank_name": a.bank_name,
                "account_type": a.account_type,
                "initial_balance": a.initial_balance,
                "current_balance": a.current_balance,
                "is_default": a.is_default,
                "status": a.status,
                "description": a.description,
                "create_time": a.create_time.isoformat() if a.create_time else None
            })
        
        return {"success": True, "data": data}
        
    except Exception as e:
        logger.error(f"获取银行账户列表失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/accounts")
async def create_bank_account(
    account_name: str,
    account_type: str = Query("bank", description="账户类型: bank/cash/alipay/wechat"),
    account_no: str = "",
    bank_name: str = "",
    initial_balance: float = 0.0,
    is_default: bool = False,
    description: str = "",
    created_by: str = "系统管理员",
    db: Session = Depends(get_db)
):
    """创建银行账户"""
    try:
        from ..models.finance import BankAccount
        
        # 如果设为默认，取消其他默认账户
        if is_default:
            db.query(BankAccount).filter(BankAccount.is_default == True).update({"is_default": False})
        
        account = BankAccount(
            account_name=account_name,
            account_no=account_no,
            bank_name=bank_name,
            account_type=account_type,
            initial_balance=initial_balance,
            current_balance=initial_balance,
            is_default=is_default,
            description=description,
            created_by=created_by
        )
        db.add(account)
        db.commit()
        db.refresh(account)
        
        logger.info(f"创建银行账户: {account_name}, 类型={account_type}, 期初余额={initial_balance}")
        
        return {
            "success": True,
            "message": "账户创建成功",
            "data": {
                "id": account.id,
                "account_name": account.account_name
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"创建银行账户失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.put("/accounts/{account_id}")
async def update_bank_account(
    account_id: int,
    account_name: Optional[str] = None,
    account_no: Optional[str] = None,
    bank_name: Optional[str] = None,
    is_default: Optional[bool] = None,
    status: Optional[str] = None,
    description: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """更新银行账户"""
    try:
        from ..models.finance import BankAccount
        
        account = db.query(BankAccount).filter(BankAccount.id == account_id).first()
        if not account:
            return {"success": False, "error": "账户不存在"}
        
        if account_name is not None:
            account.account_name = account_name
        if account_no is not None:
            account.account_no = account_no
        if bank_name is not None:
            account.bank_name = bank_name
        if description is not None:
            account.description = description
        if status is not None:
            account.status = status
        if is_default is not None:
            if is_default:
                db.query(BankAccount).filter(BankAccount.is_default == True).update({"is_default": False})
            account.is_default = is_default
        
        db.commit()
        
        return {"success": True, "message": "账户更新成功"}
        
    except Exception as e:
        db.rollback()
        logger.error(f"更新银行账户失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/cashflow")
async def get_cash_flows(
    account_id: Optional[int] = Query(None, description="账户ID"),
    flow_type: Optional[str] = Query(None, description="流水类型: income/expense"),
    category: Optional[str] = Query(None, description="分类"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """获取资金流水列表"""
    try:
        from ..models.finance import CashFlow, BankAccount
        from sqlalchemy.orm import joinedload
        from datetime import datetime
        
        query = db.query(CashFlow).options(joinedload(CashFlow.bank_account))
        
        if account_id:
            query = query.filter(CashFlow.account_id == account_id)
        if flow_type:
            query = query.filter(CashFlow.flow_type == flow_type)
        if category:
            query = query.filter(CashFlow.category == category)
        if start_date:
            query = query.filter(CashFlow.flow_date >= datetime.strptime(start_date, "%Y-%m-%d"))
        if end_date:
            query = query.filter(CashFlow.flow_date <= datetime.strptime(end_date + " 23:59:59", "%Y-%m-%d %H:%M:%S"))
        
        total = query.count()
        flows = query.order_by(CashFlow.flow_date.desc()).offset(skip).limit(limit).all()
        
        data = []
        for f in flows:
            data.append({
                "id": f.id,
                "flow_no": f.flow_no,
                "account_id": f.account_id,
                "account_name": f.bank_account.account_name if f.bank_account else None,
                "flow_type": f.flow_type,
                "category": f.category,
                "amount": f.amount,
                "balance_before": f.balance_before,
                "balance_after": f.balance_after,
                "related_type": f.related_type,
                "related_id": f.related_id,
                "flow_date": f.flow_date.isoformat() if f.flow_date else None,
                "counterparty": f.counterparty,
                "remark": f.remark,
                "created_by": f.created_by,
                "create_time": f.create_time.isoformat() if f.create_time else None
            })
        
        return {
            "success": True,
            "data": data,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"获取资金流水失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/cashflow/summary")
async def get_cashflow_summary(
    account_id: Optional[int] = Query(None, description="账户ID"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """获取资金流水汇总"""
    try:
        from ..models.finance import CashFlow, BankAccount
        from sqlalchemy import func
        from datetime import datetime
        
        # 收入汇总
        income_query = db.query(func.sum(CashFlow.amount)).filter(CashFlow.flow_type == "income")
        # 支出汇总
        expense_query = db.query(func.sum(CashFlow.amount)).filter(CashFlow.flow_type == "expense")
        
        if account_id:
            income_query = income_query.filter(CashFlow.account_id == account_id)
            expense_query = expense_query.filter(CashFlow.account_id == account_id)
        
        if start_date:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            income_query = income_query.filter(CashFlow.flow_date >= start)
            expense_query = expense_query.filter(CashFlow.flow_date >= start)
        if end_date:
            end = datetime.strptime(end_date + " 23:59:59", "%Y-%m-%d %H:%M:%S")
            income_query = income_query.filter(CashFlow.flow_date <= end)
            expense_query = expense_query.filter(CashFlow.flow_date <= end)
        
        total_income = income_query.scalar() or 0
        total_expense = expense_query.scalar() or 0
        
        # 账户余额
        if account_id:
            account = db.query(BankAccount).filter(BankAccount.id == account_id).first()
            current_balance = account.current_balance if account else 0
        else:
            current_balance = db.query(func.sum(BankAccount.current_balance)).filter(
                BankAccount.status == "active"
            ).scalar() or 0
        
        return {
            "success": True,
            "data": {
                "total_income": round(total_income, 2),
                "total_expense": round(total_expense, 2),
                "net_flow": round(total_income - total_expense, 2),
                "current_balance": round(current_balance, 2)
            }
        }
        
    except Exception as e:
        logger.error(f"获取资金流水汇总失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/cashflow/daily")
async def get_daily_cashflow(
    account_id: Optional[int] = Query(None, description="账户ID"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """获取日记账（按日汇总）"""
    try:
        from ..models.finance import CashFlow
        from sqlalchemy import func, cast, Date
        from datetime import datetime, date, timedelta
        
        # 默认最近30天
        if not end_date:
            end_date = date.today().isoformat()
        if not start_date:
            start_date = (date.today() - timedelta(days=30)).isoformat()
        
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date + " 23:59:59", "%Y-%m-%d %H:%M:%S")
        
        query = db.query(
            func.date(CashFlow.flow_date).label("date"),
            CashFlow.flow_type,
            func.sum(CashFlow.amount).label("total_amount"),
            func.count(CashFlow.id).label("count")
        ).filter(
            CashFlow.flow_date >= start,
            CashFlow.flow_date <= end
        )
        
        if account_id:
            query = query.filter(CashFlow.account_id == account_id)
        
        results = query.group_by(func.date(CashFlow.flow_date), CashFlow.flow_type).all()
        
        # 按日期整理数据
        daily_data = {}
        for r in results:
            date_str = r.date.isoformat() if hasattr(r.date, 'isoformat') else str(r.date)
            if date_str not in daily_data:
                daily_data[date_str] = {"date": date_str, "income": 0, "expense": 0, "income_count": 0, "expense_count": 0}
            
            if r.flow_type == "income":
                daily_data[date_str]["income"] = round(r.total_amount, 2)
                daily_data[date_str]["income_count"] = r.count
            else:
                daily_data[date_str]["expense"] = round(r.total_amount, 2)
                daily_data[date_str]["expense_count"] = r.count
        
        # 转为列表并排序
        data = sorted(daily_data.values(), key=lambda x: x["date"], reverse=True)
        
        return {
            "success": True,
            "data": data
        }
        
    except Exception as e:
        logger.error(f"获取日记账失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


# ==================== 费用管理 ====================

@router.get("/expense-categories")
async def get_expense_categories(
    include_inactive: bool = Query(False, description="是否包含停用的类别"),
    db: Session = Depends(get_db)
):
    """获取费用类别列表"""
    try:
        from ..models.finance import ExpenseCategory
        
        query = db.query(ExpenseCategory)
        if not include_inactive:
            query = query.filter(ExpenseCategory.is_active == True)
        
        categories = query.order_by(ExpenseCategory.sort_order.asc()).all()
        
        data = []
        for c in categories:
            data.append({
                "id": c.id,
                "code": c.code,
                "name": c.name,
                "parent_id": c.parent_id,
                "description": c.description,
                "sort_order": c.sort_order,
                "is_active": c.is_active
            })
        
        return {"success": True, "data": data}
        
    except Exception as e:
        logger.error(f"获取费用类别失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/expense-categories")
async def create_expense_category(
    code: str,
    name: str,
    parent_id: Optional[int] = None,
    description: str = "",
    sort_order: int = 0,
    db: Session = Depends(get_db)
):
    """创建费用类别"""
    try:
        from ..models.finance import ExpenseCategory
        
        # 检查编码是否重复
        existing = db.query(ExpenseCategory).filter(ExpenseCategory.code == code).first()
        if existing:
            return {"success": False, "error": "类别编码已存在"}
        
        category = ExpenseCategory(
            code=code,
            name=name,
            parent_id=parent_id,
            description=description,
            sort_order=sort_order
        )
        db.add(category)
        db.commit()
        db.refresh(category)
        
        return {
            "success": True,
            "message": "类别创建成功",
            "data": {"id": category.id, "code": category.code, "name": category.name}
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"创建费用类别失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/expense-categories/init")
async def init_expense_categories(db: Session = Depends(get_db)):
    """初始化默认费用类别"""
    try:
        from ..models.finance import ExpenseCategory, DEFAULT_EXPENSE_CATEGORIES
        
        created = 0
        for cat in DEFAULT_EXPENSE_CATEGORIES:
            existing = db.query(ExpenseCategory).filter(ExpenseCategory.code == cat["code"]).first()
            if not existing:
                category = ExpenseCategory(**cat)
                db.add(category)
                created += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"初始化完成，新增{created}个类别"
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"初始化费用类别失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/expenses")
async def get_expenses(
    category_id: Optional[int] = Query(None, description="费用类别ID"),
    status: Optional[str] = Query(None, description="状态: pending/approved/rejected"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """获取费用列表"""
    try:
        from ..models.finance import Expense, ExpenseCategory
        from sqlalchemy.orm import joinedload
        from datetime import datetime
        
        query = db.query(Expense).options(
            joinedload(Expense.category),
            joinedload(Expense.bank_account)
        )
        
        if category_id:
            query = query.filter(Expense.category_id == category_id)
        if status:
            query = query.filter(Expense.status == status)
        if start_date:
            query = query.filter(Expense.expense_date >= datetime.strptime(start_date, "%Y-%m-%d").date())
        if end_date:
            query = query.filter(Expense.expense_date <= datetime.strptime(end_date, "%Y-%m-%d").date())
        
        total = query.count()
        expenses = query.order_by(Expense.expense_date.desc()).offset(skip).limit(limit).all()
        
        data = []
        for e in expenses:
            data.append({
                "id": e.id,
                "expense_no": e.expense_no,
                "category_id": e.category_id,
                "category_name": e.category.name if e.category else None,
                "account_id": e.account_id,
                "account_name": e.bank_account.account_name if e.bank_account else None,
                "amount": e.amount,
                "expense_date": e.expense_date.isoformat() if e.expense_date else None,
                "payee": e.payee,
                "payment_method": e.payment_method,
                "status": e.status,
                "remark": e.remark,
                "created_by": e.created_by,
                "approved_by": e.approved_by,
                "approved_at": e.approved_at.isoformat() if e.approved_at else None,
                "create_time": e.create_time.isoformat() if e.create_time else None
            })
        
        return {
            "success": True,
            "data": data,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"获取费用列表失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/expenses")
async def create_expense(
    category_id: int,
    amount: float,
    expense_date: str = Query(..., description="费用日期 YYYY-MM-DD"),
    account_id: Optional[int] = Query(None, description="付款账户ID"),
    payee: str = "",
    payment_method: str = "bank_transfer",
    remark: str = "",
    created_by: str = "财务专员",
    auto_approve: bool = Query(False, description="是否自动审批"),
    db: Session = Depends(get_db)
):
    """录入费用"""
    try:
        from ..models.finance import Expense, ExpenseCategory, BankAccount, CashFlow
        from datetime import datetime
        
        # 验证类别
        category = db.query(ExpenseCategory).filter(ExpenseCategory.id == category_id).first()
        if not category:
            return {"success": False, "error": "费用类别不存在"}
        
        # 生成费用单号
        now = datetime.now()
        expense_no = _generate_safe_no(db, Expense, Expense.expense_no, "FY")
        
        # 解析日期
        exp_date = datetime.strptime(expense_date, "%Y-%m-%d").date()
        
        # 创建费用记录
        expense = Expense(
            expense_no=expense_no,
            category_id=category_id,
            account_id=account_id,
            amount=amount,
            expense_date=exp_date,
            payee=payee,
            payment_method=payment_method,
            remark=remark,
            created_by=created_by,
            status="approved" if auto_approve else "pending"
        )
        
        if auto_approve:
            expense.approved_by = created_by
            expense.approved_at = now
        
        db.add(expense)
        db.flush()
        
        # 如果自动审批且有账户，记录资金流水
        if auto_approve and account_id:
            account = db.query(BankAccount).filter(BankAccount.id == account_id).first()
            if account:
                balance_before = account.current_balance
                account.current_balance -= amount
                
                flow_no = _generate_safe_no(db, CashFlow, CashFlow.flow_no, "LS", seq_digits=4)
                
                cash_flow = CashFlow(
                    flow_no=flow_no,
                    account_id=account_id,
                    flow_type="expense",
                    category=f"费用支出-{category.name}",
                    amount=amount,
                    balance_before=balance_before,
                    balance_after=account.current_balance,
                    related_type="expense",
                    related_id=expense.id,
                    flow_date=now,
                    counterparty=payee,
                    remark=f"费用单：{expense_no}",
                    created_by=created_by
                )
                db.add(cash_flow)
        
        db.commit()
        
        logger.info(f"录入费用: {expense_no}, 类别={category.name}, 金额={amount}")
        
        return {
            "success": True,
            "message": "费用录入成功",
            "data": {
                "id": expense.id,
                "expense_no": expense_no,
                "status": expense.status
            }
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"录入费用失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.post("/expenses/{expense_id}/approve")
async def approve_expense(
    expense_id: int,
    action: str = Query(..., description="操作: approve/reject"),
    approved_by: str = "财务经理",
    reject_reason: str = "",
    db: Session = Depends(get_db)
):
    """审批费用"""
    try:
        from ..models.finance import Expense, BankAccount, CashFlow
        from datetime import datetime
        
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        if not expense:
            return {"success": False, "error": "费用记录不存在"}
        
        if expense.status != "pending":
            return {"success": False, "error": "该费用已审批"}
        
        now = datetime.now()
        
        if action == "approve":
            expense.status = "approved"
            expense.approved_by = approved_by
            expense.approved_at = now
            
            # 记录资金流水
            if expense.account_id:
                account = db.query(BankAccount).filter(BankAccount.id == expense.account_id).first()
                if account:
                    balance_before = account.current_balance
                    account.current_balance -= expense.amount
                    
                    flow_no = _generate_safe_no(db, CashFlow, CashFlow.flow_no, "LS", seq_digits=4)
                    
                    cash_flow = CashFlow(
                        flow_no=flow_no,
                        account_id=expense.account_id,
                        flow_type="expense",
                        category=f"费用支出-{expense.category.name if expense.category else '其他'}",
                        amount=expense.amount,
                        balance_before=balance_before,
                        balance_after=account.current_balance,
                        related_type="expense",
                        related_id=expense.id,
                        flow_date=now,
                        counterparty=expense.payee,
                        remark=f"费用单：{expense.expense_no}",
                        created_by=approved_by
                    )
                    db.add(cash_flow)
            
            message = "费用已审批通过"
        else:
            expense.status = "rejected"
            expense.reject_reason = reject_reason
            message = "费用已驳回"
        
        db.commit()
        
        logger.info(f"费用审批: {expense.expense_no}, 操作={action}, 审批人={approved_by}")
        
        return {"success": True, "message": message}
        
    except Exception as e:
        db.rollback()
        logger.error(f"费用审批失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@router.get("/expenses/summary")
async def get_expenses_summary(
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    group_by: str = Query("category", description="分组方式: category/month"),
    db: Session = Depends(get_db)
):
    """获取费用汇总"""
    try:
        from ..models.finance import Expense, ExpenseCategory
        from sqlalchemy import func
        from datetime import datetime, date
        
        # 默认本月
        if not start_date:
            start_date = date.today().replace(day=1).isoformat()
        if not end_date:
            end_date = date.today().isoformat()
        
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        if group_by == "category":
            results = db.query(
                ExpenseCategory.name,
                func.sum(Expense.amount).label("total_amount"),
                func.count(Expense.id).label("count")
            ).join(ExpenseCategory).filter(
                Expense.status == "approved",
                Expense.expense_date >= start,
                Expense.expense_date <= end
            ).group_by(ExpenseCategory.name).all()
            
            data = [{"category": r.name, "amount": round(r.total_amount, 2), "count": r.count} for r in results]
        else:
            results = db.query(
                func.strftime("%Y-%m", Expense.expense_date).label("month"),
                func.sum(Expense.amount).label("total_amount"),
                func.count(Expense.id).label("count")
            ).filter(
                Expense.status == "approved",
                Expense.expense_date >= start,
                Expense.expense_date <= end
            ).group_by(func.strftime("%Y-%m", Expense.expense_date)).all()
            
            data = [{"month": r.month, "amount": round(r.total_amount, 2), "count": r.count} for r in results]
        
        # 总计
        total = sum(d["amount"] for d in data)
        
        return {
            "success": True,
            "data": {
                "details": data,
                "total": round(total, 2),
                "period": {"start": start_date, "end": end_date}
            }
        }
        
    except Exception as e:
        logger.error(f"获取费用汇总失败: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


# ==================== 客户收款登记（结算角色） ====================

@router.post("/customer-payment-registration")
async def create_customer_payment_registration(
    customer_id: int,
    amount: float,
    payment_method: str = Query("bank_transfer", description="收款方式: cash/bank_transfer/wechat/alipay/card/check/other"),
    remark: str = "",
    created_by: str = "结算",
    db: Session = Depends(get_db)
):
    """创建客户收款登记（pending状态，不立即抵扣应收账款）"""
    from ..models.finance import AccountReceivable, PaymentRecord
    from ..models import Customer
    from sqlalchemy import func as sa_func
    from datetime import date
    from ..timezone_utils import china_now
    
    try:
        # 验证客户
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 验证金额
        if amount is None or amount <= 0:
            raise HTTPException(status_code=400, detail="收款金额必须大于0")
        
        # 查询该客户未收应收账款总额
        total_unpaid = db.query(sa_func.sum(AccountReceivable.unpaid_amount)).filter(
            AccountReceivable.customer_id == customer_id,
            AccountReceivable.status.in_(['unpaid', 'overdue'])
        ).scalar() or 0
        
        # 生成收款单号
        now = china_now()
        payment_no = f"SK{now.strftime('%Y%m%d%H%M%S')}"
        
        # 创建收款记录（pending状态）
        payment = PaymentRecord(
            payment_no=payment_no,
            account_receivable_id=None,
            customer_id=customer_id,
            payment_date=now.date(),
            amount=amount,
            payment_method=payment_method,
            remark=remark,
            operator=created_by,
            status="pending"
        )
        db.add(payment)
        db.commit()
        db.refresh(payment)
        
        logger.info(f"创建客户收款登记(待确认): {payment_no}, 客户: {customer.name}, 金额: ¥{amount:.2f}")
        
        return {
            "success": True,
            "data": {
                "id": payment.id,
                "payment_no": payment.payment_no,
                "customer_id": customer_id,
                "customer_name": customer.name,
                "amount": amount,
                "payment_method": payment_method,
                "status": "pending",
                "created_by": created_by,
                "create_time": payment.create_time.isoformat() if payment.create_time else None,
                "total_unpaid": round(total_unpaid, 2)
            },
            "message": f"收款登记已创建（待确认），单号：{payment_no}"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建客户收款登记失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


@router.get("/customer-payment-registrations")
async def get_customer_payment_registrations(
    customer_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """查询客户收款登记列表"""
    from ..models.finance import PaymentRecord
    from ..models import Customer
    
    try:
        query = db.query(PaymentRecord)
        if customer_id:
            query = query.filter(PaymentRecord.customer_id == customer_id)
        if status:
            query = query.filter(PaymentRecord.status == status)
        
        total = query.count()
        payments = query.order_by(PaymentRecord.create_time.desc()).offset(skip).limit(limit).all()
        
        result = []
        for p in payments:
            customer = db.query(Customer).filter(Customer.id == p.customer_id).first()
            result.append({
                "id": p.id,
                "payment_no": p.payment_no,
                "customer_id": p.customer_id,
                "customer_name": customer.name if customer else "未知",
                "amount": p.amount,
                "payment_method": p.payment_method,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "status": p.status or "confirmed",
                "remark": p.remark or "",
                "operator": p.operator or "",
                "confirmed_by": p.confirmed_by or "",
                "confirmed_at": p.confirmed_at.isoformat() if p.confirmed_at else None,
                "create_time": p.create_time.isoformat() if p.create_time else None
            })
        
        return {"success": True, "data": {"payments": result, "total": total}}
    except Exception as e:
        logger.error(f"查询客户收款列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/customer-payment-registrations/{payment_id}/confirm")
async def confirm_customer_payment_registration(
    payment_id: int,
    confirmed_by: str = "结算",
    db: Session = Depends(get_db)
):
    """确认客户收款（FIFO抵扣应收账款）"""
    from ..models.finance import PaymentRecord, AccountReceivable
    from ..models import Customer, CustomerTransaction, OrderStatusLog
    from ..timezone_utils import china_now
    
    try:
        # 加行锁防止并发确认
        payment = db.query(PaymentRecord).filter(PaymentRecord.id == payment_id).with_for_update().first()
        if not payment:
            raise HTTPException(status_code=404, detail="收款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=409, detail=f"收款状态为 {payment.status}，无法确认（可能已被其他人确认）")
        
        customer = db.query(Customer).filter(Customer.id == payment.customer_id).first()
        
        # 锁定并FIFO抵扣应收账款
        unpaid_receivables = db.query(AccountReceivable).filter(
            AccountReceivable.customer_id == payment.customer_id,
            AccountReceivable.status.in_(['unpaid', 'overdue'])
        ).order_by(AccountReceivable.due_date.asc()).with_for_update().all()
        
        remaining = payment.amount
        offset_details = []
        
        for receivable in unpaid_receivables:
            if remaining <= 0:
                break
            offset_amount = min(remaining, receivable.unpaid_amount)
            receivable.received_amount = (receivable.received_amount or 0) + offset_amount
            receivable.unpaid_amount = (receivable.unpaid_amount or 0) - offset_amount
            if receivable.unpaid_amount <= 0.01:
                receivable.unpaid_amount = 0
                receivable.status = 'paid'
            remaining -= offset_amount
            offset_details.append({
                "receivable_id": receivable.id,
                "offset_amount": round(offset_amount, 2)
            })
        
        # 更新收款状态
        now = china_now()
        payment.status = "confirmed"
        payment.confirmed_by = confirmed_by
        payment.confirmed_at = now
        
        # 创建客户往来账记录
        if customer:
            tx = CustomerTransaction(
                customer_id=customer.id,
                customer_name=customer.name,
                transaction_type='payment',
                amount=-(payment.amount),  # 负数=客户付给我们
                remark=f"收款登记确认：{payment.payment_no}"
            )
            db.add(tx)
        
        # 审计日志
        log = OrderStatusLog(
            order_type="customer_payment",
            order_id=payment.id,
            action="confirm",
            old_status="pending",
            new_status="confirmed",
            operated_by=confirmed_by,
            remark=f"确认客户收款 ¥{payment.amount:.2f}，抵扣{len(offset_details)}笔应收账款"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"确认客户收款: {payment.payment_no}, 客户: {customer.name if customer else ''}, 金额: ¥{payment.amount:.2f}, 抵扣{len(offset_details)}笔")
        
        return {
            "success": True,
            "data": {
                "id": payment.id,
                "payment_no": payment.payment_no,
                "status": "confirmed",
                "offset_details": offset_details,
                "unallocated": round(remaining, 2)
            },
            "message": f"收款已确认，抵扣{len(offset_details)}笔应收账款"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"确认客户收款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"确认失败: {str(e)}")


@router.put("/customer-payment-registrations/{payment_id}")
async def update_customer_payment_registration(
    payment_id: int,
    customer_id: Optional[int] = None,
    amount: Optional[float] = None,
    payment_method: Optional[str] = None,
    remark: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """编辑客户收款登记（仅pending状态可编辑）"""
    from ..models.finance import PaymentRecord
    from ..models import Customer
    
    try:
        payment = db.query(PaymentRecord).filter(PaymentRecord.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="收款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=400, detail="只能编辑待确认的收款记录")
        
        if customer_id is not None:
            customer = db.query(Customer).filter(Customer.id == customer_id).first()
            if not customer:
                raise HTTPException(status_code=404, detail="客户不存在")
            payment.customer_id = customer_id
        
        if amount is not None:
            if amount <= 0:
                raise HTTPException(status_code=400, detail="收款金额必须大于0")
            payment.amount = amount
        
        if payment_method is not None:
            payment.payment_method = payment_method
        
        if remark is not None:
            payment.remark = remark
        
        db.commit()
        db.refresh(payment)
        
        customer = db.query(Customer).filter(Customer.id == payment.customer_id).first()
        
        logger.info(f"编辑客户收款: {payment.payment_no}")
        
        return {
            "success": True,
            "data": {
                "id": payment.id,
                "payment_no": payment.payment_no,
                "customer_id": payment.customer_id,
                "customer_name": customer.name if customer else "未知",
                "amount": payment.amount,
                "payment_method": payment.payment_method,
                "remark": payment.remark or "",
                "status": payment.status
            },
            "message": "收款记录已更新"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"编辑客户收款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"编辑失败: {str(e)}")


@router.delete("/customer-payment-registrations/{payment_id}")
async def delete_customer_payment_registration(
    payment_id: int,
    deleted_by: str = "结算",
    db: Session = Depends(get_db)
):
    """删除客户收款登记（仅pending状态可删除）"""
    from ..models.finance import PaymentRecord
    
    try:
        payment = db.query(PaymentRecord).filter(PaymentRecord.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="收款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=400, detail="只能删除待确认的收款记录")
        
        payment_no = payment.payment_no
        db.delete(payment)
        db.commit()
        
        logger.info(f"删除客户收款: {payment_no}, 删除人: {deleted_by}")
        
        return {"success": True, "message": f"收款记录 {payment_no} 已删除"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除客户收款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.post("/customer-payment-registrations/{payment_id}/unconfirm")
async def unconfirm_customer_payment_registration(
    payment_id: int,
    reason: str = Query(..., description="反确认原因"),
    operated_by: str = "结算",
    db: Session = Depends(get_db)
):
    """反确认客户收款（回滚已抵扣的应收账款）"""
    from ..models.finance import PaymentRecord, AccountReceivable
    from ..models import Customer, OrderStatusLog
    from ..timezone_utils import china_now
    
    try:
        payment = db.query(PaymentRecord).filter(PaymentRecord.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="收款记录不存在")
        
        if (payment.status or "confirmed") != "confirmed":
            raise HTTPException(status_code=400, detail=f"收款状态为 {payment.status}，只能反确认已确认的记录")
        
        customer = db.query(Customer).filter(Customer.id == payment.customer_id).first()
        
        # 回滚：恢复被抵扣的应收账款
        receivables = db.query(AccountReceivable).filter(
            AccountReceivable.customer_id == payment.customer_id,
            AccountReceivable.status.in_(['paid', 'unpaid', 'overdue'])
        ).order_by(AccountReceivable.due_date.desc()).all()
        
        remaining = payment.amount
        restored_count = 0
        
        for receivable in receivables:
            if remaining <= 0:
                break
            restore_amount = min(remaining, receivable.received_amount or 0)
            if restore_amount <= 0:
                continue
            receivable.received_amount = (receivable.received_amount or 0) - restore_amount
            receivable.unpaid_amount = (receivable.unpaid_amount or 0) + restore_amount
            if receivable.unpaid_amount > 0 and receivable.status == 'paid':
                receivable.status = 'unpaid'
            remaining -= restore_amount
            restored_count += 1
        
        # 更新收款状态
        payment.status = "pending"
        payment.confirmed_by = None
        payment.confirmed_at = None
        
        # 审计日志
        log = OrderStatusLog(
            order_type="customer_payment",
            order_id=payment.id,
            action="unconfirm",
            old_status="confirmed",
            new_status="pending",
            operated_by=operated_by,
            remark=f"反确认客户收款 ¥{payment.amount:.2f}，恢复{restored_count}笔应收账款。原因：{reason}"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"反确认客户收款: {payment.payment_no}, 客户: {customer.name if customer else ''}, 金额: ¥{payment.amount:.2f}, 恢复{restored_count}笔, 原因: {reason}")
        
        return {
            "success": True,
            "message": f"收款已反确认，恢复{restored_count}笔应收账款"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"反确认客户收款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"反确认失败: {str(e)}")


"""
金料管理路由
- 收料单管理（从客户收取金料）
- 付料单管理（支付给供应商）
- 金料库存查询
- 客户存料管理
- 客户往来账查询
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, func, extract, and_
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict
from collections import defaultdict
import logging
from decimal import Decimal

from ..database import get_db
from ..timezone_utils import china_now
from ..models import (
    GoldMaterialTransaction,
    CustomerGoldDeposit,
    CustomerGoldDepositTransaction,
    CustomerTransaction,
    CustomerWithdrawal,
    CustomerTransfer,
    SettlementOrder,
    SalesOrder,
    InboundOrder,
    InboundDetail,
    Customer,
    Supplier,
    OrderStatusLog,
    SupplierGoldAccount,
    SupplierGoldTransaction,
)
from ..models.finance import GoldReceipt
from ..schemas import (
    GoldReceiptCreate,
    GoldReceiptUpdate,
    GoldPaymentCreate,
    GoldMaterialTransactionConfirm,
    GoldMaterialTransactionResponse,
    GoldMaterialBalanceResponse,
    CustomerGoldDepositResponse,
    CustomerGoldDepositTransactionResponse,
    CustomerTransactionResponse,
    CustomerAccountSummary,
    CustomerWithdrawalCreate,
    CustomerWithdrawalUpdate,
    CustomerWithdrawalComplete,
    CustomerWithdrawalResponse,
    CustomerTransferCreate,
    CustomerTransferConfirm,
    CustomerTransferResponse,
)
from ..middleware.permissions import has_permission
from ..utils.document_generator import (
    PDFGenerator,
    HTMLGenerator,
    build_gold_transaction_fields,
    format_datetime,
    get_current_time_str,
    get_status_label,
)
from ..dependencies.auth import get_current_role, require_permission
from ..utils.response import (
    success_response, error_response, paginated_response,
    not_found_response, conflict_response, server_error_response, ErrorCode
)
from ..utils.decimal_utils import to_decimal, round_weight, round_money, safe_float_for_json

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/gold-material", tags=["金料管理"])


# ==================== 辅助函数 ====================

def build_transaction_response(transaction: GoldMaterialTransaction, db: Session) -> dict:
    """
    构建金料流转记录响应
    
    优化说明：
    - 如果已经通过 joinedload 预加载了关联数据，直接使用，避免额外查询
    - 如果没有预加载，回退到手动查询（兼容单个记录查询场景）
    """
    settlement_no = None
    inbound_order_no = None
    
    # 优先使用预加载的数据，否则手动查询
    if transaction.settlement_order_id:
        # 检查是否已预加载
        if hasattr(transaction, 'settlement_order') and transaction.settlement_order:
            settlement_no = transaction.settlement_order.settlement_no
        else:
            settlement = db.query(SettlementOrder).filter(
                SettlementOrder.id == transaction.settlement_order_id
            ).first()
            if settlement:
                settlement_no = settlement.settlement_no
    
    if transaction.inbound_order_id:
        # 检查是否已预加载
        if hasattr(transaction, 'inbound_order') and transaction.inbound_order:
            inbound_order_no = transaction.inbound_order.order_no
        else:
            inbound = db.query(InboundOrder).filter(
                InboundOrder.id == transaction.inbound_order_id
            ).first()
            if inbound:
                inbound_order_no = inbound.order_no
    
    return {
        "id": transaction.id,
        "transaction_no": transaction.transaction_no,
        "transaction_type": transaction.transaction_type,
        "settlement_order_id": transaction.settlement_order_id,
        "settlement_no": settlement_no,
        "customer_id": transaction.customer_id,
        "customer_name": transaction.customer_name,
        "inbound_order_id": transaction.inbound_order_id,
        "inbound_order_no": inbound_order_no,
        "supplier_id": transaction.supplier_id,
        "supplier_name": transaction.supplier_name,
        "gold_weight": transaction.gold_weight,
        "status": transaction.status,
        "created_by": transaction.created_by,
        "confirmed_by": transaction.confirmed_by,
        "confirmed_at": transaction.confirmed_at,
        "created_at": transaction.created_at,
        "receipt_printed_at": transaction.receipt_printed_at,
        "payment_printed_at": transaction.payment_printed_at,
        "remark": transaction.remark,
    }


def get_gold_balance_internal(db: Session) -> dict:
    """内部函数：获取金料库存余额（统一使用新系统 GoldReceipt）"""
    from ..models.finance import GoldReceipt
    
    # 计算总收入 - 从新系统 GoldReceipt 中查询（包括期初金料和客户收料）
    total_income = db.query(func.sum(GoldReceipt.gold_weight)).filter(
        GoldReceipt.status == 'received'
    ).scalar() or 0.0
    
    # 计算总支出（已确认的付料单，仍使用 GoldMaterialTransaction，因为付料单保留在旧模型）
    total_expense = db.query(func.sum(GoldMaterialTransaction.gold_weight)).filter(
        GoldMaterialTransaction.transaction_type == 'expense',
        GoldMaterialTransaction.status == 'confirmed'
    ).scalar() or 0.0
    
    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "current_balance": total_income - total_expense
    }


def get_or_create_customer_deposit(customer_id: int, customer_name: str, db: Session) -> CustomerGoldDeposit:
    """获取或创建客户存料记录"""
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == customer_id
    ).first()
    
    if not deposit:
        deposit = CustomerGoldDeposit(
            customer_id=customer_id,
            customer_name=customer_name,
            current_balance=0.0,
            total_deposited=0.0,
            total_used=0.0
        )
        db.add(deposit)
        db.flush()
    
    return deposit


def calculate_settlement_gold_received(settlement_id: int, db: Session) -> float:
    """计算某结算单已收金料总重量（使用新系统 GoldReceipt）"""
    from ..models.finance import GoldReceipt
    
    receipts = db.query(GoldReceipt).filter(
        GoldReceipt.settlement_id == settlement_id,
        GoldReceipt.status != 'cancelled'
    ).all()
    
    return sum(r.gold_weight for r in receipts)


# ==================== 旧收料单API（已废弃，请使用 /gold-receipts 接口）====================
# 以下API已废弃，保留仅为向后兼容，请使用新的 /gold-receipts 系列API
# POST /gold-receipts - 创建收料单
# GET /gold-receipts - 获取收料单列表
# POST /gold-receipts/{receipt_id}/receive - 料部确认接收


# ==================== 付料单管理 ====================

@router.post("/payments", response_model=GoldMaterialTransactionResponse)
async def create_gold_payment(
    data: GoldPaymentCreate,
    user_role: str = Query(default="sales", description="用户角色"),
    created_by: str = Query(default="料部", description="创建人"),
    role: str = Depends(require_permission("can_create_gold_payment")),
    db: Session = Depends(get_db)
):
    """
    创建付料单（料部支付供应商）
    
    - 验证供应商是否存在
    - 校验金料库存是否足够
    - 创建金料支出记录（待确认状态，不扣供应商金料账户）
    - 确认时才更新供应商金料账户
    """
    # 权限检查
    if not has_permission(user_role, 'can_create_gold_payment'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【创建付料单】的权限")
    
    # 验证供应商
    supplier = db.query(Supplier).filter(Supplier.id == data.supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")
    
    # 验证入库单（如果提供）
    inbound_order = None
    if data.inbound_order_id:
        inbound_order = db.query(InboundOrder).filter(
            InboundOrder.id == data.inbound_order_id
        ).first()
        if not inbound_order:
            raise HTTPException(status_code=404, detail="入库单不存在")
        
        # 检查该入库单是否已创建付料单
        existing = db.query(GoldMaterialTransaction).filter(
            GoldMaterialTransaction.inbound_order_id == data.inbound_order_id,
            GoldMaterialTransaction.status != 'cancelled'
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"该入库单已创建付料单：{existing.transaction_no}")
    
    # 校验金料库存是否足够
    balance = get_gold_balance_internal(db)
    if data.gold_weight > balance["current_balance"]:
        raise HTTPException(
            status_code=400,
            detail=f"金料库存不足。当前余额：{balance['current_balance']:.2f}克，需要支付：{data.gold_weight:.2f}克"
        )
    
    # 生成付料单号（FL开头）
    now = china_now()
    count = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.transaction_no.like(f"FL{now.strftime('%Y%m%d')}%")
    ).count()
    payment_no = f"FL{now.strftime('%Y%m%d')}{count + 1:03d}"
    
    # 创建金料支出记录（付料单 - 待确认状态，不扣余额）
    transaction = GoldMaterialTransaction(
        transaction_no=payment_no,
        transaction_type='expense',
        inbound_order_id=data.inbound_order_id,
        supplier_id=supplier.id,
        supplier_name=supplier.name,
        gold_weight=data.gold_weight,
        status="pending",  # 待确认（确认后才扣供应商金料账户）
        created_by=created_by,
        remark=data.remark
    )
    db.add(transaction)
    
    db.commit()
    db.refresh(transaction)
    
    logger.info(f"创建付料单(待确认): {payment_no}, 供应商: {supplier.name}, 金料重量: {data.gold_weight}克")
    
    return build_transaction_response(transaction, db)


@router.get("/payments")
async def get_gold_payments(
    status: Optional[str] = None,
    supplier_id: Optional[int] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取付料单列表"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看金料记录】的权限")
    
    # 使用 joinedload 预加载关联数据，避免 N+1 查询问题
    query = db.query(GoldMaterialTransaction).options(
        joinedload(GoldMaterialTransaction.inbound_order),
        joinedload(GoldMaterialTransaction.supplier)
    ).filter(
        GoldMaterialTransaction.transaction_type == 'expense'
    )
    
    if status:
        query = query.filter(GoldMaterialTransaction.status == status)
    if supplier_id:
        query = query.filter(GoldMaterialTransaction.supplier_id == supplier_id)
    
    payments = query.order_by(desc(GoldMaterialTransaction.created_at)).all()
    
    return success_response(
        data={
            "payments": [build_transaction_response(p, db) for p in payments],
            "total": len(payments)
        }
    )


@router.post("/payments/{payment_id}/confirm")
async def confirm_gold_payment(
    payment_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    confirmed_by: str = Query(default="料部", description="确认人"),
    db: Session = Depends(get_db)
):
    """
    确认付料单（扣减供应商金料账户余额）
    
    - 验证状态为pending
    - 重新校验金料库存是否足够
    - 更新供应商金料账户
    - 创建供应商金料交易记录
    - 创建操作日志
    """
    if not has_permission(user_role, 'can_create_gold_payment'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【确认付料单】的权限")
    
    # 加行锁防止并发确认
    transaction = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.id == payment_id,
        GoldMaterialTransaction.transaction_type == 'expense'
    ).with_for_update().first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="付料单不存在")
    
    if transaction.status != "pending":
        raise HTTPException(status_code=409, detail=f"付料单状态为 {transaction.status}，无法确认（可能已被其他人确认）")
    
    # 重新校验金料库存
    balance = get_gold_balance_internal(db)
    if transaction.gold_weight > balance["current_balance"]:
        raise HTTPException(
            status_code=400,
            detail=f"金料库存不足。当前余额：{balance['current_balance']:.2f}克，需要支付：{transaction.gold_weight:.2f}克"
        )
    
    now = china_now()
    
    # 更新付料单状态
    transaction.status = "confirmed"
    transaction.confirmed_by = confirmed_by
    transaction.confirmed_at = now
    
    # 锁定供应商金料账户防止并发余额修改
    supplier_gold_account = db.query(SupplierGoldAccount).filter(
        SupplierGoldAccount.supplier_id == transaction.supplier_id
    ).with_for_update().first()
    
    if not supplier_gold_account:
        supplier_gold_account = SupplierGoldAccount(
            supplier_id=transaction.supplier_id,
            supplier_name=transaction.supplier_name,
            current_balance=0.0,
            total_received=0.0,
            total_paid=0.0
        )
        db.add(supplier_gold_account)
        db.flush()
    
    tx_gold_weight = to_decimal(transaction.gold_weight)
    balance_before = supplier_gold_account.current_balance
    supplier_gold_account.current_balance = round_weight(to_decimal(supplier_gold_account.current_balance) - tx_gold_weight)
    supplier_gold_account.total_paid = round_weight(to_decimal(supplier_gold_account.total_paid) + tx_gold_weight)
    supplier_gold_account.last_transaction_at = now
    
    # 创建供应商金料交易记录
    supplier_gold_tx = SupplierGoldTransaction(
        supplier_id=transaction.supplier_id,
        supplier_name=transaction.supplier_name,
        transaction_type='pay',
        payment_transaction_id=transaction.id,
        gold_weight=tx_gold_weight,
        balance_before=balance_before,
        balance_after=supplier_gold_account.current_balance,
        created_by=confirmed_by,
        remark=f"付料单确认：{transaction.transaction_no}"
    )
    db.add(supplier_gold_tx)
    
    # 创建操作日志
    log = OrderStatusLog(
        order_type="payment",
        order_id=transaction.id,
        action="confirm",
        old_status="pending",
        new_status="confirmed",
        operated_by=confirmed_by,
        operated_at=now,
        remark=f"确认付料单，扣减供应商账户 {transaction.gold_weight:.2f}克（{balance_before:.2f} → {supplier_gold_account.current_balance:.2f}）"
    )
    db.add(log)
    
    db.commit()
    db.refresh(transaction)
    
    logger.info(f"确认付料单: {transaction.transaction_no}, 供应商: {transaction.supplier_name}, 克重: {transaction.gold_weight}克, 账户: {balance_before:.2f} -> {supplier_gold_account.current_balance:.2f}")
    
    return success_response(
        message=f"付料单 {transaction.transaction_no} 已确认",
        data={"payment": build_transaction_response(transaction, db)}
    )


@router.post("/payments/{payment_id}/cancel")
async def cancel_gold_payment(
    payment_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    cancelled_by: str = Query(default="料部", description="取消人"),
    db: Session = Depends(get_db)
):
    """取消付料单（仅待确认状态可取消，pending状态不涉及账户变动）"""
    if not has_permission(user_role, 'can_create_gold_payment'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【取消付料单】的权限")
    
    transaction = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.id == payment_id,
        GoldMaterialTransaction.transaction_type == 'expense'
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="付料单不存在")
    
    if transaction.status != "pending":
        raise HTTPException(status_code=400, detail="只能取消待确认的付料单")
    
    now = china_now()
    transaction.status = "cancelled"
    
    # 创建操作日志
    log = OrderStatusLog(
        order_type="payment",
        order_id=transaction.id,
        action="cancel",
        old_status="pending",
        new_status="cancelled",
        operated_by=cancelled_by,
        operated_at=now,
        remark=f"取消付料单：{transaction.transaction_no}"
    )
    db.add(log)
    
    db.commit()
    db.refresh(transaction)
    
    logger.info(f"取消付料单: {transaction.transaction_no}, 供应商: {transaction.supplier_name}, 取消人: {cancelled_by}")
    
    return success_response(
        message="付料单已取消",
        data={"payment": build_transaction_response(transaction, db)}
    )


@router.post("/payments/{payment_id}/unconfirm")
async def unconfirm_gold_payment(
    payment_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    operated_by: str = Query(default="料部", description="操作人"),
    reason: str = Query(default="", description="反确认原因"),
    db: Session = Depends(get_db)
):
    """
    反确认付料单（回滚已确认的付料单）
    
    - 验证状态为confirmed
    - 恢复供应商金料账户余额
    - 创建反向供应商金料交易记录（留痕）
    - 创建操作日志（留痕）
    - 状态 confirmed -> pending
    """
    if not has_permission(user_role, 'can_manage_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【反确认付料单】的权限")
    
    transaction = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.id == payment_id,
        GoldMaterialTransaction.transaction_type == 'expense'
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="付料单不存在")
    
    if transaction.status != "confirmed":
        raise HTTPException(status_code=400, detail=f"付料单状态为 {transaction.status}，只能反确认已确认的付料单")
    
    # 恢复供应商金料账户
    supplier_gold_account = db.query(SupplierGoldAccount).filter(
        SupplierGoldAccount.supplier_id == transaction.supplier_id
    ).first()
    
    if not supplier_gold_account:
        raise HTTPException(status_code=500, detail="供应商金料账户不存在，无法恢复余额")
    
    now = china_now()
    tx_gold_weight = to_decimal(transaction.gold_weight)
    balance_before = supplier_gold_account.current_balance
    supplier_gold_account.current_balance = round_weight(to_decimal(supplier_gold_account.current_balance) + tx_gold_weight)
    supplier_gold_account.total_paid = round_weight(max(Decimal("0"), to_decimal(supplier_gold_account.total_paid) - tx_gold_weight))
    supplier_gold_account.last_transaction_at = now
    
    # 创建反向供应商金料交易记录（留痕）
    refund_reason = f"反确认付料单：{transaction.transaction_no}"
    if reason:
        refund_reason += f"，原因：{reason}"
    
    supplier_gold_tx = SupplierGoldTransaction(
        supplier_id=transaction.supplier_id,
        supplier_name=transaction.supplier_name,
        transaction_type='refund',
        payment_transaction_id=transaction.id,
        gold_weight=tx_gold_weight,
        balance_before=balance_before,
        balance_after=supplier_gold_account.current_balance,
        created_by=operated_by,
        remark=refund_reason
    )
    db.add(supplier_gold_tx)
    
    # 更新付料单状态
    transaction.status = "pending"
    transaction.confirmed_by = None
    transaction.confirmed_at = None
    
    # 创建操作日志（留痕）
    log = OrderStatusLog(
        order_type="payment",
        order_id=transaction.id,
        action="unconfirm",
        old_status="confirmed",
        new_status="pending",
        operated_by=operated_by,
        operated_at=now,
        remark=f"反确认付料单，恢复供应商账户 {transaction.gold_weight:.2f}克（{balance_before:.2f} → {supplier_gold_account.current_balance:.2f}）" +
               (f"，原因：{reason}" if reason else "")
    )
    db.add(log)
    
    db.commit()
    db.refresh(transaction)
    
    logger.info(f"反确认付料单: {transaction.transaction_no}, 供应商: {transaction.supplier_name}, 克重: {transaction.gold_weight}克, 账户: {balance_before:.2f} -> {supplier_gold_account.current_balance:.2f}, 操作人: {operated_by}")
    
    return success_response(
        message=f"付料单 {transaction.transaction_no} 已反确认，已恢复供应商账户 {transaction.gold_weight:.2f}克",
        data={"payment": build_transaction_response(transaction, db)}
    )


# ==================== 金料库存查询 ====================

@router.get("/balance", response_model=GoldMaterialBalanceResponse)
async def get_gold_balance(
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取金料库存余额"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看金料记录】的权限")
    
    balance = get_gold_balance_internal(db)
    
    return GoldMaterialBalanceResponse(
        total_income=balance["total_income"],
        total_expense=balance["total_expense"],
        current_balance=balance["current_balance"]
    )


# ==================== 期初金料设置 ====================

@router.post("/initial-balance")
async def set_initial_gold_balance(
    gold_weight: float = Query(..., description="期初金料克重"),
    remark: str = Query(default="期初金料库存", description="备注"),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    设置期初金料余额（仅管理层或料部可操作）
    使用新系统 GoldReceipt 创建期初金料记录
    """
    from ..models.finance import GoldReceipt
    
    # 权限检查 - 仅管理层和料部可以设置期初
    if user_role not in ['manager', 'material']:
        raise HTTPException(status_code=403, detail="仅管理层或料部可以设置期初金料")
    
    if gold_weight <= 0:
        raise HTTPException(status_code=400, detail="期初金料克重必须大于0")
    
    # 检查是否已有期初记录（在新系统中检查）
    existing = db.query(GoldReceipt).filter(
        GoldReceipt.is_initial_balance == True
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"已存在期初金料记录（{existing.receipt_no}：{existing.gold_weight}克），如需修改请联系管理员删除后重新设置"
        )
    
    # 生成期初单号（QC开头，代表"期初"）
    now = china_now()
    initial_no = f"QC{now.strftime('%Y%m%d')}001"
    
    # 创建期初金料记录（使用新系统 GoldReceipt）
    receipt = GoldReceipt(
        receipt_no=initial_no,
        customer_id=None,  # 期初金料不关联客户
        settlement_id=None,
        gold_weight=gold_weight,
        gold_fineness="足金999",
        is_initial_balance=True,  # 标记为期初金料
        status='received',  # 期初直接已接收
        created_by=user_role,
        received_by=user_role,
        received_at=now,
        remark=remark or "期初金料库存"
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)
    
    logger.info(f"设置期初金料: {gold_weight}克, 单号: {initial_no}, 操作人: {user_role}")
    
    return success_response(
        message=f"✅ 期初金料 {gold_weight}克 已成功设置",
        data={
            "transaction": {
                "transaction_no": initial_no,
                "gold_weight": gold_weight,
                "status": "received",
                "remark": remark or "期初金料库存",
                "created_at": now.isoformat()
            }
        }
    )


@router.get("/initial-balance")
async def get_initial_gold_balance(
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取期初金料信息（使用新系统 GoldReceipt）"""
    from ..models.finance import GoldReceipt
    
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足")
    
    # 查询期初记录（在新系统中查询）
    initial = db.query(GoldReceipt).filter(
        GoldReceipt.is_initial_balance == True
    ).first()
    
    if not initial:
        return success_response(
            message="尚未设置期初金料",
            data={"has_initial": False, "initial": None}
        )
    
    return success_response(
        data={
            "has_initial": True,
            "initial": {
                "transaction_no": initial.receipt_no,
                "gold_weight": initial.gold_weight,
                "status": initial.status,
                "remark": initial.remark,
                "created_at": initial.created_at.isoformat() if initial.created_at else None,
                "confirmed_at": initial.received_at.isoformat() if initial.received_at else None
            }
        }
    )


# ==================== 客户存料管理 ====================

@router.get("/customers/{customer_id}/deposit")
async def get_customer_deposit(
    customer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户存料信息（使用历史交易计算净金料，与AI分析一致）"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material') and not has_permission(user_role, 'can_view_customers'):
        raise HTTPException(status_code=403, detail="权限不足")
    
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    
    # ========== 使用统一计算函数（保证全系统口径一致） ==========
    from ..gold_balance import calculate_customer_net_gold
    net_gold = calculate_customer_net_gold(customer_id, db)  # 正数=存料，负数=欠料
    
    # 获取最近的存料交易记录
    recent_transactions = db.query(CustomerGoldDepositTransaction).filter(
        CustomerGoldDepositTransaction.customer_id == customer_id,
        CustomerGoldDepositTransaction.status == "active"
    ).order_by(desc(CustomerGoldDepositTransaction.created_at)).limit(20).all()
    
    return success_response(
        data={
            "customer": {
                "id": customer.id,
                "name": customer.name,
            },
            "customer_name": customer.name,
            "deposit": {
                "current_balance": net_gold,  # 使用统一计算函数的净金料值
                "net_gold": net_gold,         # 净金料字段（正=存料，负=欠料）
                "has_deposit": net_gold > 0,  # 是否有存料可提
            },
            "recent_transactions": [
                {
                    "id": t.id,
                    "transaction_type": t.transaction_type,
                    "amount": t.amount,
                    "balance_before": t.balance_before,
                    "balance_after": t.balance_after,
                    "created_at": t.created_at,
                    "created_by": t.created_by,
                    "remark": t.remark
                }
                for t in recent_transactions
            ]
        }
    )


# ==================== 客户往来账查询 ====================

@router.get("/customers/{customer_id}/transactions")
async def get_customer_transactions(
    customer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户往来账记录"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material') and not has_permission(user_role, 'can_view_customers'):
        raise HTTPException(status_code=403, detail="权限不足")
    
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    
    # 获取所有往来账记录
    transactions = db.query(CustomerTransaction).filter(
        CustomerTransaction.customer_id == customer_id,
        CustomerTransaction.status == "active"
    ).order_by(desc(CustomerTransaction.created_at)).all()
    
    # 计算当前欠款
    # 方法：从结算单计算
    settlements = db.query(SettlementOrder).join(SalesOrder).filter(
        SalesOrder.customer_id == customer_id,
        SettlementOrder.payment_method == 'physical_gold',
        SettlementOrder.status.in_(['pending', 'confirmed', 'printed'])
    ).all()
    
    total_due = Decimal("0")  # 应支付总重量
    total_received = Decimal("0")  # 已收总重量
    
    for settlement in settlements:
        total_due += to_decimal(settlement.physical_gold_weight)
        # 查询该结算单的所有已接收收料单（使用新系统 GoldReceipt）
        receipts = db.query(GoldReceipt).filter(
            GoldReceipt.settlement_id == settlement.id,
            GoldReceipt.status == 'received'
        ).all()
        total_received += sum(to_decimal(r.gold_weight) for r in receipts)
    
    current_due = max(Decimal("0"), total_due - total_received)  # 当前欠款（不能为负）
    
    # 获取存料信息（使用统一计算函数）
    from ..gold_balance import calculate_customer_net_gold
    current_deposit = calculate_customer_net_gold(customer_id, db)  # 正数=存料
    
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == customer_id
    ).first()
    total_deposited = deposit.total_deposited if deposit else 0.0
    total_used = deposit.total_used if deposit else 0.0
    
    # 构建交易记录响应
    transaction_list = []
    for t in transactions:
        # 获取关联单据号
        related_order_no = None
        if t.sales_order_id:
            sales = db.query(SalesOrder).filter(SalesOrder.id == t.sales_order_id).first()
            if sales:
                related_order_no = sales.order_no
        elif t.settlement_order_id:
            settlement = db.query(SettlementOrder).filter(SettlementOrder.id == t.settlement_order_id).first()
            if settlement:
                related_order_no = settlement.settlement_no
        elif t.gold_transaction_id:
            gold_trans = db.query(GoldMaterialTransaction).filter(
                GoldMaterialTransaction.id == t.gold_transaction_id
            ).first()
            if gold_trans:
                related_order_no = gold_trans.transaction_no
        
        transaction_list.append({
            "id": t.id,
            "customer_id": t.customer_id,
            "customer_name": t.customer_name,
            "transaction_type": t.transaction_type,
            "sales_order_id": t.sales_order_id,
            "settlement_order_id": t.settlement_order_id,
            "gold_transaction_id": t.gold_transaction_id,
            "related_order_no": related_order_no,
            "amount": t.amount,
            "gold_weight": t.gold_weight,
            "gold_due_before": t.gold_due_before,
            "gold_due_after": t.gold_due_after,
            "status": t.status,
            "created_at": t.created_at,
            "remark": t.remark
        })
    
    return success_response(
        data={
            "customer": {
                "id": customer.id,
                "name": customer.name,
            },
            "summary": {
                "current_gold_due": current_due,
                "total_gold_due": total_due,
                "total_gold_received": total_received,
                "current_deposit": current_deposit,
                "total_deposited": total_deposited,
                "total_used": total_used,
            },
            "transactions": transaction_list
        }
    )


# ==================== 取消金料流转记录 ====================

@router.post("/transactions/{transaction_id}/cancel")
async def cancel_transaction(
    transaction_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    取消金料流转记录（仅待确认状态可取消）
    
    注意：收料单(income)请使用新系统 GoldReceipt，此接口主要用于付料单(expense)
    """
    # 权限检查
    if not has_permission(user_role, 'can_manage_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【管理金料流转】的权限")
    
    transaction = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.id == transaction_id
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="记录不存在")
    
    if transaction.status != "pending":
        raise HTTPException(status_code=400, detail="只能取消待确认的记录")
    
    transaction.status = "cancelled"
    
    # 同时取消关联的往来账记录
    if transaction.transaction_type == 'income':
        customer_transaction = db.query(CustomerTransaction).filter(
            CustomerTransaction.gold_transaction_id == transaction.id
        ).first()
        if customer_transaction:
            customer_transaction.status = "cancelled"
    
    db.commit()
    
    logger.info(f"取消金料流转记录: {transaction.transaction_no}")
    
    return success_response(
        message="已取消",
        data={"transaction": build_transaction_response(transaction, db)}
    )


@router.post("/transactions/{transaction_id}/revoke")
async def revoke_gold_confirm(
    transaction_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    [已废弃] 撤销料部确认（仅管理层）
    
    ⚠️ 此接口用于旧系统 GoldMaterialTransaction，新系统请使用 GoldReceipt
    
    - 将confirmed状态回退到pending
    - 回滚客户存料更新（如果有）
    """
    # 权限检查（仅管理层）
    if not has_permission(user_role, 'can_manage_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：只有管理层可以撤销确认")
    
    transaction = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.id == transaction_id,
        GoldMaterialTransaction.transaction_type == 'income'
    ).first()
    
    if not transaction:
        raise HTTPException(status_code=404, detail="收料单不存在")
    
    if transaction.status != "confirmed":
        raise HTTPException(status_code=400, detail="只能撤销已确认的记录")
    
    # 回滚存料更新（如果有）
    if transaction.settlement_order_id and transaction.customer_id:
        settlement = db.query(SettlementOrder).filter(
            SettlementOrder.id == transaction.settlement_order_id
        ).first()
        
        if settlement:
            total_due = to_decimal(settlement.physical_gold_weight)
            # 计算其他已确认收料单的已收金料（不包括当前这条）
            other_receipts = db.query(GoldMaterialTransaction).filter(
                GoldMaterialTransaction.settlement_order_id == settlement.id,
                GoldMaterialTransaction.transaction_type == 'income',
                GoldMaterialTransaction.status == 'confirmed',
                GoldMaterialTransaction.id != transaction.id
            ).all()
            other_received = sum(to_decimal(r.gold_weight) for r in other_receipts)
            
            # 当前撤销后的总收金料
            total_received_after_revoke = other_received
            
            # 计算需要回滚的存料
            if total_received_after_revoke < total_due:
                # 撤销后总收金料 < 应支付，说明之前有存料，需要回滚
                # 计算之前确认时的存料
                total_received_before_revoke = other_received + to_decimal(transaction.gold_weight)
                if total_received_before_revoke > total_due:
                    deposit_to_revoke = to_decimal(total_received_before_revoke - total_due)
                    
                    # 回滚客户存料
                    deposit = db.query(CustomerGoldDeposit).filter(
                        CustomerGoldDeposit.customer_id == transaction.customer_id
                    ).first()
                    
                    if deposit and to_decimal(deposit.current_balance) >= to_decimal(deposit_to_revoke):
                        deposit.current_balance = round_weight(to_decimal(deposit.current_balance) - to_decimal(deposit_to_revoke))
                        deposit.total_deposited = round_weight(to_decimal(deposit.total_deposited) - to_decimal(deposit_to_revoke))
                        
                        # 取消存料交易记录
                        deposit_transaction = db.query(CustomerGoldDepositTransaction).filter(
                            CustomerGoldDepositTransaction.gold_transaction_id == transaction.id,
                            CustomerGoldDepositTransaction.transaction_type == 'deposit',
                            CustomerGoldDepositTransaction.status == 'active'
                        ).first()
                        if deposit_transaction:
                            deposit_transaction.status = "cancelled"
                        
                        logger.info(f"回滚客户存料: 客户={transaction.customer_name}, "
                                   f"回滚={deposit_to_revoke}克, 新余额={deposit.current_balance}克")
    
    # 回退状态
    transaction.status = "pending"
    transaction.confirmed_by = None
    transaction.confirmed_at = None
    
    db.commit()
    db.refresh(transaction)
    
    logger.info(f"撤销确认收料单: {transaction.transaction_no}, 操作人: {user_role}")
    
    return success_response(
        message="已撤销确认",
        data={"transaction": build_transaction_response(transaction, db)}
    )


# ==================== 获取所有金料流转记录 ====================

@router.get("/transactions")
async def get_all_transactions(
    transaction_type: Optional[str] = None,
    status: Optional[str] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取所有金料流转记录"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看金料记录】的权限")
    
    query = db.query(GoldMaterialTransaction)
    
    if transaction_type:
        query = query.filter(GoldMaterialTransaction.transaction_type == transaction_type)
    if status:
        query = query.filter(GoldMaterialTransaction.status == status)
    
    transactions = query.order_by(desc(GoldMaterialTransaction.created_at)).all()
    
    return success_response(
        data={
            "transactions": [build_transaction_response(t, db) for t in transactions],
            "total": len(transactions)
        }
    )


# ==================== 收料单打印和下载 ====================


@router.get("/receipts/{receipt_id}/download")
async def download_receipt(
    receipt_id: int,
    format: str = Query("pdf", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """
    [已废弃] 下载或打印收料单（旧系统）
    
    ⚠️ 此接口用于旧系统 GoldMaterialTransaction，新系统请使用 /gold-receipts/{id}/print
    """
    try:
        logger.info(f"下载收料单请求: receipt_id={receipt_id}, format={format}")
        
        # 查询收料单
        transaction = db.query(GoldMaterialTransaction).filter(
            GoldMaterialTransaction.id == receipt_id,
            GoldMaterialTransaction.transaction_type == 'income'
        ).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="收料单不存在")
        
        # 获取关联信息
        settlement_no = None
        if transaction.settlement_order_id:
            settlement = db.query(SettlementOrder).filter(
                SettlementOrder.id == transaction.settlement_order_id
            ).first()
            if settlement:
                settlement_no = settlement.settlement_no
        
        # 构建字段列表
        fields = build_gold_transaction_fields(
            transaction, 
            'receipt',
            {"settlement_no": settlement_no}
        )
        
        if format == "pdf":
            from fastapi.responses import StreamingResponse
            
            generator = PDFGenerator("收料单")
            generator.add_title()
            for label, value, _ in fields:
                if value:
                    generator.add_field(label, value)
            generator.add_footer()
            buffer = generator.generate()
            
            filename = f"receipt_{transaction.transaction_no}.pdf"
            return StreamingResponse(
                buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                }
            )
        
        elif format == "html":
            from fastapi.responses import HTMLResponse
            
            generator = HTMLGenerator("收料单", transaction.transaction_no)
            for label, value, full_width in fields:
                generator.add_field_if(label, value, full_width)
            html_content = generator.generate()
            
            return HTMLResponse(
                content=html_content
            )
        
        else:
            raise HTTPException(status_code=400, detail="不支持的格式，请使用 pdf 或 html")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成收料单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成收料单失败: {str(e)}")


# ==================== 付料单打印和下载 ====================


@router.get("/payments/{payment_id}/download")
async def download_payment(
    payment_id: int,
    format: str = Query("pdf", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """下载或打印付料单（支持PDF和HTML格式）"""
    try:
        logger.info(f"下载付料单请求: payment_id={payment_id}, format={format}")
        
        # 查询付料单
        transaction = db.query(GoldMaterialTransaction).filter(
            GoldMaterialTransaction.id == payment_id,
            GoldMaterialTransaction.transaction_type == 'expense'
        ).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="付料单不存在")
        
        # 获取关联信息
        inbound_order_no = None
        if transaction.inbound_order_id:
            inbound = db.query(InboundOrder).filter(
                InboundOrder.id == transaction.inbound_order_id
            ).first()
            if inbound:
                inbound_order_no = inbound.order_no
        
        # 构建字段列表
        fields = build_gold_transaction_fields(
            transaction, 
            'payment',
            {"inbound_order_no": inbound_order_no}
        )
        
        if format == "pdf":
            from fastapi.responses import StreamingResponse
            
            generator = PDFGenerator("付料单")
            generator.add_title()
            for label, value, _ in fields:
                if value:
                    generator.add_field(label, value)
            generator.add_footer()
            buffer = generator.generate()
            
            filename = f"payment_{transaction.transaction_no}.pdf"
            return StreamingResponse(
                buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                }
            )
        
        elif format == "html":
            from fastapi.responses import HTMLResponse
            
            generator = HTMLGenerator("付料单", transaction.transaction_no)
            for label, value, full_width in fields:
                generator.add_field_if(label, value, full_width)
            html_content = generator.generate()
            
            return HTMLResponse(
                content=html_content
            )
        
        else:
            raise HTTPException(status_code=400, detail="不支持的格式，请使用 pdf 或 html")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成付料单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成付料单失败: {str(e)}")


# ==================== 每日金料进出台账 ====================

@router.get("/ledger")
async def get_gold_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    获取每日金料进出台账（按日期分组），含收料/付料/提料统计汇总
    """
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看金料记录】的权限")

    try:
        if start_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_dt = date.today() - timedelta(days=30)
        if end_date:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_dt = date.today()
    except ValueError:
        raise HTTPException(status_code=400, detail="日期格式错误，请使用 YYYY-MM-DD 格式")

    try:
        from ..models.finance import GoldReceipt

        # ========== 1. 收料记录 ==========
        try:
            gold_receipts = db.query(GoldReceipt).filter(
                func.date(GoldReceipt.created_at) >= start_dt,
                func.date(GoldReceipt.created_at) <= end_dt
            ).order_by(GoldReceipt.created_at).all()
        except Exception as e:
            logger.warning(f"查询 gold_receipts 表失败（可能未建表）: {e}")
            db.rollback()
            gold_receipts = []

        # ========== 2. 付料记录 ==========
        try:
            expense_transactions = db.query(GoldMaterialTransaction).filter(
                GoldMaterialTransaction.transaction_type == 'expense',
                GoldMaterialTransaction.status == 'confirmed',
                func.date(GoldMaterialTransaction.confirmed_at) >= start_dt,
                func.date(GoldMaterialTransaction.confirmed_at) <= end_dt
            ).order_by(GoldMaterialTransaction.confirmed_at).all()
        except Exception as e:
            logger.warning(f"查询 gold_material_transactions 表失败: {e}")
            db.rollback()
            expense_transactions = []

        # ========== 3. 提料记录 ==========
        try:
            withdrawals = db.query(CustomerWithdrawal).filter(
                func.date(CustomerWithdrawal.created_at) >= start_dt,
                func.date(CustomerWithdrawal.created_at) <= end_dt
            ).order_by(CustomerWithdrawal.created_at).all()
        except Exception as e:
            logger.warning(f"查询 customer_withdrawals 表失败: {e}")
            db.rollback()
            withdrawals = []

        # ========== 按日期分组 ==========
        ledger_by_date: Dict[str, Dict] = defaultdict(lambda: {
            "date": None,
            "income": 0.0,
            "expense": 0.0,
            "withdrawal": 0.0,
            "net": 0.0,
            "transactions": []
        })

        # --- 收料 ---
        receipt_by_customer: Dict[str, dict] = {}
        for r in gold_receipts:
            if not r.created_at:
                continue
            date_str = r.created_at.date().strftime('%Y-%m-%d')
            weight = safe_float_for_json(r.gold_weight)
            ledger_by_date[date_str]["date"] = date_str
            ledger_by_date[date_str]["income"] += weight

            customer_name = db.query(Customer.name).filter(Customer.id == r.customer_id).scalar() or "未知客户"
            ledger_by_date[date_str]["transactions"].append({
                "id": r.id,
                "transaction_no": r.receipt_no,
                "transaction_type": "income",
                "gold_weight": weight,
                "customer_name": customer_name,
                "supplier_name": None,
                "confirmed_at": r.created_at.isoformat() if r.created_at else None,
                "source": "gold_receipt"
            })
            key = str(r.customer_id)
            if key not in receipt_by_customer:
                receipt_by_customer[key] = {"customer_id": r.customer_id, "customer_name": customer_name, "count": 0, "total_weight": 0.0}
            receipt_by_customer[key]["count"] += 1
            receipt_by_customer[key]["total_weight"] += weight

        # --- 付料 ---
        expense_by_supplier: Dict[str, dict] = {}
        for t in expense_transactions:
            if not t.confirmed_at:
                continue
            date_str = t.confirmed_at.date().strftime('%Y-%m-%d')
            weight = safe_float_for_json(t.gold_weight)
            ledger_by_date[date_str]["date"] = date_str
            ledger_by_date[date_str]["expense"] += weight

            ledger_by_date[date_str]["transactions"].append({
                "id": t.id,
                "transaction_no": t.transaction_no,
                "transaction_type": "expense",
                "gold_weight": weight,
                "customer_name": t.customer_name,
                "supplier_name": t.supplier_name,
                "confirmed_at": t.confirmed_at.isoformat() if t.confirmed_at else None,
                "source": "gold_material_transaction"
            })
            key = str(t.supplier_id or "unknown")
            if key not in expense_by_supplier:
                expense_by_supplier[key] = {"supplier_id": t.supplier_id, "supplier_name": t.supplier_name or "未知供应商", "count": 0, "total_weight": 0.0}
            expense_by_supplier[key]["count"] += 1
            expense_by_supplier[key]["total_weight"] += weight

        # --- 提料 ---
        withdrawal_by_customer: Dict[str, dict] = {}
        for w in withdrawals:
            if not w.created_at:
                continue
            date_str = w.created_at.date().strftime('%Y-%m-%d')
            weight = safe_float_for_json(w.gold_weight)
            ledger_by_date[date_str]["date"] = date_str
            ledger_by_date[date_str]["withdrawal"] += weight

            ledger_by_date[date_str]["transactions"].append({
                "id": w.id,
                "transaction_no": w.withdrawal_no,
                "transaction_type": "withdrawal",
                "gold_weight": weight,
                "customer_name": w.customer_name,
                "supplier_name": None,
                "confirmed_at": w.created_at.isoformat() if w.created_at else None,
                "source": "customer_withdrawal",
                "status": w.status
            })
            key = str(w.customer_id)
            if key not in withdrawal_by_customer:
                withdrawal_by_customer[key] = {"customer_id": w.customer_id, "customer_name": w.customer_name, "count": 0, "total_weight": 0.0}
            withdrawal_by_customer[key]["count"] += 1
            withdrawal_by_customer[key]["total_weight"] += weight

        # ========== 计算净额 ==========
        for date_str in ledger_by_date:
            d = ledger_by_date[date_str]
            d["net"] = d["income"] - d["expense"] - d["withdrawal"]

        ledger_list = sorted(
            [v for v in ledger_by_date.values() if v["date"]],
            key=lambda x: x["date"],
            reverse=True
        )

        total_income = sum(day["income"] for day in ledger_list)
        total_expense = sum(day["expense"] for day in ledger_list)
        total_withdrawal = sum(day["withdrawal"] for day in ledger_list)
        total_net = total_income - total_expense - total_withdrawal

        return success_response(
            data={
                "start_date": start_date or start_dt.strftime('%Y-%m-%d'),
                "end_date": end_date or end_dt.strftime('%Y-%m-%d'),
                "ledger": ledger_list,
                "summary": {
                    "total_income": round(total_income, 3),
                    "total_expense": round(total_expense, 3),
                    "total_withdrawal": round(total_withdrawal, 3),
                    "total_net": round(total_net, 3),
                    "days_count": len(ledger_list),
                    "income_count": len(gold_receipts),
                    "expense_count": len(expense_transactions),
                    "withdrawal_count": len(withdrawals),
                },
                "breakdown": {
                    "income_by_customer": sorted(receipt_by_customer.values(), key=lambda x: x["total_weight"], reverse=True),
                    "expense_by_supplier": sorted(expense_by_supplier.values(), key=lambda x: x["total_weight"], reverse=True),
                    "withdrawal_by_customer": sorted(withdrawal_by_customer.values(), key=lambda x: x["total_weight"], reverse=True),
                }
            }
        )
    except Exception as e:
        logger.error(f"获取金料台账失败: {e}", exc_info=True)
        db.rollback()
        return error_response(message=f"获取金料台账失败: {str(e)}")


@router.get("/ledger/export")
async def export_gold_ledger(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = Query("excel", pattern="^(excel|pdf)$"),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """导出每日金料进出台账（Excel或PDF）"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看金料记录】的权限")
    
    # 获取台账数据
    ledger_data = await get_gold_ledger(start_date, end_date, user_role, db)
    
    if format == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            import io
            
            wb = Workbook()
            ws = wb.active
            ws.title = "金料进出台账"
            
            # 标题
            ws.merge_cells('A1:F1')
            ws['A1'] = f"金料进出台账 ({ledger_data['start_date']} 至 {ledger_data['end_date']})"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 表头
            headers = ['日期', '收入（克）', '支出（克）', '净额（克）', '收入笔数', '支出笔数']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据
            row = 3
            for day in ledger_data['ledger']:
                income_count = sum(1 for t in day['transactions'] if t['transaction_type'] == 'income')
                expense_count = sum(1 for t in day['transactions'] if t['transaction_type'] == 'expense')
                
                ws.cell(row=row, column=1).value = day['date']
                ws.cell(row=row, column=2).value = day['income']
                ws.cell(row=row, column=3).value = day['expense']
                ws.cell(row=row, column=4).value = day['net']
                ws.cell(row=row, column=5).value = income_count
                ws.cell(row=row, column=6).value = expense_count
                
                # 设置数字格式
                for col in [2, 3, 4]:
                    ws.cell(row=row, column=col).number_format = '0.00'
                
                row += 1
            
            # 总计行
            ws.cell(row=row, column=1).value = "总计"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = ledger_data['summary']['total_income']
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=2).number_format = '0.00'
            ws.cell(row=row, column=3).value = ledger_data['summary']['total_expense']
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.cell(row=row, column=4).value = ledger_data['summary']['total_net']
            ws.cell(row=row, column=4).font = Font(bold=True)
            ws.cell(row=row, column=4).number_format = '0.00'
            
            # 调整列宽
            ws.column_dimensions['A'].width = 15
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 15
            
            # 保存到内存
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"gold_ledger_{ledger_data['start_date']}_{ledger_data['end_date']}.xlsx"
            
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                }
            )
        except Exception as e:
            logger.error(f"生成Excel失败: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"生成Excel失败: {str(e)}")
    
    elif format == "pdf":
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            import io
            from ..timezone_utils import format_china_time
            
            # 自定义纸张尺寸：241mm × 140mm 横向（针式打印机）
            PAGE_WIDTH = 241 * mm
            PAGE_HEIGHT = 140 * mm
            
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))
            width, height = PAGE_WIDTH, PAGE_HEIGHT
            
            # 使用 CID 字体
            try:
                pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                chinese_font = 'STSong-Light'
            except Exception:
                chinese_font = None
            
            # 页边距
            left_margin = 8 * mm
            right_margin = width - 8 * mm
            top_margin = height - 6 * mm
            
            # 标题（居中）
            if chinese_font:
                p.setFont(chinese_font, 11)
            else:
                p.setFont("Helvetica-Bold", 11)
            title = f"金料进出台账 ({ledger_data['start_date']} 至 {ledger_data['end_date']})"
            p.drawCentredString(width / 2, top_margin, title)
            
            # 表头
            y = top_margin - 16
            col_x = [left_margin, 55*mm, 100*mm, 150*mm]
            if chinese_font:
                p.setFont(chinese_font, 8)
            else:
                p.setFont("Helvetica-Bold", 8)
            
            p.drawString(col_x[0], y, "日期")
            p.drawString(col_x[1], y, "收入（克）")
            p.drawString(col_x[2], y, "支出（克）")
            p.drawString(col_x[3], y, "净额（克）")
            y -= 8
            p.line(left_margin, y, right_margin, y)
            y -= 8
            
            # 数据
            if chinese_font:
                p.setFont(chinese_font, 7)
            else:
                p.setFont("Helvetica", 7)
            
            bottom_margin = 20 * mm
            for day in ledger_data['ledger']:
                if y < bottom_margin:  # 换页
                    p.showPage()
                    y = top_margin - 10
                    # 重绘表头
                    if chinese_font:
                        p.setFont(chinese_font, 8)
                    p.drawString(col_x[0], y, "日期")
                    p.drawString(col_x[1], y, "收入（克）")
                    p.drawString(col_x[2], y, "支出（克）")
                    p.drawString(col_x[3], y, "净额（克）")
                    y -= 8
                    p.line(left_margin, y, right_margin, y)
                    y -= 8
                    if chinese_font:
                        p.setFont(chinese_font, 7)
                
                p.drawString(col_x[0], y, day['date'])
                p.setFont("Helvetica", 7)
                p.drawString(col_x[1], y, f"{day['income']:.2f}")
                p.drawString(col_x[2], y, f"{day['expense']:.2f}")
                p.drawString(col_x[3], y, f"{day['net']:.2f}")
                y -= 10
            
            # 总计
            y -= 3
            p.line(left_margin, y, right_margin, y)
            y -= 10
            if chinese_font:
                p.setFont(chinese_font, 8)
            else:
                p.setFont("Helvetica-Bold", 8)
            p.drawString(col_x[0], y, "合计")
            p.setFont("Helvetica", 8)
            p.drawString(col_x[1], y, f"{ledger_data['summary']['total_income']:.2f}")
            p.drawString(col_x[2], y, f"{ledger_data['summary']['total_expense']:.2f}")
            p.drawString(col_x[3], y, f"{ledger_data['summary']['total_net']:.2f}")
            
            # 打印时间（底部）
            print_time = format_china_time(china_now(), '%Y-%m-%d %H:%M')
            if chinese_font:
                p.setFont(chinese_font, 6)
            p.drawString(left_margin, 5*mm, f"打印时间：{print_time}")
            
            p.save()
            buffer.seek(0)
            
            filename = f"gold_ledger_{ledger_data['start_date']}_{ledger_data['end_date']}.pdf"
            
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                }
            )
        except Exception as e:
            logger.error(f"生成PDF失败: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"生成PDF失败: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail="不支持的格式，请使用 excel 或 pdf")


# ==================== 客户取料单管理 ====================

def get_or_create_customer_deposit(customer_id: int, customer_name: str, db: Session) -> CustomerGoldDeposit:
    """获取或创建客户存料记录"""
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == customer_id
    ).first()
    
    if not deposit:
        deposit = CustomerGoldDeposit(
            customer_id=customer_id,
            customer_name=customer_name,
            current_balance=0.0,
            total_deposited=0.0,
            total_used=0.0
        )
        db.add(deposit)
        db.flush()
    
    return deposit


@router.post("/withdrawals", response_model=CustomerWithdrawalResponse)
async def create_customer_withdrawal(
    data: CustomerWithdrawalCreate,
    user_role: str = Query(default="sales", description="用户角色"),
    created_by: str = Query(default="结算", description="创建人"),
    db: Session = Depends(get_db)
):
    """
    创建客户取料单
    
    - 验证客户存料余额是否足够
    - 创建取料单（待确认状态，不扣余额）
    - 确认时才扣减客户存料余额
    """
    # 权限检查
    if not has_permission(user_role, 'can_create_withdrawal'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【创建取料单】的权限")
    
    # 验证客户
    customer = db.query(Customer).filter(Customer.id == data.customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    
    # 使用统一计算函数验证存料余额（保证全系统口径一致）
    from ..gold_balance import calculate_customer_net_gold
    current_balance = calculate_customer_net_gold(data.customer_id, db)  # 正数=存料
    
    if current_balance < data.gold_weight:
        raise HTTPException(
            status_code=400,
            detail=f"客户存料余额不足。当前余额：{current_balance:.3f}克，申请取料：{data.gold_weight:.3f}克"
        )
    
    # 获取 deposit 记录用于后续日志
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == data.customer_id
    ).first()
    
    # 生成取料单号（QL开头）
    now = china_now()
    count = db.query(CustomerWithdrawal).filter(
        CustomerWithdrawal.withdrawal_no.like(f"QL{now.strftime('%Y%m%d')}%")
    ).count()
    withdrawal_no = f"QL{now.strftime('%Y%m%d')}{count + 1:03d}"
    
    # 创建取料单（待确认状态，不扣余额）
    withdrawal = CustomerWithdrawal(
        withdrawal_no=withdrawal_no,
        customer_id=customer.id,
        customer_name=customer.name,
        gold_weight=data.gold_weight,
        withdrawal_type=data.withdrawal_type,
        destination_company=data.destination_company,
        destination_address=data.destination_address,
        authorized_person=data.authorized_person,
        authorized_phone=data.authorized_phone,
        status="pending",
        created_by=created_by,
        created_at=now,
        remark=data.remark
    )
    db.add(withdrawal)
    
    db.commit()
    db.refresh(withdrawal)
    
    logger.info(f"创建取料单(待确认): {withdrawal_no}, 客户: {customer.name}, 克重: {data.gold_weight}克, 当前余额: {deposit.current_balance:.2f}克")
    
    return withdrawal


@router.get("/withdrawals")
async def get_customer_withdrawals(
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户取料单列表"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足")
    
    query = db.query(CustomerWithdrawal).options(
        joinedload(CustomerWithdrawal.customer)
    )
    
    if status:
        query = query.filter(CustomerWithdrawal.status == status)
    if customer_id:
        query = query.filter(CustomerWithdrawal.customer_id == customer_id)
    
    withdrawals = query.order_by(desc(CustomerWithdrawal.created_at)).all()
    
    return success_response(
        data={
            "withdrawals": [CustomerWithdrawalResponse.model_validate(w) for w in withdrawals],
            "total": len(withdrawals)
        }
    )


@router.post("/withdrawals/{withdrawal_id}/complete")
async def complete_customer_withdrawal(
    withdrawal_id: int,
    data: CustomerWithdrawalComplete,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    确认取料单（扣减客户存料余额）
    
    - 验证客户存料余额仍然足够
    - 扣减客户存料余额
    - 创建存料交易记录
    - 创建操作日志
    - 更新取料单状态为completed
    """
    # 权限检查
    if not has_permission(user_role, 'can_complete_withdrawal'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【确认取料】的权限")
    
    # 加行锁防止并发确认
    withdrawal = db.query(CustomerWithdrawal).filter(
        CustomerWithdrawal.id == withdrawal_id
    ).with_for_update().first()
    
    if not withdrawal:
        raise HTTPException(status_code=404, detail="取料单不存在")
    
    if withdrawal.status != "pending":
        raise HTTPException(status_code=409, detail=f"取料单状态为 {withdrawal.status}，无法确认（可能已被其他人确认）")
    
    # 锁定客户存料余额后再验证
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == withdrawal.customer_id
    ).with_for_update().first()
    
    # deposit.current_balance 负数=客户存料在我们这里，可提料量 = abs(负数)
    available_gold = abs(deposit.current_balance) if deposit and deposit.current_balance < 0 else (deposit.current_balance if deposit else 0.0)
    if not deposit or available_gold < withdrawal.gold_weight:
        raise HTTPException(
            status_code=400,
            detail=f"客户存料余额不足。当前存料：{available_gold:.2f}克，取料需要：{withdrawal.gold_weight:.2f}克"
        )
    
    now = china_now()
    
    # 扣减客户存料余额（负数存料：提料后向0靠近，所以 +gold_weight）
    w_gold_weight = to_decimal(withdrawal.gold_weight)
    balance_before = to_decimal(deposit.current_balance)
    balance_after = round_weight(balance_before + w_gold_weight)
    deposit.current_balance = balance_after
    deposit.total_used = round_weight(to_decimal(deposit.total_used) + w_gold_weight)
    deposit.last_transaction_at = now
    
    # 创建存料交易记录
    deposit_transaction = CustomerGoldDepositTransaction(
        customer_id=withdrawal.customer_id,
        customer_name=withdrawal.customer_name,
        transaction_type="use",
        amount=w_gold_weight,
        balance_before=balance_before,
        balance_after=balance_after,
        created_by=data.completed_by,
        remark=f"提料单确认：{withdrawal.withdrawal_no}"
    )
    db.add(deposit_transaction)
    
    # 更新取料单状态为已确认
    withdrawal.status = "completed"
    withdrawal.completed_by = data.completed_by
    withdrawal.completed_at = now
    
    # 创建操作日志
    log = OrderStatusLog(
        order_type="withdrawal",
        order_id=withdrawal.id,
        action="confirm",
        old_status="pending",
        new_status="completed",
        operated_by=data.completed_by,
        operated_at=now,
        remark=f"确认取料单，扣减余额 {withdrawal.gold_weight:.2f}克（{balance_before:.2f} → {balance_after:.2f}）"
    )
    db.add(log)
    
    db.commit()
    db.refresh(withdrawal)
    
    logger.info(f"确认取料单: {withdrawal.withdrawal_no}, 客户: {withdrawal.customer_name}, 克重: {withdrawal.gold_weight}克, 余额: {balance_before:.2f} -> {balance_after:.2f}, 确认人: {data.completed_by}")
    
    return success_response(
        message=f"取料单 {withdrawal.withdrawal_no} 已确认，已扣减存料 {withdrawal.gold_weight:.2f}克",
        data={"withdrawal": CustomerWithdrawalResponse.model_validate(withdrawal)}
    )


@router.post("/withdrawals/{withdrawal_id}/cancel")
async def cancel_customer_withdrawal(
    withdrawal_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    cancelled_by: str = Query(default="结算", description="取消人"),
    db: Session = Depends(get_db)
):
    """取消取料单（仅待确认状态可取消，pending状态不涉及余额变动）"""
    withdrawal = db.query(CustomerWithdrawal).filter(
        CustomerWithdrawal.id == withdrawal_id
    ).first()
    
    if not withdrawal:
        raise HTTPException(status_code=404, detail="取料单不存在")
    
    if withdrawal.status != "pending":
        raise HTTPException(status_code=400, detail="只能取消待确认的取料单")
    
    now = china_now()
    withdrawal.status = "cancelled"
    
    # 创建操作日志
    log = OrderStatusLog(
        order_type="withdrawal",
        order_id=withdrawal.id,
        action="cancel",
        old_status="pending",
        new_status="cancelled",
        operated_by=cancelled_by,
        operated_at=now,
        remark=f"取消取料单：{withdrawal.withdrawal_no}"
    )
    db.add(log)
    
    db.commit()
    db.refresh(withdrawal)
    
    logger.info(f"取消取料单: {withdrawal.withdrawal_no}, 客户: {withdrawal.customer_name}, 取消人: {cancelled_by}")
    
    return success_response(
        message="取料单已取消",
        data={"withdrawal": CustomerWithdrawalResponse.model_validate(withdrawal)}
    )


@router.post("/withdrawals/{withdrawal_id}/unconfirm")
async def unconfirm_customer_withdrawal(
    withdrawal_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    operated_by: str = Query(default="结算", description="操作人"),
    reason: str = Query(default="", description="反确认原因"),
    db: Session = Depends(get_db)
):
    """
    反确认取料单（回滚已确认的取料单）
    
    - 验证状态为completed
    - 恢复客户存料余额
    - 创建反向存料交易记录（留痕）
    - 创建操作日志（留痕）
    - 状态 completed -> pending
    """
    # 权限检查
    if not has_permission(user_role, 'can_complete_withdrawal'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【反确认取料单】的权限")
    
    withdrawal = db.query(CustomerWithdrawal).filter(
        CustomerWithdrawal.id == withdrawal_id
    ).first()
    
    if not withdrawal:
        raise HTTPException(status_code=404, detail="取料单不存在")
    
    if withdrawal.status != "completed":
        raise HTTPException(status_code=400, detail=f"取料单状态为 {withdrawal.status}，只能反确认已确认的取料单")
    
    # 恢复客户存料余额
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == withdrawal.customer_id
    ).first()
    
    if not deposit:
        raise HTTPException(status_code=500, detail="客户存料记录不存在，无法恢复余额")
    
    now = china_now()
    w_gold_weight = to_decimal(withdrawal.gold_weight)
    balance_before = to_decimal(deposit.current_balance)
    balance_after = round_weight(balance_before + w_gold_weight)
    deposit.current_balance = balance_after
    deposit.total_used = round_weight(max(Decimal("0"), to_decimal(deposit.total_used) - w_gold_weight))
    deposit.last_transaction_at = now
    
    # 创建反向存料交易记录（留痕）
    refund_reason = f"反确认提料单：{withdrawal.withdrawal_no}"
    if reason:
        refund_reason += f"，原因：{reason}"
    
    deposit_transaction = CustomerGoldDepositTransaction(
        customer_id=withdrawal.customer_id,
        customer_name=withdrawal.customer_name,
        transaction_type="refund",
        amount=w_gold_weight,
        balance_before=balance_before,
        balance_after=balance_after,
        created_by=operated_by,
        remark=refund_reason
    )
    db.add(deposit_transaction)
    
    # 更新取料单状态
    withdrawal.status = "pending"
    withdrawal.completed_by = None
    withdrawal.completed_at = None
    
    # 创建操作日志（留痕）
    log = OrderStatusLog(
        order_type="withdrawal",
        order_id=withdrawal.id,
        action="unconfirm",
        old_status="completed",
        new_status="pending",
        operated_by=operated_by,
        operated_at=now,
        remark=f"反确认取料单，恢复余额 {withdrawal.gold_weight:.2f}克（{balance_before:.2f} → {balance_after:.2f}）" + 
               (f"，原因：{reason}" if reason else "")
    )
    db.add(log)
    
    db.commit()
    db.refresh(withdrawal)
    
    logger.info(f"反确认取料单: {withdrawal.withdrawal_no}, 客户: {withdrawal.customer_name}, 克重: {withdrawal.gold_weight}克, 余额: {balance_before:.2f} -> {balance_after:.2f}, 操作人: {operated_by}")
    
    return success_response(
        message=f"取料单 {withdrawal.withdrawal_no} 已反确认，已恢复存料 {withdrawal.gold_weight:.2f}克",
        data={"withdrawal": CustomerWithdrawalResponse.model_validate(withdrawal)}
    )


@router.get("/withdrawals/{withdrawal_id}/download")
async def download_customer_withdrawal(
    withdrawal_id: int,
    format: str = Query("html", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """下载或打印取料单"""
    withdrawal = db.query(CustomerWithdrawal).filter(
        CustomerWithdrawal.id == withdrawal_id
    ).first()
    
    if not withdrawal:
        raise HTTPException(status_code=404, detail="取料单不存在")
    
    # 更新打印时间
    withdrawal.printed_at = china_now()
    db.commit()
    
    # 构建字段
    fields = [
        ("取料单号", withdrawal.withdrawal_no, False),
        ("客户名称", withdrawal.customer_name, False),
        ("取料克重", f"{withdrawal.gold_weight:.2f} 克", False),
        ("取料方式", "自取" if withdrawal.withdrawal_type == "self" else "送到其他公司", False),
    ]
    
    if withdrawal.destination_company:
        fields.append(("目的地公司", withdrawal.destination_company, False))
    if withdrawal.destination_address:
        fields.append(("目的地地址", withdrawal.destination_address, True))
    if withdrawal.authorized_person:
        fields.append(("授权取料人", withdrawal.authorized_person, False))
    if withdrawal.authorized_phone:
        fields.append(("取料人电话", withdrawal.authorized_phone, False))
    
    fields.append(("状态", get_status_label(withdrawal.status), False))
    fields.append(("创建时间", format_datetime(withdrawal.created_at), False))
    
    _ai_keywords = ("AI助手", "协同卡片确认", "系统-协同卡片", "协同卡片", "系统")
    if withdrawal.created_by and not any(k in withdrawal.created_by for k in _ai_keywords):
        fields.append(("创建人", withdrawal.created_by, False))
    if withdrawal.completed_by and not any(k in withdrawal.completed_by for k in _ai_keywords):
        fields.append(("完成人", withdrawal.completed_by, False))
    if withdrawal.completed_at:
        fields.append(("完成时间", format_datetime(withdrawal.completed_at), False))
    if withdrawal.remark and not any(k in withdrawal.remark for k in _ai_keywords):
        fields.append(("备注", withdrawal.remark, True))
    
    if format == "pdf":
        from fastapi.responses import StreamingResponse
        
        generator = PDFGenerator("客户取料单")
        generator.add_title()
        for label, value, _ in fields:
            if value:
                generator.add_field(label, value)
        generator.add_footer()
        buffer = generator.generate()
        
        filename = f"withdrawal_{withdrawal.withdrawal_no}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            }
        )
    
    else:
        from fastapi.responses import HTMLResponse
        
        generator = HTMLGenerator("客户取料单", withdrawal.withdrawal_no)
        for label, value, full_width in fields:
            generator.add_field_if(label, value, full_width)
        html_content = generator.generate()
        
        return HTMLResponse(
            content=html_content
        )


# ==================== 客户转料单管理 ====================

@router.post("/transfers", response_model=CustomerTransferResponse)
async def create_customer_transfer(
    data: CustomerTransferCreate,
    user_role: str = Query(default="sales", description="用户角色"),
    created_by: str = Query(default="结算", description="创建人"),
    db: Session = Depends(get_db)
):
    """
    创建客户转料单
    
    - 验证转出客户存料余额是否足够
    - 创建转料单（待料部确认）
    """
    # 权限检查
    if not has_permission(user_role, 'can_create_transfer'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【创建转料单】的权限")
    
    # 验证转出客户
    from_customer = db.query(Customer).filter(Customer.id == data.from_customer_id).first()
    if not from_customer:
        raise HTTPException(status_code=404, detail="转出客户不存在")
    
    # 验证转入客户
    to_customer = db.query(Customer).filter(Customer.id == data.to_customer_id).first()
    if not to_customer:
        raise HTTPException(status_code=404, detail="转入客户不存在")
    
    if data.from_customer_id == data.to_customer_id:
        raise HTTPException(status_code=400, detail="转出客户和转入客户不能相同")
    
    # 按ID顺序锁定两个客户存料记录，防止死锁
    ids_sorted = sorted([data.from_customer_id, data.to_customer_id])
    for cid in ids_sorted:
        db.query(CustomerGoldDeposit).filter(
            CustomerGoldDeposit.customer_id == cid
        ).with_for_update().first()
    
    from_deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == data.from_customer_id
    ).first()
    
    if not from_deposit or from_deposit.current_balance < data.gold_weight:
        current_balance = from_deposit.current_balance if from_deposit else 0.0
        raise HTTPException(
            status_code=400,
            detail=f"转出客户存料余额不足。当前余额：{current_balance:.2f}克，申请转料：{data.gold_weight:.2f}克"
        )
    
    # 生成转料单号（ZL开头）
    now = china_now()
    count = db.query(CustomerTransfer).filter(
        CustomerTransfer.transfer_no.like(f"ZL{now.strftime('%Y%m%d')}%")
    ).count()
    transfer_no = f"ZL{now.strftime('%Y%m%d')}{count + 1:03d}"
    
    # 获取或创建转入客户存料记录（已锁定）
    to_deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == to_customer.id
    ).first()
    if not to_deposit:
        to_deposit = get_or_create_customer_deposit(
            to_customer.id,
            to_customer.name,
            db
        )
    
    # 创建转料单（直接完成，因为只是系统调换，不涉及实物）
    transfer = CustomerTransfer(
        transfer_no=transfer_no,
        from_customer_id=from_customer.id,
        from_customer_name=from_customer.name,
        to_customer_id=to_customer.id,
        to_customer_name=to_customer.name,
        gold_weight=data.gold_weight,
        status="completed",  # 直接完成，不需要料部确认
        created_by=created_by,
        confirmed_by=created_by,  # 创建人即确认人
        confirmed_at=now,
        remark=data.remark
    )
    db.add(transfer)
    db.flush()
    
    # 扣减转出客户存料
    t_gold_weight = to_decimal(data.gold_weight)
    from_balance_before = to_decimal(from_deposit.current_balance)
    from_deposit.current_balance = round_weight(from_balance_before - t_gold_weight)
    from_deposit.total_used = round_weight(to_decimal(from_deposit.total_used) + t_gold_weight)
    from_deposit.last_transaction_at = now
    
    # 增加转入客户存料
    to_balance_before = to_decimal(to_deposit.current_balance)
    to_deposit.current_balance = round_weight(to_balance_before + t_gold_weight)
    to_deposit.total_deposited = round_weight(to_decimal(to_deposit.total_deposited) + t_gold_weight)
    to_deposit.last_transaction_at = now
    
    # 创建转出客户交易记录
    from_transaction = CustomerGoldDepositTransaction(
        customer_id=from_customer.id,
        customer_name=from_customer.name,
        transaction_type='use',
        amount=t_gold_weight,
        balance_before=from_balance_before,
        balance_after=from_deposit.current_balance,
        created_by=created_by,
        remark=f"转料单：{transfer_no}，转出至：{to_customer.name}"
    )
    db.add(from_transaction)
    
    # 创建转入客户交易记录
    to_transaction = CustomerGoldDepositTransaction(
        customer_id=to_customer.id,
        customer_name=to_customer.name,
        transaction_type='deposit',
        amount=t_gold_weight,
        balance_before=to_balance_before,
        balance_after=to_deposit.current_balance,
        created_by=created_by,
        remark=f"转料单：{transfer_no}，转入自：{from_customer.name}"
    )
    db.add(to_transaction)
    
    db.commit()
    db.refresh(transfer)
    
    logger.info(f"创建并完成转料单: {transfer_no}, "
               f"{from_customer.name}({from_deposit.current_balance}克) -> "
               f"{to_customer.name}({to_deposit.current_balance}克), "
               f"克重: {data.gold_weight}克")
    
    return transfer


@router.get("/transfers")
async def get_customer_transfers(
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户转料单列表"""
    # 权限检查
    if not has_permission(user_role, 'can_view_gold_material'):
        raise HTTPException(status_code=403, detail="权限不足")
    
    query = db.query(CustomerTransfer).options(
        joinedload(CustomerTransfer.from_customer),
        joinedload(CustomerTransfer.to_customer)
    )
    
    if status:
        query = query.filter(CustomerTransfer.status == status)
    if customer_id:
        query = query.filter(
            (CustomerTransfer.from_customer_id == customer_id) | 
            (CustomerTransfer.to_customer_id == customer_id)
        )
    
    transfers = query.order_by(desc(CustomerTransfer.created_at)).all()
    
    return success_response(
        data={
            "transfers": [CustomerTransferResponse.model_validate(t) for t in transfers],
            "total": len(transfers)
        }
    )


@router.post("/transfers/{transfer_id}/confirm")
async def confirm_customer_transfer(
    transfer_id: int,
    data: CustomerTransferConfirm,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    确认转料单（料部确认）
    
    - 更新转料单状态为completed
    - 扣减转出客户存料余额
    - 增加转入客户存料余额
    - 创建双方存料交易记录
    """
    # 权限检查
    if not has_permission(user_role, 'can_confirm_transfer'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【确认转料】的权限")
    
    # 加行锁防止并发确认
    transfer = db.query(CustomerTransfer).filter(
        CustomerTransfer.id == transfer_id
    ).with_for_update().first()
    
    if not transfer:
        raise HTTPException(status_code=404, detail="转料单不存在")
    
    if transfer.status != "pending":
        raise HTTPException(status_code=409, detail=f"转料单状态为 {transfer.status}，无法确认（可能已被其他人确认）")
    
    # 按ID顺序锁定两个客户存料记录，防止死锁
    ids_sorted = sorted([transfer.from_customer_id, transfer.to_customer_id])
    for cid in ids_sorted:
        db.query(CustomerGoldDeposit).filter(
            CustomerGoldDeposit.customer_id == cid
        ).with_for_update().first()
    
    from_deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == transfer.from_customer_id
    ).first()
    
    if not from_deposit or from_deposit.current_balance < transfer.gold_weight:
        current_balance = from_deposit.current_balance if from_deposit else 0.0
        raise HTTPException(
            status_code=400,
            detail=f"转出客户存料余额不足。当前余额：{current_balance:.2f}克，转料需要：{transfer.gold_weight:.2f}克"
        )
    
    # 获取转入客户存料记录（已锁定）
    to_deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == transfer.to_customer_id
    ).first()
    if not to_deposit:
        to_deposit = get_or_create_customer_deposit(
            transfer.to_customer_id,
            transfer.to_customer_name,
            db
        )
    
    # 更新转料单状态
    transfer.status = "completed"
    transfer.confirmed_by = data.confirmed_by
    transfer.confirmed_at = china_now()
    
    # 扣减转出客户存料
    t_gold_weight = to_decimal(transfer.gold_weight)
    from_balance_before = to_decimal(from_deposit.current_balance)
    from_deposit.current_balance = round_weight(from_balance_before - t_gold_weight)
    from_deposit.total_used = round_weight(to_decimal(from_deposit.total_used) + t_gold_weight)
    from_deposit.last_transaction_at = china_now()
    
    # 增加转入客户存料
    to_balance_before = to_decimal(to_deposit.current_balance)
    to_deposit.current_balance = round_weight(to_balance_before + t_gold_weight)
    to_deposit.total_deposited = round_weight(to_decimal(to_deposit.total_deposited) + t_gold_weight)
    to_deposit.last_transaction_at = china_now()
    
    # 创建转出客户交易记录
    from_transaction = CustomerGoldDepositTransaction(
        customer_id=transfer.from_customer_id,
        customer_name=transfer.from_customer_name,
        transaction_type='use',
        amount=t_gold_weight,
        balance_before=from_balance_before,
        balance_after=from_deposit.current_balance,
        created_by=data.confirmed_by,
        remark=f"转料单：{transfer.transfer_no}，转出至：{transfer.to_customer_name}"
    )
    db.add(from_transaction)
    
    # 创建转入客户交易记录
    to_transaction = CustomerGoldDepositTransaction(
        customer_id=transfer.to_customer_id,
        customer_name=transfer.to_customer_name,
        transaction_type='deposit',
        amount=t_gold_weight,
        balance_before=to_balance_before,
        balance_after=to_deposit.current_balance,
        created_by=data.confirmed_by,
        remark=f"转料单：{transfer.transfer_no}，转入自：{transfer.from_customer_name}"
    )
    db.add(to_transaction)
    
    db.commit()
    db.refresh(transfer)
    
    logger.info(f"确认转料单: {transfer.transfer_no}, "
               f"{transfer.from_customer_name}({from_deposit.current_balance}克) -> "
               f"{transfer.to_customer_name}({to_deposit.current_balance}克), "
               f"克重: {transfer.gold_weight}克")
    
    return success_response(
        message="转料单已确认",
        data={"transfer": CustomerTransferResponse.model_validate(transfer)}
    )


@router.post("/transfers/{transfer_id}/cancel")
async def cancel_customer_transfer(
    transfer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """取消转料单（仅待确认状态可取消）"""
    transfer = db.query(CustomerTransfer).filter(
        CustomerTransfer.id == transfer_id
    ).first()
    
    if not transfer:
        raise HTTPException(status_code=404, detail="转料单不存在")
    
    if transfer.status != "pending":
        raise HTTPException(status_code=400, detail="只能取消待确认的转料单")
    
    transfer.status = "cancelled"
    db.commit()
    db.refresh(transfer)
    
    return success_response(
        message="转料单已取消",
        data={"transfer": CustomerTransferResponse.model_validate(transfer)}
    )


@router.get("/transfers/{transfer_id}/download")
async def download_customer_transfer(
    transfer_id: int,
    format: str = Query("html", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """下载或打印转料单"""
    transfer = db.query(CustomerTransfer).filter(
        CustomerTransfer.id == transfer_id
    ).first()
    
    if not transfer:
        raise HTTPException(status_code=404, detail="转料单不存在")
    
    # 更新打印时间
    transfer.printed_at = china_now()
    db.commit()
    
    # 构建字段
    fields = [
        ("转料单号", transfer.transfer_no, False),
        ("转出客户", transfer.from_customer_name, False),
        ("转入客户", transfer.to_customer_name, False),
        ("转料克重", f"{transfer.gold_weight:.2f} 克", False),
        ("状态", get_status_label(transfer.status), False),
        ("创建时间", format_datetime(transfer.created_at), False),
    ]
    
    if transfer.created_by:
        fields.append(("创建人", transfer.created_by, False))
    if transfer.confirmed_by:
        fields.append(("确认人", transfer.confirmed_by, False))
    if transfer.confirmed_at:
        fields.append(("确认时间", format_datetime(transfer.confirmed_at), False))
    if transfer.remark:
        fields.append(("备注", transfer.remark, True))
    
    if format == "pdf":
        from fastapi.responses import StreamingResponse
        
        generator = PDFGenerator("客户转料单")
        generator.add_title()
        for label, value, _ in fields:
            if value:
                generator.add_field(label, value)
        generator.add_footer()
        buffer = generator.generate()
        
        filename = f"transfer_{transfer.transfer_no}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            }
        )
    
    else:
        from fastapi.responses import HTMLResponse
        
        generator = HTMLGenerator("客户转料单", transfer.transfer_no)
        for label, value, full_width in fields:
            generator.add_field_if(label, value, full_width)
        html_content = generator.generate()
        
        return HTMLResponse(
            content=html_content
        )


# ==================== 收料单管理（从客户收取金料）====================

@router.get("/gold-receipts")
async def get_gold_receipts(
    status: Optional[str] = Query(None, description="状态筛选: pending/received"),
    customer_id: Optional[int] = Query(None, description="客户ID"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """获取收料单列表"""
    query = db.query(GoldReceipt).options(
        joinedload(GoldReceipt.customer),
        joinedload(GoldReceipt.settlement)
    )
    
    if status:
        query = query.filter(GoldReceipt.status == status)
    if customer_id:
        query = query.filter(GoldReceipt.customer_id == customer_id)
    
    total = query.count()
    receipts = query.order_by(desc(GoldReceipt.created_at)).offset(skip).limit(limit).all()
    
    result = []
    for r in receipts:
        result.append({
            "id": r.id,
            "receipt_no": r.receipt_no,
            "settlement_id": r.settlement_id,
            "settlement_no": r.settlement.settlement_no if r.settlement else None,
            "customer_id": r.customer_id,
            "customer_name": r.customer.name if r.customer else None,
            "gold_weight": r.gold_weight,
            "gold_fineness": r.gold_fineness,
            "status": r.status,
            "created_by": r.created_by,
            "received_by": r.received_by,
            "received_at": r.received_at.isoformat() if r.received_at else None,
            "remark": r.remark,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    
    return success_response(data={"items": result, "total": total})


@router.post("/gold-receipts")
async def create_gold_receipt(
    gold_weight: float,
    gold_fineness: str = "足金999",
    customer_id: Optional[int] = None,
    settlement_id: Optional[int] = None,
    remark: str = "",
    created_by: str = "结算专员",
    role: str = Depends(require_permission("can_create_gold_receipt")),
    db: Session = Depends(get_db)
):
    """创建收料单"""
    # 如果指定了 settlement_id，可以从结算单获取客户信息
    if settlement_id:
        settlement = db.query(SettlementOrder).options(
            joinedload(SettlementOrder.sales_order)
        ).filter(SettlementOrder.id == settlement_id).first()
        if not settlement:
            raise HTTPException(status_code=404, detail="结算单不存在")
        # 从结算单关联的销售单获取客户ID
        if settlement.sales_order and not customer_id:
            customer_id = settlement.sales_order.customer_id
    
    if not customer_id:
        raise HTTPException(status_code=400, detail="请提供客户ID或关联结算单")
    
    # 验证客户存在
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="客户不存在")
    
    # 生成收料单号 SL + 时间戳
    now = china_now()
    receipt_no = f"SL{now.strftime('%Y%m%d%H%M%S')}"
    
    # 创建收料单
    receipt = GoldReceipt(
        receipt_no=receipt_no,
        settlement_id=settlement_id,
        customer_id=customer_id,
        gold_weight=gold_weight,
        gold_fineness=gold_fineness,
        status="pending",
        created_by=created_by,
        remark=remark
    )
    
    db.add(receipt)
    db.commit()
    db.refresh(receipt)
    
    logger.info(f"创建收料单: {receipt_no}, 客户={customer.name}, 克重={gold_weight}")
    
    return success_response(
        message="收料单创建成功",
        data={
            "id": receipt.id,
            "receipt_no": receipt.receipt_no,
            "customer_name": customer.name,
            "gold_weight": receipt.gold_weight,
            "gold_fineness": receipt.gold_fineness,
            "status": receipt.status
        },
        code=ErrorCode.CREATED
    )


@router.post("/gold-receipts/{receipt_id}/receive")
async def confirm_gold_receipt(
    receipt_id: int,
    received_by: str = "结算专员",
    role: str = Depends(require_permission("can_confirm_gold_receive")),
    db: Session = Depends(get_db)
):
    """
    结算确认收料单
    
    单一账户模式：直接增加 current_balance
    - 如果之前是负值（欠料），会自动"抵扣"
    - 如果之前是正值（存料），会增加存料
    """
    # 加行锁防止并发确认
    receipt = db.query(GoldReceipt).filter(GoldReceipt.id == receipt_id).with_for_update().first()
    
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    
    if receipt.status == "received":
        raise HTTPException(status_code=409, detail="该收料单已接收（可能已被其他人确认）")
    
    customer_id = receipt.customer_id
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    customer_name = customer.name if customer else "未知客户"
    gold_weight = to_decimal(receipt.gold_weight)
    
    # ========== 更新收料单状态 ==========
    receipt.status = "received"
    receipt.received_by = received_by
    receipt.received_at = china_now()
    
    # ========== 锁定客户存料记录后增加余额 ==========
    deposit = db.query(CustomerGoldDeposit).filter(
        CustomerGoldDeposit.customer_id == customer_id
    ).with_for_update().first()
    if not deposit:
        deposit = get_or_create_customer_deposit(customer_id, customer_name, db)
    balance_before = to_decimal(deposit.current_balance)
    deposit.current_balance = round_weight(balance_before + gold_weight)
    deposit.total_deposited = round_weight(to_decimal(deposit.total_deposited) + gold_weight)
    deposit.last_transaction_at = china_now()
    
    # 计算状态变化说明
    if balance_before < 0:
        # 之前是欠料状态
        if deposit.current_balance >= 0:
            remark_text = f"收料单：{receipt.receipt_no}，来料{gold_weight:.2f}克，欠料{abs(balance_before):.2f}克已清，剩余存料{deposit.current_balance:.2f}克"
        else:
            remark_text = f"收料单：{receipt.receipt_no}，来料{gold_weight:.2f}克，剩余欠料{abs(deposit.current_balance):.2f}克"
    else:
        # 之前是存料或零
        remark_text = f"收料单：{receipt.receipt_no}，来料{gold_weight:.2f}克，当前存料{deposit.current_balance:.2f}克"
    
    # 创建金料账户变动记录
    deposit_tx = CustomerGoldDepositTransaction(
        customer_id=customer_id,
        customer_name=customer_name,
        transaction_type='deposit',
        amount=gold_weight,
        balance_before=balance_before,
        balance_after=deposit.current_balance,
        created_by=received_by,
        remark=remark_text
    )
    db.add(deposit_tx)
    
    # 创建客户往来账记录
    customer_tx = CustomerTransaction(
        customer_id=customer_id,
        customer_name=customer_name,
        transaction_type='gold_receipt',
        gold_weight=gold_weight,
        gold_due_before=max(0, -balance_before),  # 之前的欠料（负值的绝对值）
        gold_due_after=max(0, -deposit.current_balance),  # 之后的欠料
        status='active',
        remark=remark_text
    )
    db.add(customer_tx)
    
    db.commit()
    db.refresh(receipt)
    
    # 记录余额变化
    if deposit.current_balance >= 0:
        status_text = f"当前存料 {deposit.current_balance:.2f}克"
    else:
        status_text = f"当前欠料 {abs(deposit.current_balance):.2f}克"
    logger.info(f"收料单已接收: {receipt.receipt_no}, 接收人={received_by}, "
                f"来料={gold_weight}克, 变动前={balance_before:.2f}克, 变动后={deposit.current_balance:.2f}克 ({status_text})")
    
    # 构建返回消息
    if deposit.current_balance >= 0:
        message = f"收料确认成功！来料 {gold_weight:.2f}克，当前存料 {deposit.current_balance:.2f}克"
    else:
        message = f"收料确认成功！来料 {gold_weight:.2f}克，剩余欠料 {abs(deposit.current_balance):.2f}克"
    
    return success_response(
        message=message,
        data={
            "id": receipt.id,
            "receipt_no": receipt.receipt_no,
            "status": receipt.status,
            "received_by": receipt.received_by,
            "received_at": receipt.received_at.isoformat() if receipt.received_at else None,
            "gold_weight": gold_weight,
            "balance_before": round(balance_before, 2),
            "balance_after": round(deposit.current_balance, 2),
            "net_gold": round(deposit.current_balance, 2)
        }
    )


@router.put("/gold-receipts/{receipt_id}")
async def update_gold_receipt(
    receipt_id: int,
    gold_weight: Optional[float] = None,
    gold_fineness: Optional[str] = None,
    customer_id: Optional[int] = None,
    remark: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """编辑收料单（仅pending状态可编辑）"""
    receipt = db.query(GoldReceipt).filter(GoldReceipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    if receipt.status != "pending":
        raise HTTPException(status_code=400, detail="只能编辑待确认的收料单")
    
    if gold_weight is not None:
        if gold_weight <= 0:
            raise HTTPException(status_code=400, detail="克重必须大于0")
        receipt.gold_weight = gold_weight
    if gold_fineness is not None:
        receipt.gold_fineness = gold_fineness
    if customer_id is not None:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        receipt.customer_id = customer_id
    if remark is not None:
        receipt.remark = remark
    
    db.commit()
    db.refresh(receipt)
    
    return success_response(
        data={
            "id": receipt.id,
            "receipt_no": receipt.receipt_no,
            "gold_weight": receipt.gold_weight,
            "gold_fineness": receipt.gold_fineness,
            "customer_id": receipt.customer_id,
            "remark": receipt.remark,
            "status": receipt.status
        },
        message="更新成功"
    )


@router.delete("/gold-receipts/{receipt_id}")
async def delete_gold_receipt(
    receipt_id: int,
    db: Session = Depends(get_db)
):
    """删除收料单（仅pending状态可删除）"""
    receipt = db.query(GoldReceipt).filter(GoldReceipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    if receipt.status != "pending":
        raise HTTPException(status_code=400, detail="只能删除待确认的收料单")
    
    receipt_no = receipt.receipt_no
    db.delete(receipt)
    db.commit()
    
    logger.info(f"删除收料单: {receipt_no}")
    return success_response(message=f"收料单 {receipt_no} 已删除")


@router.post("/gold-receipts/{receipt_id}/unconfirm")
async def unconfirm_gold_receipt(
    receipt_id: int,
    operated_by: str = "结算专员",
    reason: str = "",
    db: Session = Depends(get_db)
):
    """反确认收料单（回滚客户金料余额）"""
    receipt = db.query(GoldReceipt).filter(GoldReceipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    if receipt.status != "received":
        raise HTTPException(status_code=400, detail=f"当前状态为 {receipt.status}，无法反确认")
    
    try:
        customer_id = receipt.customer_id
        gold_weight = to_decimal(receipt.gold_weight)
        now = china_now()
        
        # 1. 回滚客户金料余额
        deposit = db.query(CustomerGoldDeposit).filter(
            CustomerGoldDeposit.customer_id == customer_id
        ).first()
        
        if deposit:
            balance_before = to_decimal(deposit.current_balance)
            deposit.current_balance = round_weight(balance_before - gold_weight)
            deposit.total_deposited = round_weight(max(Decimal("0"), to_decimal(deposit.total_deposited) - gold_weight))
            deposit.last_transaction_at = now
            
            # 创建反向变动记录
            rollback_tx = CustomerGoldDepositTransaction(
                customer_id=customer_id,
                customer_name=deposit.customer_name,
                transaction_type='withdrawal',
                amount=gold_weight,
                balance_before=balance_before,
                balance_after=deposit.current_balance,
                created_by=operated_by,
                remark=f"反确认收料单 {receipt.receipt_no}，原因: {reason}" if reason else f"反确认收料单 {receipt.receipt_no}"
            )
            db.add(rollback_tx)
        
        # 2. 恢复收料单状态
        old_status = receipt.status
        receipt.status = "pending"
        receipt.received_by = None
        receipt.received_at = None
        
        # 3. 记录操作日志
        log = OrderStatusLog(
            order_type="gold_receipt",
            order_id=receipt.id,
            action="unconfirm",
            old_status=old_status,
            new_status="pending",
            operated_by=operated_by,
            operated_at=now,
            remark=f"反确认原因: {reason}" if reason else "反确认收料单"
        )
        db.add(log)
        
        db.commit()
        
        logger.info(f"反确认收料单: {receipt.receipt_no}, 操作人: {operated_by}")
        
        return success_response(
            data={"id": receipt.id, "receipt_no": receipt.receipt_no, "status": "pending"},
            message=f"收料单 {receipt.receipt_no} 已反确认"
        )
    except Exception as e:
        db.rollback()
        logger.error(f"反确认收料单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"反确认失败: {str(e)}")


@router.get("/gold-receipts/{receipt_id}")
async def get_gold_receipt_detail(
    receipt_id: int,
    db: Session = Depends(get_db)
):
    """获取收料单详情"""
    receipt = db.query(GoldReceipt).options(
        joinedload(GoldReceipt.customer),
        joinedload(GoldReceipt.settlement)
    ).filter(GoldReceipt.id == receipt_id).first()
    
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    
    return success_response(
        data={
            "id": receipt.id,
            "receipt_no": receipt.receipt_no,
            "settlement_id": receipt.settlement_id,
            "settlement_no": receipt.settlement.settlement_no if receipt.settlement else None,
            "customer_id": receipt.customer_id,
            "customer_name": receipt.customer.name if receipt.customer else None,
            "customer_phone": receipt.customer.phone if receipt.customer else None,
            "gold_weight": receipt.gold_weight,
            "gold_fineness": receipt.gold_fineness,
            "status": receipt.status,
            "created_by": receipt.created_by,
            "received_by": receipt.received_by,
            "received_at": receipt.received_at.isoformat() if receipt.received_at else None,
            "remark": receipt.remark,
            "created_at": receipt.created_at.isoformat() if receipt.created_at else None,
        }
    )


@router.get("/gold-receipts/{receipt_id}/print")
async def print_gold_receipt(
    receipt_id: int,
    format: str = Query("html", pattern="^(pdf|html)$"),
    db: Session = Depends(get_db)
):
    """打印收料单（参考公司客户来料单 XLS 模板格式）"""
    from fastapi.responses import HTMLResponse
    
    receipt = db.query(GoldReceipt).options(
        joinedload(GoldReceipt.customer),
        joinedload(GoldReceipt.settlement)
    ).filter(GoldReceipt.id == receipt_id).first()
    
    if not receipt:
        raise HTTPException(status_code=404, detail="收料单不存在")
    
    customer_name = receipt.customer.name if receipt.customer else "-"
    created_date = receipt.created_at.strftime("%Y/%m/%d") if receipt.created_at else "-"
    now = china_now()
    print_date = now.strftime("%Y/%m/%d %H:%M")
    remark_text = receipt.remark or ""
    fineness_display = receipt.gold_fineness or "足金"
    fineness_short = fineness_display.replace("9999", "").replace("999", "") if "足金" in fineness_display else fineness_display
    
    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>客户来料单 - {receipt.receipt_no}</title>
    <style>
        @page {{
            size: 241mm 140mm;
            margin: 5mm 15mm;
        }}
        @media print {{
            body {{ margin: 0; padding: 3mm 0; }}
            html {{ overflow: hidden; }}
        }}
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: "SimSun", "宋体", serif;
            font-size: 10pt;
            line-height: 1.5;
            color: #000;
            padding: 3mm 0;
            max-width: 211mm;
            max-height: 130mm;
        }}
        .company-name {{
            text-align: center;
            font-family: "SimSun", "宋体", serif;
            font-size: 18pt;
            margin-bottom: 1px;
        }}
        .doc-title {{
            text-align: center;
            font-family: "SimSun", "宋体", serif;
            font-size: 16pt;
            margin-bottom: 4px;
        }}
        .info-row {{
            display: flex;
            font-family: "SimSun", "宋体", serif;
            font-size: 10pt;
            line-height: 1.6;
        }}
        .info-row .left {{
            flex: 1;
        }}
        .info-row .right {{
            text-align: left;
        }}
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 4px;
            margin-bottom: 8px;
        }}
        .data-table th {{
            font-family: Arial, sans-serif;
            font-size: 9pt;
            font-weight: normal;
            text-align: center;
            vertical-align: middle;
            border: 1px solid #000;
            padding: 4px 6px;
        }}
        .data-table td {{
            font-family: Arial, sans-serif;
            font-size: 8pt;
            vertical-align: middle;
            border: 1px solid #000;
            padding: 4px 6px;
        }}
        .data-table td.num {{
            text-align: right;
        }}
        .data-table td.center {{
            text-align: center;
        }}
        .data-table td.left {{
            text-align: left;
        }}
        .signature-area {{
            display: flex;
            justify-content: flex-start;
            gap: 120px;
            margin-top: 12px;
            font-family: Arial, sans-serif;
            font-size: 9pt;
        }}
        .signature-box {{
            display: flex;
            align-items: center;
        }}
        .signature-line {{
            display: inline-block;
            width: 120px;
            border-bottom: 1px solid #000;
            margin-left: 4px;
        }}
    </style>
</head>
<body>
    <div class="company-name">深圳市梵贝琳珠宝有限公司</div>
    <div class="doc-title">客户来料单</div>

    <div class="info-row">
        <div class="left">来料单号：{receipt.receipt_no}</div>
        <div class="right" style="margin-right:40px;">来料日期：{created_date}</div>
        <div class="right">打 印 人：{receipt.created_by or "-"}</div>
    </div>
    <div class="info-row">
        <div class="left">来料客户：{customer_name}</div>
    </div>
    <div class="info-row">
        <div class="left">来料备注：{remark_text}</div>
        <div class="right">打印日期：{print_date}</div>
    </div>

    <table class="data-table">
        <thead>
            <tr>
                <th style="width:40px">序号</th>
                <th style="width:80px">来料名称</th>
                <th style="width:80px">来料重量</th>
                <th style="width:80px">实测成色</th>
                <th style="width:70px">折算率</th>
                <th style="width:70px">结算成色</th>
                <th style="width:80px">结算重量</th>
                <th>小备注</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td class="center">1</td>
                <td class="left">{fineness_display}</td>
                <td class="num">{receipt.gold_weight:.3f}</td>
                <td class="center">100.00%</td>
                <td class="center">100.00%</td>
                <td class="num">{fineness_short}</td>
                <td class="num">{receipt.gold_weight:.3f}</td>
                <td class="left">{remark_text}</td>
            </tr>
            <tr>
                <td></td>
                <td></td>
                <td class="num"><b>{receipt.gold_weight:.3f}</b></td>
                <td></td>
                <td></td>
                <td></td>
                <td class="num"><b>{receipt.gold_weight:.3f}</b></td>
                <td></td>
            </tr>
        </tbody>
    </table>

    <div class="signature-area">
        <div class="signature-box">
            <span>来料人：</span>
            <span class="signature-line"></span>
        </div>
        <div class="signature-box">
            <span>收料人：</span>
            <span class="signature-line"></span>
        </div>
    </div>

    <script>
        window.onload = function() {{
            window.print();
        }}
    </script>
</body>
</html>
"""
    
    return HTMLResponse(
        content=html_content
    )


# ==================== 金料每日统计 ====================

@router.get("/daily-summary")
async def get_gold_daily_summary(
    period: str = Query("today", pattern="^(today|week|month)$", description="统计周期：today/week/month"),
    db: Session = Depends(get_db)
):
    """
    获取金料每日统计汇总
    - period: today=今日, week=本周, month=本月
    返回：收料汇总、付料汇总、明细列表
    """
    from datetime import timedelta
    from sqlalchemy import func, cast, Date
    
    now = china_now()
    today = now.date()
    
    # 根据 period 计算日期范围
    if period == "today":
        start_date = today
        end_date = today
    elif period == "week":
        # 本周一到今天
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    else:  # month
        # 本月1号到今天
        start_date = today.replace(day=1)
        end_date = today
    
    # ========== 收料统计（使用新系统 GoldReceipt）==========
    from ..models.finance import GoldReceipt
    
    # 收料总计
    receipt_stats = db.query(
        func.count(GoldReceipt.id).label("count"),
        func.coalesce(func.sum(GoldReceipt.gold_weight), 0).label("total_weight")
    ).filter(
        func.date(GoldReceipt.created_at) >= start_date,
        func.date(GoldReceipt.created_at) <= end_date
    ).first()
    
    # 收料明细（按客户分组）
    receipt_by_customer = db.query(
        GoldReceipt.customer_id,
        Customer.name.label("customer_name"),
        func.count(GoldReceipt.id).label("count"),
        func.sum(GoldReceipt.gold_weight).label("total_weight")
    ).join(
        Customer, GoldReceipt.customer_id == Customer.id
    ).filter(
        func.date(GoldReceipt.created_at) >= start_date,
        func.date(GoldReceipt.created_at) <= end_date
    ).group_by(
        GoldReceipt.customer_id, Customer.name
    ).all()
    
    # 收料每笔明细（新增）
    receipt_details = db.query(GoldReceipt).join(
        Customer, GoldReceipt.customer_id == Customer.id
    ).filter(
        func.date(GoldReceipt.created_at) >= start_date,
        func.date(GoldReceipt.created_at) <= end_date
    ).order_by(GoldReceipt.created_at.desc()).all()
    
    # ========== 付料统计（使用 GoldMaterialTransaction）==========
    payment_stats = db.query(
        func.count(GoldMaterialTransaction.id).label("count"),
        func.coalesce(func.sum(GoldMaterialTransaction.gold_weight), 0).label("total_weight")
    ).filter(
        GoldMaterialTransaction.transaction_type == "expense",
        func.date(GoldMaterialTransaction.created_at) >= start_date,
        func.date(GoldMaterialTransaction.created_at) <= end_date
    ).first()
    
    # 付料明细（按供应商分组）
    payment_by_supplier = db.query(
        GoldMaterialTransaction.supplier_id,
        Supplier.name.label("supplier_name"),
        func.count(GoldMaterialTransaction.id).label("count"),
        func.sum(GoldMaterialTransaction.gold_weight).label("total_weight")
    ).outerjoin(
        Supplier, GoldMaterialTransaction.supplier_id == Supplier.id
    ).filter(
        GoldMaterialTransaction.transaction_type == "expense",
        func.date(GoldMaterialTransaction.created_at) >= start_date,
        func.date(GoldMaterialTransaction.created_at) <= end_date
    ).group_by(
        GoldMaterialTransaction.supplier_id, Supplier.name
    ).all()
    
    # ========== 客户提料统计 ==========
    # 提料总计
    withdrawal_stats = db.query(
        func.count(CustomerWithdrawal.id).label("count"),
        func.coalesce(func.sum(CustomerWithdrawal.gold_weight), 0).label("total_weight")
    ).filter(
        func.date(CustomerWithdrawal.created_at) >= start_date,
        func.date(CustomerWithdrawal.created_at) <= end_date
    ).first()
    
    # 提料明细（按客户分组）
    withdrawal_by_customer = db.query(
        CustomerWithdrawal.customer_id,
        Customer.name.label("customer_name"),
        func.count(CustomerWithdrawal.id).label("count"),
        func.sum(CustomerWithdrawal.gold_weight).label("total_weight")
    ).join(
        Customer, CustomerWithdrawal.customer_id == Customer.id
    ).filter(
        func.date(CustomerWithdrawal.created_at) >= start_date,
        func.date(CustomerWithdrawal.created_at) <= end_date
    ).group_by(
        CustomerWithdrawal.customer_id, Customer.name
    ).all()
    
    # 提料每笔明细
    withdrawal_details = db.query(CustomerWithdrawal).filter(
        func.date(CustomerWithdrawal.created_at) >= start_date,
        func.date(CustomerWithdrawal.created_at) <= end_date
    ).order_by(CustomerWithdrawal.created_at.desc()).all()
    
    # 构建响应
    return success_response(
        data={
            "period": period,
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "receipt_summary": {
                "total_count": receipt_stats.count or 0,
                "total_weight": safe_float_for_json(receipt_stats.total_weight),
                "by_customer": [
                    {
                        "customer_id": r.customer_id,
                        "customer_name": r.customer_name,
                        "count": r.count,
                        "total_weight": safe_float_for_json(r.total_weight)
                    }
                    for r in receipt_by_customer
                ],
                "receipts": [
                    {
                        "id": r.id,
                        "receipt_no": r.receipt_no,
                        "customer_id": r.customer_id,
                        "customer_name": db.query(Customer.name).filter(Customer.id == r.customer_id).scalar() or "未知",
                        "gold_weight": safe_float_for_json(r.gold_weight),
                        "gold_fineness": r.gold_fineness,
                        "status": r.status,
                        "created_at": r.created_at.isoformat() if r.created_at else None,
                        "remark": r.remark
                    }
                    for r in receipt_details
                ]
            },
            "payment_summary": {
                "total_count": payment_stats.count or 0,
                "total_weight": safe_float_for_json(payment_stats.total_weight),
                "by_supplier": [
                    {
                        "supplier_id": p.supplier_id,
                        "supplier_name": p.supplier_name or "未知供应商",
                        "count": p.count,
                        "total_weight": safe_float_for_json(p.total_weight)
                    }
                    for p in payment_by_supplier
                ]
            },
            "withdrawal_summary": {
                "total_count": withdrawal_stats.count or 0,
                "total_weight": safe_float_for_json(withdrawal_stats.total_weight),
                "by_customer": [
                    {
                        "customer_id": w.customer_id,
                        "customer_name": w.customer_name,
                        "count": w.count,
                        "total_weight": safe_float_for_json(w.total_weight)
                    }
                    for w in withdrawal_by_customer
                ],
                "withdrawals": [
                    {
                        "id": w.id,
                        "withdrawal_no": w.withdrawal_no,
                        "customer_id": w.customer_id,
                        "customer_name": w.customer_name,
                        "gold_weight": safe_float_for_json(w.gold_weight),
                        "status": w.status,
                        "created_at": w.created_at.isoformat() if w.created_at else None,
                        "remark": w.remark
                    }
                    for w in withdrawal_details
                ]
            }
        }
    )


# ==================== 数据修复：同步 GoldReceipt 到往来账 ====================

@router.post("/fix-receipt-transactions")
async def fix_receipt_transactions(
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    修复已确认的 GoldReceipt 数据：为缺失往来账记录的收料单补充记录
    
    - 查询所有 status='received' 但没有对应 CustomerTransaction 的 GoldReceipt
    - 为每条补充 CustomerTransaction 和 CustomerGoldDeposit 记录
    """
    if user_role != 'manager':
        raise HTTPException(status_code=403, detail="只有管理员可以执行数据修复")
    
    try:
        # 查询所有已接收的 GoldReceipt
        received_receipts = db.query(GoldReceipt).filter(
            GoldReceipt.status == 'received'
        ).all()
        
        fixed_count = 0
        skipped_count = 0
        results = []
        
        for receipt in received_receipts:
            # 检查是否已有对应的 CustomerTransaction
            existing_tx = db.query(CustomerTransaction).filter(
                CustomerTransaction.remark.contains(receipt.receipt_no)
            ).first()
            
            if existing_tx:
                skipped_count += 1
                continue
            
            customer_id = receipt.customer_id
            customer = db.query(Customer).filter(Customer.id == customer_id).first()
            customer_name = customer.name if customer else "未知客户"
            gold_weight = to_decimal(receipt.gold_weight)
            
            # 查询客户当时的欠料情况（简化处理：全部存入存料）
            # 因为历史欠料状态难以准确还原，这里将所有来料作为存料处理
            deposit_amount = gold_weight
            
            # 创建往来账记录
            customer_tx = CustomerTransaction(
                customer_id=customer_id,
                customer_name=customer_name,
                transaction_type='gold_receipt',
                gold_weight=gold_weight,
                gold_due_before=0,
                gold_due_after=0,
                status='active',
                remark=f"收料单：{receipt.receipt_no}（数据修复），存入存料：{deposit_amount:.2f}克",
                created_at=receipt.received_at or receipt.created_at
            )
            db.add(customer_tx)
            
            # 更新存料余额
            deposit = get_or_create_customer_deposit(customer_id, customer_name, db)
            balance_before = to_decimal(deposit.current_balance)
            deposit.current_balance = round_weight(balance_before + deposit_amount)
            deposit.total_deposited = round_weight(to_decimal(deposit.total_deposited) + deposit_amount)
            
            # 创建存料交易记录
            deposit_tx = CustomerGoldDepositTransaction(
                customer_id=customer_id,
                customer_name=customer_name,
                transaction_type='deposit',
                amount=deposit_amount,
                balance_before=balance_before,
                balance_after=deposit.current_balance,
                created_by="系统修复",
                remark=f"收料单：{receipt.receipt_no}（数据修复）",
                created_at=receipt.received_at or receipt.created_at
            )
            db.add(deposit_tx)
            
            fixed_count += 1
            results.append({
                "receipt_no": receipt.receipt_no,
                "customer_name": customer_name,
                "gold_weight": gold_weight,
                "action": "存入存料"
            })
            
            logger.info(f"数据修复: 收料单 {receipt.receipt_no}, 客户={customer_name}, "
                       f"金料={gold_weight}克 存入存料")
        
        db.commit()
        
        return success_response(
            message=f"数据修复完成：修复 {fixed_count} 条，跳过 {skipped_count} 条（已有记录）",
            data={
                "fixed_count": fixed_count,
                "skipped_count": skipped_count,
                "details": results
            }
        )
        
    except Exception as e:
        db.rollback()
        logger.error(f"数据修复失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"数据修复失败: {str(e)}")


# ==================== 供应商金料账户管理 ====================

@router.get("/supplier-gold-accounts")
async def get_supplier_gold_accounts(
    supplier_id: Optional[int] = None,
    supplier_name: Optional[str] = None,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    查询供应商金料账户列表（仅料部和管理层可查看）
    
    返回每个供应商的金料账户余额：
    - 正数 = 我们欠供应商的料（供应商发货了，我们还没付料）
    - 负数 = 供应商欠我们的料（我们付料了，供应商还没发货）
    """
    # 权限检查：仅料部和管理层可查看
    if not has_permission(user_role, 'can_view_supplier_gold_account'):
        raise HTTPException(
            status_code=403, 
            detail="权限不足：您没有查看供应商金料账户的权限。请联系料部或管理层。"
        )
    
    from ..models import SupplierGoldAccount
    
    query = db.query(SupplierGoldAccount)
    
    if supplier_id:
        query = query.filter(SupplierGoldAccount.supplier_id == supplier_id)
    
    if supplier_name:
        query = query.filter(SupplierGoldAccount.supplier_name.ilike(f"%{supplier_name}%"))
    
    accounts = query.order_by(desc(SupplierGoldAccount.last_transaction_at)).all()
    
    result = []
    total_balance = 0.0
    
    for account in accounts:
        balance = account.current_balance or 0
        total_balance += balance
        
        # 确定状态描述
        if balance > 0:
            status_text = f"我们欠供应商 {balance:.2f}克"
        elif balance < 0:
            status_text = f"供应商欠我们 {abs(balance):.2f}克"
        else:
            status_text = "已结清"
        
        result.append({
            "id": account.id,
            "supplier_id": account.supplier_id,
            "supplier_name": account.supplier_name,
            "current_balance": balance,
            "total_received": account.total_received or 0,
            "total_paid": account.total_paid or 0,
            "status_text": status_text,
            "last_transaction_at": account.last_transaction_at.isoformat() if account.last_transaction_at else None,
            "created_at": account.created_at.isoformat() if account.created_at else None
        })
    
    return success_response(
        data={
            "items": result,
            "total": len(result),
            "total_balance": total_balance,
            "summary": {
                "total_we_owe": sum(a["current_balance"] for a in result if a["current_balance"] > 0),
                "total_they_owe": abs(sum(a["current_balance"] for a in result if a["current_balance"] < 0))
            }
        }
    )


@router.get("/supplier-gold-accounts/{supplier_id}")
async def get_supplier_gold_account_detail(
    supplier_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    查询供应商统一往来账明细（金料+工费双 running balance）
    
    数据来源合并为 unified_transactions：
    - 入库记录 → 金料变动(+)、工费变动(+)
    - 付料记录 → 金料变动(-)
    - 退货记录 → 金料变动(-)
    - 工费付款 → 工费变动(-)
    """
    if not has_permission(user_role, 'can_view_supplier_gold_account'):
        raise HTTPException(
            status_code=403, 
            detail="权限不足：您没有查看供应商金料账户的权限。请联系料部或管理层。"
        )
    
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="供应商不存在")
    
    from ..models.finance import AccountPayable, SupplierPayment
    from ..models import SupplierGoldTransaction
    
    unified = []
    total_inbound = 0.0
    total_paid = 0.0
    total_return = 0.0
    
    # ===== 1. 入库记录（按入库单聚合） =====
    inbound_records = db.query(
        InboundDetail.id,
        InboundDetail.weight,
        InboundDetail.product_name,
        InboundOrder.order_no,
        InboundOrder.id.label('order_id'),
        InboundOrder.create_time,
        InboundOrder.operator
    ).join(
        InboundOrder, InboundDetail.order_id == InboundOrder.id
    ).filter(
        InboundDetail.supplier_id == supplier_id
    ).order_by(InboundOrder.create_time).all()
    
    # 工费总额从 InboundDetail.total_cost 汇总
    labor_cost_total = db.query(
        func.sum(InboundDetail.total_cost)
    ).filter(
        InboundDetail.supplier_id == supplier_id
    ).scalar() or 0.0
    labor_cost_total = to_decimal(labor_cost_total)

    # 工费已付从 SupplierPayment 汇总（仅已确认的）
    labor_paid_total = db.query(
        func.sum(SupplierPayment.amount)
    ).filter(
        SupplierPayment.supplier_id == supplier_id,
        SupplierPayment.status == 'confirmed'
    ).scalar() or 0.0
    labor_paid_total = to_decimal(labor_paid_total)

    labor_debt_total = labor_cost_total - labor_paid_total

    # AccountPayable 按入库单分组（用于明细展示）
    payables = db.query(AccountPayable).filter(
        AccountPayable.supplier_id == supplier_id
    ).all()
    payable_map = {}
    for p in payables:
        if p.inbound_order_id:
            payable_map.setdefault(p.inbound_order_id, 0)
            payable_map[p.inbound_order_id] += safe_float_for_json(p.total_amount)
    
    # 按入库单号聚合明细
    order_groups = {}
    for record in inbound_records:
        oid = record.order_id
        if oid not in order_groups:
            order_groups[oid] = {
                "order_no": record.order_no,
                "create_time": record.create_time,
                "total_weight": 0,
                "products": []
            }
        weight = safe_float_for_json(record.weight)
        order_groups[oid]["total_weight"] += weight
        total_inbound += weight
        if record.product_name:
            order_groups[oid]["products"].append(record.product_name)
    
    for oid, data in order_groups.items():
        labor_cost = payable_map.get(oid)
        products_str = "、".join(data["products"]) if data["products"] else ""
        remark = products_str if products_str else f"入库单 {data['order_no']}"
        unified.append({
            "date": data["create_time"].isoformat() if data["create_time"] else None,
            "type": "inbound",
            "type_text": "供应商发货",
            "order_no": data["order_no"],
            "gold_change": round(data["total_weight"], 2),
            "gold_balance": 0,
            "labor_change": round(labor_cost, 2) if labor_cost else None,
            "labor_balance": 0,
            "remark": remark
        })
    
    # ===== 2. 付料记录 =====
    pay_records = db.query(GoldMaterialTransaction).filter(
        GoldMaterialTransaction.supplier_id == supplier_id,
        GoldMaterialTransaction.transaction_type == 'expense',
        GoldMaterialTransaction.status == 'confirmed'
    ).all()
    
    for record in pay_records:
        weight = safe_float_for_json(record.gold_weight)
        total_paid += weight
        unified.append({
            "date": record.created_at.isoformat() if record.created_at else None,
            "type": "pay",
            "type_text": "我们付料",
            "order_no": record.transaction_no,
            "gold_change": round(-weight, 2),
            "gold_balance": 0,
            "labor_change": None,
            "labor_balance": 0,
            "remark": record.remark or f"付料单 {record.transaction_no}"
        })
    
    # ===== 3. 退货记录 =====
    return_records = db.query(SupplierGoldTransaction).filter(
        SupplierGoldTransaction.supplier_id == supplier_id,
        SupplierGoldTransaction.transaction_type == 'return',
        SupplierGoldTransaction.status == 'active'
    ).all()
    
    for record in return_records:
        weight = safe_float_for_json(record.gold_weight)
        total_return += weight
        unified.append({
            "date": record.created_at.isoformat() if record.created_at else None,
            "type": "return",
            "type_text": "退货给供应商",
            "order_no": None,
            "gold_change": round(-weight, 2),
            "gold_balance": 0,
            "labor_change": None,
            "labor_balance": 0,
            "remark": record.remark or "退货"
        })
    
    # ===== 4. 工费付款（已确认的） =====
    sp_records = db.query(SupplierPayment).filter(
        SupplierPayment.supplier_id == supplier_id,
        SupplierPayment.status == "confirmed"
    ).all()
    
    for sp in sp_records:
        pay_date = sp.confirmed_at or sp.create_time
        unified.append({
            "date": pay_date.isoformat() if pay_date else None,
            "type": "labor_payment",
            "type_text": "工费付款",
            "order_no": sp.payment_no,
            "gold_change": None,
            "gold_balance": 0,
            "labor_change": round(-safe_float_for_json(sp.amount), 2),
            "labor_balance": 0,
            "remark": sp.remark or f"工费付款 {sp.payment_no}"
        })
    
    # ===== 插入期初余额行 =====
    init_gold = to_decimal(supplier.initial_gold_debt)
    init_labor = to_decimal(supplier.initial_labor_debt)
    if init_gold != 0 or init_labor != 0:
        unified.insert(0, {
            "date": "2026-02-25T00:00:00",
            "type": "initial_balance",
            "type_text": "期初余额",
            "order_no": "-",
            "gold_change": round(init_gold, 2) if init_gold != 0 else None,
            "gold_balance": 0,
            "labor_change": round(init_labor, 2) if init_labor != 0 else None,
            "labor_balance": 0,
            "remark": "旧系统迁移期初数据"
        })

    # ===== 按时间升序排列，计算双 running balance =====
    unified.sort(key=lambda x: x['date'] or '')
    
    gold_bal = 0.0
    labor_bal = 0.0
    for tx in unified:
        if tx['gold_change'] is not None:
            gold_bal += tx['gold_change']
        tx['gold_balance'] = round(gold_bal, 2)
        if tx['labor_change'] is not None:
            labor_bal += tx['labor_change']
        tx['labor_balance'] = round(labor_bal, 2)
    
    # 反转为降序（最新在前）
    unified.reverse()
    
    # ===== 生成状态文本（含期初） =====
    current_balance = total_inbound - total_paid - total_return + init_gold
    labor_debt_total = labor_debt_total + init_labor
    
    if current_balance > 0:
        status_text = f"我们欠供应商 {current_balance:.2f}克"
    elif current_balance < 0:
        status_text = f"供应商欠我们 {abs(current_balance):.2f}克"
    else:
        status_text = "已结清"
    
    if not unified:
        status_text = "无往来记录"
    
    return success_response(
        data={
            "account": {
                "supplier_id": supplier_id,
                "supplier_name": supplier.name,
                "current_balance": round(current_balance, 2),
                "total_received": round(total_inbound, 2),
                "total_paid": round(total_paid, 2),
                "total_return": round(total_return, 2),
                "labor_debt_total": round(labor_debt_total, 2),
                "labor_paid_total": round(labor_paid_total, 2),
                "initial_gold_debt": round(init_gold, 2),
                "initial_labor_debt": round(init_labor, 2),
                "status_text": status_text
            },
            "unified_transactions": unified
        }
    )


# ==================== 客料回仓单 ====================

@router.post("/customer-gold-transfers")
async def create_customer_gold_transfer(
    gold_weight: float = Query(..., description="转交克重"),
    gold_fineness: str = Query("足金999", description="成色"),
    remark: str = Query("", description="备注"),
    created_by: str = Query("结算", description="开单人"),
    db: Session = Depends(get_db)
):
    """创建客料回仓单（结算将收到的客户金料转交给料部）"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        if gold_weight <= 0:
            raise HTTPException(status_code=400, detail="转交克重必须大于0")
        
        now = china_now()
        count = db.query(CustomerGoldTransfer).filter(
            CustomerGoldTransfer.create_time >= now.replace(hour=0, minute=0, second=0)
        ).count()
        transfer_no = f"KHH{now.strftime('%Y%m%d')}{count + 1:03d}"
        
        transfer = CustomerGoldTransfer(
            transfer_no=transfer_no,
            gold_weight=gold_weight,
            gold_fineness=gold_fineness,
            status="pending",
            created_by=created_by,
            remark=remark
        )
        db.add(transfer)
        db.commit()
        db.refresh(transfer)
        
        logger.info(f"创建客料回仓单: {transfer_no}, 克重: {gold_weight}g")
        
        return success_response(
            data={
                "id": transfer.id,
                "transfer_no": transfer.transfer_no,
                "gold_weight": transfer.gold_weight,
                "gold_fineness": transfer.gold_fineness,
                "status": transfer.status,
                "created_by": transfer.created_by,
                "create_time": transfer.create_time.isoformat() if transfer.create_time else None,
                "remark": transfer.remark
            },
            message=f"客料回仓单已创建，单号：{transfer_no}"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


@router.get("/customer-gold-transfers")
async def get_customer_gold_transfers(
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """查询客料回仓单列表"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        query = db.query(CustomerGoldTransfer)
        if status:
            query = query.filter(CustomerGoldTransfer.status == status)
        
        total = query.count()
        transfers = query.order_by(desc(CustomerGoldTransfer.create_time)).offset(skip).limit(limit).all()
        
        result = []
        for t in transfers:
            result.append({
                "id": t.id,
                "transfer_no": t.transfer_no,
                "gold_weight": t.gold_weight,
                "gold_fineness": t.gold_fineness,
                "status": t.status,
                "created_by": t.created_by,
                "create_time": t.create_time.isoformat() if t.create_time else None,
                "confirmed_by": t.confirmed_by,
                "confirmed_at": t.confirmed_at.isoformat() if t.confirmed_at else None,
                "remark": t.remark
            })
        
        return success_response(data={"items": result, "total": total})
    except Exception as e:
        logger.error(f"查询客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/customer-gold-transfers/{transfer_id}/confirm")
async def confirm_customer_gold_transfer(
    transfer_id: int,
    confirmed_by: str = Query("料部", description="确认人"),
    db: Session = Depends(get_db)
):
    """料部确认客料回仓单"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        transfer = db.query(CustomerGoldTransfer).filter(CustomerGoldTransfer.id == transfer_id).first()
        if not transfer:
            raise HTTPException(status_code=404, detail="回仓单不存在")
        if transfer.status != "pending":
            raise HTTPException(status_code=400, detail=f"当前状态为 {transfer.status}，无法确认")
        
        now = china_now()
        old_status = transfer.status
        transfer.status = "confirmed"
        transfer.confirmed_by = confirmed_by
        transfer.confirmed_at = now
        
        # 记录操作日志
        log = OrderStatusLog(
            order_type="customer_gold_transfer",
            order_id=transfer.id,
            action="confirm",
            old_status=old_status,
            new_status="confirmed",
            operated_by=confirmed_by,
            operated_at=now,
            remark=f"料部确认接收客料 {transfer.gold_weight}克"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"确认客料回仓单: {transfer.transfer_no}, 克重: {transfer.gold_weight}g")
        
        return success_response(
            data={
                "id": transfer.id,
                "transfer_no": transfer.transfer_no,
                "status": "confirmed",
                "confirmed_by": confirmed_by,
                "confirmed_at": now.isoformat()
            },
            message=f"客料回仓单 {transfer.transfer_no} 已确认"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"确认客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"确认失败: {str(e)}")


@router.post("/customer-gold-transfers/{transfer_id}/unconfirm")
async def unconfirm_customer_gold_transfer(
    transfer_id: int,
    operated_by: str = Query("料部", description="操作人"),
    reason: str = Query("", description="反确认原因"),
    db: Session = Depends(get_db)
):
    """反确认客料回仓单"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        transfer = db.query(CustomerGoldTransfer).filter(CustomerGoldTransfer.id == transfer_id).first()
        if not transfer:
            raise HTTPException(status_code=404, detail="回仓单不存在")
        if transfer.status != "confirmed":
            raise HTTPException(status_code=400, detail=f"当前状态为 {transfer.status}，无法反确认")
        
        now = china_now()
        old_status = transfer.status
        transfer.status = "pending"
        transfer.confirmed_by = None
        transfer.confirmed_at = None
        
        # 记录操作日志
        log = OrderStatusLog(
            order_type="customer_gold_transfer",
            order_id=transfer.id,
            action="unconfirm",
            old_status=old_status,
            new_status="pending",
            operated_by=operated_by,
            operated_at=now,
            remark=f"反确认原因: {reason}" if reason else "反确认客料回仓单"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"反确认客料回仓单: {transfer.transfer_no}")
        
        return success_response(
            data={"id": transfer.id, "transfer_no": transfer.transfer_no, "status": "pending"},
            message=f"客料回仓单 {transfer.transfer_no} 已反确认"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"反确认客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"反确认失败: {str(e)}")


@router.put("/customer-gold-transfers/{transfer_id}")
async def update_customer_gold_transfer(
    transfer_id: int,
    gold_weight: Optional[float] = None,
    gold_fineness: Optional[str] = None,
    remark: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """编辑客料回仓单（仅pending状态可编辑）"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        transfer = db.query(CustomerGoldTransfer).filter(CustomerGoldTransfer.id == transfer_id).first()
        if not transfer:
            raise HTTPException(status_code=404, detail="回仓单不存在")
        if transfer.status != "pending":
            raise HTTPException(status_code=400, detail="只能编辑待确认的回仓单")
        
        if gold_weight is not None:
            if gold_weight <= 0:
                raise HTTPException(status_code=400, detail="克重必须大于0")
            transfer.gold_weight = gold_weight
        if gold_fineness is not None:
            transfer.gold_fineness = gold_fineness
        if remark is not None:
            transfer.remark = remark
        
        db.commit()
        db.refresh(transfer)
        
        return success_response(
            data={
                "id": transfer.id,
                "transfer_no": transfer.transfer_no,
                "gold_weight": transfer.gold_weight,
                "gold_fineness": transfer.gold_fineness,
                "remark": transfer.remark,
                "status": transfer.status
            },
            message="更新成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


@router.delete("/customer-gold-transfers/{transfer_id}")
async def delete_customer_gold_transfer(
    transfer_id: int,
    db: Session = Depends(get_db)
):
    """删除客料回仓单（仅pending状态可删除）"""
    from ..models.finance import CustomerGoldTransfer
    
    try:
        transfer = db.query(CustomerGoldTransfer).filter(CustomerGoldTransfer.id == transfer_id).first()
        if not transfer:
            raise HTTPException(status_code=404, detail="回仓单不存在")
        if transfer.status != "pending":
            raise HTTPException(status_code=400, detail="只能删除待确认的回仓单")
        
        transfer_no = transfer.transfer_no
        db.delete(transfer)
        db.commit()
        
        logger.info(f"删除客料回仓单: {transfer_no}")
        
        return success_response(message=f"回仓单 {transfer_no} 已删除")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除客料回仓单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


# ==================== 供应商工费付款登记 ====================

@router.post("/supplier-cash-payments")
async def create_supplier_cash_payment(
    supplier_id: int,
    amount: float,
    payment_method: str = Query("bank_transfer", description="付款方式: bank_transfer/cash/check"),
    remark: str = "",
    created_by: str = "料部",
    db: Session = Depends(get_db)
):
    """创建供应商工费付款登记（pending状态，不立即抵扣应付账款）"""
    from ..models.finance import SupplierPayment, AccountPayable
    
    try:
        # 验证供应商
        supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="供应商不存在")
        
        # 验证金额
        if amount is None or amount <= 0:
            raise HTTPException(status_code=400, detail="付款金额必须大于0")
        
        # 检查该供应商是否有未付应付账款
        total_unpaid = db.query(func.sum(AccountPayable.unpaid_amount)).filter(
            AccountPayable.supplier_id == supplier_id,
            AccountPayable.status.in_(['unpaid', 'partial'])
        ).scalar() or 0
        
        if amount > total_unpaid and total_unpaid > 0:
            logger.warning(f"付款金额 {amount} 超过未付总额 {total_unpaid}，供应商: {supplier.name}")
        
        # 生成付款单号
        now = china_now()
        count = db.query(SupplierPayment).filter(
            SupplierPayment.create_time >= now.replace(hour=0, minute=0, second=0)
        ).count()
        payment_no = f"FK{now.strftime('%Y%m%d')}{count + 1:03d}"
        
        # 创建付款记录（pending状态）
        payment = SupplierPayment(
            payment_no=payment_no,
            supplier_id=supplier_id,
            payment_date=now.date(),
            amount=amount,
            payment_method=payment_method,
            remark=remark,
            created_by=created_by,
            status="pending"
        )
        db.add(payment)
        db.commit()
        db.refresh(payment)
        
        logger.info(f"创建供应商付款登记(待确认): {payment_no}, 供应商: {supplier.name}, 金额: ¥{amount:.2f}")
        
        return success_response(
            data={
                "id": payment.id,
                "payment_no": payment.payment_no,
                "supplier_id": supplier_id,
                "supplier_name": supplier.name,
                "amount": amount,
                "payment_method": payment_method,
                "status": "pending",
                "created_by": created_by,
                "create_time": payment.create_time.isoformat() if payment.create_time else None,
                "total_unpaid": round(total_unpaid, 2)
            },
            message=f"付款登记已创建（待确认），单号：{payment_no}"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建供应商付款登记失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


@router.get("/supplier-cash-payments")
async def get_supplier_cash_payments(
    supplier_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """查询供应商工费付款列表"""
    from ..models.finance import SupplierPayment
    
    try:
        query = db.query(SupplierPayment)
        if supplier_id:
            query = query.filter(SupplierPayment.supplier_id == supplier_id)
        if status:
            query = query.filter(SupplierPayment.status == status)
        
        total = query.count()
        payments = query.order_by(desc(SupplierPayment.create_time)).offset(skip).limit(limit).all()
        
        result = []
        for p in payments:
            supplier = db.query(Supplier).filter(Supplier.id == p.supplier_id).first()
            result.append({
                "id": p.id,
                "payment_no": p.payment_no,
                "supplier_id": p.supplier_id,
                "supplier_name": supplier.name if supplier else "未知",
                "amount": p.amount,
                "payment_method": p.payment_method,
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "status": p.status or "confirmed",
                "remark": p.remark or "",
                "created_by": p.created_by or "",
                "confirmed_by": p.confirmed_by or "",
                "confirmed_at": p.confirmed_at.isoformat() if p.confirmed_at else None,
                "create_time": p.create_time.isoformat() if p.create_time else None
            })
        
        return success_response(data={"payments": result, "total": total})
    except Exception as e:
        logger.error(f"查询供应商付款列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/supplier-cash-payments/{payment_id}/confirm")
async def confirm_supplier_cash_payment(
    payment_id: int,
    confirmed_by: str = "料部",
    db: Session = Depends(get_db)
):
    """确认供应商工费付款（FIFO抵扣应付账款）"""
    from ..models.finance import SupplierPayment, AccountPayable
    
    try:
        payment = db.query(SupplierPayment).filter(SupplierPayment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="付款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=400, detail=f"付款状态为 {payment.status}，无法确认")
        
        supplier = db.query(Supplier).filter(Supplier.id == payment.supplier_id).first()
        
        # FIFO 抵扣应付账款
        unpaid_payables = db.query(AccountPayable).filter(
            AccountPayable.supplier_id == payment.supplier_id,
            AccountPayable.status.in_(['unpaid', 'partial'])
        ).order_by(AccountPayable.due_date.asc()).all()
        
        remaining = payment.amount
        offset_details = []
        
        for payable in unpaid_payables:
            if remaining <= 0:
                break
            offset_amount = min(remaining, payable.unpaid_amount)
            payable.paid_amount = (payable.paid_amount or 0) + offset_amount
            payable.unpaid_amount = (payable.unpaid_amount or 0) - offset_amount
            if payable.unpaid_amount <= 0.01:
                payable.unpaid_amount = 0
                payable.status = 'paid'
            else:
                payable.status = 'partial'
            remaining -= offset_amount
            offset_details.append({
                "payable_id": payable.id,
                "payable_no": payable.payable_no,
                "offset_amount": round(offset_amount, 2)
            })
        
        # 更新付款状态
        now = china_now()
        payment.status = "confirmed"
        payment.confirmed_by = confirmed_by
        payment.confirmed_at = now
        
        # 审计日志
        log = OrderStatusLog(
            order_type="supplier_cash_payment",
            order_id=payment.id,
            action="confirm",
            old_status="pending",
            new_status="confirmed",
            operated_by=confirmed_by,
            remark=f"确认供应商付款 ¥{payment.amount:.2f}，抵扣{len(offset_details)}笔应付账款"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"确认供应商付款: {payment.payment_no}, 供应商: {supplier.name if supplier else ''}, 金额: ¥{payment.amount:.2f}, 抵扣{len(offset_details)}笔")
        
        return success_response(
            data={
                "id": payment.id,
                "payment_no": payment.payment_no,
                "status": "confirmed",
                "offset_details": offset_details,
                "unallocated": round(remaining, 2)
            },
            message=f"付款已确认，抵扣{len(offset_details)}笔应付账款"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"确认供应商付款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"确认失败: {str(e)}")


@router.put("/supplier-cash-payments/{payment_id}")
async def update_supplier_cash_payment(
    payment_id: int,
    supplier_id: Optional[int] = None,
    amount: Optional[float] = None,
    payment_method: Optional[str] = None,
    remark: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """编辑供应商工费付款记录（仅pending状态可编辑）"""
    from ..models.finance import SupplierPayment
    
    try:
        payment = db.query(SupplierPayment).filter(SupplierPayment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="付款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=400, detail="只能编辑待确认的付款记录")
        
        if supplier_id is not None:
            supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
            if not supplier:
                raise HTTPException(status_code=404, detail="供应商不存在")
            payment.supplier_id = supplier_id
        
        if amount is not None:
            if amount <= 0:
                raise HTTPException(status_code=400, detail="付款金额必须大于0")
            payment.amount = amount
        
        if payment_method is not None:
            payment.payment_method = payment_method
        
        if remark is not None:
            payment.remark = remark
        
        db.commit()
        db.refresh(payment)
        
        supplier = db.query(Supplier).filter(Supplier.id == payment.supplier_id).first()
        
        logger.info(f"编辑供应商付款: {payment.payment_no}")
        
        return success_response(
            data={
                "id": payment.id,
                "payment_no": payment.payment_no,
                "supplier_id": payment.supplier_id,
                "supplier_name": supplier.name if supplier else "未知",
                "amount": payment.amount,
                "payment_method": payment.payment_method,
                "remark": payment.remark or "",
                "status": payment.status
            },
            message="付款记录已更新"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"编辑供应商付款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"编辑失败: {str(e)}")


@router.delete("/supplier-cash-payments/{payment_id}")
async def delete_supplier_cash_payment(
    payment_id: int,
    deleted_by: str = "料部",
    db: Session = Depends(get_db)
):
    """删除供应商工费付款记录（仅pending状态可删除）"""
    from ..models.finance import SupplierPayment
    
    try:
        payment = db.query(SupplierPayment).filter(SupplierPayment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="付款记录不存在")
        
        if (payment.status or "confirmed") != "pending":
            raise HTTPException(status_code=400, detail="只能删除待确认的付款记录")
        
        payment_no = payment.payment_no
        db.delete(payment)
        db.commit()
        
        logger.info(f"删除供应商付款: {payment_no}, 删除人: {deleted_by}")
        
        return success_response(message=f"付款记录 {payment_no} 已删除")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除供应商付款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.post("/supplier-cash-payments/{payment_id}/unconfirm")
async def unconfirm_supplier_cash_payment(
    payment_id: int,
    reason: str = Query(..., description="反确认原因"),
    operated_by: str = "料部",
    db: Session = Depends(get_db)
):
    """反确认供应商工费付款（回滚已抵扣的应付账款）"""
    from ..models.finance import SupplierPayment, AccountPayable
    
    try:
        payment = db.query(SupplierPayment).filter(SupplierPayment.id == payment_id).first()
        if not payment:
            raise HTTPException(status_code=404, detail="付款记录不存在")
        
        if (payment.status or "confirmed") != "confirmed":
            raise HTTPException(status_code=400, detail=f"付款状态为 {payment.status}，只能反确认已确认的记录")
        
        supplier = db.query(Supplier).filter(Supplier.id == payment.supplier_id).first()
        
        # 回滚：查找该供应商最近被抵扣的应付账款，恢复金额
        # 从最近付清的开始反向恢复
        payables = db.query(AccountPayable).filter(
            AccountPayable.supplier_id == payment.supplier_id,
            AccountPayable.status.in_(['paid', 'partial'])
        ).order_by(AccountPayable.due_date.desc()).all()
        
        remaining = payment.amount
        restored_count = 0
        
        for payable in payables:
            if remaining <= 0:
                break
            # 恢复该笔应付的已付金额
            restore_amount = min(remaining, payable.paid_amount or 0)
            if restore_amount <= 0:
                continue
            payable.paid_amount = (payable.paid_amount or 0) - restore_amount
            payable.unpaid_amount = (payable.unpaid_amount or 0) + restore_amount
            if payable.paid_amount <= 0.01:
                payable.status = 'unpaid'
            else:
                payable.status = 'partial'
            remaining -= restore_amount
            restored_count += 1
        
        # 更新付款状态
        payment.status = "pending"
        payment.confirmed_by = None
        payment.confirmed_at = None
        
        # 审计日志
        log = OrderStatusLog(
            order_type="supplier_cash_payment",
            order_id=payment.id,
            action="unconfirm",
            old_status="confirmed",
            new_status="pending",
            operated_by=operated_by,
            remark=f"反确认供应商付款 ¥{payment.amount:.2f}，恢复{restored_count}笔应付账款。原因：{reason}"
        )
        db.add(log)
        db.commit()
        
        logger.info(f"反确认供应商付款: {payment.payment_no}, 供应商: {supplier.name if supplier else ''}, 金额: ¥{payment.amount:.2f}, 恢复{restored_count}笔, 原因: {reason}")
        
        return success_response(
            message=f"付款已反确认，恢复{restored_count}笔应付账款"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"反确认供应商付款失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"反确认失败: {str(e)}")


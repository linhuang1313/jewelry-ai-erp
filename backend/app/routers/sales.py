"""
销售单管理路由
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse, HTMLResponse, Response
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import func, desc
from datetime import datetime, timezone, timedelta
from typing import Optional
import logging
import io

from ..database import get_db
from ..utils.response import success_response, error_response
from ..models import SalesOrder, SalesDetail, Customer, Inventory, ProductCode, LocationInventory, Location, OrderStatusLog, InboundDetail
from ..schemas import SalesOrderCreate, SalesOrderResponse, SalesDetailResponse

from ..timezone_utils import china_now
from ..utils.product_utils import resolve_product_code
from ..dependencies.auth import get_current_role, require_permission

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/sales", tags=["销售单管理"])


@router.post("/orders")
async def create_sales_order(order_data: SalesOrderCreate, db: Session = Depends(get_db)):
    """创建销售单"""
    try:
        # ==================== 数据验证 ====================
        # 验证商品明细数据
        for item in order_data.items:
            if item.weight <= 0:
                return error_response(
                    message=f"商品 {item.product_name} 的重量必须大于0",
                    data={
                        "validation_error": {
                            "product_name": item.product_name,
                            "field": "weight",
                            "value": item.weight
                        }
                    }
                )
            if item.labor_cost < 0:
                return error_response(
                    message=f"商品 {item.product_name} 的工费不能为负数",
                    data={
                        "validation_error": {
                            "product_name": item.product_name,
                            "field": "labor_cost",
                            "value": item.labor_cost
                        }
                    }
                )
        # ==================== 数据验证结束 ====================
        
        # ==================== 商品编码转换 ====================
        # 如果输入的是商品编码或「编码+名称」格式，自动转换为标准商品名称
        for item in order_data.items:
            product_name = item.product_name
            if not product_name or not product_name.strip():
                continue
            product_name = product_name.strip()
            code_record = None
            # 1. 精确匹配：纯编码（全大写或包含数字）
            if product_name.isupper() or any(c.isdigit() for c in product_name):
                code_record = db.query(ProductCode).filter(ProductCode.code == product_name).first()
            # 2. 「编码 名称」格式：提取首段作为编码再查（如 "GFDZ 足金古法999吊坠" -> "GFDZ"）
            if not code_record and " " in product_name:
                possible_code = product_name.split()[0]
                if possible_code and (possible_code.isupper() or any(c.isdigit() for c in possible_code)):
                    code_record = db.query(ProductCode).filter(ProductCode.code == possible_code).first()
            if code_record and code_record.name:
                logger.info(f"商品编码转换: {product_name} -> {code_record.name}")
                item.product_name = code_record.name
        # ==================== 商品编码转换结束 ====================
        
        # ==================== 库存检查 ====================
        # 使用展厅库存（LocationInventory）检查，与展厅库存页面展示一致
        # 避免 Inventory 与 LocationInventory 不同步导致"展厅有货但开单失败"的问题
        showroom_location = db.query(Location).filter(
            Location.code == "showroom", Location.is_active == 1
        ).first()
        if not showroom_location:
            return error_response(message="未配置展厅位置，无法创建销售单")

        inventory_errors = []
        # 预解析商品编码（支持用户输入编码代替商品名称）
        resolved_items = []
        for item in order_data.items:
            resolved_name, code = resolve_product_code(item.product_name, db)
            resolved_items.append((item, resolved_name, code))

        for item, resolved_name, code in resolved_items:
            # 查询展厅库存（对同名商品所有记录求和）
            showroom_weight_result = db.query(func.sum(LocationInventory.weight)).filter(
                LocationInventory.product_name == resolved_name,
                LocationInventory.location_id == showroom_location.id,
            ).scalar()
            showroom_weight = float(showroom_weight_result) if showroom_weight_result else 0.0

            # 计算可用库存：展厅库存 - 待结算销售单占用的库存
            reserved_weight = float(db.query(func.sum(SalesDetail.weight)).join(
                SalesOrder
            ).filter(
                SalesDetail.product_name == resolved_name,
                SalesOrder.status == "draft"
            ).scalar() or 0)

            available_weight = showroom_weight - reserved_weight

            if available_weight < item.weight:
                inventory_errors.append({
                    "product_name": resolved_name,
                    "error": "库存不足" if showroom_weight > 0 else "商品不存在于展厅库存中",
                    "required_weight": item.weight,
                    "available_weight": available_weight,
                    "total_weight": showroom_weight,
                    "reserved_weight": reserved_weight
                })
        
        # 如果有任何商品库存不足，拒绝创建销售单
        if inventory_errors:
            return error_response(
                message="库存检查失败，无法创建销售单",
                data={"inventory_errors": inventory_errors}
            )
        # ==================== 库存检查结束 ====================
        
        # 处理客户（在库存检查通过后）
        customer_id = order_data.customer_id
        customer_name = order_data.customer_name
        
        # 如果没有提供customer_id，尝试根据姓名查找
        if not customer_id:
            customer = db.query(Customer).filter(
                Customer.name == customer_name,
                Customer.status == "active"
            ).first()
            if customer:
                customer_id = customer.id
            else:
                # 客户不存在，自动创建
                customer_no = f"KH{china_now().strftime('%Y%m%d%H%M%S')}"
                customer = Customer(
                    customer_no=customer_no,
                    name=customer_name,
                    customer_type="个人"
                )
                db.add(customer)
                db.flush()
                customer_id = customer.id
        
        # 计算总工费和总克重
        # 总工费 = (克重 × 克工费) + (件数 × 件工费)
        def calc_item_total(item):
            gram_cost = item.labor_cost * item.weight
            piece_cost = (item.piece_count or 0) * (item.piece_labor_cost or 0)
            return gram_cost + piece_cost
        
        total_labor_cost = sum(calc_item_total(item) for item in order_data.items)
        total_weight = sum(item.weight for item in order_data.items)
        
        # 生成销售单号（使用中国时间）
        order_no = f"XS{china_now().strftime('%Y%m%d%H%M%S')}"
        
        # 创建销售单
        sales_order = SalesOrder(
            order_no=order_no,
            order_date=order_data.order_date or datetime.now(),
            customer_id=customer_id,
            customer_name=customer_name,
            salesperson=order_data.salesperson,
            store_code=order_data.store_code,
            remark=order_data.remark,
            total_labor_cost=total_labor_cost,
            total_weight=total_weight,
            status="draft"
        )
        db.add(sales_order)
        db.flush()
        
        # 创建销售明细（使用解析后的商品名称）
        resolved_map = {id(item): (rn, c) for item, rn, c in resolved_items}
        details = []
        for item in order_data.items:
            resolved_name, resolved_code = resolved_map.get(id(item), (item.product_name, None))
            code = item.product_code or resolved_code
            # 计算单项总工费：(克重 × 克工费) + (件数 × 件工费)
            gram_cost = item.labor_cost * item.weight
            piece_cost = (item.piece_count or 0) * (item.piece_labor_cost or 0)
            item_total_cost = gram_cost + piece_cost
            
            detail = SalesDetail(
                order_id=sales_order.id,
                product_code=code,
                product_name=resolved_name,
                weight=item.weight,
                labor_cost=item.labor_cost,
                piece_count=item.piece_count,
                piece_labor_cost=item.piece_labor_cost,
                total_labor_cost=item_total_cost
            )
            db.add(detail)
            details.append(detail)
        
        # 库存将在确认(confirm)时扣减，创建时不影响库存
        
        # 更新客户统计信息
        if customer_id:
            customer = db.query(Customer).filter(Customer.id == customer_id).first()
            if customer:
                customer.total_purchase_amount = (customer.total_purchase_amount or 0) + total_labor_cost
                customer.total_purchase_weight = (customer.total_purchase_weight or 0) + total_weight
                customer.total_purchase_count = (customer.total_purchase_count or 0) + 1
                customer.last_purchase_time = sales_order.order_date
        
        db.commit()
        db.refresh(sales_order)
        for detail in details:
            db.refresh(detail)
        
        # 构建响应
        order_response = SalesOrderResponse.model_validate(sales_order)
        order_response.details = [SalesDetailResponse.model_validate(d).model_dump(mode='json') for d in details]
        
        return success_response(
            data={"order": order_response.model_dump(mode='json')},
            message=f"销售单创建成功：{order_no}"
        )
    
    except Exception as e:
        db.rollback()
        logger.error(f"创建销售单失败: {e}", exc_info=True)
        return error_response(message=f"创建销售单失败: {str(e)}")


@router.get("/orders")
async def get_sales_orders(
    order_no: Optional[str] = Query(None, description="销售单号（模糊匹配）"),
    customer_name: Optional[str] = None,
    salesperson: Optional[str] = None,
    status: Optional[str] = Query(None, description="状态筛选：draft/confirmed/cancelled"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    min_weight: Optional[float] = Query(None, description="最小克重"),
    max_weight: Optional[float] = Query(None, description="最大克重"),
    page: int = Query(1, ge=1, description="页码（从1开始）"),
    page_size: int = Query(20, ge=1, le=200, description="每页数量"),
    db: Session = Depends(get_db)
):
    """获取销售单列表（分页 + 批量查询避免 N+1）"""
    from collections import defaultdict
    
    try:
        query = db.query(SalesOrder)
        
        if order_no:
            query = query.filter(SalesOrder.order_no.contains(order_no))
        if customer_name:
            query = query.filter(SalesOrder.customer_name.contains(customer_name))
        if salesperson:
            query = query.filter(SalesOrder.salesperson == salesperson)
        if status:
            query = query.filter(SalesOrder.status == status)
        if start_date:
            try:
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(SalesOrder.order_date >= start_dt)
            except (ValueError, TypeError):
                pass
        if end_date:
            try:
                end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(SalesOrder.order_date <= end_dt)
            except (ValueError, TypeError):
                pass
        if min_weight is not None:
            query = query.filter(SalesOrder.total_weight >= min_weight)
        if max_weight is not None:
            query = query.filter(SalesOrder.total_weight <= max_weight)
        
        total = query.count()

        from sqlalchemy import case
        status_agg = query.with_entities(
            func.count(case((SalesOrder.status == 'draft', 1))).label('draft_count'),
            func.count(case((SalesOrder.status.in_(['confirmed', '待结算', '已结算', 'completed']), 1))).label('confirmed_count'),
            func.count(case((SalesOrder.status.in_(['cancelled', '已取消']), 1))).label('cancelled_count'),
            func.coalesce(func.sum(SalesOrder.total_weight), 0).label('sum_weight'),
            func.coalesce(func.sum(SalesOrder.total_labor_cost), 0).label('sum_labor'),
        ).first()

        offset = (page - 1) * page_size
        orders = query.order_by(desc(SalesOrder.order_date)).offset(offset).limit(page_size).all()
        
        if not orders:
            return success_response(data={
                "orders": [], "total": total, "page": page, "page_size": page_size,
                "total_pages": 0, "draft_count": 0, "confirmed_count": 0,
                "cancelled_count": 0, "sum_weight": 0, "sum_labor": 0
            }, message="查询成功")
        
        order_ids = [o.id for o in orders]
        
        all_details = db.query(SalesDetail).filter(
            SalesDetail.order_id.in_(order_ids)
        ).all()
        
        details_map = defaultdict(list)
        for d in all_details:
            details_map[d.order_id].append(d)
        
        f_codes = [d.product_code for d in all_details if d.product_code and d.product_code.startswith('F')]
        inbound_map = {}
        if f_codes:
            inbound_rows = db.query(InboundDetail).filter(InboundDetail.product_code.in_(f_codes)).all()
            for ib in inbound_rows:
                inbound_map[ib.product_code] = ib

        INLAY_FIELDS = ['main_stone_weight', 'main_stone_count', 'sub_stone_weight', 'sub_stone_count',
                        'main_stone_mark', 'sub_stone_mark', 'pearl_weight', 'bearing_weight',
                        'sale_labor_cost', 'sale_piece_labor_cost']

        def _build_detail_dict(d):
            dd = {
                "id": d.id,
                "product_code": d.product_code,
                "product_name": d.product_name,
                "weight": float(d.weight or 0),
                "labor_cost": float(d.labor_cost or 0),
                "piece_count": int(d.piece_count or 0),
                "piece_labor_cost": float(d.piece_labor_cost or 0),
                "total_labor_cost": float(d.total_labor_cost or 0)
            }
            if d.product_code and d.product_code.startswith('F'):
                ib = inbound_map.get(d.product_code)
                if ib:
                    for f in INLAY_FIELDS:
                        val = getattr(ib, f, None)
                        dd[f] = float(val) if val is not None and isinstance(val, (int, float)) else val
            return dd

        result = []
        for order in orders:
            details = details_map.get(order.id, [])
            order_dict = {
                "id": order.id,
                "order_no": order.order_no,
                "order_date": order.order_date.isoformat() if order.order_date else None,
                "customer_name": order.customer_name,
                "customer_id": order.customer_id,
                "salesperson": order.salesperson,
                "store_code": order.store_code,
                "total_weight": float(order.total_weight or 0),
                "total_labor_cost": float(order.total_labor_cost or 0),
                "remark": order.remark,
                "status": order.status,
                "create_time": order.create_time.isoformat() if order.create_time else None,
                "operator": order.operator,
                "details": [_build_detail_dict(d) for d in details]
            }
            result.append(order_dict)
        
        total_pages = (total + page_size - 1) // page_size if page_size > 0 else 0
        return success_response(data={
            "orders": result,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "draft_count": status_agg.draft_count if status_agg else 0,
            "confirmed_count": status_agg.confirmed_count if status_agg else 0,
            "cancelled_count": status_agg.cancelled_count if status_agg else 0,
            "sum_weight": float(status_agg.sum_weight) if status_agg else 0,
            "sum_labor": float(status_agg.sum_labor) if status_agg else 0,
        }, message="查询成功")
    except Exception as e:
        logger.error(f"查询销售单失败: {e}", exc_info=True)
        return error_response(message=f"查询销售单失败: {str(e)}")


@router.get("/orders/{order_id}")
async def get_sales_order(order_id: int, db: Session = Depends(get_db)):
    """获取销售单详情"""
    try:
        order = db.query(SalesOrder).filter(
            SalesOrder.id == order_id
        ).options(
            selectinload(SalesOrder.details)
        ).first()
        
        if not order:
            return error_response(message="销售单不存在")
        
        order_response = SalesOrderResponse.model_validate(order)

        f_codes = [d.product_code for d in order.details if d.product_code and d.product_code.startswith('F')]
        inbound_map = {}
        if f_codes:
            inbound_rows = db.query(InboundDetail).filter(InboundDetail.product_code.in_(f_codes)).all()
            for ib in inbound_rows:
                inbound_map[ib.product_code] = ib

        INLAY_FIELDS = ['main_stone_weight', 'main_stone_count', 'sub_stone_weight', 'sub_stone_count',
                        'main_stone_mark', 'sub_stone_mark', 'pearl_weight', 'bearing_weight',
                        'sale_labor_cost', 'sale_piece_labor_cost']

        detail_list = []
        for d in order.details:
            dd = SalesDetailResponse.model_validate(d).model_dump(mode='json')
            if d.product_code and d.product_code.startswith('F'):
                ib = inbound_map.get(d.product_code)
                if ib:
                    for f in INLAY_FIELDS:
                        val = getattr(ib, f, None)
                        dd[f] = float(val) if val is not None and isinstance(val, (int, float)) else val
            detail_list.append(dd)

        order_dict = order_response.model_dump(mode='json')
        order_dict['details'] = detail_list

        return success_response(data={"order": order_dict}, message="查询成功")
    except Exception as e:
        logger.error(f"查询销售单详情失败: {e}", exc_info=True)
        return error_response(message=f"查询销售单详情失败: {str(e)}")



@router.get("/orders/{order_id}/download")
async def download_sales_order(
    order_id: int,
    format: str = Query("pdf", pattern="^(pdf|html|excel)$"),
    db: Session = Depends(get_db)
):
    """下载或打印销售单（支持PDF、HTML、Excel格式）"""
    try:
        logger.info(f"下载销售单请求: order_id={order_id}, format={format}")
        
        # 查询销售单（含明细，避免额外查询）
        order = db.query(SalesOrder).filter(
            SalesOrder.id == order_id
        ).options(
            selectinload(SalesOrder.details)
        ).first()
        if not order:
            raise HTTPException(status_code=404, detail="销售单不存在")
        
        details = order.details
        if not details:
            raise HTTPException(status_code=404, detail="销售单明细不存在")
        
        logger.info(f"找到销售单: order_no={order.order_no}, 明细数={len(details)}")
        
        # 时间格式化
        from ..timezone_utils import to_china_time, format_china_time
        if order.order_date:
            china_time = to_china_time(order.order_date)
            order_date_str = format_china_time(china_time, '%Y-%m-%d %H:%M:%S')
        else:
            order_date_str = "未知"
        
        if format == "pdf":
            try:
                import math
                from reportlab.pdfgen import canvas
                from reportlab.lib.units import mm
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                
                # ========== 动态高度计算（针式打印机 241mm x 140mm倍数） ==========
                PAGE_WIDTH = 241 * mm
                # 基础高度（页头页尾固定部分）+ 每行明细高度
                base_height = 80 * mm    # 页头页尾固定部分
                row_height = 9 * mm      # 每行明细高度
                detail_count = len(details)
                content_height = base_height + (row_height * detail_count)
                
                # 按140mm倍数向上取整（最小140mm）
                min_unit = 140 * mm
                PAGE_HEIGHT = max(min_unit, math.ceil(content_height / min_unit) * min_unit)
                # ========== 动态高度计算完成 ==========
                
                buffer = io.BytesIO()
                p = canvas.Canvas(buffer, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))
                width, height = PAGE_WIDTH, PAGE_HEIGHT
                
                # 使用 CID 字体
                try:
                    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                    chinese_font = 'STSong-Light'
                except Exception as cid_error:
                    logger.warning(f"注册CID字体失败: {cid_error}")
                    chinese_font = None
                
                # 页边距
                left_margin = 8 * mm
                right_margin = width - 8 * mm
                top_margin = height - 6 * mm
                
                # 标题（居中）
                if chinese_font:
                    p.setFont(chinese_font, 12)
                else:
                    p.setFont("Helvetica-Bold", 12)
                p.drawCentredString(width / 2, top_margin, "销售单")
                
                # 基本信息（紧凑两列布局）
                y = top_margin - 14
                if chinese_font:
                    p.setFont(chinese_font, 8)
                else:
                    p.setFont("Helvetica", 8)
                
                customer_name = (order.customer_name or '未知')[:10]
                salesperson = (order.salesperson or '未知')[:6]
                p.drawString(left_margin, y, f"单号：{order.order_no}")
                p.drawString(width/2, y, f"日期：{order_date_str}")
                y -= 10
                p.drawString(left_margin, y, f"客户：{customer_name}  业务员：{salesperson}")
                p.drawString(width/2, y, f"状态：{order.status}")
                y -= 12
                
                # 分隔线
                p.line(left_margin, y, right_margin, y)
                y -= 10
                
                # 商品明细表头
                col_x = [left_margin, 55*mm, 85*mm, 115*mm, 145*mm]
                if chinese_font:
                    p.setFont(chinese_font, 7)
                else:
                    p.setFont("Helvetica-Bold", 7)
                p.drawString(col_x[0], y, "商品名称")
                p.drawString(col_x[1], y, "克重(g)")
                p.drawString(col_x[2], y, "工费/克")
                p.drawString(col_x[3], y, "总工费")
                y -= 8
                p.line(left_margin, y, right_margin, y)
                y -= 8
                
                # 商品明细行
                for detail in details:
                    product_name = detail.product_name[:12] if len(detail.product_name) > 12 else detail.product_name
                    if chinese_font:
                        p.setFont(chinese_font, 7)
                    p.drawString(col_x[0], y, product_name)
                    p.setFont("Helvetica", 7)
                    p.drawString(col_x[1], y, f"{detail.weight:.2f}")
                    p.drawString(col_x[2], y, f"{detail.labor_cost:.1f}")
                    p.drawString(col_x[3], y, f"{detail.total_labor_cost:.2f}")
                    y -= 9
                
                y -= 3
                p.line(left_margin, y, right_margin, y)
                y -= 10
                
                # 汇总
                if chinese_font:
                    p.setFont(chinese_font, 8)
                else:
                    p.setFont("Helvetica-Bold", 8)
                p.drawString(left_margin, y, f"合计：总克重 {order.total_weight:.2f}g  |  总工费 ¥{order.total_labor_cost:.2f}")
                y -= 10
                
                # 备注
                if order.remark:
                    if chinese_font:
                        p.setFont(chinese_font, 7)
                    remark_text = order.remark[:30] if len(order.remark) > 30 else order.remark
                    p.drawString(left_margin, y, f"备注：{remark_text}")
                
                p.save()
                buffer.seek(0)
                
                filename = f"sales_order_{order.order_no}.pdf"
                return StreamingResponse(
                    buffer,
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f'attachment; filename="{filename}"',
                    }
                )
            except Exception as pdf_error:
                logger.error(f"生成销售单PDF失败: {pdf_error}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"生成PDF失败: {str(pdf_error)}")
        
        elif format == "excel":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from urllib.parse import quote

                has_f_code = any(d.product_code and d.product_code.startswith('F') for d in details)
                inbound_map = {}
                inlay_fields = [
                    ('main_stone_weight', '主石重'), ('main_stone_count', '主石粒数'),
                    ('sub_stone_weight', '副石重'), ('sub_stone_count', '副石粒数'),
                    ('main_stone_mark', '主石字印'), ('sub_stone_mark', '副石字印'),
                    ('pearl_weight', '珍珠重'), ('bearing_weight', '轴承重'),
                    ('sale_labor_cost', '销售克工费'), ('sale_piece_labor_cost', '销售件工费'),
                ]
                if has_f_code:
                    f_codes = [d.product_code for d in details if d.product_code and d.product_code.startswith('F')]
                    inbound_rows = db.query(InboundDetail).filter(InboundDetail.product_code.in_(f_codes)).all()
                    for ib in inbound_rows:
                        inbound_map[ib.product_code] = ib

                wb = Workbook()
                ws = wb.active
                ws.title = "销售单"

                title_font = Font(bold=True, size=14)
                header_font = Font(bold=True, color="FFFFFF", size=10)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                base_headers = ["商品编码", "商品名称", "克重(g)", "工费(元/克)", "件数", "件工费(元/件)", "总工费(元)"]
                extra_headers = [label for _, label in inlay_fields] if has_f_code else []
                headers = base_headers + extra_headers
                total_cols = len(headers)
                last_col_letter = get_column_letter(total_cols)

                ws.merge_cells(f"A1:{last_col_letter}1")
                ws["A1"] = "销售单"
                ws["A1"].font = title_font
                ws["A1"].alignment = Alignment(horizontal="center")

                info_labels = [
                    ("A3", "销售单号："), ("B3", order.order_no),
                    ("D3", "客户："), ("E3", order.customer_name or "未知"),
                    ("A4", "业务员："), ("B4", order.salesperson or "未知"),
                    ("D4", "门店代码："), ("E4", order.store_code or "未填写"),
                    ("A5", "日期："), ("B5", order_date_str),
                    ("D5", "状态："), ("E5", order.status),
                ]
                label_font = Font(bold=True, size=10)
                for cell_ref, value in info_labels:
                    ws[cell_ref] = value
                    if cell_ref[0] in ("A", "D"):
                        ws[cell_ref].font = label_font

                header_row = 7
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for row_idx, detail in enumerate(details, header_row + 1):
                    values = [
                        detail.product_code or "",
                        detail.product_name,
                        float(detail.weight) if detail.weight else 0,
                        float(detail.labor_cost) if detail.labor_cost else 0,
                        detail.piece_count if detail.piece_count else "",
                        float(detail.piece_labor_cost) if detail.piece_labor_cost else "",
                        float(detail.total_labor_cost) if detail.total_labor_cost else 0,
                    ]
                    if has_f_code:
                        ib = inbound_map.get(detail.product_code) if detail.product_code else None
                        for field_name, _ in inlay_fields:
                            val = getattr(ib, field_name, None) if ib else None
                            if val is not None and isinstance(val, (int, float)):
                                values.append(float(val))
                            else:
                                values.append(val if val else "")
                    for col_idx, val in enumerate(values, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.border = thin_border
                        if col_idx >= 3:
                            cell.alignment = center_align

                summary_row = header_row + len(details) + 1
                ws.merge_cells(f"A{summary_row}:B{summary_row}")
                ws.cell(row=summary_row, column=1, value="合计").font = Font(bold=True, size=10)
                ws.cell(row=summary_row, column=3, value=float(order.total_weight) if order.total_weight else 0).font = Font(bold=True)
                ws.cell(row=summary_row, column=3).border = thin_border
                ws.cell(row=summary_row, column=7, value=float(order.total_labor_cost) if order.total_labor_cost else 0).font = Font(bold=True)
                ws.cell(row=summary_row, column=7).border = thin_border

                if order.remark:
                    remark_row = summary_row + 1
                    ws.merge_cells(f"A{remark_row}:{last_col_letter}{remark_row}")
                    ws.cell(row=remark_row, column=1, value=f"备注：{order.remark}")

                col_widths = [16, 20, 12, 14, 8, 14, 14]
                if has_f_code:
                    col_widths += [10, 10, 10, 10, 12, 12, 10, 10, 12, 12]
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                filename = f"销售单_{order.order_no}.xlsx"
                encoded_filename = quote(filename, safe="")

                return Response(
                    content=output.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                    },
                )
            except Exception as excel_error:
                logger.error(f"生成销售单Excel失败: {excel_error}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"生成Excel失败: {str(excel_error)}")

        elif format == "html":
            # HTML 打印格式 - 针式打印机 241mm 宽度
            print_time = format_china_time(china_now(), '%Y/%m/%d %H:%M')
            html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>销售单 - {order.order_no}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'Microsoft YaHei', 'SimHei', sans-serif; padding: 10px; background: #f5f5f5; font-size: 11px; }}
        .container {{ width: 241mm; min-height: 140mm; margin: 0 auto; background: white; padding: 6mm 8mm; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; border-bottom: 1px solid #333; padding-bottom: 8px; margin-bottom: 10px; }}
        .header h1 {{ font-size: 16px; margin-bottom: 5px; }}
        .header-info {{ display: flex; justify-content: space-between; font-size: 10px; margin-top: 5px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 5px 15px; margin-bottom: 10px; font-size: 10px; }}
        .info-item label {{ color: #666; }}
        .info-item span {{ font-weight: 500; }}
        .details-table {{ width: 100%; border-collapse: collapse; margin: 8px 0; font-size: 10px; }}
        .details-table th, .details-table td {{ border: 1px solid #999; padding: 4px 6px; }}
        .details-table th {{ background: #f0f0f0; font-weight: bold; text-align: center; }}
        .details-table td {{ text-align: center; }}
        .details-table td.left {{ text-align: left; }}
        .details-table td.right {{ text-align: right; }}
        .summary {{ display: flex; justify-content: flex-end; gap: 30px; margin: 10px 0; font-size: 11px; }}
        .summary .total {{ font-weight: bold; color: #c00; }}
        .remark {{ margin-top: 8px; padding: 6px 10px; background: #fffef0; border: 1px solid #e0d080; border-radius: 3px; font-size: 10px; }}
        .footer {{ margin-top: 15px; display: flex; justify-content: space-between; font-size: 10px; }}
        .signature {{ border-bottom: 1px solid #333; width: 100px; display: inline-block; margin-left: 5px; }}
        .print-btn {{ display: block; margin: 15px auto 0; padding: 8px 25px; background: #3498db; color: white; border: none; border-radius: 5px; font-size: 12px; cursor: pointer; }}
        .print-btn:hover {{ background: #2980b9; }}
        @media print {{
            @page {{ size: 241mm auto; margin: 0; }}
            body {{ background: white; padding: 0; }}
            .container {{ box-shadow: none; width: 241mm; padding: 5mm 8mm; }}
            .print-btn {{ display: none; }}
        }}
        @media screen {{
            body {{ background: #f0f0f0; padding: 15px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>销 售 单</h1>
            <div class="header-info">
                <span>单号：{order.order_no}</span>
                <span>日期：{order_date_str}</span>
                <span>状态：{order.status}</span>
            </div>
        </div>
        <div class="info-grid">
            <div class="info-item">
                <label>客户：</label>
                <span>{order.customer_name or '未知'}</span>
            </div>
            <div class="info-item">
                <label>业务员：</label>
                <span>{order.salesperson or '未知'}</span>
            </div>
        </div>
        <table class="details-table">
            <thead>
                <tr>
                    <th style="width:8%">序号</th>
                    <th style="width:35%">商品名称</th>
                    <th style="width:15%">克重(g)</th>
                    <th style="width:17%">工费(元/克)</th>
                    <th style="width:20%">总工费(元)</th>
                </tr>
            </thead>
            <tbody>
                {"".join(f'''
                <tr>
                    <td>{idx}</td>
                    <td class="left">{detail.product_name}</td>
                    <td class="right">{detail.weight:.2f}</td>
                    <td class="right">{detail.labor_cost:.2f}</td>
                    <td class="right">{detail.total_labor_cost:.2f}</td>
                </tr>
                ''' for idx, detail in enumerate(details, 1))}
                <tr style="font-weight: bold; background: #f8f8f8;">
                    <td colspan="2" class="left">合计：{len(details)} 件</td>
                    <td class="right">{order.total_weight:.2f}</td>
                    <td></td>
                    <td class="right" style="color: #c00;">¥{order.total_labor_cost:.2f}</td>
                </tr>
            </tbody>
        </table>
        {f'<div class="remark"><strong>备注：</strong>{order.remark}</div>' if order.remark else ''}
        <div class="footer">
            <span>制单人：{order.salesperson or ''}</span>
            <span>客户签字：<span class="signature"></span></span>
            <span>打印时间：{print_time}</span>
        </div>
    </div>
    <button class="print-btn" onclick="window.print()">打印销售单</button>
</body>
</html>
"""
            return HTMLResponse(
                content=html_content,
            )
        
        else:
            raise HTTPException(status_code=400, detail="不支持的格式，请使用 pdf 或 html")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成销售单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成销售单失败: {str(e)}")


@router.post("/orders/{order_id}/confirm")
async def confirm_sales_order(
    order_id: int,
    confirmed_by: str = Query(default="系统管理员", description="确认人"),
    user_role: str = Query(default="sales", description="用户角色"),
    role: str = Depends(require_permission("can_create_sales")),
    db: Session = Depends(get_db)
):
    """确认销售单（库存生效）"""
    try:
        # 加行锁防止并发确认
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).with_for_update().first()
        if not order:
            return error_response(message="销售单不存在")
        if order.status != "draft":
            return error_response(message=f"销售单状态为 {order.status}，只有未确认的销售单才能确认")
        
        details = db.query(SalesDetail).filter(SalesDetail.order_id == order_id).all()
        if not details:
            return error_response(message="销售单没有商品明细")
        
        # 验证展厅库存（与 create_sales_order 一致）
        showroom_location = db.query(Location).filter(Location.code == "showroom", Location.is_active == 1).first()
        if not showroom_location:
            return error_response(message="未配置展厅位置，无法确认销售单")
        
        # 预解析商品名称（支持商品编码 fallback 到商品名称）
        for detail in details:
            resolved_name, code = resolve_product_code(detail.product_name, db)
            if code and detail.product_name != resolved_name:
                detail.product_code = code
                detail.product_name = resolved_name
        
        # 验证展厅库存是否充足（使用 SUM 聚合，防御重复记录）
        for detail in details:
            showroom_weight_result = db.query(func.sum(LocationInventory.weight)).filter(
                LocationInventory.product_name == detail.product_name,
                LocationInventory.location_id == showroom_location.id
            ).scalar()
            showroom_weight = float(showroom_weight_result) if showroom_weight_result else 0.0
            if showroom_weight == 0.0:
                return error_response(message=f"商品不存在于展厅库存中：{detail.product_name}")
            if showroom_weight < detail.weight:
                return error_response(message=f"展厅库存不足：{detail.product_name} 仅有 {showroom_weight:.2f}g，需要 {detail.weight}g")
        
        # 锁定并扣减库存
        for detail in details:
            inventory = db.query(Inventory).filter(Inventory.product_name == detail.product_name).with_for_update().first()
            if inventory:
                inventory.total_weight = round(max(0, inventory.total_weight - detail.weight), 3)
            
            location_inv = db.query(LocationInventory).filter(
                LocationInventory.product_name == detail.product_name,
                LocationInventory.location_id == showroom_location.id
            ).with_for_update().first()
            if location_inv:
                location_inv.weight = round(max(0, float(location_inv.weight or 0) - float(detail.weight or 0)), 3)
        
        order.status = "待结算"
        
        status_log = OrderStatusLog(order_type="sales", order_id=order_id, action="confirm", old_status="draft", new_status="待结算", operated_by=confirmed_by, operated_at=china_now())
        db.add(status_log)
        
        db.commit()
        logger.info(f"销售单已确认: {order.order_no}, 确认人: {confirmed_by}")
        
        return success_response(message=f"销售单 {order.order_no} 已确认，库存已扣减")
    
    except Exception as e:
        db.rollback()
        logger.error(f"确认销售单失败: {e}", exc_info=True)
        return error_response(message=f"确认销售单失败: {str(e)}")


@router.post("/orders/{order_id}/unconfirm")
async def unconfirm_sales_order(
    order_id: int,
    operated_by: str = Query(default="系统管理员", description="操作人"),
    user_role: str = Query(default="sales", description="用户角色"),
    remark: str = Query(default="", description="反确认原因"),
    role: str = Depends(require_permission("can_create_sales")),
    db: Session = Depends(get_db)
):
    """反确认销售单（回滚库存）"""
    try:
        # 加行锁防止并发反确认
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).with_for_update().first()
        if not order:
            return error_response(message="销售单不存在")
        if order.status == "已结算":
            return error_response(
                message="该销售单已结算，不能直接反确认。正确流程：1. 对结算单点击「反确认」撤销结算 2. 对结算单点击「撤单」3. 销售单回到待开结算状态后，方可反确认销售单。"
            )
        if order.status not in ("confirmed", "待结算"):
            return error_response(message=f"销售单状态为 {order.status}，只有已确认的销售单才能反确认")
        
        details = db.query(SalesDetail).filter(SalesDetail.order_id == order_id).all()
        
        showroom_location = db.query(Location).filter(Location.code == "showroom", Location.is_active == 1).first()
        
        for detail in details:
            # 锁定并回滚总库存
            inventory = db.query(Inventory).filter(Inventory.product_name == detail.product_name).with_for_update().first()
            if inventory:
                inventory.total_weight = round(inventory.total_weight + detail.weight, 3)
            
            # 锁定并回滚展厅库存
            if showroom_location:
                location_inv = db.query(LocationInventory).filter(
                    LocationInventory.product_name == detail.product_name,
                    LocationInventory.location_id == showroom_location.id
                ).with_for_update().first()
                if location_inv:
                    location_inv.weight += detail.weight
        
        order.status = "draft"
        
        status_log = OrderStatusLog(order_type="sales", order_id=order_id, action="unconfirm", old_status=order.status, new_status="draft", operated_by=operated_by, operated_at=china_now(), remark=remark or None)
        db.add(status_log)
        
        db.commit()
        logger.info(f"销售单已反确认: {order.order_no}, 操作人: {operated_by}")
        
        return success_response(message=f"销售单 {order.order_no} 已反确认，库存已回滚")
    
    except Exception as e:
        db.rollback()
        logger.error(f"反确认销售单失败: {e}", exc_info=True)
        return error_response(message=f"反确认销售单失败: {str(e)}")


@router.post("/orders/{order_id}/cancel")
async def cancel_sales_order(order_id: int, db: Session = Depends(get_db)):
    """取消销售单（仅draft状态可直接取消，confirmed需先反确认）"""
    try:
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
        if not order:
            return error_response(message="销售单不存在")
        
        if order.status == "cancelled":
            return error_response(message="销售单已经是取消状态")
        
        if order.status in ("confirmed", "待结算"):
            return error_response(message="已确认的销售单请先反确认再取消")
        
        if order.status != "draft":
            return error_response(message=f"销售单状态为 {order.status}，无法取消")
        
        # draft状态直接取消，无需回滚库存（创建时未扣减库存）
        order.status = "cancelled"
        
        # 更新客户统计信息（回滚）
        if order.customer_id:
            customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
            if customer:
                customer.total_purchase_amount = max(0, (customer.total_purchase_amount or 0) - (order.total_labor_cost or 0))
                customer.total_purchase_weight = max(0, (customer.total_purchase_weight or 0) - (order.total_weight or 0))
                customer.total_purchase_count = max(0, (customer.total_purchase_count or 0) - 1)
        
        status_log = OrderStatusLog(order_type="sales", order_id=order_id, action="cancel", old_status="draft", new_status="cancelled", operated_by="系统管理员", operated_at=china_now())
        db.add(status_log)
        
        db.commit()
        
        logger.info(f"销售单已取消: {order.order_no}")
        
        return success_response(message=f"销售单 {order.order_no} 已取消")
    
    except Exception as e:
        db.rollback()
        logger.error(f"取消销售单失败: {e}", exc_info=True)
        return error_response(message=f"取消销售单失败: {str(e)}")


@router.delete("/orders/{order_id}")
async def delete_sales_order(order_id: int, db: Session = Depends(get_db)):
    """彻底删除销售单（仅draft状态可删除）"""
    try:
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
        if not order:
            return error_response(message="销售单不存在")

        if order.status != "draft":
            return error_response(message=f"只有未确认的销售单才能删除，当前状态：{order.status}")

        order_no = order.order_no

        if order.customer_id:
            customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
            if customer:
                customer.total_purchase_amount = max(0, (customer.total_purchase_amount or 0) - (order.total_labor_cost or 0))
                customer.total_purchase_weight = max(0, (customer.total_purchase_weight or 0) - (order.total_weight or 0))
                customer.total_purchase_count = max(0, (customer.total_purchase_count or 0) - 1)

        db.query(SalesDetail).filter(SalesDetail.order_id == order_id).delete()

        status_log = OrderStatusLog(order_type="sales", order_id=order_id, action="delete", old_status="draft", new_status="deleted", operated_by="系统管理员", operated_at=china_now(), remark=f"删除销售单 {order_no}")
        db.add(status_log)

        db.delete(order)
        db.commit()

        logger.info(f"销售单已删除: {order_no}")
        return success_response(message=f"销售单 {order_no} 已删除")

    except Exception as e:
        db.rollback()
        logger.error(f"删除销售单失败: {e}", exc_info=True)
        return error_response(message=f"删除销售单失败: {str(e)}")


@router.put("/orders/{order_id}")
async def update_sales_order(
    order_id: int,
    updates: dict,
    user_role: str = Query(default="sales", description="用户角色"),
    role: str = Depends(require_permission("can_create_sales")),
    db: Session = Depends(get_db)
):
    """编辑销售单（仅未确认状态可编辑）"""
    try:
        order = db.query(SalesOrder).filter(SalesOrder.id == order_id).first()
        if not order:
            return error_response(message="销售单不存在")
        if order.status != "draft":
            return error_response(message="只有未确认的销售单才能编辑，请先反确认")
        
        # 更新基本字段
        if "customer_id" in updates and updates["customer_id"]:
            customer = db.query(Customer).filter(Customer.id == updates["customer_id"]).first()
            if customer:
                order.customer_id = customer.id
                order.customer_name = customer.name
        if "customer_name" in updates:
            order.customer_name = updates["customer_name"]
        if "salesperson" in updates:
            order.salesperson = updates["salesperson"]
        if "remark" in updates:
            order.remark = updates["remark"]
        
        # 更新商品明细（删除旧的，重建新的）
        if "items" in updates and isinstance(updates["items"], list) and len(updates["items"]) > 0:
            # 删除旧明细
            db.query(SalesDetail).filter(SalesDetail.order_id == order_id).delete()
            
            # 重建新明细
            total_weight = 0.0
            total_labor_cost = 0.0
            for item in updates["items"]:
                weight = float(item.get("weight", 0))
                labor_cost = float(item.get("labor_cost", 0))
                piece_count = item.get("piece_count")
                piece_labor_cost = float(item.get("piece_labor_cost")) if item.get("piece_labor_cost") is not None else None
                # 单项总工费 = (克重 × 克工费) + (件数 × 件工费)，与创建逻辑一致
                gram_cost = weight * labor_cost
                piece_cost = (piece_count or 0) * (piece_labor_cost or 0)
                item_total_cost = gram_cost + piece_cost
                detail = SalesDetail(
                    order_id=order_id,
                    product_name=item.get("product_name", ""),
                    weight=weight,
                    labor_cost=labor_cost,
                    piece_count=int(piece_count) if piece_count is not None else None,
                    piece_labor_cost=piece_labor_cost,
                    total_labor_cost=round(item_total_cost, 2)
                )
                db.add(detail)
                total_weight += weight
                total_labor_cost += item_total_cost
            
            # 更新主表汇总
            order.total_weight = round(total_weight, 3)
            order.total_labor_cost = round(total_labor_cost, 2)
        
        db.commit()
        return success_response(message="销售单已更新")
    
    except Exception as e:
        db.rollback()
        logger.error(f"编辑销售单失败: {e}", exc_info=True)
        return error_response(message=f"编辑销售单失败: {str(e)}")


@router.get("/frequent-products")
async def get_frequent_products(
    limit: int = Query(8, ge=1, le=20, description="返回数量"),
    db: Session = Depends(get_db)
):
    """获取常用商品（根据历史销售和入库频率排序）"""
    try:
        from sqlalchemy import union_all, literal_column
        
        # 从销售明细和入库明细中统计商品使用频率
        sales_q = db.query(
            SalesDetail.product_name.label("product_name"),
            SalesDetail.labor_cost.label("labor_cost")
        )
        inbound_q = db.query(
            InboundDetail.product_name.label("product_name"),
            InboundDetail.labor_cost.label("labor_cost")
        )
        
        # 合并查询，按商品名分组统计
        combined = sales_q.union_all(inbound_q).subquery()
        
        results = db.query(
            combined.c.product_name,
            func.count().label("frequency"),
            func.round(func.avg(combined.c.labor_cost), 1).label("avg_labor_cost")
        ).group_by(
            combined.c.product_name
        ).order_by(
            desc("frequency")
        ).limit(limit).all()
        
        products = [
            {
                "name": r.product_name,
                "labor_cost": str(r.avg_labor_cost or 0),
                "frequency": r.frequency
            }
            for r in results
            if r.product_name  # 排除空名称
        ]
        
        return products
    except Exception as e:
        logger.error(f"获取常用商品失败: {e}", exc_info=True)
        return []


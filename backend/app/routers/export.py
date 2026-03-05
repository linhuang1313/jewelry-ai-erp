"""
数据导出路由
"""

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response
from slowapi import Limiter
from slowapi.util import get_remote_address

_limiter = Limiter(key_func=get_remote_address)
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from typing import Optional
import logging
import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..dependencies.auth import require_permission, get_current_role
from ..middleware.permissions import has_permission
from ..models import (
    ChatLog, Inventory, InboundOrder, InboundDetail,
    SalesOrder, SalesDetail, Customer, Supplier
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["数据导出"])


def create_excel_response(wb: Workbook, filename: str):
    """创建 Excel 文件响应"""
    from urllib.parse import quote
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 对中文文件名进行 URL 编码
    encoded_filename = quote(filename, safe='')
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        }
    )


def style_header(ws, row=1):
    """为表头添加样式"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def auto_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except (ValueError, TypeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@router.get("/chat-logs")
@_limiter.limit("10/minute")
async def export_chat_logs(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出对话日志为 Excel"""
    try:
        query = db.query(ChatLog).order_by(ChatLog.created_at.desc())
        
        if start_date:
            query = query.filter(func.date(ChatLog.created_at) >= start_date)
        if end_date:
            query = query.filter(func.date(ChatLog.created_at) <= end_date)
        
        logs = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "对话日志"
        
        # 表头
        headers = ["ID", "会话ID", "用户角色", "消息类型", "内容", "意图", "响应时间(ms)", "是否成功", "创建时间"]
        ws.append(headers)
        style_header(ws)
        
        # 数据
        role_names = {"sales": "业务员", "finance": "财务", "product": "商品专员", "manager": "管理层"}
        for log in logs:
            ws.append([
                log.id,
                log.session_id,
                role_names.get(log.user_role, log.user_role),
                "用户" if log.message_type == "user" else "AI助手",
                log.content[:500] if log.content else "",
                log.intent or "",
                log.response_time_ms,
                "是" if log.is_successful else "否",
                log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"对话日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出对话日志失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/inventory")
@_limiter.limit("10/minute")
async def export_inventory(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出库存数据为 Excel"""
    try:
        inventories = db.query(Inventory).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "库存数据"
        
        headers = ["ID", "商品名称", "库存重量(克)", "最后更新时间"]
        ws.append(headers)
        style_header(ws)
        
        for inv in inventories:
            ws.append([
                inv.id,
                inv.product_name,
                inv.total_weight,
                inv.last_update.strftime("%Y-%m-%d %H:%M:%S") if inv.last_update else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"库存数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出库存数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/inbound")
@_limiter.limit("10/minute")
async def export_inbound(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出入库记录为 Excel"""
    try:
        query = db.query(InboundOrder).order_by(InboundOrder.create_time.desc())
        
        if start_date:
            query = query.filter(func.date(InboundOrder.create_time) >= start_date)
        if end_date:
            query = query.filter(func.date(InboundOrder.create_time) <= end_date)
        
        orders = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "入库记录"
        
        headers = ["入库单号", "商品名称", "重量(克)", "工费(元/克)", "总成本(元)", "供应商", "入库时间", "操作员"]
        ws.append(headers)
        style_header(ws)
        
        for order in orders:
            details = db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all()
            for detail in details:
                ws.append([
                    order.order_no,
                    detail.product_name,
                    detail.weight,
                    detail.labor_cost,
                    detail.total_cost,
                    detail.supplier or "",
                    order.create_time.strftime("%Y-%m-%d %H:%M:%S") if order.create_time else "",
                    order.operator or ""
                ])
        
        auto_column_width(ws)
        
        filename = f"入库记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出入库记录失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/sales")
@_limiter.limit("10/minute")
async def export_sales(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出销售订单为 Excel"""
    try:
        query = db.query(SalesOrder).order_by(SalesOrder.order_date.desc())
        
        if start_date:
            query = query.filter(func.date(SalesOrder.order_date) >= start_date)
        if end_date:
            query = query.filter(func.date(SalesOrder.order_date) <= end_date)
        
        orders = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "销售订单"
        
        headers = ["订单号", "客户名称", "商品名称", "重量(克)", "工费(元/克)", "总工费(元)", "业务员", "门店代码", "订单日期"]
        ws.append(headers)
        style_header(ws)
        
        for order in orders:
            details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
            customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
            for detail in details:
                ws.append([
                    order.order_no,
                    customer.name if customer else "",
                    detail.product_name,
                    detail.weight,
                    detail.labor_cost,
                    detail.total_labor_cost,
                    order.salesperson or "",
                    order.store_code or "",
                    order.order_date.strftime("%Y-%m-%d") if order.order_date else ""
                ])
        
        auto_column_width(ws)
        
        filename = f"销售订单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出销售订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/customer-transactions/{customer_id}")
@_limiter.limit("10/minute")
async def export_customer_transactions(
    request: Request,
    customer_id: int,
    role: str = Depends(require_permission("can_export")),
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """导出指定客户的往来账目为 Excel（三栏式会计格式，数据与页面完全一致）"""
    from .customers import get_customer_detail as _get_customer_detail
    
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 直接调用 detail 接口的内部函数，复用完全一致的数据
        detail_response = await _get_customer_detail(
            customer_id=customer_id,
            user_role="manager",
            date_start=date_start,
            date_end=date_end,
            db=db
        )
        if not detail_response.get('success'):
            raise HTTPException(status_code=500, detail="获取客户详情数据失败")
        data = detail_response.get('data', {})
        transactions = data.get('transactions', [])
        opening_balance = data.get('opening_balance')
        summary = data.get('transaction_summary', {})
        
        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        safe_name = customer.name[:15].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
        ws.title = f"{safe_name}往来账"
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        sub_header_fill_cash = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        sub_header_fill_gold = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        sub_header_font = Font(bold=True, size=9)
        summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        summary_font = Font(bold=True, size=10)
        
        # ===== 表头区域 =====
        # 第1行：公司名称
        ws.append(["深圳市梵贝琳珠宝有限公司"])
        ws.merge_cells('A1:K1')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 第2行：报表标题
        ws.append(["客户往来账明细表（三栏式）"])
        ws.merge_cells('A2:K2')
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # 第3-5行：信息
        date_range_text = "统计日期："
        if date_start and date_end:
            date_range_text += f"{date_start} 至 {date_end}"
        elif date_start:
            date_range_text += f"{date_start} 起"
        elif date_end:
            date_range_text += f"截至 {date_end}"
        else:
            date_range_text += "全部"
        ws.append([date_range_text])
        ws.merge_cells('A3:K3')
        ws['A3'].alignment = Alignment(horizontal='center')
        
        ws.append([f"往来客户：{customer.name}（{customer.customer_no or '-'}）"])
        ws.merge_cells('A4:K4')
        ws['A4'].alignment = Alignment(horizontal='center')
        
        ws.append([f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.merge_cells('A5:K5')
        ws['A5'].alignment = Alignment(horizontal='center')
        
        ws.append([])  # 空行
        
        # ===== 第7行：分组表头 =====
        row7 = ["序号", "日期", "类型", "单号", "现金借方(应收)", "现金贷方(已收)", "现金余额", "金料借方(克)", "金料贷方(克)", "金料余额", "备注"]
        ws.append(row7)
        for col_idx in range(1, 12):
            cell = ws.cell(row=7, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # ===== 数据行 =====
        current_row = 8
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        def format_cash_balance(val):
            if val is None:
                return ""
            return f"{abs(val):.2f}{'借' if val >= 0 else '贷'}"
        
        def format_gold_balance(val):
            if val is None:
                return ""
            if val == 0:
                return "0.000"
            return f"{abs(val):.3f}{'欠' if val > 0 else '存'}"
        
        # 期初余额行
        if opening_balance:
            date_str = str(opening_balance.get('created_at', ''))[:10]
            row_data = [
                "-",
                date_str,
                "期初余额",
                "-",
                opening_balance.get('cash_debit') or "",
                opening_balance.get('cash_credit') or "",
                format_cash_balance(opening_balance.get('running_cash_balance')),
                opening_balance.get('gold_debit') or "",
                opening_balance.get('gold_credit') or "",
                format_gold_balance(opening_balance.get('running_gold_balance')),
                opening_balance.get('description', '')
            ]
            ws.append(row_data)
            for col_idx in range(1, 12):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.font = Font(bold=True)
            current_row += 1
        
        # 交易记录
        for idx, tx in enumerate(transactions, 1):
            date_str = str(tx.get('created_at', ''))[:10]
            
            cash_balance = tx.get('running_cash_balance')
            gold_balance = tx.get('running_gold_balance')
            
            cash_bal_str = format_cash_balance(cash_balance)
            gold_bal_str = format_gold_balance(gold_balance)
            
            row_data = [
                idx,
                date_str,
                tx.get('type_label', ''),
                tx.get('order_no', ''),
                tx.get('cash_debit') or "",
                tx.get('cash_credit') or "",
                cash_bal_str,
                tx.get('gold_debit') or "",
                tx.get('gold_credit') or "",
                gold_bal_str,
                tx.get('remark') or tx.get('description') or ""
            ]
            ws.append(row_data)
            for col_idx in range(1, 12):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                if col_idx in [5, 8]:  # 借方列红色
                    cell.font = Font(color="CC0000")
                elif col_idx in [6, 9]:  # 贷方列绿色
                    cell.font = Font(color="008000")
                elif col_idx == 7:  # 现金余额列
                    cell.font = Font(bold=True, color="CC0000" if (cash_balance or 0) >= 0 else "008000")
                elif col_idx == 10:  # 金料余额列
                    cell.font = Font(bold=True, color="CC0000" if (gold_balance or 0) > 0 else "008000")
            current_row += 1
        
        # ===== 汇总行 =====
        ws.append([])
        current_row += 1
        
        summary_data = [
            "",
            "",
            "本期合计",
            "",
            summary.get('total_cash_debit', 0),
            summary.get('total_cash_credit', 0),
            "",
            summary.get('total_gold_debit', 0),
            summary.get('total_gold_credit', 0),
            "",
            ""
        ]
        ws.append(summary_data)
        for col_idx in range(1, 12):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.fill = summary_fill
            cell.font = summary_font
        current_row += 1
        
        # 期末余额行
        net_cash = summary.get('net_cash', 0)
        net_gold = summary.get('net_gold', 0)
        cash_label = f"{'应收' if net_cash >= 0 else '预收'} ¥{abs(net_cash):.2f}"
        gold_label = ""
        if net_gold != 0:
            gold_label = f"{'欠料' if net_gold > 0 else '存料'} {abs(net_gold):.3f}g"
        
        balance_data = ["", "", "期末余额", "", "", "", cash_label, "", "", gold_label, ""]
        ws.append(balance_data)
        for col_idx in range(1, 12):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            cell.font = Font(bold=True, size=11)
        
        # 列宽
        col_widths = [6, 12, 10, 20, 14, 14, 16, 12, 12, 14, 25]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        # K列
        ws.column_dimensions['K'].width = 25
        
        filename = f"{customer.name}_往来账目_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出客户往来账目失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")


@router.get("/customers")
@_limiter.limit("10/minute")
async def export_customers(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出客户列表为 Excel"""
    try:
        customers = db.query(Customer).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "客户列表"
        
        headers = ["ID", "客户编号", "客户名称", "电话", "地址", "创建时间"]
        ws.append(headers)
        style_header(ws)
        
        for cust in customers:
            ws.append([
                cust.id,
                cust.customer_no,
                cust.name,
                cust.phone or "",
                cust.address or "",
                cust.create_time.strftime("%Y-%m-%d %H:%M:%S") if cust.create_time else ""
            ])
        
        auto_column_width(ws)
        
        filename = f"客户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出客户列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/suppliers")
@_limiter.limit("10/minute")
async def export_suppliers(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出供应商列表为 Excel"""
    try:
        suppliers = db.query(Supplier).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商列表"
        
        headers = ["ID", "供应商编号", "供应商名称", "联系人", "电话", "地址", "总供货金额(元)", "总供货重量(克)", "供货次数", "最后供货时间", "状态"]
        ws.append(headers)
        style_header(ws)
        
        for sup in suppliers:
            ws.append([
                sup.id,
                sup.supplier_no,
                sup.name,
                sup.contact_person or "",
                sup.phone or "",
                sup.address or "",
                sup.total_supply_amount or 0,
                sup.total_supply_weight or 0,
                sup.total_supply_count or 0,
                sup.last_supply_time.strftime("%Y-%m-%d %H:%M:%S") if sup.last_supply_time else "",
                "活跃" if sup.status == "active" else "停用"
            ])
        
        auto_column_width(ws)
        
        filename = f"供应商列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except Exception as e:
        logger.error(f"导出供应商列表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/customer-sales/{customer_id}")
@_limiter.limit("10/minute")
async def export_customer_sales(
    request: Request,
    customer_id: int,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出指定客户的销售记录为 Excel"""
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 查询该客户的销售订单（优先 customer_id，兜底 customer_name）
        _exp_has_id = db.query(SalesOrder.id).filter(
            SalesOrder.customer_id == customer_id, SalesOrder.status != "已取消"
        ).first()
        _exp_filter = (SalesOrder.customer_id == customer_id) if _exp_has_id else (SalesOrder.customer_name == customer.name)
        sales_orders = db.query(SalesOrder).filter(
            _exp_filter,
            SalesOrder.status != "已取消"
        ).order_by(SalesOrder.create_time.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "销售记录"
        
        headers = ["销售单号", "产品名称", "克重(g)", "工费单价(元/克)", "件数", "件工费(元/件)", "总工费(元)", "业务员", "日期"]
        ws.append(headers)
        style_header(ws)
        
        total_weight = 0.0
        total_labor = 0.0
        
        for order in sales_orders:
            details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
            for detail in details:
                ws.append([
                    order.order_no,
                    detail.product_name,
                    detail.weight or 0,
                    detail.labor_cost or 0,
                    detail.piece_count or "",
                    detail.piece_labor_cost or "",
                    detail.total_labor_cost or 0,
                    order.salesperson or "",
                    order.create_time.strftime("%Y-%m-%d %H:%M:%S") if order.create_time else ""
                ])
                total_weight += float(detail.weight or 0)
                total_labor += float(detail.total_labor_cost or 0)
        
        # 添加合计行
        ws.append([])
        ws.append(["合计", "", round(total_weight, 2), "", "", "", round(total_labor, 2), "", ""])
        summary_row = ws.max_row
        for cell in ws[summary_row]:
            cell.font = Font(bold=True)
        
        auto_column_width(ws)
        
        filename = f"{customer.name}_销售记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出客户销售记录失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/customer-returns/{customer_id}")
@_limiter.limit("10/minute")
async def export_customer_returns(
    request: Request,
    customer_id: int,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出指定客户的退货记录为 Excel"""
    from ..models import ReturnOrder
    
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 查询与该客户关联的销售单（优先 customer_id，兜底 customer_name）
        _ret_has_id = db.query(SalesOrder.id).filter(
            SalesOrder.customer_id == customer_id, SalesOrder.status != "已取消"
        ).first()
        _ret_filter = (SalesOrder.customer_id == customer_id) if _ret_has_id else (SalesOrder.customer_name == customer.name)
        sales_orders = db.query(SalesOrder).filter(
            _ret_filter,
            SalesOrder.status != "已取消"
        ).all()
        
        returns_list = []
        for order in sales_orders:
            try:
                related_returns = db.query(ReturnOrder).filter(
                    ReturnOrder.remark.contains(order.order_no) if hasattr(ReturnOrder, 'remark') else False
                ).all()
                for ret in related_returns:
                    returns_list.append(ret)
            except Exception:
                pass
        
        wb = Workbook()
        ws = wb.active
        ws.title = "退货记录"
        
        headers = ["退货单号", "产品名称", "退货克重(g)", "退货原因", "状态", "日期"]
        ws.append(headers)
        style_header(ws)
        
        status_map = {"draft": "草稿", "confirmed": "已确认", "cancelled": "已取消"}
        total_weight = 0.0
        
        for ret in returns_list:
            ws.append([
                ret.return_no,
                ret.product_name,
                ret.return_weight or 0,
                ret.return_reason or "",
                status_map.get(ret.status, ret.status or ""),
                ret.created_at.strftime("%Y-%m-%d %H:%M:%S") if ret.created_at else ""
            ])
            total_weight += (ret.return_weight or 0)
        
        # 添加合计行
        if returns_list:
            ws.append([])
            ws.append(["合计", "", round(total_weight, 2), "", "", ""])
            summary_row = ws.max_row
            for cell in ws[summary_row]:
                cell.font = Font(bold=True)
        
        auto_column_width(ws)
        
        filename = f"{customer.name}_退货记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出客户退货记录失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/all")
@_limiter.limit("10/minute")
async def export_all_data(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """一键导出全部数据为 ZIP 包"""
    try:
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 导出对话日志
            wb = Workbook()
            ws = wb.active
            ws.title = "对话日志"
            ws.append(["ID", "会话ID", "用户角色", "消息类型", "内容", "意图", "响应时间(ms)", "是否成功", "创建时间"])
            style_header(ws)
            role_names = {"sales": "业务员", "finance": "财务", "product": "商品专员", "manager": "管理层"}
            for log in db.query(ChatLog).order_by(ChatLog.created_at.desc()).all():
                ws.append([
                    log.id, log.session_id, role_names.get(log.user_role, log.user_role),
                    "用户" if log.message_type == "user" else "AI助手",
                    log.content[:500] if log.content else "", log.intent or "", log.response_time_ms,
                    "是" if log.is_successful else "否",
                    log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
                ])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("对话日志.xlsx", excel_buffer.getvalue())
            
            # 2. 导出库存数据
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"
            ws.append(["ID", "商品名称", "库存重量(克)", "最后更新时间"])
            style_header(ws)
            for inv in db.query(Inventory).all():
                ws.append([inv.id, inv.product_name, inv.total_weight,
                          inv.last_update.strftime("%Y-%m-%d %H:%M:%S") if inv.last_update else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("库存数据.xlsx", excel_buffer.getvalue())
            
            # 3. 导出入库记录
            wb = Workbook()
            ws = wb.active
            ws.title = "入库记录"
            ws.append(["入库单号", "商品名称", "重量(克)", "工费(元/克)", "总成本(元)", "供应商", "入库时间", "操作员"])
            style_header(ws)
            for order in db.query(InboundOrder).all():
                for detail in db.query(InboundDetail).filter(InboundDetail.order_id == order.id).all():
                    ws.append([order.order_no, detail.product_name, detail.weight, detail.labor_cost,
                              detail.total_cost, detail.supplier or "",
                              order.create_time.strftime("%Y-%m-%d %H:%M:%S") if order.create_time else "",
                              order.operator or ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("入库记录.xlsx", excel_buffer.getvalue())
            
            # 4. 导出销售订单
            wb = Workbook()
            ws = wb.active
            ws.title = "销售订单"
            ws.append(["订单号", "客户名称", "商品名称", "重量(克)", "工费(元/克)", "总工费(元)", "业务员", "门店代码", "订单日期"])
            style_header(ws)
            for order in db.query(SalesOrder).all():
                customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
                for detail in db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all():
                    ws.append([order.order_no, customer.name if customer else "", detail.product_name,
                              detail.weight, detail.labor_cost, detail.total_labor_cost,
                              order.salesperson or "", order.store_code or "",
                              order.order_date.strftime("%Y-%m-%d") if order.order_date else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("销售订单.xlsx", excel_buffer.getvalue())
            
            # 5. 导出客户列表
            wb = Workbook()
            ws = wb.active
            ws.title = "客户列表"
            ws.append(["ID", "客户编号", "客户名称", "电话", "地址", "创建时间"])
            style_header(ws)
            for cust in db.query(Customer).all():
                ws.append([cust.id, cust.customer_no, cust.name, cust.phone or "", cust.address or "",
                          cust.create_time.strftime("%Y-%m-%d %H:%M:%S") if cust.create_time else ""])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("客户列表.xlsx", excel_buffer.getvalue())
            
            # 6. 导出供应商列表
            wb = Workbook()
            ws = wb.active
            ws.title = "供应商列表"
            ws.append(["ID", "供应商编号", "供应商名称", "联系人", "电话", "地址", "总供货金额(元)", "总供货重量(克)", "供货次数", "最后供货时间", "状态"])
            style_header(ws)
            for sup in db.query(Supplier).all():
                ws.append([sup.id, sup.supplier_no, sup.name, sup.contact_person or "", sup.phone or "",
                          sup.address or "", sup.total_supply_amount or 0, sup.total_supply_weight or 0,
                          sup.total_supply_count or 0,
                          sup.last_supply_time.strftime("%Y-%m-%d %H:%M:%S") if sup.last_supply_time else "",
                          "活跃" if sup.status == "active" else "停用"])
            auto_column_width(ws)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            zip_file.writestr("供应商列表.xlsx", excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        
        filename = f"珠宝ERP数据备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
            }
        )
        
    except Exception as e:
        logger.error(f"导出全部数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats")
@_limiter.limit("10/minute")
async def get_export_stats(request: Request, db: Session = Depends(get_db)):
    """获取可导出数据的统计信息"""
    try:
        return {
            "success": True,
            "data": {
                "chat_logs": db.query(func.count(ChatLog.id)).scalar() or 0,
                "inventory": db.query(func.count(Inventory.id)).scalar() or 0,
                "inbound_orders": db.query(func.count(InboundOrder.id)).scalar() or 0,
                "sales_orders": db.query(func.count(SalesOrder.id)).scalar() or 0,
                "customers": db.query(func.count(Customer.id)).scalar() or 0,
                "suppliers": db.query(func.count(Supplier.id)).scalar() or 0,
            }
        }
    except Exception as e:
        logger.error(f"获取导出统计失败: {e}", exc_info=True)
        return {"success": False, "message": str(e)}


@router.get("/customer-account-summary")
@_limiter.limit("10/minute")
async def export_customer_account_summary(
    request: Request,
    role: str = Depends(require_permission("can_export")),
    db: Session = Depends(get_db)
):
    """导出客户实时账目表为 Excel"""
    try:
        from ..models.finance import AccountReceivable
        from ..models import (
            SettlementOrder, GoldReceipt, CustomerWithdrawal
        )
        from sqlalchemy import desc as sql_desc
        
        # 查询所有活跃客户
        customers = db.query(Customer).filter(Customer.status == "active").all()
        customer_ids = [c.id for c in customers]
        customer_names = [c.name for c in customers]
        
        # 1. 批量查询现金欠款
        cash_debt_map = {}
        if customer_ids:
            cash_results = db.query(
                AccountReceivable.customer_id,
                func.sum(AccountReceivable.unpaid_amount).label('total_debt')
            ).filter(
                AccountReceivable.customer_id.in_(customer_ids),
                AccountReceivable.status.in_(["unpaid", "overdue"])
            ).group_by(AccountReceivable.customer_id).all()
            for row in cash_results:
                cash_debt_map[row.customer_id] = float(row.total_debt or 0)
            
            # 加入期初现金余额（数据迁移写入的 initial_balance 交易）
            from ..models import CustomerTransaction
            initial_cash_results = db.query(
                CustomerTransaction.customer_id,
                func.sum(CustomerTransaction.amount).label('initial_amount')
            ).filter(
                CustomerTransaction.customer_id.in_(customer_ids),
                CustomerTransaction.transaction_type == 'initial_balance',
                CustomerTransaction.status == 'active'
            ).group_by(CustomerTransaction.customer_id).all()
            for row in initial_cash_results:
                cid = row.customer_id
                cash_debt_map[cid] = cash_debt_map.get(cid, 0.0) + float(row.initial_amount or 0)
        
        # 2. 批量查询金料净值（使用统一计算函数，保证全系统口径一致）
        net_gold_map = {}
        if customer_ids:
            from ..gold_balance import calculate_batch_net_gold
            batch_net_gold = calculate_batch_net_gold(customer_ids, db)
            # 统一函数返回正数=存料，导出需要正数=欠料，取反
            for cid, val in batch_net_gold.items():
                net_gold_map[cid] = -val
        
        # 3. 批量查询业务员
        salesperson_map = {}
        if customer_names:
            salesperson_results = db.query(
                SalesOrder.customer_name,
                SalesOrder.salesperson,
                func.count(SalesOrder.id).label('order_count')
            ).filter(
                SalesOrder.customer_name.in_(customer_names),
                SalesOrder.status != "已取消",
                SalesOrder.salesperson.isnot(None)
            ).group_by(
                SalesOrder.customer_name,
                SalesOrder.salesperson
            ).order_by(
                SalesOrder.customer_name,
                sql_desc('order_count')
            ).all()
            for row in salesperson_results:
                if row.customer_name not in salesperson_map:
                    salesperson_map[row.customer_name] = row.salesperson
        
        # 构建数据
        wb = Workbook()
        ws = wb.active
        ws.title = "客户实时账目表"
        
        headers = ["序号", "客户名称", "欠料重量(克)", "欠款金额(元)", "业务员"]
        ws.append(headers)
        style_header(ws)
        
        total_gold_debt = 0.0
        total_cash_debt = 0.0
        row_num = 0
        
        # 按欠款金额降序排列
        customer_data = []
        for customer in customers:
            cash_debt = cash_debt_map.get(customer.id, 0.0)
            net_gold = net_gold_map.get(customer.id, 0.0)
            if abs(cash_debt) < 0.01 and abs(net_gold) < 0.001:
                continue
            customer_data.append({
                "name": customer.name,
                "net_gold": net_gold,
                "cash_debt": cash_debt,
                "salesperson": salesperson_map.get(customer.name, "")
            })
        
        customer_data.sort(key=lambda x: x["cash_debt"], reverse=True)
        
        for item in customer_data:
            row_num += 1
            ws.append([
                row_num,
                item["name"],
                round(item["net_gold"], 3),
                round(item["cash_debt"], 2),
                item["salesperson"]
            ])
            total_gold_debt += item["net_gold"]
            total_cash_debt += item["cash_debt"]
        
        # 合计行
        sum_row = ["合计", f"共 {row_num} 位客户", round(total_gold_debt, 3), round(total_cash_debt, 2), ""]
        ws.append(sum_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        
        auto_column_width(ws)
        
        filename = f"客户实时账目表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
    except Exception as e:
        logger.error(f"导出客户实时账目表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/supplier-debt-summary")
@_limiter.limit("10/minute")
async def export_supplier_debt_summary(
    request: Request,
    role: str = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出供应商款料汇总为 Excel"""
    try:
        # 该报表属于供应商金料账户场景：
        # 管理层/财务可通过 can_export 导出，料部可通过 can_view_supplier_gold_account 导出。
        if not (
            has_permission(role, "can_export")
            or has_permission(role, "can_view_supplier_gold_account")
        ):
            raise HTTPException(status_code=403, detail="权限不足：您没有【数据导出】的权限")

        from ..models.finance import AccountPayable
        from ..models import InboundDetail, InboundOrder, GoldMaterialTransaction
        
        # 查询入库重量
        inbound_weights = db.query(
            InboundDetail.supplier_id,
            func.sum(InboundDetail.weight).label('total_weight')
        ).filter(InboundDetail.supplier_id.isnot(None)).group_by(InboundDetail.supplier_id).all()
        
        # 查询已付料重量
        paid_weights = db.query(
            GoldMaterialTransaction.supplier_id,
            func.sum(GoldMaterialTransaction.gold_weight).label('total_weight')
        ).filter(
            GoldMaterialTransaction.transaction_type == 'expense',
            GoldMaterialTransaction.status == 'confirmed',
            GoldMaterialTransaction.supplier_id.isnot(None)
        ).group_by(GoldMaterialTransaction.supplier_id).all()
        
        # 查询工费欠款
        labor_debts = db.query(
            AccountPayable.supplier_id,
            func.sum(AccountPayable.unpaid_amount).label('total_unpaid')
        ).filter(
            AccountPayable.status.in_(["unpaid", "partial"])
        ).group_by(AccountPayable.supplier_id).all()
        
        inbound_map = {row.supplier_id: float(row.total_weight or 0) for row in inbound_weights}
        paid_map = {row.supplier_id: float(row.total_weight or 0) for row in paid_weights}
        labor_debt_map = {row.supplier_id: float(row.total_unpaid or 0) for row in labor_debts}
        
        suppliers = db.query(Supplier).filter(Supplier.status == "active").all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "供应商款料汇总"
        
        headers = ["供应商名称", "供应商编号", "入库重量(克)", "已付料(克)", "欠料(克)", "工费欠款(元)", "状态"]
        ws.append(headers)
        style_header(ws)
        
        total_inbound = 0
        total_paid = 0
        total_debt = 0
        total_labor = 0
        
        for supplier in suppliers:
            inbound_weight = inbound_map.get(supplier.id, 0.0)
            paid_weight = paid_map.get(supplier.id, 0.0)
            debt_weight = inbound_weight - paid_weight
            labor_debt = labor_debt_map.get(supplier.id, 0.0)
            
            if inbound_weight > 0 or paid_weight > 0 or labor_debt > 0:
                status = "欠款" if (debt_weight > 0 or labor_debt > 0) else ("多付" if debt_weight < 0 else "结清")
                ws.append([
                    supplier.name,
                    supplier.supplier_no,
                    round(inbound_weight, 2),
                    round(paid_weight, 2),
                    round(debt_weight, 2),
                    round(labor_debt, 2),
                    status
                ])
                total_inbound += inbound_weight
                total_paid += paid_weight
                total_debt += debt_weight
                total_labor += labor_debt
        
        # 汇总行
        sum_row = ["合计", "", round(total_inbound, 2), round(total_paid, 2), round(total_debt, 2), round(total_labor, 2), ""]
        ws.append(sum_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        
        auto_column_width(ws)
        
        filename = f"供应商款料汇总_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return create_excel_response(wb, filename)
    except Exception as e:
        logger.error(f"导出供应商款料汇总失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

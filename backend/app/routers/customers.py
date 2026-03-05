"""
客户管理路由
"""
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, selectinload, joinedload
from sqlalchemy import desc, func
from datetime import datetime
from ..timezone_utils import china_now
from typing import Optional, List, Dict, Any
import logging
import io
import csv
import time

from ..database import get_db
from ..utils.response import (
    success_response, error_response, paginated_response,
    not_found_response, conflict_response, server_error_response, ErrorCode
)
from ..models import (
    Customer, SalesOrder, SalesDetail, ReturnOrder,
    AccountReceivable, CustomerTransaction, CustomerGoldDeposit,
    CustomerGoldDepositTransaction, SettlementOrder, CustomerWithdrawal,
    SalesReturnOrder, SalesReturnSettlement
)
from ..models.finance import PaymentRecord, GoldReceipt
from ..schemas import CustomerCreate, CustomerResponse
from ..dependencies.auth import get_current_role, require_permission

logger = logging.getLogger(__name__)

import re
_SUPPLIER_PATTERN = re.compile(r'\s*\[[实實]出供[应應]商[^\]]*\]', re.UNICODE)

def _sanitize_remark(remark: str) -> str:
    """过滤备注中的供应商敏感信息"""
    if not remark:
        return ""
    return _SUPPLIER_PATTERN.sub('', remark).strip()

router = APIRouter(prefix="/api/customers", tags=["客户管理"])


@router.post("")
async def create_customer(
    customer_data: CustomerCreate,
    user_role: str = Query(default="sales", description="用户角色"),
    role: str = Depends(require_permission("can_manage_customers")),
    db: Session = Depends(get_db)
):
    """创建客户"""
    # 权限检查 - 需要 can_manage_customers 权限
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_manage_customers'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【客户管理】的权限（创建/编辑/删除）")
    
    try:
        # 检查客户是否已存在
        existing = db.query(Customer).filter(
            Customer.name == customer_data.name,
            Customer.status == "active"
        ).first()
        
        if existing:
            return conflict_response(
                message=f"客户 {customer_data.name} 已存在",
            )
        
        # 生成客户编号
        customer_no = f"KH{china_now().strftime('%Y%m%d%H%M%S')}"
        
        customer = Customer(
            customer_no=customer_no,
            **customer_data.model_dump()
        )
        db.add(customer)
        db.commit()
        db.refresh(customer)
        
        return success_response(
            data={"customer": CustomerResponse.model_validate(customer).model_dump(mode='json')},
            message=f"客户创建成功：{customer.name}",
            code=ErrorCode.CREATED
        )
    except Exception as e:
        db.rollback()
        logger.error(f"创建客户失败: {e}", exc_info=True)
        return server_error_response(message=f"创建客户失败: {str(e)}")


@router.get("")
async def get_customers(
    name: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(default=1, ge=1, description="页码"),
    page_size: int = Query(default=20, ge=1, le=10000, description="每页数量"),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户列表（分页）"""
    # 权限检查 - 需要 can_view_customers 或 can_manage_customers 权限
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_view_customers') and not has_permission(user_role, 'can_manage_customers'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看客户】的权限")
    
    try:
        query = db.query(Customer).filter(Customer.status == "active")
        
        # 模糊搜索：支持姓名、电话、客户编号
        keyword = search or name
        if keyword:
            from sqlalchemy import or_
            query = query.filter(
                or_(
                    Customer.name.contains(keyword),
                    Customer.phone.contains(keyword),
                    Customer.customer_no.contains(keyword)
                )
            )
        
        # 获取总数
        total = query.count()
        
        # 分页查询
        # NOTE: CustomerResponse does not serialize relationship data (sales_orders, gold_deposit).
        # If the response schema expands to include relationships, add eager loading here:
        #   .options(selectinload(Customer.sales_orders), joinedload(Customer.gold_deposit))
        offset = (page - 1) * page_size
        if keyword:
            from sqlalchemy import case
            exact_match = case(
                (Customer.name == keyword, 0),
                else_=1
            )
            customers = query.order_by(exact_match, desc(Customer.create_time)).offset(offset).limit(page_size).all()
        else:
            customers = query.order_by(desc(Customer.create_time)).offset(offset).limit(page_size).all()
        
        return success_response(
            data={
                "customers": [CustomerResponse.model_validate(c).model_dump(mode='json') for c in customers],
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size
            },
            message="查询成功"
        )
    except Exception as e:
        logger.error(f"查询客户失败: {e}", exc_info=True)
        return server_error_response(message=f"查询客户失败: {str(e)}")


@router.get("/suggest-salesperson")
async def suggest_salesperson(customer_name: str, db: Session = Depends(get_db)):
    """根据客户名智能推荐业务员（基于历史销售记录）"""
    try:
        if not customer_name or not customer_name.strip():
            return success_response(
                data={"salesperson": None, "hint": "请输入客户名", "is_new_customer": None}
            )
        
        customer_name = customer_name.strip()
        
        # 查找该客户最近一次的销售单
        latest_order = db.query(SalesOrder).filter(
            SalesOrder.customer_name == customer_name,
            SalesOrder.status != "已取消"
        ).order_by(SalesOrder.create_time.desc()).first()
        
        if latest_order and latest_order.salesperson:
            last_date = latest_order.create_time.strftime('%Y-%m-%d') if latest_order.create_time else "未知"
            return success_response(
                data={
                    "salesperson": latest_order.salesperson,
                    "hint": f"已自动匹配业务员（上次服务：{last_date}）",
                    "is_new_customer": False
                }
            )
        
        # 如果没有历史记录，返回空
        return success_response(
            data={
                "salesperson": None,
                "hint": "新客户，请手动输入业务员",
                "is_new_customer": True
            }
        )
    
    except Exception as e:
        logger.error(f"查询业务员推荐失败: {e}", exc_info=True)
        return server_error_response(message=f"查询业务员推荐失败: {str(e)}")


# ============= 客户欠款查询 API =============
# 注意：此路由必须在 /{customer_id} 之前定义，否则会被错误匹配

@router.get("/debt-summary")
async def get_customer_debt_summary(
    search: Optional[str] = Query(None, description="搜索客户名称"),
    sort_by: str = Query("total_debt", description="排序字段: cash_debt/gold_debt/total_debt/name"),
    sort_order: str = Query("desc", description="排序方向: asc/desc"),
    hide_zero: bool = Query(True, description="隐藏无欠款客户"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    获取客户欠款汇总列表
    
    返回所有客户的现金欠款和金料欠款情况，支持搜索和排序。
    业务员和结算专员都可以查询所有客户。
    """
    from ..middleware.permissions import has_permission
    # 业务员和结算专员都可以查询
    can_view = (
        has_permission(user_role, 'can_view_customers') or 
        has_permission(user_role, 'can_query_customer_sales') or
        has_permission(user_role, 'can_create_settlement')
    )
    if not can_view:
        raise HTTPException(status_code=403, detail="权限不足：您没有查看客户欠款的权限")
    
    try:
        # 查询所有活跃客户
        query = db.query(Customer).filter(Customer.status == "active")
        
        if search:
            query = query.filter(Customer.name.contains(search))
        
        customers = query.all()
        customer_ids = [c.id for c in customers]
        customer_names = [c.name for c in customers]
        
        # ========== 批量查询优化：4次查询代替 N*4 次查询 ==========
        
        # 1. 批量查询现金欠款（按 customer_id 聚合）
        cash_debt_map = {}
        if customer_ids:
            cash_results = db.query(
                AccountReceivable.customer_id,
                func.sum(AccountReceivable.unpaid_amount).label('total_debt')
            ).filter(
                AccountReceivable.customer_id.in_(customer_ids),
                AccountReceivable.status.in_(["unpaid", "overdue"])
            ).group_by(AccountReceivable.customer_id).all()
            
            for row in cash_results:
                cash_debt_map[row.customer_id] = float(row.total_debt or 0)
            
            # 加入期初现金余额（数据迁移写入的 initial_balance 交易）
            initial_cash_results = db.query(
                CustomerTransaction.customer_id,
                func.sum(CustomerTransaction.amount).label('initial_amount')
            ).filter(
                CustomerTransaction.customer_id.in_(customer_ids),
                CustomerTransaction.transaction_type == 'initial_balance',
                CustomerTransaction.status == 'active'
            ).group_by(CustomerTransaction.customer_id).all()
            
            for row in initial_cash_results:
                cid = row.customer_id
                cash_debt_map[cid] = cash_debt_map.get(cid, 0.0) + float(row.initial_amount or 0)
        
        # 2. 批量查询金料账户净值（使用统一计算函数，正数=存料，负数=欠料）
        net_gold_map = {}
        if customer_ids:
            try:
                from ..gold_balance import calculate_batch_net_gold
                # 返回 {cid: net_gold}，正数=存料，负数=欠料
                batch_net_gold = calculate_batch_net_gold(customer_ids, db)
                # 取反：统一函数返回正数=存料，这里 net_gold_map 需要正数=欠料
                for cid, val in batch_net_gold.items():
                    net_gold_map[cid] = -val  # 取反：正数变欠料
            except Exception as e:
                logger.warning(f"批量查询金料账户时出错: {e}")
        
        # 4. 批量查询最后交易时间（按客户名称）
        last_tx_map = {}
        if customer_names:
            # 使用子查询获取每个客户名称的最后交易时间
            last_orders = db.query(
                SalesOrder.customer_name,
                func.max(SalesOrder.create_time).label('last_time')
            ).filter(
                SalesOrder.customer_name.in_(customer_names)
            ).group_by(SalesOrder.customer_name).all()
            
            for row in last_orders:
                if row.last_time:
                    last_tx_map[row.customer_name] = row.last_time.strftime("%Y-%m-%d")
        
        # 5. 批量查询每个客户最常用的业务员
        salesperson_map = {}
        if customer_names:
            # 查询每个客户最常用的业务员（按销售次数统计）
            salesperson_results = db.query(
                SalesOrder.customer_name,
                SalesOrder.salesperson,
                func.count(SalesOrder.id).label('order_count')
            ).filter(
                SalesOrder.customer_name.in_(customer_names),
                SalesOrder.status != "已取消",
                SalesOrder.salesperson.isnot(None)
            ).group_by(
                SalesOrder.customer_name,
                SalesOrder.salesperson
            ).order_by(
                SalesOrder.customer_name,
                desc('order_count')
            ).all()
            
            # 为每个客户选择订单数最多的业务员
            for row in salesperson_results:
                if row.customer_name not in salesperson_map:
                    salesperson_map[row.customer_name] = row.salesperson
        
        # ========== 构建欠款数据 ==========
        # 单一账户模式：net_gold 正数=欠料，负数=存料
        debt_list = []
        total_cash_debt = 0.0
        total_net_gold = 0.0
        
        for customer in customers:
            cash_debt = cash_debt_map.get(customer.id, 0.0)
            net_gold = net_gold_map.get(customer.id, 0.0)  # 直接获取净金料值
            last_transaction_date = last_tx_map.get(customer.name)
            
            # 单一账户模式：net_gold 就是最终值（正数=欠料）
            gold_balance = net_gold
            
            # 如果隐藏无欠款客户（现金欠款为0且金料净额为0）
            if hide_zero and abs(cash_debt) < 0.01 and abs(gold_balance) < 0.001:
                continue
            
            debt_list.append({
                "customer_id": customer.id,
                "customer_no": customer.customer_no,
                "customer_name": customer.name,
                "phone": customer.phone,
                "salesperson": salesperson_map.get(customer.name, ""),  # 业务员
                "cash_debt": round(cash_debt, 2),
                "net_gold": round(net_gold, 3),  # 单一账户净值（正=欠料，负=存料）
                "gold_balance": round(gold_balance, 3),  # 兼容字段（正=欠料）
                # 保留原字段用于兼容
                "gold_debt": round(max(0, net_gold), 3),     # 欠料（正值部分）
                "gold_deposit": round(max(0, -net_gold), 3),  # 存料（负值绝对值）
                "total_debt": round(cash_debt, 2),
                "last_transaction_date": last_transaction_date
            })
            
            total_cash_debt += cash_debt
            total_net_gold += net_gold
        
        # 排序
        reverse = sort_order == "desc"
        if sort_by == "cash_debt":
            debt_list.sort(key=lambda x: x["cash_debt"], reverse=reverse)
        elif sort_by == "gold_debt" or sort_by == "gold_balance":
            debt_list.sort(key=lambda x: x["gold_balance"], reverse=reverse)
        elif sort_by == "name":
            debt_list.sort(key=lambda x: x["customer_name"], reverse=reverse)
        else:  # total_debt
            debt_list.sort(key=lambda x: (x["cash_debt"], x["gold_balance"]), reverse=reverse)
        
        # 分页
        total = len(debt_list)
        debt_list = debt_list[skip:skip + limit]
        
        return success_response(
            data={
                "items": debt_list,
                "total": total,
                "summary": {
                    "total_cash_debt": round(total_cash_debt, 2),
                    "total_net_gold": round(total_net_gold, 3),
                    "total_gold_balance": round(total_net_gold, 3),
                    "customer_count": total
                }
            },
            message="查询成功"
        )
        
    except Exception as e:
        logger.error(f"查询客户欠款汇总失败: {e}", exc_info=True)
        return server_error_response(message=f"查询失败: {str(e)}")


@router.get("/chat-debt-query")
async def chat_debt_query(
    customer_name: str = Query(..., description="客户名称（支持模糊匹配）"),
    query_type: str = Query(default="all", description="查询类型：all/cash_debt/gold_debt/gold_deposit"),
    date_start: Optional[str] = Query(default=None, description="开始日期 YYYY-MM-DD"),
    date_end: Optional[str] = Query(default=None, description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """
    聊天查询客户账务（供AI聊天使用）
    返回客户的欠款、欠料、存料等财务信息
    """
    try:
        # 1. 优先精确匹配客户名
        customer = db.query(Customer).filter(
            Customer.name == customer_name,
            Customer.status == "active"
        ).first()
        
        if not customer:
            # 2. 精确匹配失败，尝试模糊匹配（优先匹配最短名称，避免错误匹配）
            candidates = db.query(Customer).filter(
                Customer.name.contains(customer_name),
                Customer.status == "active"
            ).all()
            if candidates:
                # 优先选择名称长度最接近搜索词的客户
                customer = min(candidates, key=lambda c: abs(len(c.name) - len(customer_name)))
        
        if not customer:
            return not_found_response(message=f"未找到客户：{customer_name}")
        
        customer_id = customer.id
        
        # 2. 解析日期范围
        start_date = None
        end_date = None
        if date_start:
            try:
                start_date = datetime.strptime(date_start, "%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        if date_end:
            try:
                end_date = datetime.strptime(date_end, "%Y-%m-%d")
                # 设置为当天结束时间
                end_date = end_date.replace(hour=23, minute=59, second=59)
            except (ValueError, TypeError):
                pass
        
        result = {
            "success": True,
            "customer": {
                "id": customer.id,
                "name": customer.name,
                "phone": customer.phone,
                "customer_no": customer.customer_no
            },
            "query_period": {
                "start": date_start,
                "end": date_end
            }
        }
        
        # 3. 查询现金欠款（使用历史交易汇总方式，与财务对账单一致）
        # Bug #30 fix: 始终查询 cash_debt，确保 AI 能看到完整账务数据
        logger.info(f"[诊断-chat_debt_query] query_type = '{query_type}'")
        # 不再根据 query_type 过滤，始终查询现金账户
        if True:  # 原条件: query_type in ["all", "cash_debt"]
            cash_debt = 0.0
            cash_transactions = []
            use_name_fallback = False
            try:
                # 优先用 customer_id 关联；若无记录则用 customer_name 兜底
                if customer.name:
                    has_id_sales = db.query(SalesOrder.id).filter(
                        SalesOrder.customer_id == customer_id,
                        SalesOrder.status != "已取消"
                    ).first()
                    if not has_id_sales:
                        use_name_fallback = True

                sales_orders_query = db.query(SalesOrder).filter(
                    SalesOrder.status != "已取消"
                )
                if use_name_fallback:
                    sales_orders_query = sales_orders_query.filter(
                        SalesOrder.customer_name == customer.name
                    )
                else:
                    sales_orders_query = sales_orders_query.filter(
                        SalesOrder.customer_id == customer_id
                    )

                if start_date:
                    sales_orders_query = sales_orders_query.filter(SalesOrder.create_time >= start_date)
                if end_date:
                    sales_orders_query = sales_orders_query.filter(SalesOrder.create_time <= end_date)

                sales_orders = sales_orders_query.all()
                total_sales_labor = sum(o.total_labor_cost or 0 for o in sales_orders)

                # 计算结算金额（cash_price + mixed的现金部分 + labor_amount）
                total_settlement_cash = 0.0
                settlements_query = db.query(SettlementOrder).join(SalesOrder).filter(
                    SalesOrder.customer_name == customer.name if use_name_fallback else SalesOrder.customer_id == customer_id,
                    SettlementOrder.status.in_(['confirmed', 'printed'])
                )
                
                if start_date:
                    settlements_query = settlements_query.filter(
                        SettlementOrder.created_at >= start_date
                    )
                if end_date:
                    settlements_query = settlements_query.filter(
                        SettlementOrder.created_at <= end_date
                    )
                
                settlements = settlements_query.all()
                for s in settlements:
                    if s.payment_method == 'cash_price':
                        total_settlement_cash += float(s.total_amount or 0)
                    elif s.payment_method == 'mixed':
                        if s.cash_payment_weight and s.gold_price:
                            total_settlement_cash += float(s.cash_payment_weight) * float(s.gold_price)
                        total_settlement_cash += float(s.labor_amount or 0)
                
                # 查询收款记录
                payments_query = db.query(PaymentRecord).filter(
                    PaymentRecord.customer_id == customer_id
                )
                
                if start_date:
                    payments_query = payments_query.filter(
                        PaymentRecord.payment_date >= start_date.date()
                    )
                if end_date:
                    payments_query = payments_query.filter(
                        PaymentRecord.payment_date <= end_date.date()
                    )
                
                total_payments = db.query(func.coalesce(func.sum(PaymentRecord.amount), 0)).filter(
                    PaymentRecord.customer_id == customer_id
                )
                if start_date:
                    total_payments = total_payments.filter(PaymentRecord.payment_date >= start_date.date())
                if end_date:
                    total_payments = total_payments.filter(PaymentRecord.payment_date <= end_date.date())
                
                total_payments = total_payments.scalar() or 0
                
                # 查询期初现金余额（数据迁移写入的 initial_balance 交易）
                initial_cash_q = db.query(func.coalesce(func.sum(CustomerTransaction.amount), 0)).filter(
                    CustomerTransaction.customer_id == customer_id,
                    CustomerTransaction.transaction_type == 'initial_balance',
                    CustomerTransaction.status == 'active'
                )
                initial_cash = float(initial_cash_q.scalar() or 0)
                
                # 净欠款 = 期初余额 + 销售工费 + 结算现金 - 收款
                cash_debt = initial_cash + float(total_sales_labor) + float(total_settlement_cash) - float(total_payments)
                
                # 获取交易明细用于展示
                payments = payments_query.order_by(desc(PaymentRecord.create_time)).limit(20).all()
                for p in payments:
                    payment_method_label = {
                        'bank_transfer': '银行转账',
                        'cash': '现金',
                        'wechat': '微信',
                        'alipay': '支付宝',
                        'card': '刷卡'
                    }.get(p.payment_method, p.payment_method)
                    
                    cash_transactions.append({
                        "id": p.id,
                        "type": "payment",
                        "description": f"{payment_method_label} 收款",
                        "amount": p.amount,
                        "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                        "created_at": p.create_time.isoformat() if p.create_time else None
                    })
                
                # 添加销售工费明细（按销售单汇总）
                for order in sales_orders[:20]:
                    cash_transactions.append({
                        "id": f"sales_{order.id}",
                        "type": "sales_labor",
                        "description": f"销售结算（销售单号: {order.order_no}）",
                        "amount": order.total_labor_cost or 0,
                        "created_at": order.create_time.isoformat() if order.create_time else None
                    })

                # 添加结算单明细
                for s in settlements[:20]:
                    cash_transactions.append({
                        "id": s.id,
                        "type": "settlement",
                        "description": f"销售结算（结算单号: {s.settlement_no}）",
                        "amount": s.total_amount if s.payment_method == 'cash_price' else (s.cash_payment_weight * s.gold_price if s.payment_method == 'mixed' and s.cash_payment_weight and s.gold_price else 0) + (s.labor_amount or 0),
                        "created_at": s.created_at.isoformat() if s.created_at else None
                    })
                
            except Exception as e:
                logger.warning(f"查询现金欠款出错: {e}", exc_info=True)
            
            result["cash_debt"] = cash_debt
            result["cash_transactions"] = cash_transactions
        
        # 4. 查询金料账户（使用统一计算函数，与全系统口径一致）
        # Bug #30 fix: 始终查询 net_gold，确保 AI 能看到完整账务数据
        if True:  # 原条件: query_type in ["all", "gold_debt", "gold_deposit"]
            net_gold = 0.0  # 净金料值（正数=欠料，负数=存料）
            gold_transactions = []
            deposit_transactions = []
            
            try:
                # 使用统一函数计算净金料（无日期筛选时直接用；有日期筛选时仍需逐项计算）
                if not start_date and not end_date:
                    from ..gold_balance import calculate_customer_net_gold
                    # 统一函数返回正数=存料，这里取反得到正数=欠料
                    net_gold = -calculate_customer_net_gold(customer_id, db)
                else:
                    # 有日期筛选时，逐项计算（与统一函数公式一致）
                    total_settlement_gold = 0.0
                    settlements_date_q = db.query(SettlementOrder).join(SalesOrder).filter(
                        SalesOrder.customer_name == customer.name if use_name_fallback else SalesOrder.customer_id == customer_id,
                        SettlementOrder.status.in_(['confirmed', 'printed'])
                    )
                    if start_date:
                        settlements_date_q = settlements_date_q.filter(SettlementOrder.created_at >= start_date)
                    if end_date:
                        settlements_date_q = settlements_date_q.filter(SettlementOrder.created_at <= end_date)
                    for s in settlements_date_q.all():
                        if s.payment_method == 'physical_gold':
                            total_settlement_gold += float(s.physical_gold_weight or 0)
                        elif s.payment_method == 'mixed':
                            total_settlement_gold += float(s.gold_payment_weight or 0)
                    
                    total_receipts_gold_q = db.query(func.coalesce(func.sum(GoldReceipt.gold_weight), 0)).filter(
                        GoldReceipt.customer_id == customer_id, GoldReceipt.status == 'received'
                    )
                    if start_date:
                        total_receipts_gold_q = total_receipts_gold_q.filter(GoldReceipt.created_at >= start_date)
                    if end_date:
                        total_receipts_gold_q = total_receipts_gold_q.filter(GoldReceipt.created_at <= end_date)
                    total_receipts_gold = total_receipts_gold_q.scalar() or 0
                    
                    from ..models import CustomerWithdrawal
                    total_withdrawals_gold_q = db.query(func.coalesce(func.sum(CustomerWithdrawal.gold_weight), 0)).filter(
                        CustomerWithdrawal.customer_id == customer_id, CustomerWithdrawal.status.in_(['pending', 'completed'])
                    )
                    if start_date:
                        total_withdrawals_gold_q = total_withdrawals_gold_q.filter(CustomerWithdrawal.created_at >= start_date)
                    if end_date:
                        total_withdrawals_gold_q = total_withdrawals_gold_q.filter(CustomerWithdrawal.created_at <= end_date)
                    total_withdrawals_gold = total_withdrawals_gold_q.scalar() or 0
                    
                    from ..models import DepositSettlement
                    total_deposit_settled_q = db.query(func.coalesce(func.sum(DepositSettlement.gold_weight), 0)).filter(
                        DepositSettlement.customer_id == customer_id, DepositSettlement.status == 'confirmed'
                    )
                    if start_date:
                        total_deposit_settled_q = total_deposit_settled_q.filter(DepositSettlement.confirmed_at >= start_date)
                    if end_date:
                        total_deposit_settled_q = total_deposit_settled_q.filter(DepositSettlement.confirmed_at <= end_date)
                    total_deposit_settled = total_deposit_settled_q.scalar() or 0
                    
                    net_gold = float(total_settlement_gold) + float(total_withdrawals_gold) + float(total_deposit_settled) - float(total_receipts_gold)

                # 获取结算单明细用于展示（始终需要，不依赖上面的计算方式）
                settlements_query = db.query(SettlementOrder).join(SalesOrder).filter(
                    SalesOrder.customer_name == customer.name if use_name_fallback else SalesOrder.customer_id == customer_id,
                    SettlementOrder.status.in_(['confirmed', 'printed'])
                )
                if start_date:
                    settlements_query = settlements_query.filter(SettlementOrder.created_at >= start_date)
                if end_date:
                    settlements_query = settlements_query.filter(SettlementOrder.created_at <= end_date)
                settlements = settlements_query.all()
                
                # 查询来料明细
                receipts_query = db.query(GoldReceipt).filter(
                    GoldReceipt.customer_id == customer_id, GoldReceipt.status == 'received'
                )
                if start_date:
                    receipts_query = receipts_query.filter(GoldReceipt.created_at >= start_date)
                if end_date:
                    receipts_query = receipts_query.filter(GoldReceipt.created_at <= end_date)
                
                for s in settlements[:20]:
                    gold_amount = 0.0
                    if s.payment_method == 'physical_gold':
                        gold_amount = s.physical_gold_weight or 0
                    elif s.payment_method == 'mixed':
                        gold_amount = s.gold_payment_weight or 0
                    
                    if gold_amount > 0:
                        gold_transactions.append({
                            "id": s.id,
                            "type": "settlement",
                            "type_label": "销售结算",
                            "gold_weight": gold_amount,
                            "order_no": s.settlement_no,
                            "created_at": s.created_at.isoformat() if s.created_at else None,
                            "remark": s.remark or ""
                        })
                
                # 获取来料记录明细用于展示
                receipts = receipts_query.order_by(desc(GoldReceipt.created_at)).limit(20).all()
                for r in receipts:
                    deposit_transactions.append({
                        "id": r.id,
                        "type": "receipt",
                        "type_label": "客户来料",
                        "amount": r.gold_weight,
                        "gold_weight": r.gold_weight,
                        "order_no": r.receipt_no,
                        "gold_fineness": r.gold_fineness,
                        "created_at": r.created_at.isoformat() if r.created_at else None,
                        "remark": r.remark or ""
                    })
                    
            except Exception as e:
                logger.warning(f"查询金料账户出错: {e}")
            
            # 单一账户模式：从 net_gold 计算兼容字段
            # 语义：net_gold = 结算欠料 - 来料（正数=欠料，负数=存料）
            result["net_gold"] = net_gold  # 净金料值（核心字段）
            result["gold_debt"] = max(0, net_gold)  # 欠料（正数部分）
            result["gold_deposit"] = max(0, -net_gold)  # 存料（负数的绝对值）
            result["gold_transactions"] = gold_transactions
            result["deposit_transactions"] = deposit_transactions
        
        # 6. 查询客户销售历史表现（新增）
        try:
            # === 第一部分：查询指定时间段的销售（今日/本周/本月等，只统计已结算）===
            _perf_has_id = db.query(SalesOrder.id).filter(
                SalesOrder.customer_id == customer_id, SalesOrder.status == "已结算"
            ).first()
            _perf_filter = (SalesOrder.customer_id == customer_id) if _perf_has_id else (SalesOrder.customer_name == customer.name)
            period_sales_query = db.query(SalesOrder).filter(
                _perf_filter,
                SalesOrder.status == "已结算"
            )
            
            if start_date:
                period_sales_query = period_sales_query.filter(SalesOrder.order_date >= start_date)
            if end_date:
                period_sales_query = period_sales_query.filter(SalesOrder.order_date <= end_date)
            
            period_orders = period_sales_query.all()
            
            # 统计指定时间段的销售
            period_sales_weight = 0.0
            period_labor_cost = 0.0
            period_order_count = len(period_orders)
            
            for order in period_orders:
                details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
                for detail in details:
                    period_sales_weight += float(detail.weight or 0)
                    period_labor_cost += float(detail.total_labor_cost or 0)
            
            # === 第二部分：查询历史总览（不受日期限制，只统计已结算）===
            all_sales_query = db.query(SalesOrder).filter(
                _perf_filter,
                SalesOrder.status == "已结算"
            )
            all_orders = all_sales_query.all()
            
            # 统计历史总销售克重和工费
            total_sales_weight = 0.0
            total_labor_cost = 0.0
            total_order_count = len(all_orders)
            
            # 品类统计（基于全部历史记录）
            category_stats = {}
            
            for order in all_orders:
                details = db.query(SalesDetail).filter(SalesDetail.order_id == order.id).all()
                for detail in details:
                    weight = float(detail.weight or 0)
                    labor = float(detail.total_labor_cost or 0)
                    total_sales_weight += weight
                    total_labor_cost += labor
                    
                    # 按品类统计
                    product_name = detail.product_name or "其他"
                    if product_name not in category_stats:
                        category_stats[product_name] = {"weight": 0, "labor": 0, "count": 0}
                    category_stats[product_name]["weight"] += weight
                    category_stats[product_name]["labor"] += labor
                    category_stats[product_name]["count"] += 1
            
            # 将品类统计转为列表并按销售克重排序
            category_breakdown = [
                {"name": name, "weight": stats["weight"], "labor": stats["labor"], "count": stats["count"]}
                for name, stats in category_stats.items()
            ]
            category_breakdown.sort(key=lambda x: x["weight"], reverse=True)
            
            # 计算客户排名（按总购买金额）
            all_customers = db.query(Customer).filter(Customer.status == "active").order_by(
                desc(Customer.total_purchase_amount)
            ).all()
            customer_rank = 1
            for idx, c in enumerate(all_customers, 1):
                if c.id == customer.id:
                    customer_rank = idx
                    break
            total_customer_count = len(all_customers)
            
            result["sales_history"] = {
                # 指定时间段销售（今日/本周/本月等）
                "period_sales_weight": period_sales_weight,
                "period_labor_cost": period_labor_cost,
                "period_order_count": period_order_count,
                # 历史总览（全部记录）
                "total_sales_weight": total_sales_weight,
                "total_labor_cost": total_labor_cost,
                "order_count": total_order_count,
                "category_breakdown": category_breakdown[:5],  # 只返回前5个品类
                "customer_rank": customer_rank,
                "total_customer_count": total_customer_count,
                "last_purchase_time": str(customer.last_purchase_time) if customer.last_purchase_time else None
            }
        except Exception as e:
            logger.warning(f"查询客户销售历史出错: {e}")
            result["sales_history"] = None
        
        # 将result转换为统一响应格式
        return success_response(data=result)
        
    except Exception as e:
        logger.error(f"聊天查询客户账务失败: {e}", exc_info=True)
        return server_error_response(message=f"查询失败: {str(e)}")


@router.get("/{customer_id}")
async def get_customer(
    customer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """获取客户详情"""
    # 权限检查 - 需要 can_view_customers 或 can_manage_customers 权限
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_view_customers') and not has_permission(user_role, 'can_manage_customers'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【查看客户】的权限")
    
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        
        if not customer:
            return not_found_response(message="客户不存在")
        
        return success_response(
            data={"customer": CustomerResponse.model_validate(customer).model_dump(mode='json')}
        )
    except Exception as e:
        logger.error(f"查询客户详情失败: {e}", exc_info=True)
        return server_error_response(message=f"查询客户详情失败: {str(e)}")


@router.put("/{customer_id}")
async def update_customer(
    customer_id: int,
    data: CustomerCreate,
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """更新客户信息"""
    # 权限检查 - 需要 can_manage_customers 权限
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_manage_customers'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【客户管理】的权限（创建/编辑/删除）")
    
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return not_found_response(message="客户不存在")
        
        # 更新字段
        if data.name:
            customer.name = data.name
        if data.phone is not None:
            customer.phone = data.phone
        if data.wechat is not None:
            customer.wechat = data.wechat
        if data.address is not None:
            customer.address = data.address
        if data.remark is not None:
            customer.remark = data.remark
        
        db.commit()
        db.refresh(customer)
        
        return success_response(
            data={"customer": CustomerResponse.model_validate(customer).model_dump(mode='json')},
            message=f"客户【{customer.name}】信息已更新"
        )
    except Exception as e:
        db.rollback()
        logger.error(f"更新客户失败: {e}", exc_info=True)
        return server_error_response(message=str(e))


@router.delete("/{customer_id}")
async def delete_customer(
    customer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    role: str = Depends(require_permission("can_delete")),
    db: Session = Depends(get_db)
):
    """删除客户（软删除）"""
    # 权限检查 - 只有管理层可以删除
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_delete'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【删除数据】的权限")
    
    try:
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return not_found_response(message="客户不存在")
        
        customer.status = "inactive"
        db.commit()
        
        return success_response(message=f"客户【{customer.name}】已删除")
    except Exception as e:
        db.rollback()
        logger.error(f"删除客户失败: {e}", exc_info=True)
        return server_error_response(message=str(e))


@router.get("/{customer_id}/detail")
async def get_customer_detail(
    customer_id: int,
    user_role: str = Query(default="sales", description="用户角色"),
    date_start: Optional[str] = Query(default=None, description="开始日期 YYYY-MM-DD"),
    date_end: Optional[str] = Query(default=None, description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db)
):
    """
    获取客户详情（销售记录、退货记录、欠款/存料余额、往来账目）
    业务员角色可以查看客户的完整往来信息
    支持日期范围筛选往来账目
    """
    # 权限检查 - 需要查看客户或查询客户销售权限
    from ..middleware.permissions import has_permission
    can_view = (
        has_permission(user_role, 'can_view_customers') or 
        has_permission(user_role, 'can_manage_customers') or
        has_permission(user_role, 'can_query_customer_sales')
    )
    if not can_view:
        raise HTTPException(status_code=403, detail="权限不足：您没有查看客户详情的权限")
    
    try:
        # 获取客户基本信息
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return not_found_response(message="客户不存在")
        
        # 获取销售记录（优先 customer_id，兼容迁移数据用 customer_name 兜底）
        _has_id_sales = db.query(SalesOrder.id).filter(
            SalesOrder.customer_id == customer_id,
            SalesOrder.status.notin_(["cancelled", "已取消", "draft"])
        ).first()
        _sales_filter = (SalesOrder.customer_id == customer_id) if _has_id_sales else (SalesOrder.customer_name == customer.name)
        sales_orders = db.query(SalesOrder).filter(
            _sales_filter,
            SalesOrder.status.notin_(["cancelled", "已取消", "draft"])
        ).options(
            selectinload(SalesOrder.details)
        ).order_by(desc(SalesOrder.create_time)).limit(50).all()
        
        sales_list = []
        for order in sales_orders:
            for detail in order.details:
                sales_list.append({
                    "id": detail.id,
                    "order_no": order.order_no,
                    "product_name": detail.product_name,
                    "weight": detail.weight,
                    "labor_cost": detail.labor_cost,
                    "total_amount": detail.total_labor_cost,  # 使用 total_labor_cost 字段
                    "status": order.status,
                    "created_at": order.create_time.isoformat() if order.create_time else None
                })
        
        # 获取销退记录（客户退货 - 来自 SalesReturnOrder）
        returns_list = []
        try:
            sales_return_orders = db.query(SalesReturnOrder).filter(
                SalesReturnOrder.customer_id == customer_id
            ).order_by(desc(SalesReturnOrder.create_time)).all()
            
            for sro in sales_return_orders:
                # 获取明细商品名
                detail_names = [d.product_name for d in sro.details] if sro.details else []
                product_text = "、".join(detail_names) if detail_names else "未知商品"
                item_count = len(detail_names)
                
                # 状态映射
                status_map = {
                    'draft': '未确认',
                    'confirmed': '已确认',
                    '待结算': '待结算',
                    '已结算': '已结算',
                }
                
                returns_list.append({
                    "id": sro.id,
                    "return_no": sro.return_no,
                    "product_name": f"{product_text}（{item_count}项）" if item_count > 1 else product_text,
                    "return_weight": float(sro.total_weight or 0),
                    "return_reason": sro.return_reason or "未知",
                    "status": status_map.get(sro.status, sro.status),
                    "return_to": "展厅" if sro.return_to == "showroom" else "商品部",
                    "created_at": sro.create_time.isoformat() if sro.create_time else None
                })
        except Exception as e:
            logger.warning(f"查询客户销退记录时出错: {e}")
            returns_list = []
        
        # 获取往来账目（支持日期范围筛选）
        transactions_list = []
        
        # 解析日期参数
        from datetime import datetime as dt
        filter_start = None
        filter_end = None
        if date_start:
            try:
                filter_start = dt.strptime(date_start, "%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        if date_end:
            try:
                filter_end = dt.strptime(date_end, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            except (ValueError, TypeError):
                pass
        
        # 销售交易（产生工费，客户欠款增加，正数=客户欠我们）
        for order in sales_orders[:50]:  # 增加限制数量
            # 日期过滤
            if filter_start and order.create_time and order.create_time < filter_start:
                continue
            if filter_end and order.create_time and order.create_time > filter_end:
                continue
            # 计算工费明细（平均工费单价）
            labor_detail = ""
            total_w = order.total_weight or 0
            total_lc = order.total_labor_cost or 0
            if total_w > 0 and total_lc > 0:
                avg_labor_per_gram = total_lc / total_w
                labor_detail = f"（工费 {total_w:.1f}g × ¥{avg_labor_per_gram:.0f}/g）"
            transactions_list.append({
                "id": order.id,
                "type": "sales_labor",
                "type_label": "销售结算",
                "order_no": order.order_no or "",
                "description": labor_detail.strip("（）") if labor_detail else "工费",
                "amount": total_lc,  # 正数表示客户欠款增加
                "gold_weight": total_w if total_w > 0 else None,
                "created_at": order.create_time.isoformat() if order.create_time else None,
                "remark": _sanitize_remark(order.remark)
            })
        
        # 金料收料记录（客户来料=我们欠客户，负数）
        try:
            receipts_query = db.query(GoldReceipt).filter(
                GoldReceipt.customer_id == customer_id,
                GoldReceipt.status == 'received'
            )
            if filter_start:
                receipts_query = receipts_query.filter(GoldReceipt.received_at >= filter_start)
            if filter_end:
                receipts_query = receipts_query.filter(GoldReceipt.received_at <= filter_end)
            gold_receipts = receipts_query.order_by(desc(GoldReceipt.received_at)).limit(50).all()
            
            for receipt in gold_receipts:
                transactions_list.append({
                    "id": receipt.id,
                    "type": "customer_receipt",
                    "type_label": "客户来料",
                    "order_no": receipt.receipt_no or "",
                    "description": f"收料 {receipt.gold_weight or 0:.2f}克",
                    "amount": None,
                    "gold_weight": -(receipt.gold_weight or 0),
                    "created_at": (receipt.received_at or receipt.created_at).isoformat() if (receipt.received_at or receipt.created_at) else None,
                    "remark": receipt.remark or ""
                })
        except Exception as e:
            logger.warning(f"查询收料记录时出错: {e}")
        
        # 金料提料记录（客户提料=客户欠我们，正数）
        try:
            from ..models.finance import CustomerWithdrawal
            withdrawals_query = db.query(CustomerWithdrawal).filter(
                CustomerWithdrawal.customer_id == customer_id,
                CustomerWithdrawal.status == 'completed'
            )
            if filter_start:
                withdrawals_query = withdrawals_query.filter(CustomerWithdrawal.completed_at >= filter_start)
            if filter_end:
                withdrawals_query = withdrawals_query.filter(CustomerWithdrawal.completed_at <= filter_end)
            withdrawals = withdrawals_query.order_by(desc(CustomerWithdrawal.completed_at)).limit(50).all()
            
            for withdrawal in withdrawals:
                transactions_list.append({
                    "id": withdrawal.id,
                    "type": "customer_withdrawal",
                    "type_label": "客户提料",
                    "order_no": withdrawal.withdrawal_no or "",
                    "description": f"提料 {withdrawal.gold_weight or 0:.2f}克",
                    "amount": None,
                    "gold_weight": (withdrawal.gold_weight or 0),
                    "created_at": (withdrawal.completed_at or withdrawal.created_at).isoformat() if (withdrawal.completed_at or withdrawal.created_at) else None,
                    "remark": withdrawal.remark or ""
                })
        except Exception as e:
            logger.warning(f"查询提料记录时出错: {e}")
        
        # 结算记录（从SettlementOrder表获取）
        try:
            for order in sales_orders[:50]:
                settlement_query = db.query(SettlementOrder).filter(
                    SettlementOrder.sales_order_id == order.id,
                    SettlementOrder.status.in_(['confirmed', 'printed'])
                )
                if filter_start:
                    settlement_query = settlement_query.filter(SettlementOrder.created_at >= filter_start)
                if filter_end:
                    settlement_query = settlement_query.filter(SettlementOrder.created_at <= filter_end)
                settlements = settlement_query.all()
                
                for s in settlements:
                    # 根据支付方式生成描述和类型
                    settle_weight = s.total_weight or 0
                    gold_price = s.gold_price or 0
                    if s.payment_method == 'cash_price':
                        type_code = "settle_cash"
                        type_label = "结价"
                        if settle_weight > 0 and gold_price > 0:
                            total_val = settle_weight * gold_price
                            method_desc = f"{settle_weight:.1f}g × ¥{gold_price:.0f}/g = ¥{total_val:,.0f}"
                        else:
                            method_desc = f"结价 ¥{gold_price}/克"
                        gold_change = -settle_weight if settle_weight > 0 else None
                        amount_change = float(s.total_amount or 0)
                    elif s.payment_method == 'physical_gold':
                        type_code = "settle_gold"
                        type_label = "结料"
                        method_desc = f"结料 {s.physical_gold_weight or 0:.1f}克"
                        gold_change = 0
                        amount_change = float(s.labor_amount or 0)
                    else:  # mixed
                        type_code = "settle_mixed"
                        type_label = "混合结算"
                        gold_part = s.gold_payment_weight or 0
                        cash_part = s.cash_payment_weight or 0
                        cash_amount = cash_part * gold_price
                        method_desc = f"结料{gold_part:.1f}g + 结价{cash_part:.1f}g×¥{gold_price:.0f}/g=¥{cash_amount:,.0f}"
                        gold_change = 0
                        amount_change = float(s.total_amount or 0)
                    
                    transactions_list.append({
                        "id": f"settlement_{s.id}",
                        "type": type_code,
                        "type_label": type_label,
                        "order_no": s.settlement_no or "",
                        "description": method_desc,
                        "amount": amount_change,
                        "gold_weight": gold_change,
                        "created_at": s.created_at.isoformat() if s.created_at else None,
                        "remark": _sanitize_remark(order.remark)
                    })
        except Exception as e:
            logger.warning(f"查询结算记录时出错: {e}")
        
        # 收款记录（客户来款=我们欠客户，负数）
        # 排除由存料结价自动生成的来款记录，避免与"存料结价"记录重复计算
        try:
            from ..models import PaymentRecord
            payments_query = db.query(PaymentRecord).filter(
                PaymentRecord.customer_id == customer_id,
                ~PaymentRecord.remark.like("%存料结价抵扣%")
            )
            if filter_start:
                payments_query = payments_query.filter(PaymentRecord.create_time >= filter_start)
            if filter_end:
                payments_query = payments_query.filter(PaymentRecord.create_time <= filter_end)
            payments = payments_query.order_by(desc(PaymentRecord.create_time)).limit(50).all()
            
            for p in payments:
                method_map = {'bank_transfer': '银行转账', 'cash': '现金', 'wechat': '微信', 'alipay': '支付宝', 'card': '刷卡', 'check': '支票'}
                method_text = method_map.get(p.payment_method, p.payment_method or '')
                transactions_list.append({
                    "id": f"payment_{p.id}",
                    "type": "customer_payment",
                    "type_label": "客户来款",
                    "order_no": p.payment_no or "",
                    "description": method_text,
                    "amount": -(p.amount or 0),
                    "gold_weight": None,
                    "created_at": p.create_time.isoformat() if p.create_time else None,
                    "remark": p.remark or "",
                    "voucher_images": p.voucher_images or None
                })
        except Exception as e:
            logger.warning(f"查询收款记录时出错: {e}")
        
        # 退货记录（注：ReturnOrder 是展厅退商品部/商品部退供应商的内部退货，客户往来账不包含）
        
        # 销退记录（客户退货 → 退款到客户账户）
        try:
            sr_query = db.query(SalesReturnSettlement).join(
                SalesReturnOrder,
                SalesReturnSettlement.sales_return_order_id == SalesReturnOrder.id
            ).filter(
                SalesReturnOrder.customer_id == customer_id,
                SalesReturnSettlement.status.in_(["confirmed", "printed"])
            )
            if filter_start:
                sr_query = sr_query.filter(SalesReturnSettlement.created_at >= filter_start)
            if filter_end:
                sr_query = sr_query.filter(SalesReturnSettlement.created_at <= filter_end)
            sales_returns = sr_query.order_by(desc(SalesReturnSettlement.created_at)).limit(50).all()
            
            for sr in sales_returns:
                # 获取关联的销退单
                return_order = db.query(SalesReturnOrder).filter(
                    SalesReturnOrder.id == sr.sales_return_order_id
                ).first()
                
                method_map = {
                    'cash_price': '销退退价',
                    'physical_gold': '销退退料',
                    'mixed': '销退混合退款'
                }
                type_label = method_map.get(sr.payment_method, '销退')
                
                # 描述信息
                if sr.payment_method == 'cash_price':
                    desc_text = f"退价 ¥{sr.total_amount:.2f}" if sr.total_amount else "退价"
                elif sr.payment_method == 'physical_gold':
                    gold_w = sr.physical_gold_weight or sr.total_weight or 0
                    desc_text = f"退料 {gold_w:.3f}g"
                else:
                    desc_text = "混合退款"
                
                # 金料退还（退料/混合时）
                gold_weight_val = None
                if sr.payment_method == 'physical_gold':
                    gold_weight_val = -(sr.physical_gold_weight or sr.total_weight or 0)
                elif sr.payment_method == 'mixed' and sr.gold_payment_weight:
                    gold_weight_val = -(sr.gold_payment_weight or 0)
                
                transactions_list.append({
                    "id": f"sales_return_{sr.id}",
                    "type": "sales_refund",
                    "type_label": type_label,
                    "order_no": sr.settlement_no or "",
                    "description": desc_text,
                    "amount": -(sr.total_amount or 0),  # 负数：退给客户
                    "gold_weight": gold_weight_val,
                    "created_at": sr.confirmed_at.isoformat() if sr.confirmed_at else (sr.created_at.isoformat() if sr.created_at else None),
                    "remark": sr.remark or (return_order.remark if return_order else "") or ""
                })
        except Exception as e:
            logger.warning(f"查询销退记录时出错: {e}")
        
        # 存料结价记录（客户存料折现抵扣欠款）
        try:
            from ..models import DepositSettlement
            ds_query = db.query(DepositSettlement).filter(
                DepositSettlement.customer_id == customer_id,
                DepositSettlement.status == "confirmed"
            )
            if filter_start:
                ds_query = ds_query.filter(DepositSettlement.confirmed_at >= filter_start)
            if filter_end:
                ds_query = ds_query.filter(DepositSettlement.confirmed_at <= filter_end)
            deposit_settlements = ds_query.order_by(desc(DepositSettlement.confirmed_at)).limit(50).all()
            
            for ds in deposit_settlements:
                transactions_list.append({
                    "id": f"deposit_settle_{ds.id}",
                    "type": "deposit_settle",
                    "type_label": "存料结价",
                    "order_no": ds.settlement_no or "",
                    "description": f"{ds.gold_weight:.3f}g × ¥{ds.gold_price:.0f}/g = ¥{ds.total_amount:,.2f}",
                    "amount": -(ds.total_amount or 0),
                    "gold_weight": (ds.gold_weight or 0),  # 正数：客户存料减少，我们少欠客户
                    "created_at": ds.confirmed_at.isoformat() if ds.confirmed_at else (ds.created_at.isoformat() if ds.created_at else None),
                    "remark": ds.remark or ""
                })
        except Exception as e:
            logger.warning(f"查询存料结价记录时出错: {e}")
        
        # ========== 三栏式会计账目处理 ==========
        # 按时间正序排列（旧→新），方便计算逐笔余额
        transactions_list.sort(key=lambda x: x["created_at"] or "", reverse=False)
        
        # 为每笔交易拆分借方(debit)和贷方(credit)
        # 会计规则：借方=应收增加（客户欠我们更多），贷方=应收减少（收到客户的钱/料）
        for tx in transactions_list:
            tx_amount = float(tx.get("amount") or 0)
            tx_gold = float(tx.get("gold_weight") or 0)
            
            # 现金：正数→借方(应收增加)，负数→贷方(已收/冲抵)
            if tx_amount > 0:
                tx["cash_debit"] = round(tx_amount, 2)
                tx["cash_credit"] = None
            elif tx_amount < 0:
                tx["cash_debit"] = None
                tx["cash_credit"] = round(abs(tx_amount), 2)
            else:
                tx["cash_debit"] = None
                tx["cash_credit"] = None
            
            # 金料：正数→借方(欠料增加)，负数→贷方(来料/退料)
            if tx_gold > 0:
                tx["gold_debit"] = round(tx_gold, 3)
                tx["gold_credit"] = None
            elif tx_gold < 0:
                tx["gold_debit"] = None
                tx["gold_credit"] = round(abs(tx_gold), 3)
            else:
                tx["gold_debit"] = None
                tx["gold_credit"] = None
        
        # 计算期初余额（如果有日期筛选）
        opening_balance = None
        if filter_start:
            # 计算开始日期之前的累计余额
            opening_cash = 0.0  # 现金欠款
            opening_gold = 0.0  # 金料余额
            
            # 查询开始日期前的销售工费（增加欠款）
            for order in sales_orders:
                if order.create_time and order.create_time < filter_start:
                    opening_cash += (order.total_labor_cost or 0)
            
            # 查询开始日期前的来料（我们欠客户，负数）
            try:
                prev_receipts = db.query(GoldReceipt).filter(
                    GoldReceipt.customer_id == customer_id,
                    GoldReceipt.status == 'received',
                    GoldReceipt.received_at < filter_start
                ).all()
                for r in prev_receipts:
                    opening_gold -= (r.gold_weight or 0)
            except (ValueError, TypeError):
                pass
            
            # 查询开始日期前的结算（正数=客户欠我们）
            try:
                for order in sales_orders:
                    if order.create_time and order.create_time < filter_start:
                        prev_settlements = db.query(SettlementOrder).filter(
                            SettlementOrder.sales_order_id == order.id,
                            SettlementOrder.status.in_(['confirmed', 'printed']),
                            SettlementOrder.created_at < filter_start
                        ).all()
                        for s in prev_settlements:
                            if s.payment_method == 'cash_price':
                                opening_cash += float(s.material_amount or 0)
                            elif s.payment_method == 'physical_gold':
                                opening_gold += float(s.physical_gold_weight or 0)
                            else:  # mixed
                                opening_gold += float(s.gold_payment_weight or 0)
                                cash_amount = float(s.cash_payment_weight or 0) * float(s.gold_price or 0)
                                opening_cash += cash_amount
            except (ValueError, TypeError):
                pass
            
            # 查询开始日期前的收款（我们欠客户，负数）
            try:
                prev_payments = db.query(PaymentRecord).filter(
                    PaymentRecord.customer_id == customer_id,
                    PaymentRecord.create_time < filter_start,
                    ~PaymentRecord.remark.like("%存料结价抵扣%")
                ).all()
                for p in prev_payments:
                    opening_cash -= float(p.amount or 0)
            except (ValueError, TypeError):
                pass

            # 查询开始日期前的提料（客户欠我们，正数）
            try:
                prev_withdrawals = db.query(CustomerWithdrawal).filter(
                    CustomerWithdrawal.customer_id == customer_id,
                    CustomerWithdrawal.status.in_(['pending', 'completed']),
                    CustomerWithdrawal.created_at < filter_start
                ).all()
                for w in prev_withdrawals:
                    opening_gold += float(w.gold_weight or 0)
            except (ValueError, TypeError):
                pass
            
            ob_cash_debit = round(opening_cash, 2) if opening_cash > 0 else None
            ob_cash_credit = round(abs(opening_cash), 2) if opening_cash < 0 else None
            ob_gold_debit = round(opening_gold, 3) if opening_gold > 0 else None
            ob_gold_credit = round(abs(opening_gold), 3) if opening_gold < 0 else None
            
            opening_balance = {
                "id": "opening_balance",
                "type": "opening_balance",
                "type_label": "期初余额",
                "description": f"期初余额（{date_start}之前）",
                "amount": opening_cash if opening_cash != 0 else None,
                "gold_weight": opening_gold if opening_gold != 0 else None,
                "cash_debit": ob_cash_debit,
                "cash_credit": ob_cash_credit,
                "gold_debit": ob_gold_debit,
                "gold_credit": ob_gold_credit,
                "created_at": date_start,
                "remark": ""
            }

        # 加入期初现金余额（数据迁移写入的 initial_balance 交易）
        initial_cash = 0.0
        try:
            initial_cash_q = db.query(func.coalesce(func.sum(CustomerTransaction.amount), 0)).filter(
                CustomerTransaction.customer_id == customer_id,
                CustomerTransaction.transaction_type == 'initial_balance',
                CustomerTransaction.status == 'active'
            )
            initial_cash = float(initial_cash_q.scalar() or 0)
        except Exception:
            pass

        # 计算逐笔余额（running balance）
        running_cash = initial_cash
        running_gold = 0.0
        
        if opening_balance:
            running_cash += float(opening_balance.get("amount") or 0)
            running_gold = float(opening_balance.get("gold_weight") or 0)
            opening_balance["running_cash_balance"] = round(running_cash, 2)
            opening_balance["running_gold_balance"] = round(running_gold, 3)
        
        for tx in transactions_list:
            tx_amount = float(tx.get("amount") or 0)
            tx_gold = float(tx.get("gold_weight") or 0)
            running_cash += tx_amount
            running_gold += tx_gold
            tx["running_cash_balance"] = round(running_cash, 2)
            tx["running_gold_balance"] = round(running_gold, 3)

        # 无日期筛选时，金料余额使用统一计算函数
        if not date_start and not date_end:
            from ..gold_balance import calculate_customer_net_gold
            unified_net_gold = calculate_customer_net_gold(customer_id, db)
            running_gold = -unified_net_gold

        cash_total = running_cash
        gold_total = running_gold

        balance = {
            "cash_debt": round(cash_total, 2),
            "gold_debt": round(max(0, gold_total), 3),
            "gold_deposit": round(max(0, -gold_total), 3),
            "net_gold": round(gold_total, 3)
        }
        
        # 三栏式汇总
        total_cash_debit = 0.0
        total_cash_credit = 0.0
        total_gold_debit = 0.0
        total_gold_credit = 0.0
        
        for tx in transactions_list:
            total_cash_debit += float(tx.get("cash_debit") or 0)
            total_cash_credit += float(tx.get("cash_credit") or 0)
            total_gold_debit += float(tx.get("gold_debit") or 0)
            total_gold_credit += float(tx.get("gold_credit") or 0)
        
        transaction_summary = {
            "total_cash_debit": round(total_cash_debit, 2),
            "total_cash_credit": round(total_cash_credit, 2),
            "total_gold_debit": round(total_gold_debit, 3),
            "total_gold_credit": round(total_gold_credit, 3),
            "net_cash": round(cash_total, 2),
            "net_gold": round(gold_total, 3),
            "total_labor_fee": 0, "total_settlement_cash": 0,
            "total_settlement_gold": 0, "total_gold_receipt": 0,
            "total_payment": 0, "total_withdrawal": 0,
            "total_deposit_settle": 0,
        }

        return success_response(
            data={
                "customer": CustomerResponse.model_validate(customer).model_dump(mode='json'),
                "sales": sales_list,
                "returns": returns_list,
                "balance": balance,
                "transactions": transactions_list[:50],
                "opening_balance": opening_balance,
                "transaction_summary": transaction_summary,
                "date_range": {"start": date_start, "end": date_end} if (date_start or date_end) else None
            }
        )
    except Exception as e:
        logger.error(f"查询客户详情失败: {e}", exc_info=True)
        return server_error_response(message=f"查询客户详情失败: {str(e)}")


@router.post("/batch-import")
async def batch_import_customers(
    file: UploadFile = File(...),
    user_role: str = Query(default="sales", description="用户角色"),
    encoding: Optional[str] = Query(default=None, description="文件编码（可选）"),
    role: str = Depends(require_permission("can_manage_customers")),
    db: Session = Depends(get_db)
):
    """
    批量导入客户（支持2000+条数据）
    支持格式：
    1. Excel (.xlsx, .xls) - 第一列必须是姓名，其他列可选（电话、微信、地址、类型、备注）
    2. CSV (.csv) - 第一列必须是姓名，其他列可选
    3. 纯文本 (.txt) - 每行一个姓名
    """
    from ..middleware.permissions import has_permission
    if not has_permission(user_role, 'can_manage_customers'):
        raise HTTPException(status_code=403, detail="权限不足：您没有【客户管理】的权限")
    
    results = {
        "success": True,
        "total": 0,
        "created": 0,
        "skipped": 0,
        "errors": [],
        "details": []
    }
    
    try:
        # 读取文件内容
        content = await file.read()
        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        
        customers_data = []
        
        def _score_text(text: str) -> int:
            if not text:
                return -1
            chinese = sum(1 for c in text if '\u4e00' <= c <= '\u9fff')
            alnum = sum(1 for c in text if c.isalnum())
            bad = text.count('\ufffd')
            return chinese * 2 + alnum - bad * 5

        def _decode_with_auto_detect(raw: bytes) -> tuple[str, str]:
            candidates = ['utf-8-sig', 'gb18030', 'gbk', 'big5']
            best_text = None
            best_encoding = None
            best_score = -1
            for enc in candidates:
                try:
                    text = raw.decode(enc, errors='replace')
                except Exception:
                    continue
                score = _score_text(text)
                if score > best_score:
                    best_text = text
                    best_encoding = enc
                    best_score = score
            if best_text is None or best_encoding is None:
                raise ValueError("无法识别文件编码，请尝试手动指定 encoding 参数")
            return best_text, best_encoding

        if file_extension in ['xlsx', 'xls']:
            # Excel 文件处理
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                
                # 跳过表头（第一行）
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row and row[0]:
                        name = str(row[0]).strip()
                        if name:
                            customers_data.append({
                                "name": name,
                                "phone": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                                "wechat": str(row[2]).strip() if len(row) > 2 and row[2] else None,
                                "address": str(row[3]).strip() if len(row) > 3 and row[3] else None,
                                "customer_type": str(row[4]).strip() if len(row) > 4 and row[4] else "个人",
                                "remark": str(row[5]).strip() if len(row) > 5 and row[5] else None,
                            })
                wb.close()
            except Exception as e:
                error_msg = str(e).lower()
                # 检测是否是 .xls 格式导致的错误
                if "zip file" in error_msg or "not a zip file" in error_msg or file_extension == 'xls':
                    return error_response(
                        message="不支持旧版 Excel (.xls) 格式，请将文件另存为 .xlsx 格式或转换为 CSV 格式后重试"
                    )
                else:
                    return error_response(
                        message=f"Excel 文件解析失败: {str(e)[:200]}\n\n提示：请确保文件格式正确，或尝试转换为 CSV 格式上传"
                    )
        
        elif file_extension == 'csv':
            # CSV 文件处理
            try:
                if encoding:
                    try:
                        content_str = content.decode(encoding, errors='strict')
                        results["encoding"] = encoding
                    except Exception:
                        return error_response(message=f"指定编码解析失败：{encoding}")
                else:
                    content_str, used_encoding = _decode_with_auto_detect(content)
                    results["encoding"] = used_encoding
            
                csv_reader = csv.reader(io.StringIO(content_str))
                
                # 跳过表头
                next(csv_reader, None)
                
                for row in csv_reader:
                    if row and row[0]:
                        name = str(row[0]).strip()
                        if name:
                            customers_data.append({
                                "name": name,
                                "phone": row[1].strip() if len(row) > 1 and row[1] else None,
                                "wechat": row[2].strip() if len(row) > 2 and row[2] else None,
                                "address": row[3].strip() if len(row) > 3 and row[3] else None,
                                "customer_type": row[4].strip() if len(row) > 4 and row[4] else "个人",
                                "remark": row[5].strip() if len(row) > 5 and row[5] else None,
                            })
            except Exception as e:
                return error_response(message=f"CSV 文件解析失败: {str(e)[:200]}")
        
        elif file_extension == 'txt':
            # 纯文本文件（每行一个姓名）
            try:
                if encoding:
                    try:
                        content_str = content.decode(encoding, errors='strict')
                        results["encoding"] = encoding
                    except Exception:
                        return error_response(message=f"指定编码解析失败：{encoding}")
                else:
                    content_str, used_encoding = _decode_with_auto_detect(content)
                    results["encoding"] = used_encoding
            
                for line in content_str.split('\n'):
                    name = line.strip()
                    if name:
                        customers_data.append({
                            "name": name,
                            "phone": None,
                            "wechat": None,
                            "address": None,
                            "customer_type": "个人",
                            "remark": None,
                        })
            except Exception as e:
                return error_response(message=f"文本文件解析失败: {str(e)[:200]}")
        else:
            return error_response(
                message=f"不支持的文件格式：{file_extension}。支持格式：.xlsx, .xls, .csv, .txt"
            )
        
        results["total"] = len(customers_data)
        
        if results["total"] == 0:
            return error_response(message="文件中没有找到有效的客户数据")
        
        # 批量创建客户（性能优化：批量提交）
        start_time = time.time()
        batch_size = 100  # 每100条提交一次
        
        # 先批量查询已存在的客户（避免重复查询）
        existing_names = set()
        existing_customers = db.query(Customer.name).filter(
            Customer.status == "active"
        ).all()
        existing_names = {c[0] for c in existing_customers}
        
        for idx, customer_data in enumerate(customers_data, 1):
            try:
                # 检查是否已存在
                if customer_data["name"] in existing_names:
                    results["skipped"] += 1
                    if idx <= 10:  # 只记录前10个跳过的详情
                        results["details"].append({
                            "row": idx,
                            "name": customer_data["name"],
                            "status": "skipped",
                            "message": "客户已存在"
                        })
                    continue
                
                # 生成客户编号（使用时间戳+序号，确保唯一）
                timestamp = china_now().strftime('%Y%m%d%H%M%S')
                customer_no = f"KH{timestamp}{idx:06d}"
                
                # 创建客户对象
                customer = Customer(
                    customer_no=customer_no,
                    name=customer_data["name"],
                    phone=customer_data.get("phone"),
                    wechat=customer_data.get("wechat"),
                    address=customer_data.get("address"),
                    customer_type=customer_data.get("customer_type", "个人"),
                    remark=customer_data.get("remark"),
                    status="active"
                )
                db.add(customer)
                existing_names.add(customer_data["name"])  # 添加到已存在集合
                results["created"] += 1
                
                # 每 batch_size 条提交一次（提高性能）
                if idx % batch_size == 0:
                    db.commit()
                    logger.info(f"已导入 {idx}/{results['total']} 条客户数据")
                
                # 只记录前10个成功的详情
                if results["created"] <= 10:
                    results["details"].append({
                        "row": idx,
                        "name": customer_data["name"],
                        "status": "created",
                        "customer_no": customer_no
                    })
                    
            except Exception as e:
                results["errors"].append({
                    "row": idx,
                    "name": customer_data.get("name", "未知"),
                    "error": str(e)[:100]  # 限制错误信息长度
                })
                logger.error(f"导入第 {idx} 行失败: {e}")
                # 如果错误太多，停止导入
                if len(results["errors"]) > 100:
                    db.rollback()
                    return error_response(
                        message=f"导入过程中错误过多（超过100个），已停止导入。已成功导入 {results['created']} 条",
                        data=results
                    )
        
        # 最终提交
        db.commit()
        
        elapsed_time = time.time() - start_time
        message = f"导入完成！成功创建 {results['created']} 个客户，跳过 {results['skipped']} 个已存在客户"
        if results["errors"]:
            message += f"，失败 {len(results['errors'])} 个"
        message += f"。耗时 {elapsed_time:.2f} 秒"
        results["elapsed_time"] = elapsed_time
        
        return success_response(data=results, message=message)
        
    except Exception as e:
        db.rollback()
        logger.error(f"批量导入客户失败: {e}", exc_info=True)
        return server_error_response(message=f"批量导入失败: {str(e)}")


@router.get("/{customer_id}/debt-history")
async def get_customer_debt_history(
    customer_id: int,
    limit: int = Query(50, ge=1, le=200),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    获取客户欠款交易历史
    
    返回该客户的所有交易记录，包括销售、结算、收款、金料收付等。
    """
    from ..middleware.permissions import has_permission
    from ..models import SettlementOrder, PaymentRecord
    
    can_view = (
        has_permission(user_role, 'can_view_customers') or 
        has_permission(user_role, 'can_query_customer_sales') or
        has_permission(user_role, 'can_create_settlement')
    )
    if not can_view:
        raise HTTPException(status_code=403, detail="权限不足")
    
    try:
        # 获取客户信息
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            return not_found_response(message="客户不存在")
        
        transactions = []
        
        # 1. 查询销售单（仅用于关联结算记录，不直接加入欠款明细）
        # 业务规则：客户欠款只在结算确认时产生，销售单本身不产生应收
        _debt_has_id = db.query(SalesOrder.id).filter(
            SalesOrder.customer_id == customer_id, SalesOrder.status != "已取消"
        ).first()
        _debt_filter = (SalesOrder.customer_id == customer_id) if _debt_has_id else (SalesOrder.customer_name == customer.name)
        sales_orders = db.query(SalesOrder).filter(
            _debt_filter,
            SalesOrder.status != "已取消"
        ).order_by(desc(SalesOrder.create_time)).limit(limit).all()
        
        # 2. 结算记录
        try:
            for order in sales_orders:
                settlements = db.query(SettlementOrder).filter(
                    SettlementOrder.sales_order_id == order.id
                ).all()
                for s in settlements:
                    payment_method_label = {
                        "cash_price": "结价",
                        "physical_gold": "结料",
                        "mixed": "混合支付"
                    }.get(s.payment_method, s.payment_method)
                    
                    if s.payment_method == 'physical_gold':
                        cash_amt = float(s.labor_amount or 0)
                    else:
                        cash_amt = float(s.total_amount or 0)
                    
                    transactions.append({
                        "id": f"settlement_{s.id}",
                        "type": "settlement",
                        "type_label": "结算",
                        "order_no": s.settlement_no,
                        "description": f"结算单 {s.settlement_no}（{payment_method_label}）",
                        "cash_amount": cash_amt,
                        "gold_amount": 0,
                        "status": s.status,
                        "created_at": s.created_at.isoformat() if s.created_at else None,
                        "operator": s.created_by
                    })
        except Exception as e:
            logger.warning(f"查询结算记录出错: {e}")
        
        # 3. 收款记录
        try:
            payments = db.query(PaymentRecord).filter(
                PaymentRecord.customer_id == customer_id
            ).order_by(desc(PaymentRecord.create_time)).limit(limit).all()
            
            for p in payments:
                transactions.append({
                    "id": f"payment_{p.id}",
                    "type": "payment",
                    "type_label": "收款",
                    "order_no": f"PY{p.id:06d}",
                    "description": f"收款 ¥{p.amount:.2f}",
                    "cash_amount": -(p.amount or 0),  # 负数表示减少欠款
                    "gold_amount": 0,
                    "status": "completed",
                    "created_at": p.create_time.isoformat() if p.create_time else None,
                    "operator": p.operator
                })
        except Exception as e:
            logger.warning(f"查询收款记录出错: {e}")
        
        # 4. 金料存取记录（正数=客户欠我们，负数=我们欠客户）
        try:
            gold_txs = db.query(CustomerGoldDepositTransaction).filter(
                CustomerGoldDepositTransaction.customer_id == customer_id
            ).order_by(desc(CustomerGoldDepositTransaction.created_at)).limit(limit).all()
            
            for tx in gold_txs:
                tx_type_label = {
                    "deposit": "存料",
                    "use": "用料",
                    "refund": "退料"
                }.get(tx.transaction_type, tx.transaction_type)
                
                amount = tx.amount or 0
                if tx.transaction_type == "deposit":
                    amount = -abs(amount)  # 存料=我们欠客户
                elif tx.transaction_type == "use":
                    amount = abs(amount)   # 用料=客户欠我们
                elif tx.transaction_type == "refund":
                    amount = -abs(amount)  # 退料=我们欠客户
                
                transactions.append({
                    "id": f"gold_deposit_{tx.id}",
                    "type": "gold_deposit",
                    "type_label": tx_type_label,
                    "order_no": f"GD{tx.id:06d}",
                    "description": tx.remark or f"金料{tx_type_label} {abs(tx.amount):.2f}克",
                    "cash_amount": 0,
                    "gold_amount": amount,
                    "status": tx.status,
                    "created_at": tx.created_at.isoformat() if tx.created_at else None,
                    "operator": tx.created_by
                })
        except Exception as e:
            logger.warning(f"查询金料交易记录出错: {e}")
        
        # 5. 客户往来账记录
        try:
            customer_txs = db.query(CustomerTransaction).filter(
                CustomerTransaction.customer_id == customer_id,
                CustomerTransaction.status == "active"
            ).order_by(desc(CustomerTransaction.created_at)).limit(limit).all()
            
            for tx in customer_txs:
                tx_type_label = {
                    "sales": "销售",
                    "settlement": "结算",
                    "gold_receipt": "收料",
                    "payment": "付款"
                }.get(tx.transaction_type, tx.transaction_type)
                
                transactions.append({
                    "id": f"tx_{tx.id}",
                    "type": "transaction",
                    "type_label": tx_type_label,
                    "order_no": f"TX{tx.id:06d}",
                    "description": tx.remark or f"往来账：{tx_type_label}",
                    "cash_amount": tx.amount or 0,
                    "gold_amount": tx.gold_weight or 0,
                    "gold_debt_before": tx.gold_due_before or 0,
                    "gold_debt_after": tx.gold_due_after or 0,
                    "status": tx.status,
                    "created_at": tx.created_at.isoformat() if tx.created_at else None,
                    "operator": None
                })
        except Exception as e:
            logger.warning(f"查询往来账记录出错: {e}")
        
        # 按时间排序（去重）
        seen_ids = set()
        unique_transactions = []
        for tx in transactions:
            if tx["id"] not in seen_ids:
                seen_ids.add(tx["id"])
                unique_transactions.append(tx)
        
        unique_transactions.sort(key=lambda x: x["created_at"] or "", reverse=True)
        
        # 获取当前欠款余额
        current_balance = {
            "cash_debt": 0.0,
            "gold_debt": 0.0,
            "gold_deposit": 0.0
        }
        
        try:
            receivables = db.query(AccountReceivable).filter(
                AccountReceivable.customer_id == customer_id,
                AccountReceivable.status.in_(["unpaid", "overdue"])
            ).all()
            ar_debt = sum(r.unpaid_amount or 0 for r in receivables)
            
            # 加入期初现金余额（数据迁移写入的 initial_balance 交易）
            initial_cash = float(db.query(
                func.coalesce(func.sum(CustomerTransaction.amount), 0)
            ).filter(
                CustomerTransaction.customer_id == customer_id,
                CustomerTransaction.transaction_type == 'initial_balance',
                CustomerTransaction.status == 'active'
            ).scalar() or 0)
            
            current_balance["cash_debt"] = ar_debt + initial_cash
        except (ValueError, TypeError):
            pass
        
        # 计算净金料（使用统一计算函数，保证全系统口径一致）
        try:
            from ..gold_balance import calculate_customer_net_gold
            net_gold = calculate_customer_net_gold(customer_id, db)  # 正数=存料，负数=欠料
            
            current_balance["net_gold"] = net_gold  # 净金料值（正=存料，负=欠料）
            current_balance["gold_debt"] = max(0, -net_gold)  # 兼容字段：欠料
            current_balance["gold_deposit"] = max(0, net_gold)  # 兼容字段：存料
        except Exception as e:
            logger.warning(f"计算客户金料余额失败: {e}")
            current_balance["net_gold"] = 0
            current_balance["gold_debt"] = 0
            current_balance["gold_deposit"] = 0
        
        return success_response(
            data={
                "customer": {
                    "id": customer.id,
                    "name": customer.name,
                    "phone": customer.phone,
                    "customer_no": customer.customer_no
                },
                "current_balance": current_balance,
                "transactions": unique_transactions[:limit]
            }
        )
        
    except Exception as e:
        logger.error(f"查询客户欠款历史失败: {e}", exc_info=True)
        return server_error_response(message=f"查询失败: {str(e)}")


# ============= 客户往来账明细表 API =============

@router.get("/{customer_id}/transactions-detail")
async def get_customer_transactions_detail(
    customer_id: int,
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    获取客户往来账明细表
    
    合并显示：销售结算、客户来料、客户来款
    按发生日期排序
    """
    from ..middleware.permissions import has_permission
    from ..models import SettlementOrder
    from ..models.finance import GoldReceipt, PaymentRecord, AccountReceivable
    
    # 权限检查
    can_view = (
        has_permission(user_role, 'can_view_customers') or 
        has_permission(user_role, 'can_create_settlement') or
        user_role == 'manager'
    )
    if not can_view:
        raise HTTPException(status_code=403, detail="权限不足")
    
    try:
        # 验证客户存在
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail="客户不存在")
        
        # 解析日期范围
        date_start = None
        date_end = None
        if start_date:
            try:
                date_start = datetime.strptime(start_date, "%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        if end_date:
            try:
                date_end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            except (ValueError, TypeError):
                pass
        
        transactions = []
        
        # ========== 1. 销售结算记录 ==========
        settlement_query = db.query(SettlementOrder).join(SalesOrder).filter(
            SalesOrder.customer_id == customer_id,
            SettlementOrder.status.in_(['confirmed', 'printed'])
        )
        if date_start:
            settlement_query = settlement_query.filter(SettlementOrder.create_time >= date_start)
        if date_end:
            settlement_query = settlement_query.filter(SettlementOrder.create_time <= date_end)
        
        settlements = settlement_query.all()
        
        for s in settlements:
            # 足金：结料时的应付金料重量
            gold_amount = 0.0
            if s.payment_method == 'physical_gold':
                gold_amount = s.physical_gold_weight or 0.0
            
            # 欠款金额：结价时的应收金额
            cash_amount = 0.0
            if s.payment_method == 'cash_price':
                cash_amount = s.total_amount or 0.0
            elif s.payment_method == 'mixed':
                cash_amount = s.cash_payment_weight or 0.0  # 混合支付的现金部分
            
            transactions.append({
                "id": f"settlement_{s.id}",
                "date": s.create_time.strftime("%Y-%m-%d") if s.create_time else None,
                "datetime": s.create_time.isoformat() if s.create_time else None,
                "type": "销售结算",
                "order_no": s.settlement_no,
                "gold_amount": round(gold_amount, 3),  # 正数=客户欠料
                "cash_amount": round(cash_amount, 2),  # 正数=客户欠款
                "remark": s.remark or ""
            })
        
        # ========== 2. 客户来料记录 ==========
        receipt_query = db.query(GoldReceipt).filter(
            GoldReceipt.customer_id == customer_id,
            GoldReceipt.status == 'received'
        )
        if date_start:
            receipt_query = receipt_query.filter(GoldReceipt.created_at >= date_start)
        if date_end:
            receipt_query = receipt_query.filter(GoldReceipt.created_at <= date_end)
        
        receipts = receipt_query.all()
        
        for r in receipts:
            transactions.append({
                "id": f"receipt_{r.id}",
                "date": r.received_at.strftime("%Y-%m-%d") if r.received_at else (r.created_at.strftime("%Y-%m-%d") if r.created_at else None),
                "datetime": r.received_at.isoformat() if r.received_at else (r.created_at.isoformat() if r.created_at else None),
                "type": "客户来料",
                "order_no": r.receipt_no,
                "gold_amount": round(-(r.gold_weight or 0), 3),  # 负数=客户给料
                "cash_amount": 0.0,
                "remark": f"{r.gold_fineness or ''} {r.remark or ''}".strip()
            })
        
        # ========== 3. 客户来款记录 ==========
        payment_query = db.query(PaymentRecord).filter(
            PaymentRecord.customer_id == customer_id
        )
        if date_start:
            payment_query = payment_query.filter(PaymentRecord.payment_date >= date_start.date())
        if date_end:
            payment_query = payment_query.filter(PaymentRecord.payment_date <= date_end.date())
        
        payments = payment_query.all()
        
        for p in payments:
            payment_method_label = {
                'bank_transfer': '银行转账',
                'cash': '现金',
                'wechat': '微信',
                'alipay': '支付宝',
                'card': '刷卡'
            }.get(p.payment_method, p.payment_method)
            
            transactions.append({
                "id": f"payment_{p.id}",
                "date": p.payment_date.strftime("%Y-%m-%d") if p.payment_date else None,
                "datetime": p.create_time.isoformat() if p.create_time else None,
                "type": "客户来款",
                "order_no": p.payment_no or f"SK{p.id}",
                "gold_amount": 0.0,
                "cash_amount": round(-(p.amount or 0), 2),  # 负数=客户付款
                "remark": f"{payment_method_label} {p.remark or ''}".strip()
            })
        
        # ========== 按日期排序 ==========
        transactions.sort(key=lambda x: x.get("datetime") or "0000-00-00")
        
        # ========== 计算合计 ==========
        total_gold = sum(t["gold_amount"] for t in transactions)
        total_cash = sum(t["cash_amount"] for t in transactions)
        
        # 添加序号
        for i, t in enumerate(transactions, 1):
            t["seq"] = i
        
        return {
            "success": True,
            "customer": {
                "id": customer.id,
                "name": customer.name,
                "customer_no": customer.customer_no
            },
            "date_range": {
                "start": start_date,
                "end": end_date
            },
            "transactions": transactions,
            "summary": {
                "total_gold": round(total_gold, 3),  # 正数=客户欠料，负数=客户存料
                "total_cash": round(total_cash, 2),  # 正数=客户欠款，负数=预收款
                "count": len(transactions)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取客户往来账明细失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取往来账明细失败: {str(e)}")


# ========== 金料账户诊断和修复 API ==========

@router.get("/{customer_id}/gold-account-diagnosis")
async def diagnose_gold_account(
    customer_id: int,
    db: Session = Depends(get_db)
):
    """
    诊断客户金料账户，计算正确的净值
    
    通过汇总所有金料交易记录，计算应该的净值：
    - 收料/存料 = 增加（正）
    - 结算欠料/提料 = 减少（负）
    """
    try:
        # 查找客户
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail=f"客户ID {customer_id} 不存在")
        
        # 当前账户余额
        deposit = db.query(CustomerGoldDeposit).filter(
            CustomerGoldDeposit.customer_id == customer_id
        ).first()
        current_balance = deposit.current_balance if deposit else 0.0
        
        # 汇总所有金料交易
        # 1. 从 CustomerGoldDepositTransaction 获取存料记录
        deposit_txs = db.query(CustomerGoldDepositTransaction).filter(
            CustomerGoldDepositTransaction.customer_id == customer_id,
            CustomerGoldDepositTransaction.status == "active"
        ).all()
        
        total_deposited = 0.0  # 总存入
        total_used = 0.0  # 总使用
        deposit_details = []
        
        for tx in deposit_txs:
            if tx.transaction_type == "deposit":
                total_deposited += float(tx.amount or 0)
            elif tx.transaction_type == "use":
                total_used += float(tx.amount or 0)
            
            deposit_details.append({
                "id": tx.id,
                "type": tx.transaction_type,
                "amount": tx.amount,
                "balance_after": tx.balance_after,
                "remark": tx.remark,
                "created_at": tx.created_at.isoformat() if tx.created_at else None
            })
        
        # 2. 从 CustomerTransaction 获取金料欠款记录
        gold_txs = db.query(CustomerTransaction).filter(
            CustomerTransaction.customer_id == customer_id,
            CustomerTransaction.status == "active"
        ).order_by(CustomerTransaction.created_at).all()
        
        latest_gold_due = 0.0
        gold_tx_details = []
        
        for tx in gold_txs:
            latest_gold_due = tx.gold_due_after or 0
            gold_tx_details.append({
                "id": tx.id,
                "type": tx.transaction_type,
                "gold_weight": tx.gold_weight,
                "gold_due_before": tx.gold_due_before,
                "gold_due_after": tx.gold_due_after,
                "remark": tx.remark,
                "created_at": tx.created_at.isoformat() if tx.created_at else None
            })
        
        # 计算正确的净值
        # 方法1：根据交易记录计算
        calculated_from_deposits = total_deposited - total_used
        
        # 方法2：根据最后一条存料交易的 balance_after
        last_deposit_balance = deposit_txs[-1].balance_after if deposit_txs else 0.0
        
        # 推荐的正确净值：使用最后一条存料交易的 balance_after
        # 如果没有存料交易，则使用 -latest_gold_due（欠料转为负值）
        if deposit_txs:
            recommended_balance = last_deposit_balance
        else:
            recommended_balance = -latest_gold_due  # 欠料转为负值
        
        return {
            "success": True,
            "customer": {
                "id": customer.id,
                "name": customer.name
            },
            "current_balance": current_balance,
            "diagnosis": {
                "total_deposited": total_deposited,
                "total_used": total_used,
                "calculated_net": calculated_from_deposits,
                "latest_gold_due": latest_gold_due,
                "last_deposit_balance": last_deposit_balance
            },
            "recommended_balance": recommended_balance,
            "status": "存料" if recommended_balance > 0 else ("欠料" if recommended_balance < 0 else "清账"),
            "deposit_transactions": deposit_details[-10:],  # 最近10条
            "gold_transactions": gold_tx_details[-10:]  # 最近10条
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"诊断客户金料账户失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"诊断失败: {str(e)}")


@router.post("/{customer_id}/fix-gold-account")
async def fix_customer_gold_account(
    customer_id: int,
    correct_balance: float = Query(..., description="正确的净金料值（正=存料，负=欠料）"),
    db: Session = Depends(get_db)
):
    """
    手动修复客户金料账户余额
    
    参数:
    - customer_id: 客户ID
    - correct_balance: 正确的净金料值（正数=存料，负数=欠料，0=清账）
    
    此API用于一次性修复数据，不会自动执行。
    """
    try:
        # 查找客户
        customer = db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise HTTPException(status_code=404, detail=f"客户ID {customer_id} 不存在")
        
        # 查找或创建金料账户记录
        deposit = db.query(CustomerGoldDeposit).filter(
            CustomerGoldDeposit.customer_id == customer_id
        ).first()
        
        old_balance = 0.0
        if deposit:
            old_balance = deposit.current_balance or 0.0
            deposit.current_balance = correct_balance
        else:
            # 创建新记录
            deposit = CustomerGoldDeposit(
                customer_id=customer_id,
                customer_name=customer.name,
                current_balance=correct_balance,
                total_deposited=max(0, correct_balance),
                total_used=max(0, -correct_balance)
            )
            db.add(deposit)
        
        db.commit()
        
        logger.info(f"[金料账户修复] 客户 {customer.name}(ID:{customer_id}): {old_balance:.2f}克 -> {correct_balance:.2f}克")
        
        return {
            "success": True,
            "message": f"客户 {customer.name} 的金料账户已修复",
            "customer_id": customer_id,
            "customer_name": customer.name,
            "old_balance": old_balance,
            "new_balance": correct_balance,
            "status": "存料" if correct_balance > 0 else ("欠料" if correct_balance < 0 else "清账")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"修复客户金料账户失败: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail=f"修复失败: {str(e)}")


@router.post("/recalculate-stats", summary="重新计算所有客户的累计购买统计")
async def recalculate_customer_stats(
    user_role: str = Query(default="sales", description="用户角色"),
    db: Session = Depends(get_db)
):
    """
    根据 sales_orders 实际数据重新计算每个客户的统计字段。
    修复迁移数据导致统计不一致，同时关联 customer_id 为空的销售单。
    """
    if user_role not in ["manager", "admin"]:
        raise HTTPException(status_code=403, detail="权限不足：仅管理层可执行此操作")

    try:
        orphan_orders = db.query(SalesOrder).filter(
            SalesOrder.customer_id.is_(None),
            SalesOrder.status != "cancelled"
        ).all()

        linked_count = 0
        for order in orphan_orders:
            if order.customer_name:
                customer = db.query(Customer).filter(
                    Customer.name == order.customer_name,
                    Customer.status == "active"
                ).first()
                if customer:
                    order.customer_id = customer.id
                    linked_count += 1

        stats = db.query(
            SalesOrder.customer_id,
            func.count(SalesOrder.id).label("order_count"),
            func.coalesce(func.sum(SalesOrder.total_labor_cost), 0).label("total_amount"),
            func.coalesce(func.sum(SalesOrder.total_weight), 0).label("total_weight"),
            func.max(SalesOrder.order_date).label("last_purchase")
        ).filter(
            SalesOrder.customer_id.isnot(None),
            SalesOrder.status != "cancelled"
        ).group_by(SalesOrder.customer_id).all()

        stats_map = {
            s.customer_id: {
                "count": s.order_count,
                "amount": float(s.total_amount or 0),
                "weight": float(s.total_weight or 0),
                "last": s.last_purchase
            }
            for s in stats
        }

        customers = db.query(Customer).filter(Customer.status == "active").all()
        updated_count = 0
        for c in customers:
            s = stats_map.get(c.id)
            if s:
                c.total_purchase_amount = round(s["amount"], 2)
                c.total_purchase_count = s["count"]
                c.total_purchase_weight = round(s["weight"], 3)
                c.last_purchase_time = s["last"]
            else:
                c.total_purchase_amount = 0.0
                c.total_purchase_count = 0
                c.total_purchase_weight = 0.0
                c.last_purchase_time = None
            updated_count += 1

        db.commit()
        logger.info(f"客户统计重算完成: 更新{updated_count}个客户, 关联{linked_count}条无主销售单")

        return success_response(
            message=f"统计重算完成：更新{updated_count}个客户，关联{linked_count}条无主销售单",
            data={
                "updated_customers": updated_count,
                "linked_orphan_orders": linked_count,
                "customers_with_sales": len(stats_map)
            }
        )
    except Exception as e:
        db.rollback()
        logger.error(f"重算客户统计失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"重算失败: {str(e)}")


@router.post("/merge-duplicates", summary="合并重复客户（按姓名）")
async def merge_duplicate_customers(
    user_role: str = Query(default="sales", description="用户角色"),
    dry_run: bool = Query(default=True, description="预览模式（不实际执行）"),
    db: Session = Depends(get_db)
):
    """
    查找并合并同名客户。保留最早创建的，将关联数据迁移过去。
    dry_run=true 时只预览不执行。
    """
    if user_role not in ["manager", "admin"]:
        raise HTTPException(status_code=403, detail="权限不足：仅管理层可执行此操作")

    try:
        dup_names = db.query(Customer.name, func.count(Customer.id).label("cnt")).filter(
            Customer.status == "active"
        ).group_by(Customer.name).having(func.count(Customer.id) > 1).all()

        if not dup_names:
            return success_response(message="没有发现重复客户", data={"duplicates": []})

        merge_results = []
        total_merged = 0

        for dup in dup_names:
            name = dup.name
            duplicates = db.query(Customer).filter(
                Customer.name == name,
                Customer.status == "active"
            ).order_by(Customer.create_time.asc()).all()

            keep = duplicates[0]
            to_merge = duplicates[1:]

            merge_info = {
                "name": name,
                "keep_id": keep.id,
                "keep_no": keep.customer_no,
                "merge_ids": [c.id for c in to_merge],
                "merge_nos": [c.customer_no for c in to_merge]
            }

            if not dry_run:
                for dup_customer in to_merge:
                    db.query(SalesOrder).filter(
                        SalesOrder.customer_id == dup_customer.id
                    ).update({"customer_id": keep.id}, synchronize_session=False)
                    db.query(CustomerTransaction).filter(
                        CustomerTransaction.customer_id == dup_customer.id
                    ).update({"customer_id": keep.id}, synchronize_session=False)
                    db.query(CustomerGoldDeposit).filter(
                        CustomerGoldDeposit.customer_id == dup_customer.id
                    ).update({"customer_id": keep.id}, synchronize_session=False)
                    db.query(CustomerGoldDepositTransaction).filter(
                        CustomerGoldDepositTransaction.customer_id == dup_customer.id
                    ).update({"customer_id": keep.id}, synchronize_session=False)
                    db.query(CustomerWithdrawal).filter(
                        CustomerWithdrawal.customer_id == dup_customer.id
                    ).update({"customer_id": keep.id}, synchronize_session=False)

                    dup_customer.status = "merged"
                    total_merged += 1

            merge_results.append(merge_info)

        if not dry_run:
            db.commit()
            logger.info(f"合并重复客户完成: {len(merge_results)}组, 合并{total_merged}个重复记录")

        return success_response(
            message=f"{'预览' if dry_run else '执行'}完成：发现{len(merge_results)}组重复客户" +
                    (f"，已合并{total_merged}个" if not dry_run else ""),
            data={
                "dry_run": dry_run,
                "duplicate_groups": len(merge_results),
                "total_to_merge": sum(len(r["merge_ids"]) for r in merge_results),
                "details": merge_results
            }
        )
    except Exception as e:
        db.rollback()
        logger.error(f"合并重复客户失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"合并失败: {str(e)}")
